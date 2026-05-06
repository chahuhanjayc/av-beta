"""TDS filing pack generation and export helpers."""

import csv
import json
from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO

from django.db import transaction
from django.urls import reverse
from django.utils import timezone

from .models import TDSFilingPack, TDSReturnWorkpaper
from .workbench import build_tds_return_workbench, financial_year_label, parse_workbench_filters


ZERO = Decimal("0.00")


def build_tds_filing_pack(company, fy_start, quarter, form_type):
    workbench = build_tds_return_workbench(company, fy_start, quarter, form_type)
    workpaper = workbench["workpaper"]
    pack_record = TDSFilingPack.objects.filter(
        company=company,
        form_type=form_type,
        financial_year_start=fy_start,
        quarter=quarter,
    ).select_related("workpaper", "generated_by", "filed_by").first()

    export_data = _export_data(company, workbench)
    validations = list(workbench["validations"]) + _pack_validations(company, workbench, export_data)
    critical_count = sum(1 for item in validations if item["severity"] == "critical")
    warning_count = sum(1 for item in validations if item["severity"] == "warning")
    workpaper_ready = bool(workpaper and workpaper.status in {
        TDSReturnWorkpaper.STATUS_READY_FOR_REVIEW,
        TDSReturnWorkpaper.STATUS_FILED,
    })
    can_generate = workpaper_ready and critical_count == 0

    if pack_record and pack_record.is_filed:
        status = "Filed"
    elif can_generate:
        status = "Ready for filing pack"
    elif not workpaper:
        status = "Workpaper not saved"
    elif not workpaper_ready:
        status = "Workpaper review pending"
    else:
        status = "Blocked by validations"

    return {
        "company": company,
        "filters": {
            "fy_start": fy_start,
            "fy_label": financial_year_label(fy_start),
            "quarter": quarter,
            "form_type": form_type,
            "period_start": workbench["summary"]["period_start"],
            "period_end": workbench["summary"]["period_end"],
            "due_date": workbench["summary"]["due_date"],
        },
        "workpaper": workpaper,
        "pack_record": pack_record,
        "status": status,
        "can_generate": can_generate,
        "critical_count": critical_count,
        "warning_count": warning_count,
        "validations": validations,
        "workbench": workbench,
        "export_data": export_data,
        "summary_snapshot": _summary_snapshot(workbench, export_data, critical_count, warning_count),
        "validation_snapshot": {
            "critical_count": critical_count,
            "warning_count": warning_count,
            "validations": validations,
            "generated_at": timezone.now().isoformat(),
        },
    }


def build_tds_filing_pack_from_params(company, params):
    filters = parse_workbench_filters(params)
    return build_tds_filing_pack(
        company=company,
        fy_start=filters["fy_start"],
        quarter=filters["quarter"],
        form_type=filters["form_type"],
    )


@transaction.atomic
def save_tds_filing_pack(pack, user, notes=""):
    if not pack["workpaper"]:
        raise ValueError("Save and mark the TDS workpaper ready before generating the filing pack.")
    if pack["workpaper"].status not in {TDSReturnWorkpaper.STATUS_READY_FOR_REVIEW, TDSReturnWorkpaper.STATUS_FILED}:
        raise ValueError("Mark the TDS workpaper ready for review before generating the filing pack.")
    if pack["critical_count"]:
        raise ValueError("Resolve critical TDS filing pack validations before generating the pack.")

    filters = pack["filters"]
    record, _ = TDSFilingPack.objects.get_or_create(
        company=pack["company"],
        form_type=filters["form_type"],
        financial_year_start=filters["fy_start"],
        quarter=filters["quarter"],
        defaults={
            "workpaper": pack["workpaper"],
            "period_start": filters["period_start"],
            "period_end": filters["period_end"],
            "due_date": filters["due_date"],
            "generated_by": user,
        },
    )
    record.workpaper = pack["workpaper"]
    record.period_start = filters["period_start"]
    record.period_end = filters["period_end"]
    record.due_date = filters["due_date"]
    record.status = TDSFilingPack.STATUS_READY
    record.summary_snapshot = pack["summary_snapshot"]
    record.validation_snapshot = pack["validation_snapshot"]
    record.export_snapshot = pack["export_data"]
    record.notes = notes.strip()
    if not record.generated_by_id:
        record.generated_by = user
    record.save()
    return record


@transaction.atomic
def mark_tds_filing_pack_filed(pack_record, user, ack_number, notes=""):
    ack_number = (ack_number or "").strip()
    if not ack_number:
        raise ValueError("Acknowledgement number is required.")
    pack_record.status = TDSFilingPack.STATUS_FILED
    pack_record.ack_number = ack_number
    pack_record.filed_by = user
    pack_record.filed_at = timezone.now()
    if notes.strip():
        pack_record.notes = notes.strip()
    pack_record.save()
    if pack_record.workpaper_id:
        pack_record.workpaper.status = TDSReturnWorkpaper.STATUS_FILED
        pack_record.workpaper.ack_number = ack_number
        pack_record.workpaper.filed_by = user
        pack_record.workpaper.filed_at = pack_record.filed_at
        pack_record.workpaper.save(update_fields=["status", "ack_number", "filed_by", "filed_at", "updated_at"])
    return pack_record


def reopen_tds_filing_pack(pack_record):
    pack_record.status = TDSFilingPack.STATUS_REOPENED
    pack_record.ack_number = ""
    pack_record.filed_by = None
    pack_record.filed_at = None
    pack_record.save()
    return pack_record


def tds_pack_xlsx_bytes(pack):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True)

    ws = wb.active
    ws.title = "Summary"
    filters = pack["filters"]
    summary = pack["workbench"]["summary"]
    ws["A1"] = "TDS Filing Pack"
    ws["A1"].font = title_font
    ws["A2"] = pack["company"].name
    ws["A3"] = f"{filters['form_type']} {filters['quarter']} FY {filters['fy_label']}"
    summary_rows = [
        ("Status", pack["status"]),
        ("TAN", summary["company_tan"]),
        ("Period", f"{filters['period_start']:%d %b %Y} to {filters['period_end']:%d %b %Y}"),
        ("Due Date", filters["due_date"]),
        ("Deductee Rows", len(pack["export_data"]["deductee_rows"])),
        ("Challan Rows", len(pack["export_data"]["challan_rows"])),
        ("Total TDS", summary["total_tds"]),
        ("Critical Validations", pack["critical_count"]),
        ("Warnings", pack["warning_count"]),
    ]
    for idx, (label, value) in enumerate(summary_rows, start=5):
        ws.cell(row=idx, column=1, value=label).font = label_font
        ws.cell(row=idx, column=2, value=_excel_value(value))
        ws.cell(row=idx, column=1).border = border
        ws.cell(row=idx, column=2).border = border
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 34

    _write_table(
        wb.create_sheet("Pre-FVU Checks"),
        ["Severity", "Check", "Count", "Result"],
        [[item["severity"], item["title"], item["count"], item["description"]] for item in pack["validations"]],
        header_fill,
        header_font,
        border,
    )
    _write_table(
        wb.create_sheet("RPU Deductee"),
        _deductee_headers(),
        [[row.get(header, "") for header in _deductee_headers()] for row in pack["export_data"]["deductee_rows"]],
        header_fill,
        header_font,
        border,
    )
    _write_table(
        wb.create_sheet("RPU Challan"),
        _challan_headers(),
        [[row.get(header, "") for header in _challan_headers()] for row in pack["export_data"]["challan_rows"]],
        header_fill,
        header_font,
        border,
    )
    _write_table(
        wb.create_sheet("Sections"),
        ["Section", "Description", "Entries", "Base Amount", "TDS Amount"],
        [
            [row["section"], row["description"], row["entry_count"], row["base_amount"], row["tds_amount"]]
            for row in pack["workbench"]["sections"]
        ],
        header_fill,
        header_font,
        border,
    )
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def deductee_csv_bytes(pack):
    return _csv_bytes(_deductee_headers(), pack["export_data"]["deductee_rows"])


def challan_csv_bytes(pack):
    return _csv_bytes(_challan_headers(), pack["export_data"]["challan_rows"])


def draft_json_bytes(pack):
    payload = {
        "company": {
            "name": pack["company"].name,
            "tan": pack["workbench"]["summary"]["company_tan"],
            "responsible_person": pack["export_data"]["deductor"]["Responsible Person"],
            "responsible_designation": pack["export_data"]["deductor"]["Responsible Designation"],
        },
        "return": pack["filters"],
        "summary": pack["summary_snapshot"],
        "validations": pack["validation_snapshot"],
        "rpu": pack["export_data"],
    }
    return json.dumps(payload, indent=2, ensure_ascii=False, default=_json_default).encode("utf-8")


def _export_data(company, workbench):
    challan_serials = {}
    challan_rows = []
    for idx, row in enumerate(workbench["challans"], start=1):
        challan_serials[(row["bsr_code"], row["challan_number"], row["deposit_date"])] = idx
        challan_rows.append({
            "Challan Serial": idx,
            "challan_serial": idx,
            "BSR Code": row["bsr_code"],
            "bsr_code": row["bsr_code"],
            "Deposit Date": row["deposit_date"],
            "deposit_date": row["deposit_date"],
            "Challan Number": row["challan_number"],
            "challan_number": row["challan_number"],
            "TDS Amount": row["tds_amount"],
            "tds_amount": row["tds_amount"],
            "Entry Count": row["entry_count"],
            "entry_count": row["entry_count"],
            "Issue Flag": "YES" if row["has_issue"] else "NO",
            "issue_flag": "YES" if row["has_issue"] else "NO",
        })

    deductee_rows = []
    for idx, row in enumerate(workbench["rows"], start=1):
        key = (row["bsr_code"] or "Missing BSR", row["challan_number"] or "Missing Challan", row["deposit_date"])
        voucher = row["entry"].voucher
        deductee_rows.append({
            "Deductee Serial": idx,
            "deductee_serial": idx,
            "Challan Serial": challan_serials.get(key, ""),
            "challan_serial": challan_serials.get(key, ""),
            "Section": row["section"],
            "section": row["section"],
            "Deductee Name": row["party"],
            "deductee_name": row["party"],
            "Deductee PAN": row["pan"],
            "deductee_pan": row["pan"],
            "Deductee Type": row["entry"].deductee_type,
            "deductee_type": row["entry"].deductee_type,
            "Payment Date": row["date"],
            "payment_date": row["date"],
            "Amount Paid": row["base_amount"],
            "amount_paid": row["base_amount"],
            "TDS Rate": row["rate"],
            "tds_rate": row["rate"],
            "TDS Amount": row["tds_amount"],
            "tds_amount": row["tds_amount"],
            "Deposit Date": row["deposit_date"] or "",
            "deposit_date": row["deposit_date"] or "",
            "BSR Code": row["bsr_code"],
            "bsr_code": row["bsr_code"],
            "Challan Number": row["challan_number"],
            "challan_number": row["challan_number"],
            "Voucher Number": getattr(voucher, "number", "") if voucher else "",
            "voucher_number": getattr(voucher, "number", "") if voucher else "",
            "Issue Flags": ", ".join(row["issues"]) if row["issues"] else "",
            "issue_flags": ", ".join(row["issues"]) if row["issues"] else "",
        })

    filters = {
        "Form Type": workbench["summary"]["form_type"],
        "Financial Year": workbench["summary"]["fy_label"],
        "Quarter": workbench["summary"]["quarter"],
        "Period Start": workbench["summary"]["period_start"],
        "Period End": workbench["summary"]["period_end"],
    }
    deductor = {
        "Company": company.name,
        "TAN": workbench["summary"]["company_tan"],
        "Responsible Person": company.tds_responsible_person or "",
        "Responsible Designation": company.tds_responsible_designation or "",
    }
    return {
        "deductor": _json_row(deductor),
        "return": _json_row(filters),
        "challan_rows": [_json_row(row) for row in challan_rows],
        "deductee_rows": [_json_row(row) for row in deductee_rows],
    }


def _pack_validations(company, workbench, export_data):
    validations = []

    def add(severity, code, title, count, description, action_url=""):
        validations.append({
            "severity": severity,
            "code": code,
            "title": title,
            "count": count,
            "description": description,
            "action_url": action_url,
        })

    rows = workbench["rows"]
    non_positive = [
        row for row in rows
        if (row["base_amount"] or ZERO) <= ZERO or (row["tds_amount"] or ZERO) <= ZERO
    ]
    add(
        "ok" if not non_positive else "critical",
        "positive_amounts",
        "RPU Amount Controls",
        len(non_positive),
        "RPU rows have positive paid and TDS amounts." if not non_positive else "Some RPU rows have zero or negative paid/TDS amounts.",
        reverse("tds:entry_list"),
    )

    unmapped = [row for row in export_data["deductee_rows"] if row["BSR Code"] and not row["Challan Serial"]]
    add(
        "ok" if not unmapped else "critical",
        "challan_serial_map",
        "Challan Serial Mapping",
        len(unmapped),
        "Every deposited deductee row maps to a challan serial." if not unmapped else "Some deposited rows could not be mapped to a challan serial.",
    )

    missing_responsible = 0 if company.tds_responsible_person and company.tds_responsible_designation else 1
    add(
        "ok" if not missing_responsible else "warning",
        "responsible_person",
        "Responsible Person",
        missing_responsible,
        "Responsible person and designation are configured." if not missing_responsible else "Responsible person or designation is missing in Company Settings.",
        reverse("core:company_settings"),
    )

    duplicate_keys = {}
    for row in export_data["deductee_rows"]:
        key = (
            row["Deductee PAN"],
            row["Payment Date"],
            row["Section"],
            row["Amount Paid"],
            row["TDS Amount"],
        )
        duplicate_keys[key] = duplicate_keys.get(key, 0) + 1
    duplicate_count = sum(count - 1 for count in duplicate_keys.values() if count > 1)
    add(
        "ok" if duplicate_count == 0 else "warning",
        "duplicate_deductee_rows",
        "Duplicate Deductee Rows",
        duplicate_count,
        "No likely duplicate deductee rows detected." if duplicate_count == 0 else "Likely duplicate deductee rows detected; review before RPU entry.",
        reverse("tds:entry_list"),
    )

    workpaper = workbench["workpaper"]
    workpaper_ready = bool(workpaper and workpaper.status in {
        TDSReturnWorkpaper.STATUS_READY_FOR_REVIEW,
        TDSReturnWorkpaper.STATUS_FILED,
    })
    add(
        "ok" if workpaper_ready else "critical",
        "workpaper_approval",
        "Workpaper Approval",
        0 if workpaper_ready else 1,
        "TDS workpaper is ready for review or filed." if workpaper_ready else "Mark the TDS workpaper ready for review before generating final exports.",
        reverse("tds:return_workbench"),
    )

    if workbench["summary"]["form_type"] == TDSReturnWorkpaper.FORM_24Q and workbench["summary"]["quarter"] == TDSReturnWorkpaper.Q4:
        add(
            "warning",
            "salary_annexure",
            "Salary Annexure",
            1,
            "Q4 Form 24Q salary annexure details are not modelled yet; verify salary schedule outside this Phase 1 pack.",
        )

    return validations


def _summary_snapshot(workbench, export_data, critical_count, warning_count):
    summary = dict(workbench["summary_snapshot"])
    summary.update({
        "deductee_export_rows": len(export_data["deductee_rows"]),
        "challan_export_rows": len(export_data["challan_rows"]),
        "critical_count": critical_count,
        "warning_count": warning_count,
    })
    return summary


def _write_table(ws, headers, rows, header_fill, header_font, border):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_excel_value(value))
            cell.border = border
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(len(str(header)) + 4, 14), 32)
    ws.freeze_panes = "A2"


def _deductee_headers():
    return [
        "Deductee Serial",
        "Challan Serial",
        "Section",
        "Deductee Name",
        "Deductee PAN",
        "Deductee Type",
        "Payment Date",
        "Amount Paid",
        "TDS Rate",
        "TDS Amount",
        "Deposit Date",
        "BSR Code",
        "Challan Number",
        "Voucher Number",
        "Issue Flags",
    ]


def _challan_headers():
    return [
        "Challan Serial",
        "BSR Code",
        "Deposit Date",
        "Challan Number",
        "TDS Amount",
        "Entry Count",
        "Issue Flag",
    ]


def _csv_bytes(headers, rows):
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({header: _csv_value(row.get(header, "")) for header in headers})
    return output.getvalue().encode("utf-8-sig")


def _json_row(row):
    return {key: _json_default(value) for key, value in row.items()}


def _excel_value(value):
    if isinstance(value, Decimal):
        return float(value)
    return value


def _csv_value(value):
    if isinstance(value, Decimal):
        return str(value.quantize(Decimal("0.01")))
    if isinstance(value, date):
        return value.isoformat()
    return value


def _json_default(value):
    if isinstance(value, Decimal):
        return str(value.quantize(Decimal("0.01")))
    if isinstance(value, date):
        return value.isoformat()
    return value
