"""
core/views.py
Company selection, dashboard, company switch, and company settings.
"""

import json
import logging
import calendar
import csv
import io
import secrets
import zipfile
from datetime import date as _date, timedelta
from decimal import Decimal
from email.utils import formataddr
from pathlib import Path, PurePosixPath
from urllib.parse import quote, urlencode

from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.mail import EmailMessage
from django.db import connection
from django.db.models import Q, Sum
from django.views.decorators.http import require_GET, require_POST
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.exceptions import SuspiciousFileOperation
from django.http import FileResponse, HttpResponse, HttpResponseForbidden, Http404, JsonResponse
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils._os import safe_join
from inventory.models import StockItem
from ledger.models import Ledger
from vouchers.models import Voucher
from .models import ClientEngagement, Company, CompanyStatutoryProfile, StatutoryRuleOverride, UserCompanyAccess, AuditLog, ComplianceFiling, ComplianceNotice, FilingReview, GSTEvidenceDocument, GSTFilingPack, GSTPeriodReview, GSTPostFilingTracker, MarketProofCaseStudy, MarketProofExternalEvidence, PilotFeedback, PracticeTask
from .decorators import admin_required, write_required
from .forms import (
    AppSettingsForm,
    ClientEngagementForm,
    CompanyStatutoryProfileForm,
    CompanySettingsForm,
    ComplianceCalendarGenerationForm,
    ComplianceFilingForm,
    ComplianceNoticeForm,
    MarketProofCaseStudyForm,
    MarketProofExternalEvidenceForm,
    PilotFeedbackForm,
    PracticeTaskForm,
    StatutoryRuleOverrideForm,
)
from .compliance_calendar import generate_compliance_calendar
from .compliance_autopilot import normalize_autopilot_months, run_compliance_autopilot
from .demo_workspace import demo_workspace_snapshot, seed_demo_workspace
from .approval_inbox import approval_inbox_csv_response, build_ca_approval_inbox
from .close_workbench import build_close_workbench, create_close_tasks
from .statutory_exposure import (
    build_statutory_exposure,
    create_statutory_exposure_tasks,
    statutory_exposure_csv_response,
)
from .operating_readiness import (
    build_operating_readiness,
    create_operating_readiness_tasks,
    operating_readiness_csv_response,
)
from .pilot_launch import (
    build_pilot_launch_control,
    create_pilot_launch_tasks,
    pilot_launch_csv_response,
)
from .client_success import (
    build_client_success_cockpit,
    client_success_csv_response,
    create_client_success_tasks,
)
from .client_portal_health import (
    build_client_portal_health,
    client_portal_health_csv_response,
    create_client_portal_health_tasks,
)
from .pilot_adoption import (
    build_pilot_adoption_evidence,
    create_pilot_adoption_tasks,
    pilot_adoption_csv_response,
)
from .pilot_feedback import (
    build_pilot_feedback_register,
    create_pilot_feedback_follow_up,
    pilot_feedback_csv_response,
    pilot_feedback_filter_query,
    reopen_pilot_feedback,
    resolve_pilot_feedback,
)
from .market_proof import (
    build_market_proof_evidence_pack,
    build_market_proof_pack,
    create_market_proof_tasks,
    market_proof_csv_response,
    market_proof_evidence_pack_response,
)
from .market_case_studies import (
    approve_case_study,
    build_market_case_study_register,
    create_case_study_follow_up,
    market_case_study_csv_response,
    market_case_study_filter_query,
    publish_case_study,
)
from .market_external_evidence import (
    build_market_external_evidence_register,
    create_external_evidence_follow_up,
    market_external_evidence_csv_response,
    market_external_evidence_filter_query,
    reject_external_evidence,
    reopen_external_evidence,
    verify_external_evidence,
)
from .operations_monitor import build_operations_monitor, create_operations_monitor_tasks
from .go_live_certificate import (
    build_go_live_certificate,
    create_go_live_remediation_tasks,
    go_live_certificate_payload,
)
from .go_live_evidence_pack import (
    build_go_live_evidence_pack,
    go_live_evidence_pack_bytes,
    go_live_evidence_pack_filename,
)
from .production_trust import (
    build_production_trust_context,
    create_backup_policy_tasks,
    create_scheduled_backup_tasks,
    record_restore_drill,
    run_operational_backup,
    run_scheduled_backup,
    verify_backup_restore_rehearsal,
)
from .security_control import build_security_control, create_security_control_tasks
from .system_observability import (
    build_system_observability,
    create_system_observability_tasks,
    observability_public_payload,
)
from .statutory_exports import (
    build_statutory_export_center,
    create_statutory_export_tasks,
    parse_gst_export_period,
    parse_tds_export_filters,
    statutory_export_csv_response,
)
from .collections_center import (
    build_collections_center,
    collections_csv_response,
    create_collection_tasks as create_collection_center_tasks,
    risk_filter_choices as collection_risk_filter_choices,
    send_collection_emails,
)
from .bank_reco_autopilot import (
    bank_reco_autopilot_csv_response,
    bank_reco_working_paper_zip_response,
    bank_reco_focus_choices,
    build_bank_reco_autopilot,
    close_bank_reco_uploaded_evidence,
    create_bank_reco_client_requests,
    create_bank_reco_tasks,
    post_bank_reco_auto_ready_vouchers,
)
from .filing_readiness import (
    build_filing_readiness,
    create_filing_readiness_tasks,
    save_filing_readiness_review,
)
from .review_center import (
    approve_review,
    build_filing_review,
    build_filing_review_rows,
    create_review_blocker_tasks,
    mark_reviewed,
    normalise_review_type,
    reopen_review,
    review_type_choices,
    send_back_review,
    start_review,
    unwaive_blocker,
    waive_blocker,
)
from .filing_pack import (
    GSTIN_RE,
    build_gst_filing_pack,
    draft_json_bytes,
    filing_pack_xlsx_bytes,
    mark_gst_filing_pack_filed,
    portal_gstr1_json_bytes,
    save_gst_filing_pack,
)
from .gst_post_filing import (
    build_gst_post_filing_center,
    build_gst_post_filing_dashboard,
    create_gst_notice_from_post_filing,
    save_gst_post_filing_tracker,
    update_gst_notice_from_post_filing,
    upload_gst_evidence,
)
from .compliance_workflow import set_filing_status, set_notice_status, sync_task_for_filing, sync_task_for_notice
from reports.utils import get_monthly_cash_flow
from .utils.chatbot_logic import get_bot_reply
from django.core.paginator import Paginator

import difflib
from django.db.models import Q

logger = logging.getLogger(__name__)


@require_GET
def healthz(request):
    """Lightweight deployment health check with a database round trip."""
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            cursor.fetchone()
    except Exception as exc:
        return JsonResponse({"ok": False, "database": "error", "error": str(exc)}, status=503)

    return JsonResponse({"ok": True, "database": "ok"})

def _rank_and_fuzzy(qs, query, field_name='name'):
    """
    Rank results by:
    1. Exact match (100)
    2. Starts with (50)
    3. Contains (20)
    4. Fuzzy match score (0-100)
    """
    results = []
    q_lower = query.lower()
    
    # Base filter already applied in queryset
    for obj in qs:
        val = getattr(obj, field_name)
        val_lower = val.lower()
        score = 0
        
        if val_lower == q_lower:
            score = 100
        elif val_lower.startswith(q_lower):
            score = 50
        elif q_lower in val_lower:
            score = 20
        else:
            # Fuzzy match using difflib
            similarity = difflib.SequenceMatcher(None, q_lower, val_lower).ratio()
            if similarity > 0.6:  # Threshold for typos
                score = int(similarity * 10) # Lower priority than direct matches
        
        if score > 0:
            results.append((score, obj))
            
    # Sort by score DESC
    results.sort(key=lambda x: x[0], reverse=True)
    return [r[1] for r in results[:5]]


def _navigation_entry(name, url_name, category, *, aliases=(), default=False):
    return {
        "name": name,
        "url": reverse(url_name),
        "category": category,
        "aliases": aliases,
        "default": default,
    }


def _navigation_index():
    return [
        _navigation_entry("Dashboard", "core:dashboard", "Dashboard", aliases=("home",), default=True),
        _navigation_entry("Operations Monitor", "core:operations_monitor", "Operations", aliases=("incident center", "production incidents", "ops monitor", "risk monitor"), default=True),
        _navigation_entry("System Observability", "core:system_observability", "Operations", aliases=("diagnostics", "system health", "deep health", "runtime health", "observability")),
        _navigation_entry("Go-Live Certificate", "core:go_live_certificate", "Operations", aliases=("go live", "deployment certificate", "production certificate", "launch readiness", "readiness certificate")),
        _navigation_entry("Production Trust", "core:production_trust_center", "Operations", aliases=("backup", "preflight", "restore", "trust center", "production checks")),
        _navigation_entry("Market Proof Pack", "core:market_proof_pack", "Operations", aliases=("market proof", "product depth", "external proof", "100 percent", "no 1", "pilot proof", "certification proof")),
        _navigation_entry("Market Proof Evidence Pack", "core:market_proof_evidence_pack", "Operations", aliases=("proof dossier", "evidence dossier", "market evidence export", "signed product proof", "market proof json")),
        _navigation_entry("External Evidence Register", "core:market_external_evidence", "Operations", aliases=("external evidence", "live proof", "remaining proof", "provider credential proof", "real client proof", "100 percent proof")),
        _navigation_entry("Security Control", "core:security_control", "Operations", aliases=("access review", "user access", "mfa", "roles", "security center"), default=True),
        _navigation_entry("CA Command Center", "core:ca_command_center", "CA Practice", aliases=("practice dashboard", "command"), default=True),
        _navigation_entry("CA Approval Inbox", "core:ca_approval_inbox", "CA Practice", aliases=("approval inbox", "ca approvals", "maker checker", "blocked imports", "gst approvals")),
        _navigation_entry("Statutory Exposure", "core:statutory_exposure", "CA Practice", aliases=("penalty engine", "deadline exposure", "late fee", "statutory risk", "due date exposure")),
        _navigation_entry("Statutory Export Center", "core:statutory_export_center", "CA Practice", aliases=("export center", "statutory exports", "filing exports", "gst tds exports", "rpu exports", "gstr json")),
        _navigation_entry("Operating Readiness", "core:client_operating_readiness", "CA Practice", aliases=("client readiness", "tally exit", "erp exit", "operating readiness", "go live")),
        _navigation_entry("Pilot Launch Control", "core:client_pilot_launch", "CA Practice", aliases=("pilot launch", "client launch", "rollout", "implementation", "go live clients")),
        _navigation_entry("Client Success Cockpit", "core:client_success_cockpit", "CA Practice", aliases=("client success", "adoption", "renewal risk", "customer success", "health score")),
        _navigation_entry("Client Portal Health", "core:client_portal_health", "CA Practice", aliases=("portal health", "client portal", "portal adoption", "upload links", "client reachability", "whatsapp clients", "email clients")),
        _navigation_entry("Pilot Adoption Evidence", "core:pilot_adoption_evidence", "CA Practice", aliases=("pilot evidence", "pilot adoption", "usage evidence", "client pilot proof", "product market fit", "pilot feedback")),
        _navigation_entry("Pilot Feedback Register", "core:pilot_feedback_register", "CA Practice", aliases=("pilot feedback", "client feedback", "pilot objections", "tally objections", "product feedback", "nps", "confidence score")),
        _navigation_entry("Market Case Studies", "core:market_case_studies", "CA Practice", aliases=("case studies", "testimonial", "market story", "client proof story", "published proof", "tally case study")),
        _navigation_entry("Partner Review Cockpit", "core:partner_review_cockpit", "CA Practice", aliases=("partner review", "sign off", "signoff", "approval cockpit", "partner cockpit")),
        _navigation_entry("Client Engagements", "core:client_engagements", "CA Practice", aliases=("retainers", "engagement", "scope", "renewals", "pricing")),
        _navigation_entry("Client Profitability", "core:ca_client_profitability", "CA Practice", aliases=("profitability", "workload", "pricing", "client control")),
        _navigation_entry("Filing Review Center", "core:filing_review_center", "CA Practice", aliases=("review center", "ca sign off", "partner signoff", "approve filing", "ready to file"), default=True),
        _navigation_entry("GST Filing Pack", "core:gst_filing_pack", "CA Practice", aliases=("filing pack", "gstr pack", "gst return pack", "gstr1 pack", "gstr3b pack"), default=True),
        _navigation_entry("GST Post-Filing Center", "core:gst_post_filing", "CA Practice", aliases=("gst arn", "gst post filing", "gst notice room", "drc 03", "challan", "gst evidence"), default=True),
        _navigation_entry("GST Post-Filing Dashboard", "core:gst_post_filing_dashboard", "CA Practice", aliases=("multi client gst", "gst closure dashboard", "gst evidence dashboard", "partner gst status")),
        _navigation_entry("Filing Readiness", "core:filing_readiness", "CA Practice", aliases=("filing room", "return readiness", "sign off", "signoff"), default=True),
        _navigation_entry("Client Requests", "portal:client_requests", "CA Practice", aliases=("document chase", "client documents", "request room"), default=True),
        _navigation_entry("New Client Request", "portal:client_request_create", "CA Practice", aliases=("create request", "request upload link", "new document request"), default=True),
        _navigation_entry("Client Request Reminders", "portal:client_request_reminders", "CA Practice", aliases=("reminders", "client reminders", "request reminders", "follow up", "follow-up"), default=True),
        _navigation_entry("Compliance Calendar", "core:compliance_calendar", "CA Practice", aliases=("due dates", "filing calendar")),
        _navigation_entry("GST Workbench", "core:gst_workbench", "CA Practice", aliases=("gst review", "gstr", "2b", "ims"), default=True),
        _navigation_entry("Practice Work Queue", "core:practice_tasks", "CA Practice", aliases=("tasks", "work queue", "practice tasks"), default=True),
        _navigation_entry("Compliance Filings", "core:compliance_filings", "CA Practice", aliases=("filings", "returns")),
        _navigation_entry("Compliance Notices", "core:compliance_notices", "CA Practice", aliases=("notices", "notice management")),
        _navigation_entry("Integration Control", "integrations:statutory_control", "CA Practice", aliases=("statutory integrations", "gst api status", "traces status", "bank feed status", "connector control"), default=True),
        _navigation_entry("Bank Feed Import", "integrations:bank_feed_import", "CA Practice", aliases=("connected banking", "bank feed", "bank sync", "bank csv import", "bank statement feed"), default=True),
        _navigation_entry("TRACES Results", "integrations:traces_result_import", "CA Practice", aliases=("traces import", "tds result import", "traces status", "tds acknowledgement", "justification report"), default=True),
        _navigation_entry("Integrations", "integrations:dashboard", "CA Practice", aliases=("gst api", "e invoice", "eway bill", "whatsapp")),

        _navigation_entry("Ledgers", "ledger:list", "Accounting", aliases=("ledger list", "accounts")),
        _navigation_entry("New Ledger", "ledger:create", "Accounting", aliases=("create ledger", "new account", "party")),
        _navigation_entry("Vouchers", "vouchers:list", "Accounting", aliases=("voucher list", "entries", "old entries")),
        _navigation_entry("New Voucher", "vouchers:create", "Accounting", aliases=("create voucher", "new entry", "entry"), default=True),
        _navigation_entry("Voucher Quality", "vouchers:quality", "Accounting", aliases=("quality engine", "voucher audit", "risk vouchers"), default=True),
        _navigation_entry("Collections Center", "core:collections_command_center", "Accounting", aliases=("collections", "receivables", "payment reminders", "debtors", "invoice chase"), default=True),
        _navigation_entry("Bank Reco Autopilot", "core:bank_reco_autopilot", "Accounting", aliases=("bank autopilot", "bank reconciliation autopilot", "bank reco", "unreconciled bank", "bank exceptions"), default=True),
        _navigation_entry("Accounting Close", "core:accounting_close", "Accounting", aliases=("close workbench", "month close", "books close"), default=True),
        _navigation_entry("Outstanding Statement", "vouchers:outstanding", "Accounting", aliases=("receivables", "payables", "outstanding")),
        _navigation_entry("Bulk Settlement", "receivables:bulk_settlement", "Accounting", aliases=("settlement", "receipt matching")),
        _navigation_entry("GSTR-2B Upload", "gstr2b:upload", "Accounting", aliases=("2b upload", "gst 2b import")),
        _navigation_entry("GSTR-2B Results", "gstr2b:results", "Accounting", aliases=("2b reconciliation", "gstr2b results")),

        _navigation_entry("Reports Home", "reports:home", "Reports", aliases=("reports",)),
        _navigation_entry("Financial Dashboard", "reports:dashboard_financials", "Reports", aliases=("financials",)),
        _navigation_entry("Day Book", "reports:day_book", "Reports", aliases=("daybook", "transactions")),
        _navigation_entry("Trial Balance", "reports:trial_balance_simple", "Reports", aliases=("tb",)),
        _navigation_entry("Profit & Loss", "reports:profit_loss_simple", "Reports", aliases=("p&l", "pl", "profit loss")),
        _navigation_entry("Balance Sheet", "reports:balance_sheet_simple", "Reports", aliases=("bs",)),
        _navigation_entry("GST Report", "reports:gst_report", "Reports", aliases=("gst summary", "gstr 1", "gstr 3b")),
        _navigation_entry("Cash Flow", "reports:cash_flow", "Reports", aliases=("cashflow",)),
        _navigation_entry("Cash Flow Forecast", "reports:cash_flow_forecast", "Reports", aliases=("forecast",)),
        _navigation_entry("Receivables Aging", "reports:receivables_aging", "Reports", aliases=("aging", "debtors aging")),
        _navigation_entry("MSME Overdue", "reports:msme_overdue", "Reports", aliases=("msme",)),

        _navigation_entry("Inventory Items", "inventory:list", "Inventory", aliases=("stock items", "items")),
        _navigation_entry("New Stock Item", "inventory:create", "Inventory", aliases=("create item", "new item")),
        _navigation_entry("Stock Summary", "inventory:summary", "Inventory", aliases=("inventory summary",)),
        _navigation_entry("Stock Valuation", "inventory:valuation", "Inventory", aliases=("valuation",)),
        _navigation_entry("Low Stock", "inventory:low_stock", "Inventory", aliases=("reorder",)),
        _navigation_entry("Batch Summary", "inventory:batch_summary", "Inventory", aliases=("batch report",)),
        _navigation_entry("Godowns", "inventory:godown_list", "Inventory", aliases=("warehouse", "locations")),
        _navigation_entry("Batches", "inventory:batch_list", "Inventory", aliases=("batch list",)),

        _navigation_entry("Orders", "orders:order_list", "Orders", aliases=("sales order", "purchase order")),
        _navigation_entry("New Order", "orders:order_create", "Orders", aliases=("create order",)),
        _navigation_entry("Open Orders", "orders:open_orders", "Orders", aliases=("pending orders",)),
        _navigation_entry("Cost Centers", "costcenter:cost_center_list", "Cost Centers", aliases=("projects", "departments")),
        _navigation_entry("New Cost Center", "costcenter:cost_center_create", "Cost Centers", aliases=("create project",)),
        _navigation_entry("Budget Variance", "costcenter:budget_variance", "Cost Centers", aliases=("variance", "budget report")),
        _navigation_entry("Cost Center Report", "costcenter:cost_center_report", "Cost Centers", aliases=("project report",)),

        _navigation_entry("Payroll Employees", "payroll:employee_list", "Payroll", aliases=("employees", "staff")),
        _navigation_entry("New Employee", "payroll:employee_create", "Payroll", aliases=("create employee",)),
        _navigation_entry("Payroll Runs", "payroll:payroll_run_list", "Payroll", aliases=("salary run", "payroll processing")),
        _navigation_entry("New Payroll Run", "payroll:payroll_run_create", "Payroll", aliases=("process payroll",)),
        _navigation_entry("Payroll Summary", "payroll:payroll_summary", "Payroll", aliases=("salary report",)),

        _navigation_entry("Fixed Assets", "fixedassets:asset_list", "Fixed Assets", aliases=("assets",)),
        _navigation_entry("New Fixed Asset", "fixedassets:asset_create", "Fixed Assets", aliases=("create asset",)),
        _navigation_entry("Asset Register", "fixedassets:asset_register", "Fixed Assets", aliases=("fixed asset register",)),
        _navigation_entry("Asset Groups", "fixedassets:asset_group_list", "Fixed Assets", aliases=("asset categories",)),

        _navigation_entry("TDS Return Workbench", "tds:return_workbench", "TDS", aliases=("24q", "26q", "27q", "traces", "fvu", "tds return")),
        _navigation_entry("TDS Filing Pack", "tds:filing_pack", "TDS", aliases=("tds filing", "rpu", "tds csv", "form 16a")),
        _navigation_entry("TDS Post-Filing Center", "tds:post_filing_center", "TDS", aliases=("traces status", "justification report", "conso file", "form 16a issue", "tds certificates")),
        _navigation_entry("TDS Entries", "tds:entry_list", "TDS", aliases=("tds", "tds list")),
        _navigation_entry("New TDS Entry", "tds:entry_create", "TDS", aliases=("create tds",)),
        _navigation_entry("TDS Sections", "tds:section_list", "TDS", aliases=("tds setup",)),
        _navigation_entry("TDS Register", "tds:tds_register", "TDS", aliases=("tds report",)),

        _navigation_entry("Scan Bill / OCR", "ocr:upload", "Document Inbox", aliases=("ocr upload", "scan", "bill scan"), default=True),
        _navigation_entry("OCR Submissions", "ocr:list", "Document Inbox", aliases=("document inbox", "ocr list")),
        _navigation_entry("GST Certificate Scan", "ocr:gst_certificate_scan", "Document Inbox", aliases=("gst certificate", "gstin scan")),

        _navigation_entry("Compliance Health", "core:compliance_health", "Settings", aliases=("health check",)),
        _navigation_entry("Bank Reconciliation", "core:bank_statement_list", "Settings", aliases=("bank rec", "bank statements")),
        _navigation_entry("Bank Reconciliation Report", "core:bank_reconciliation_report", "Settings", aliases=("recon summary",)),
        _navigation_entry("Audit Trail", "core:audit_log", "Settings", aliases=("audit log",)),
        _navigation_entry("Tally Exit Control", "migration:exit_control", "Settings", aliases=("tally exit", "erp exit", "migration control", "migration readiness")),
        _navigation_entry("Move from Tally", "migration:sessions", "Settings", aliases=("tally import", "migration")),
        _navigation_entry("Setup Wizard", "core:setup_wizard", "Settings", aliases=("first run", "onboarding", "setup checklist", "readiness setup", "client setup"), default=True),
        _navigation_entry("Demo Mode", "core:demo_workspace", "Settings", aliases=("sales mode", "demo workspace", "sample data", "partner demo", "showcase"), default=True),
        _navigation_entry("App Settings", "core:app_settings", "Settings", aliases=("whatsapp settings", "client whatsapp", "settings"), default=True),
        _navigation_entry("Company Settings", "core:company_settings", "Settings", aliases=("settings", "period locks")),
        _navigation_entry("Switch Company", "core:select_company", "Settings", aliases=("select company",)),
    ]


def _rank_navigation(query):
    q_lower = query.lower().strip()
    ranked = []
    for entry in _navigation_index():
        names = [entry["name"], *entry.get("aliases", ()), entry.get("category", "")]
        haystack = " ".join(names).lower()
        name_lower = entry["name"].lower()
        score = 0
        if name_lower == q_lower or q_lower in [alias.lower() for alias in entry.get("aliases", ())]:
            score = 120
        elif name_lower.startswith(q_lower):
            score = 90
        elif any(alias.lower().startswith(q_lower) for alias in entry.get("aliases", ())):
            score = 75
        elif q_lower in haystack:
            score = 45
        else:
            similarity = max(
                difflib.SequenceMatcher(None, q_lower, value.lower()).ratio()
                for value in names
                if value
            )
            if similarity > 0.62:
                score = int(similarity * 35)
        if score:
            item = {
                "name": entry["name"],
                "url": entry["url"],
                "category": entry["category"],
                "score": score,
            }
            ranked.append(item)
    ranked.sort(key=lambda item: (item["score"], item["name"]), reverse=True)
    return ranked


def _default_navigation_entries():
    priority = {
        "New Voucher": 0,
        "Filing Review Center": 1,
        "GST Filing Pack": 2,
        "New Client Request": 3,
        "Client Request Reminders": 4,
        "Dashboard": 5,
        "CA Command Center": 6,
        "GST Workbench": 7,
        "Practice Work Queue": 8,
        "Operations Monitor": 9,
        "Security Control": 10,
        "App Settings": 11,
    }
    defaults = [entry for entry in _navigation_index() if entry.get("default")]
    return sorted(defaults, key=lambda entry: (priority.get(entry["name"], 100), entry["category"], entry["name"]))


@login_required
@require_POST
def chatbot_query(request):
    """Rule-based chatbot endpoint."""
    try:
        data = json.loads(request.body)
        message = data.get("message", "").strip()
        
        reply = get_bot_reply(message)
        return JsonResponse({"reply": reply})
    except Exception as e:
        logger.exception("Chatbot error: %s", e)
        # View fallback (slightly different text to identify it)
        return JsonResponse({"reply": "System Error. Try: sales entry, gst, voucher."}, status=200)

@login_required
def universal_search(request):
    """
    Enhanced Universal Search (Ctrl+K).
    Features: Priority Ranking, Recent Items, Fuzzy Matching.
    """
    query = request.GET.get("q", "").strip()
    company = request.current_company

    # Show recent items and suggestions if query is empty
    if not query:
        recent = request.session.get('recent_search_items', {})
        
        # Also suggest top ledgers
        from django.db.models import Count
        top_ledgers = (
            Ledger.objects.filter(company=company, is_active=True)
            .annotate(usage_count=Count('voucher_items'))
            .order_by('-usage_count', 'name')[:5]
        )
        suggestions = [
            {"id": l.id, "name": l.name, "url": reverse("reports:ledger_detail", args=[l.id])}
            for l in top_ledgers
        ]

        return JsonResponse({
            "recent": True,
            "navigation": [
                {
                    "name": entry["name"],
                    "url": entry["url"],
                    "category": entry["category"],
                }
                for entry in _default_navigation_entries()
            ][:12],
            "ledgers": recent.get('ledgers', []) or suggestions,
            "vouchers": recent.get('vouchers', []),
            "items": recent.get('items', []),
            "tasks": [],
            "client_requests": [],
        })

    if len(query) < 2:
        return JsonResponse({
            "navigation": _rank_navigation(query)[:10],
            "ledgers": [],
            "vouchers": [],
            "items": [],
            "tasks": [],
            "client_requests": [],
        })

    # 1. Navigation (Ranked by priority, with aliases)
    navigation = _rank_navigation(query)

    # 2. Ledgers (Fuzzy)
    # Get more than 5 to allow for fuzzy ranking
    ledgers_qs = Ledger.objects.filter(
        Q(company=company) & 
        (Q(name__icontains=query) | Q(name__icontains=query[0:3])) # Broaden for fuzzy
    ).distinct()[:20]
    ledgers_ranked = _rank_and_fuzzy(ledgers_qs, query, 'name')
    ledgers_out = [
        {"id": l.id, "name": l.name, "url": reverse("reports:ledger_detail", args=[l.id])}
        for l in ledgers_ranked
    ]

    # 3. Vouchers (number, narration, and imported/source reference)
    vouchers_qs = Voucher.objects.filter(
        Q(company=company)
        & (
            Q(number__icontains=query)
            | Q(narration__icontains=query)
            | Q(source_reference__icontains=query)
            | Q(voucher_type__icontains=query)
        )
    ).order_by("-date", "-created_at")[:50]
    vouchers_ranked = _rank_and_fuzzy(vouchers_qs, query, 'number')
    if len(vouchers_ranked) < 5:
        seen_voucher_ids = {voucher.pk for voucher in vouchers_ranked}
        for voucher in vouchers_qs:
            if voucher.pk not in seen_voucher_ids:
                vouchers_ranked.append(voucher)
            if len(vouchers_ranked) >= 5:
                break
    vouchers_out = [
        {
            "id": v.id,
            "name": f"{v.number or 'Voucher'} - {v.voucher_type} - {v.date:%d %b %Y}",
            "number": v.number,
            "url": reverse("vouchers:detail", args=[v.id]),
        }
        for v in vouchers_ranked
    ]

    # 4. Items (Fuzzy)
    items_qs = StockItem.objects.filter(
        Q(company=company) &
        (Q(name__icontains=query) | Q(name__icontains=query[0:3]))
    ).distinct()[:20]
    items_ranked = _rank_and_fuzzy(items_qs, query, 'name')
    items_out = [
        {"id": i.id, "name": i.name, "url": reverse("inventory:edit", args=[i.id])}
        for i in items_ranked
    ]

    # 5. Practice tasks and client request records added by CA workflows.
    tasks_qs = PracticeTask.objects.filter(
        Q(company=company)
        & (
            Q(title__icontains=query)
            | Q(reference__icontains=query)
            | Q(description__icontains=query)
        )
    ).order_by("status", "due_date", "-priority")[:10]
    tasks_out = [
        {
            "id": task.id,
            "name": task.title,
            "url": reverse("core:practice_task_update", args=[task.id]),
        }
        for task in tasks_qs[:5]
    ]

    try:
        from portal.models import ClientDocumentRequest

        client_requests_qs = ClientDocumentRequest.objects.filter(
            Q(company=company)
            & (
                Q(title__icontains=query)
                | Q(source_reference__icontains=query)
                | Q(notes__icontains=query)
                | Q(response_note__icontains=query)
                | Q(portal_user__name__icontains=query)
                | Q(portal_user__email__icontains=query)
            )
        ).order_by("status", "due_date", "-created_at")[:10]
        client_requests_out = [
            {
                "id": doc.id,
                "name": doc.title,
                "url": f"{reverse('portal:client_requests')}?{urlencode({'q': doc.source_reference or doc.title})}",
            }
            for doc in client_requests_qs[:5]
        ]
    except Exception:
        client_requests_out = []

    return JsonResponse({
        "navigation": navigation[:12],
        "ledgers": ledgers_out,
        "vouchers": vouchers_out,
        "items": items_out,
        "tasks": tasks_out,
        "client_requests": client_requests_out,
    })

@login_required
def select_company(request):
    """Show list of companies the user can access and let them choose one."""
    access_list = (
        UserCompanyAccess.objects.filter(user=request.user)
        .select_related("company")
        .order_by("company__name")
    )

    if request.method == "POST":
        company_id = request.POST.get("company_id")
        if access_list.filter(company_id=company_id).exists():
            request.session["current_company_id"] = int(company_id)
            messages.success(request, "Company selected successfully.")
            return redirect("core:dashboard")
        else:
            messages.error(request, "You do not have access to that company.")

    return render(request, "core/select_company.html", {"access_list": access_list})


@login_required
def switch_company(request, company_id):
    """Switch active company and redirect back."""
    access = UserCompanyAccess.objects.filter(
        user=request.user, company_id=company_id
    ).first()
    if access:
        request.session["current_company_id"] = company_id
        messages.success(request, f"Switched to {access.company.name}.")
    else:
        messages.error(request, "Access denied.")
    
    # Try redirecting back to where the user came from
    next_url = request.GET.get('next') or request.META.get('HTTP_REFERER')
    if (
        next_url
        and not next_url.endswith(f"/switch/{company_id}/")
        and url_has_allowed_host_and_scheme(
            next_url,
            allowed_hosts={request.get_host()},
            require_https=request.is_secure(),
        )
    ):
        return redirect(next_url)
    return redirect("core:dashboard")


def _companies_for_user(user):
    if user.is_superuser:
        return Company.objects.all().order_by("name")
    return (
        Company.objects.filter(user_access__user=user)
        .distinct()
        .order_by("name")
    )


def _task_users_for_companies(companies):
    return (
        get_user_model().objects.filter(company_access__company__in=companies)
        .distinct()
        .order_by("email")
    )


def _can_manage_company(user, company):
    if user.is_superuser:
        return True
    return UserCompanyAccess.objects.filter(
        user=user,
        company=company,
        role__in=["Admin", "Accountant"],
    ).exists()


def _manageable_companies_for_user(user):
    if user.is_superuser:
        return Company.objects.all().order_by("name")
    return (
        Company.objects.filter(
            user_access__user=user,
            user_access__role__in=["Admin", "Accountant"],
        )
        .distinct()
        .order_by("name")
    )


def _parse_month_period(raw_period=None):
    today = timezone.localdate()
    if not raw_period:
        start = today.replace(day=1)
    else:
        try:
            year_text, month_text = raw_period.split("-", 1)
            start = _date(int(year_text), int(month_text), 1)
        except (TypeError, ValueError):
            start = today.replace(day=1)
    end = start.replace(day=calendar.monthrange(start.year, start.month)[1])
    return start, end


def _add_months(value, months):
    month_index = (value.month - 1) + months
    year = value.year + (month_index // 12)
    month = (month_index % 12) + 1
    day = min(value.day, calendar.monthrange(year, month)[1])
    return _date(year, month, day)


def _period_due_date(period_start, day, month_offset=1):
    due_month = _add_months(period_start, month_offset)
    return due_month.replace(day=min(day, calendar.monthrange(due_month.year, due_month.month)[1]))


def _gst_filing_templates(period_start):
    return [
        (ComplianceFiling.TYPE_GST_IMS, "GST IMS Review", _period_due_date(period_start, 10)),
        (ComplianceFiling.TYPE_GSTR1, "GSTR-1", _period_due_date(period_start, 11)),
        (ComplianceFiling.TYPE_GSTR3B, "GSTR-3B", _period_due_date(period_start, 20)),
    ]


def _gst_filings_for_period(company, period_start, period_end):
    qs = ComplianceFiling.objects.filter(
        company=company,
        filing_type__in=[
            ComplianceFiling.TYPE_GST_IMS,
            ComplianceFiling.TYPE_GSTR1,
            ComplianceFiling.TYPE_GSTR3B,
        ],
        period_start=period_start,
        period_end=period_end,
    )
    return {filing.filing_type: filing for filing in qs}


def _ensure_gst_period_filings(company, period_start, period_end, user=None):
    from .compliance_workflow import sync_task_for_filing

    created = []
    for filing_type, label, due_date in _gst_filing_templates(period_start):
        filing, was_created = ComplianceFiling.objects.get_or_create(
            company=company,
            filing_type=filing_type,
            period_start=period_start,
            period_end=period_end,
            defaults={
                "title": f"{label} - {period_start:%b %Y}",
                "status": ComplianceFiling.STATUS_NOT_STARTED,
                "priority": PracticeTask.PRIORITY_NORMAL,
                "due_date": due_date,
                "created_by": user,
                "source": ComplianceFiling.SOURCE_CALENDAR,
            },
        )
        if was_created:
            sync_task_for_filing(filing, user=user)
            created.append(filing)
    return created


def _gst_party_ledger(voucher, *, entry_type="DR", preferred_nature="Asset"):
    fallback = None
    for item in voucher.items.all():
        ledger = item.ledger
        if item.entry_type != entry_type:
            continue
        if ledger.account_group.nature == "Tax":
            continue
        if fallback is None:
            fallback = ledger
        if ledger.account_group.nature == preferred_nature:
            return ledger
    return fallback


def _voucher_line_gross(voucher):
    debit_total = Decimal("0.00")
    credit_total = Decimal("0.00")
    for item in voucher.items.all():
        if item.entry_type == "DR":
            debit_total += item.amount
        elif item.entry_type == "CR":
            credit_total += item.amount
    return max(debit_total, credit_total, Decimal("0.00"))


def _voucher_has_registered_gst_recipient(voucher):
    party = _gst_party_ledger(voucher, entry_type="DR", preferred_nature="Asset")
    party_gstin = (party.gstin or "").strip().upper() if party and party.gstin else ""
    return bool(party_gstin and GSTIN_RE.match(party_gstin))


def _gst_sales_readiness(company, period_start, period_end):
    own_state = (company.gstin or "")[:2]
    vouchers = list(
        Voucher.objects.filter(
            company=company,
            voucher_type__in=["Sales", "Sales Return"],
            status="APPROVED",
            date__gte=period_start,
            date__lte=period_end,
        )
        .prefetch_related("items__ledger__account_group", "items__stock_item__hsn_sac")
        .order_by("date", "id")
    )

    missing_pos = []
    invalid_party_gstin = []
    tax_pos_mismatch = []
    missing_hsn = []
    for voucher in vouchers:
        taxable = (voucher.total_tax or Decimal("0.00")) > Decimal("0.00")
        if taxable and not voucher.place_of_supply:
            missing_pos.append(voucher)

        party = _gst_party_ledger(voucher, entry_type="DR", preferred_nature="Asset")
        party_gstin = (party.gstin or "").strip().upper() if party and party.gstin else ""
        if party_gstin and not GSTIN_RE.match(party_gstin):
            invalid_party_gstin.append({"voucher": voucher, "party": party, "party_gstin": party_gstin})

        pos = (voucher.place_of_supply or "").strip()
        if taxable and own_state and pos:
            if pos == own_state and (voucher.igst_amount or Decimal("0.00")) > Decimal("0.00"):
                tax_pos_mismatch.append({"voucher": voucher, "party": party, "party_gstin": party_gstin})
            elif pos != own_state and (
                (voucher.cgst_amount or Decimal("0.00")) > Decimal("0.00")
                or (voucher.sgst_amount or Decimal("0.00")) > Decimal("0.00")
            ):
                tax_pos_mismatch.append({"voucher": voucher, "party": party, "party_gstin": party_gstin})

        stock_lines_missing_hsn = [
            item for item in voucher.items.all()
            if (
                item.stock_item_id
                and item.ledger.account_group.nature == "Income"
                and item.entry_type == "CR"
                and not item.stock_item.hsn_sac_id
            )
        ]
        if stock_lines_missing_hsn:
            missing_hsn.append({
                "voucher": voucher,
                "party": party,
                "items": [item.stock_item for item in stock_lines_missing_hsn],
            })

    return {
        "missing_pos": missing_pos,
        "invalid_party_gstin": invalid_party_gstin,
        "tax_pos_mismatch": tax_pos_mismatch,
        "missing_hsn": missing_hsn,
        "missing_pos_count": len(missing_pos),
        "invalid_party_gstin_count": len(invalid_party_gstin),
        "tax_pos_mismatch_count": len(tax_pos_mismatch),
        "missing_hsn_count": len(missing_hsn),
    }


def _gst_itc_180_day_watch(company, *, as_of_date=None, window_end=None, limit=100):
    as_of_date = as_of_date or timezone.localdate()
    window_end = window_end or as_of_date
    horizon = max(as_of_date, window_end)
    cutoff_date = horizon - timedelta(days=180)
    vouchers = (
        Voucher.objects.filter(
            company=company,
            voucher_type="Purchase",
            status="APPROVED",
            is_itc_claimed=True,
            reverse_charge=False,
            total_tax__gt=0,
            outstanding_amount__gt=0,
            date__lte=cutoff_date,
        )
        .prefetch_related("items__ledger__account_group")
        .order_by("date", "id")
    )

    rows = []
    summary = {
        "overdue_count": 0,
        "due_soon_count": 0,
        "attention_count": 0,
        "overdue_itc": Decimal("0.00"),
        "due_soon_itc": Decimal("0.00"),
        "total_reversal_itc": Decimal("0.00"),
    }

    for voucher in vouchers:
        reversal_due_date = voucher.date + timedelta(days=180)
        if reversal_due_date > horizon:
            continue

        gross_amount = _voucher_line_gross(voucher)
        outstanding_amount = voucher.outstanding_amount or Decimal("0.00")
        tax_amount = voucher.total_tax or Decimal("0.00")
        if gross_amount > 0:
            unpaid_ratio = min(Decimal("1.00"), outstanding_amount / gross_amount)
        else:
            unpaid_ratio = Decimal("1.00")
        reversal_itc = (tax_amount * unpaid_ratio).quantize(Decimal("0.01"))
        vendor = _gst_party_ledger(voucher, entry_type="CR", preferred_nature="Liability")
        status = "overdue" if reversal_due_date <= as_of_date else "due_soon"

        if status == "overdue":
            summary["overdue_count"] += 1
            summary["overdue_itc"] += reversal_itc
        else:
            summary["due_soon_count"] += 1
            summary["due_soon_itc"] += reversal_itc
        summary["attention_count"] += 1
        summary["total_reversal_itc"] += reversal_itc

        rows.append({
            "voucher": voucher,
            "vendor": vendor,
            "reversal_due_date": reversal_due_date,
            "days_to_due": (reversal_due_date - as_of_date).days,
            "days_overdue": max((as_of_date - reversal_due_date).days, 0),
            "gross_amount": gross_amount,
            "outstanding_amount": outstanding_amount,
            "tax_amount": tax_amount,
            "reversal_itc": reversal_itc,
            "status": status,
        })

    rows.sort(key=lambda row: (row["reversal_due_date"], row["voucher"].pk))
    return {
        "rows": rows[:limit],
        "summary": summary,
    }


def _gst_rcm_watch(company, period_start, period_end):
    vouchers = list(
        Voucher.objects.filter(
            company=company,
            voucher_type="Purchase",
            status="APPROVED",
            reverse_charge=True,
            date__gte=period_start,
            date__lte=period_end,
        )
        .prefetch_related("items__ledger__account_group")
        .order_by("date", "id")
    )
    missing_tax = [
        voucher for voucher in vouchers
        if (voucher.total_tax or Decimal("0.00")) <= Decimal("0.00")
    ]
    rcm_tax = sum(((voucher.total_tax or Decimal("0.00")) for voucher in vouchers), Decimal("0.00"))
    return {
        "vouchers": vouchers,
        "missing_tax": missing_tax,
        "count": len(vouchers),
        "missing_tax_count": len(missing_tax),
        "tax_amount": rcm_tax,
    }


def _gst_eway_bill_watch(company, period_start, period_end):
    now = timezone.now()
    due_soon_cutoff = now + timedelta(hours=24)
    vouchers = list(
        Voucher.objects.filter(
            company=company,
            voucher_type__in=["Sales", "Purchase", "Sales Return", "Purchase Return"],
            status="APPROVED",
            date__gte=period_start,
            date__lte=period_end,
        )
        .prefetch_related("items__ledger__account_group")
        .order_by("date", "id")
    )

    missing = []
    expired = []
    due_soon = []
    for voucher in vouchers:
        has_movement_signal = bool(
            voucher.dispatch_pincode
            or voucher.ship_to_pincode
            or voucher.vehicle_number
            or voucher.transporter_id
            or voucher.transport_doc_no
        )
        if has_movement_signal and not voucher.e_way_bill_no and _voucher_line_gross(voucher) >= Decimal("50000.00"):
            missing.append(voucher)

        valid_until = voucher.e_way_bill_valid_until
        if voucher.e_way_bill_no and valid_until:
            if valid_until < now:
                expired.append(voucher)
            elif valid_until <= due_soon_cutoff:
                due_soon.append(voucher)

    return {
        "missing": missing,
        "expired": expired,
        "due_soon": due_soon,
        "missing_count": len(missing),
        "expired_count": len(expired),
        "due_soon_count": len(due_soon),
    }


def _gst_workbench_snapshot(company, period_start, period_end):
    from gstr2b.models import PortalGSTR2BEntry
    from portal.models import ClientDocumentRequest
    from reports.utils import get_gst_report

    today = timezone.localdate()
    gst_report = get_gst_report(company, period_start, period_end)
    portal_qs = PortalGSTR2BEntry.objects.filter(
        company=company,
        invoice_date__gte=period_start,
        invoice_date__lte=period_end,
    )
    matched_2b = portal_qs.filter(match_status="matched").count()
    missing_in_books = portal_qs.filter(match_status="missing_in_books").count()
    pending_2b = portal_qs.filter(action_status="pending").count()
    rejected_2b = portal_qs.filter(action_status="rejected").count()
    missing_in_portal = Voucher.objects.filter(
        company=company,
        voucher_type="Purchase",
        status="APPROVED",
        is_itc_claimed=False,
        date__gte=period_start,
        date__lte=period_end,
    ).count()
    draft_gst_vouchers = Voucher.objects.filter(
        company=company,
        voucher_type__in=["Sales", "Purchase"],
        date__gte=period_start,
        date__lte=period_end,
    ).exclude(status="APPROVED").count()
    sales_readiness = _gst_sales_readiness(company, period_start, period_end)
    itc_180_watch = _gst_itc_180_day_watch(
        company,
        as_of_date=min(today, period_end),
        window_end=period_end,
    )
    itc_180_summary = itc_180_watch["summary"]
    rcm_watch = _gst_rcm_watch(company, period_start, period_end)
    eway_watch = _gst_eway_bill_watch(company, period_start, period_end)
    e_invoice_enabled = bool(company.e_invoice_enabled)
    e_invoice_deadline_days = max(1, int(company.e_invoice_reporting_deadline_days or 30))
    e_invoice_warning_days = max(0, int(company.e_invoice_warning_days or 0))
    e_invoice_warning_days = min(e_invoice_warning_days, e_invoice_deadline_days)
    e_invoice_missing_irn = 0
    e_invoice_expired = 0
    e_invoice_due_soon = 0
    if e_invoice_enabled:
        e_invoice_candidates = list(
            Voucher.objects.filter(
                company=company,
                voucher_type__in=["Sales", "Sales Return"],
                status="APPROVED",
                date__gte=period_start,
                date__lte=period_end,
            )
            .filter(Q(e_invoice_irn="") | Q(e_invoice_irn__isnull=True))
            .prefetch_related("items__ledger__account_group")
            .order_by("date", "id")
        )
        e_invoice_missing_vouchers = [
            voucher for voucher in e_invoice_candidates
            if _voucher_has_registered_gst_recipient(voucher)
        ]
        e_invoice_missing_irn = len(e_invoice_missing_vouchers)
        deadline_cutoff = today - timedelta(days=e_invoice_deadline_days)
        warning_cutoff = today - timedelta(days=e_invoice_warning_days)
        e_invoice_expired = sum(1 for voucher in e_invoice_missing_vouchers if voucher.date <= deadline_cutoff)
        e_invoice_due_soon = sum(
            1 for voucher in e_invoice_missing_vouchers
            if voucher.date <= warning_cutoff and voucher.date > deadline_cutoff
        )
    filings = _gst_filings_for_period(company, period_start, period_end)
    open_filings = [filing for filing in filings.values() if filing.is_open]
    overdue_filings = [
        filing for filing in open_filings
        if filing.due_date and filing.due_date < today
    ]
    open_gst_tasks = PracticeTask.objects.filter(
        company=company,
        task_type=PracticeTask.TYPE_GST,
        period_start=period_start,
        period_end=period_end,
    ).exclude(status__in=[PracticeTask.STATUS_DONE, PracticeTask.STATUS_CANCELLED])
    overdue_gst_tasks = open_gst_tasks.filter(due_date__lt=today).count()
    open_notices = ComplianceNotice.objects.filter(
        company=company,
        notice_type=ComplianceNotice.TYPE_GST,
    ).exclude(status=ComplianceNotice.STATUS_CLOSED)
    overdue_notices = open_notices.filter(response_due_date__lt=today).count()
    client_chase = ClientDocumentRequest.objects.filter(
        company=company,
        related_task__task_type=PracticeTask.TYPE_GST,
        related_task__period_start=period_start,
        related_task__period_end=period_end,
    ).exclude(status__in=[
        ClientDocumentRequest.STATUS_CLOSED,
        ClientDocumentRequest.STATUS_CANCELLED,
    ])
    client_chase_open = client_chase.filter(status=ClientDocumentRequest.STATUS_OPEN).count()
    client_chase_uploaded = client_chase.filter(status=ClientDocumentRequest.STATUS_UPLOADED).count()
    client_chase_overdue = client_chase.filter(
        status=ClientDocumentRequest.STATUS_OPEN,
        due_date__lt=today,
    ).count()
    review = GSTPeriodReview.objects.filter(
        company=company,
        period_start=period_start,
        period_end=period_end,
    ).select_related("reviewed_by", "prepared_by").first()

    risk_score = 0
    risk_score += min(35, missing_in_books * 5)
    risk_score += min(25, missing_in_portal * 5)
    risk_score += min(15, pending_2b * 3)
    risk_score += min(20, rejected_2b * 8)
    risk_score += min(20, draft_gst_vouchers * 2)
    risk_score += min(20, len(overdue_filings) * 10)
    risk_score += min(15, overdue_gst_tasks * 5)
    risk_score += min(15, overdue_notices * 5)
    risk_score += min(15, client_chase_overdue * 5)
    risk_score += min(10, client_chase_uploaded * 2)
    risk_score += min(25, e_invoice_expired * 10)
    risk_score += min(15, e_invoice_due_soon * 5)
    risk_score += min(10, e_invoice_missing_irn * 1)
    risk_score += min(20, sales_readiness["missing_pos_count"] * 8)
    risk_score += min(20, sales_readiness["invalid_party_gstin_count"] * 10)
    risk_score += min(15, sales_readiness["tax_pos_mismatch_count"] * 5)
    risk_score += min(15, sales_readiness["missing_hsn_count"] * 5)
    risk_score += min(25, itc_180_summary["overdue_count"] * 10)
    risk_score += min(15, itc_180_summary["due_soon_count"] * 5)
    risk_score += min(20, rcm_watch["missing_tax_count"] * 10)
    risk_score += min(15, eway_watch["missing_count"] * 8)
    risk_score += min(10, eway_watch["expired_count"] * 5)
    if not company.gstin:
        risk_score += 15
    risk_score = min(100, risk_score)

    row = {
        "company": company,
        "period_start": period_start,
        "period_end": period_end,
        "filings": filings,
        "review": review,
        "risk_score": risk_score,
        "health_score": max(0, 100 - risk_score),
        "sales_invoices": len(gst_report["gstr1_rows"]),
        "taxable_sales": gst_report["tot_taxable_sales"],
        "output_tax": gst_report["tot_out_tax"],
        "taxable_purchases": gst_report["tot_taxable_purchases"],
        "itc": gst_report["tot_itc"],
        "net_tax_payable": gst_report["net_tax_payable"],
        "matched_2b": matched_2b,
        "missing_in_books": missing_in_books,
        "missing_in_portal": missing_in_portal,
        "pending_2b": pending_2b,
        "rejected_2b": rejected_2b,
        "draft_gst_vouchers": draft_gst_vouchers,
        "open_filings": len(open_filings),
        "overdue_filings": len(overdue_filings),
        "open_gst_tasks": open_gst_tasks.count(),
        "overdue_gst_tasks": overdue_gst_tasks,
        "open_notices": open_notices.count(),
        "overdue_notices": overdue_notices,
        "client_chase_open": client_chase_open,
        "client_chase_uploaded": client_chase_uploaded,
        "client_chase_overdue": client_chase_overdue,
        "client_chase_total": client_chase.count(),
        "e_invoice_enabled": e_invoice_enabled,
        "e_invoice_deadline_days": e_invoice_deadline_days,
        "e_invoice_warning_days": e_invoice_warning_days,
        "e_invoice_missing_irn": e_invoice_missing_irn,
        "e_invoice_expired": e_invoice_expired,
        "e_invoice_due_soon": e_invoice_due_soon,
        "sales_missing_pos": sales_readiness["missing_pos_count"],
        "sales_invalid_party_gstin": sales_readiness["invalid_party_gstin_count"],
        "sales_tax_pos_mismatch": sales_readiness["tax_pos_mismatch_count"],
        "sales_missing_hsn": sales_readiness["missing_hsn_count"],
        "itc_180_overdue": itc_180_summary["overdue_count"],
        "itc_180_due_soon": itc_180_summary["due_soon_count"],
        "itc_180_reversal_due": itc_180_summary["attention_count"],
        "itc_180_reversal_itc": itc_180_summary["total_reversal_itc"],
        "itc_180_overdue_itc": itc_180_summary["overdue_itc"],
        "itc_180_due_soon_itc": itc_180_summary["due_soon_itc"],
        "rcm_purchase_count": rcm_watch["count"],
        "rcm_missing_tax": rcm_watch["missing_tax_count"],
        "rcm_tax_amount": rcm_watch["tax_amount"],
        "eway_missing": eway_watch["missing_count"],
        "eway_expired": eway_watch["expired_count"],
        "eway_due_soon": eway_watch["due_soon_count"],
        "snapshot": {
            "risk_score": risk_score,
            "sales_invoices": len(gst_report["gstr1_rows"]),
            "taxable_sales": str(gst_report["tot_taxable_sales"]),
            "output_tax": str(gst_report["tot_out_tax"]),
            "taxable_purchases": str(gst_report["tot_taxable_purchases"]),
            "itc": str(gst_report["tot_itc"]),
            "net_tax_payable": str(gst_report["net_tax_payable"]),
            "matched_2b": matched_2b,
            "missing_in_books": missing_in_books,
            "missing_in_portal": missing_in_portal,
            "pending_2b": pending_2b,
            "rejected_2b": rejected_2b,
            "draft_gst_vouchers": draft_gst_vouchers,
            "open_filings": len(open_filings),
            "overdue_filings": len(overdue_filings),
            "open_gst_tasks": open_gst_tasks.count(),
            "open_notices": open_notices.count(),
            "client_chase_open": client_chase_open,
            "client_chase_uploaded": client_chase_uploaded,
            "client_chase_overdue": client_chase_overdue,
            "e_invoice_enabled": e_invoice_enabled,
            "e_invoice_deadline_days": e_invoice_deadline_days,
            "e_invoice_warning_days": e_invoice_warning_days,
            "e_invoice_missing_irn": e_invoice_missing_irn,
            "e_invoice_expired": e_invoice_expired,
            "e_invoice_due_soon": e_invoice_due_soon,
            "sales_missing_pos": sales_readiness["missing_pos_count"],
            "sales_invalid_party_gstin": sales_readiness["invalid_party_gstin_count"],
            "sales_tax_pos_mismatch": sales_readiness["tax_pos_mismatch_count"],
            "sales_missing_hsn": sales_readiness["missing_hsn_count"],
            "itc_180_overdue": itc_180_summary["overdue_count"],
            "itc_180_due_soon": itc_180_summary["due_soon_count"],
            "itc_180_reversal_due": itc_180_summary["attention_count"],
            "itc_180_reversal_itc": str(itc_180_summary["total_reversal_itc"]),
            "rcm_purchase_count": rcm_watch["count"],
            "rcm_missing_tax": rcm_watch["missing_tax_count"],
            "rcm_tax_amount": str(rcm_watch["tax_amount"]),
            "eway_missing": eway_watch["missing_count"],
            "eway_expired": eway_watch["expired_count"],
            "eway_due_soon": eway_watch["due_soon_count"],
        },
    }
    row["signoff_blockers"] = _gst_signoff_blockers(row)
    row["can_sign_off"] = not row["signoff_blockers"]
    row["snapshot"]["can_sign_off"] = row["can_sign_off"]
    row["snapshot"]["signoff_blockers"] = [
        {
            "code": blocker["code"],
            "title": blocker["title"],
            "count": blocker["count"],
        }
        for blocker in row["signoff_blockers"]
    ]
    return row


def _gst_signoff_blockers(row):
    blockers = []

    def add(code, title, count, description):
        if not count:
            return
        blockers.append({
            "code": code,
            "title": title,
            "count": count,
            "description": description,
        })

    add(
        "missing_gstin",
        "Company GSTIN missing",
        0 if row["company"].gstin else 1,
        "Add GSTIN before final GST review sign-off.",
    )
    add(
        "missing_in_books",
        "2B entries missing in books",
        row["missing_in_books"],
        "Book the purchase invoice or reject/mark pending with a documented reason.",
    )
    add(
        "missing_in_portal",
        "Book purchases missing in 2B",
        row["missing_in_portal"],
        "Resolve vendor filing gaps before claiming ITC.",
    )
    add(
        "pending_2b",
        "Pending IMS / 2B actions",
        row["pending_2b"],
        "Accept, reject, or document pending treatment before sign-off.",
    )
    add(
        "rejected_2b",
        "Rejected IMS / 2B records",
        row["rejected_2b"],
        "Review rejected records and keep evidence for ITC treatment.",
    )
    add(
        "draft_gst_vouchers",
        "Draft GST vouchers",
        row["draft_gst_vouchers"],
        "Approve or remove GST-sensitive draft vouchers before sign-off.",
    )
    add(
        "sales_missing_pos",
        "Sales invoices missing place of supply",
        row["sales_missing_pos"],
        "Add place of supply before GSTR-1 and 3B review sign-off.",
    )
    add(
        "sales_invalid_party_gstin",
        "Invalid customer GSTIN values",
        row["sales_invalid_party_gstin"],
        "Correct party master GSTIN values used in sales vouchers.",
    )
    add(
        "sales_tax_pos_mismatch",
        "Tax split does not match place of supply",
        row["sales_tax_pos_mismatch"],
        "Review CGST/SGST/IGST split against the place of supply.",
    )
    add(
        "sales_missing_hsn",
        "Sales stock items missing HSN/SAC",
        row["sales_missing_hsn"],
        "Add HSN/SAC classification on stock items before GSTR-1 sign-off.",
    )
    add(
        "itc_180_reversal_due",
        "ITC 180-day payment reversal due",
        row["itc_180_reversal_due"],
        "Reverse proportionate ITC in GSTR-3B or record payment proof before GST sign-off.",
    )
    add(
        "rcm_missing_tax",
        "RCM purchases missing GST tax",
        row["rcm_missing_tax"],
        "Add reverse-charge GST liability details before GSTR-3B sign-off.",
    )
    add(
        "eway_bill_missing",
        "E-way bill details missing",
        row["eway_missing"],
        "Add e-way bill details for high-value movement vouchers before GST review sign-off.",
    )
    add(
        "eway_bill_expired",
        "E-way bill validity expired",
        row["eway_expired"],
        "Review expired e-way bill validity before GST review sign-off.",
    )
    add(
        "e_invoice_expired",
        "E-invoice IRP deadline crossed",
        row["e_invoice_expired"],
        "Generate IRN or document treatment before signing off GST.",
    )
    add(
        "e_invoice_missing_irn",
        "Sales invoices without IRN",
        row["e_invoice_missing_irn"],
        "Applicable e-invoice sales documents must have IRN before GST review sign-off.",
    )
    add(
        "client_chase_open",
        "Open client document requests",
        row["client_chase_open"],
        "Client upload requests must be uploaded and reviewed or cancelled.",
    )
    add(
        "client_chase_uploaded",
        "Uploaded documents awaiting review",
        row["client_chase_uploaded"],
        "Review uploaded evidence and close the client request.",
    )
    add(
        "open_notices",
        "Open GST notices",
        row["open_notices"],
        "Close or document active GST notice handling before final sign-off.",
    )
    return blockers


def _gst_period_detail_context(company, period_start, period_end):
    from integrations.gst import build_gst_voucher_execution_context
    from gstr2b.models import PortalGSTR2BEntry
    from portal.models import ClientDocumentRequest
    from reports.utils import get_gst_report

    snapshot = _gst_workbench_snapshot(company, period_start, period_end)
    itc_180_watch = _gst_itc_180_day_watch(
        company,
        as_of_date=min(timezone.localdate(), period_end),
        window_end=period_end,
    )
    rcm_watch = _gst_rcm_watch(company, period_start, period_end)
    eway_watch = _gst_eway_bill_watch(company, period_start, period_end)
    gst_report = get_gst_report(company, period_start, period_end)
    portal_entries = PortalGSTR2BEntry.objects.filter(
        company=company,
        invoice_date__gte=period_start,
        invoice_date__lte=period_end,
    ).select_related("matched_voucher").order_by("-invoice_date", "gstin", "invoice_number")
    matched_2b = portal_entries.filter(match_status="matched")
    missing_in_books = portal_entries.filter(match_status="missing_in_books")
    pending_or_rejected = portal_entries.filter(action_status__in=["pending", "rejected"])
    sales_vouchers = Voucher.objects.filter(
        company=company,
        voucher_type="Sales",
        date__gte=period_start,
        date__lte=period_end,
    ).order_by("-date", "-id")
    e_invoice_missing_vouchers = []
    if snapshot["e_invoice_enabled"]:
        e_invoice_candidates = (
            Voucher.objects.filter(
                company=company,
                voucher_type__in=["Sales", "Sales Return"],
                status="APPROVED",
                date__gte=period_start,
                date__lte=period_end,
            )
            .filter(Q(e_invoice_irn="") | Q(e_invoice_irn__isnull=True))
            .prefetch_related("items__ledger__account_group")
            .order_by("date", "id")
        )
        e_invoice_missing_vouchers = [
            voucher for voucher in e_invoice_candidates
            if _voucher_has_registered_gst_recipient(voucher)
        ]
    draft_gst_vouchers = Voucher.objects.filter(
        company=company,
        voucher_type__in=["Sales", "Purchase"],
        date__gte=period_start,
        date__lte=period_end,
    ).exclude(status="APPROVED").order_by("-date", "-id")
    sales_readiness = _gst_sales_readiness(company, period_start, period_end)
    missing_in_portal = Voucher.objects.filter(
        company=company,
        voucher_type="Purchase",
        status="APPROVED",
        is_itc_claimed=False,
        date__gte=period_start,
        date__lte=period_end,
    ).order_by("-date", "-id")
    filings = list(ComplianceFiling.objects.filter(
        company=company,
        filing_type__in=[
            ComplianceFiling.TYPE_GST_IMS,
            ComplianceFiling.TYPE_GSTR1,
            ComplianceFiling.TYPE_GSTR3B,
        ],
        period_start=period_start,
        period_end=period_end,
    ).select_related("assigned_to", "reviewer", "related_task").order_by("due_date", "filing_type"))
    open_tasks = PracticeTask.objects.filter(
        company=company,
        task_type=PracticeTask.TYPE_GST,
        period_start=period_start,
        period_end=period_end,
    ).exclude(status__in=[PracticeTask.STATUS_DONE, PracticeTask.STATUS_CANCELLED]).select_related("assigned_to").order_by("due_date", "-priority")
    notices = ComplianceNotice.objects.filter(
        company=company,
        notice_type=ComplianceNotice.TYPE_GST,
    ).exclude(status=ComplianceNotice.STATUS_CLOSED).select_related("assigned_to", "related_filing").order_by("response_due_date", "-priority")

    exception_items = []
    for entry in missing_in_books[:50]:
        exception_items.append({
            "kind": "2b_missing_in_books",
            "severity": "danger",
            "label": "2B Missing in Books",
            "title": f"{entry.supplier_name or entry.gstin} - {entry.invoice_number}",
            "reference": entry.gstin,
            "date": entry.invoice_date,
            "amount": entry.tax_amount,
            "object": entry,
            "task_reference": f"GSTEX:2B:{entry.pk}",
        })
    for voucher in missing_in_portal[:50]:
        exception_items.append({
            "kind": "purchase_missing_in_2b",
            "severity": "warning",
            "label": "Purchase Missing in 2B",
            "title": f"{voucher.number or 'Purchase voucher'} - {voucher.narration[:60] if voucher.narration else voucher.date}",
            "reference": voucher.number,
            "date": voucher.date,
            "amount": voucher.total_tax,
            "object": voucher,
            "task_reference": f"GSTEX:VCH:{voucher.pk}",
        })
    for item in itc_180_watch["rows"][:50]:
        voucher = item["voucher"]
        exception_items.append({
            "kind": "itc_180_reversal_due",
            "severity": "danger" if item["status"] == "overdue" else "warning",
            "label": "ITC 180-Day Reversal",
            "title": f"{item['vendor'].name if item['vendor'] else 'Vendor'} - {voucher.number or voucher.pk}",
            "reference": voucher.number,
            "date": item["reversal_due_date"],
            "amount": item["reversal_itc"],
            "object": voucher,
            "task_reference": f"GSTEX:ITC180:{voucher.pk}",
        })
    for voucher in rcm_watch["missing_tax"][:50]:
        exception_items.append({
            "kind": "rcm_missing_tax",
            "severity": "danger",
            "label": "RCM Tax Missing",
            "title": f"Purchase {voucher.number or voucher.pk}",
            "reference": voucher.number,
            "date": voucher.date,
            "amount": voucher.total_tax,
            "object": voucher,
            "task_reference": f"GSTEX:RCM:{voucher.pk}",
        })
    for voucher in eway_watch["missing"][:50]:
        exception_items.append({
            "kind": "eway_bill_missing",
            "severity": "danger",
            "label": "E-Way Bill Missing",
            "title": f"{voucher.voucher_type} {voucher.number or voucher.pk}",
            "reference": voucher.number,
            "date": voucher.date,
            "amount": _voucher_line_gross(voucher),
            "object": voucher,
            "task_reference": f"GSTEX:EWAY:{voucher.pk}",
        })
    for voucher in eway_watch["expired"][:50]:
        exception_items.append({
            "kind": "eway_bill_expired",
            "severity": "danger",
            "label": "E-Way Bill Expired",
            "title": f"{voucher.voucher_type} {voucher.number or voucher.pk}",
            "reference": voucher.e_way_bill_no,
            "date": voucher.e_way_bill_valid_until.date() if voucher.e_way_bill_valid_until else voucher.date,
            "amount": _voucher_line_gross(voucher),
            "object": voucher,
            "task_reference": f"GSTEX:EWAYEXP:{voucher.pk}",
        })
    for voucher in draft_gst_vouchers[:50]:
        exception_items.append({
            "kind": "draft_gst_voucher",
            "severity": "warning",
            "label": "Draft GST Voucher",
            "title": f"{voucher.voucher_type} {voucher.number or voucher.pk}",
            "reference": voucher.number,
            "date": voucher.date,
            "amount": voucher.total_tax,
            "object": voucher,
            "task_reference": f"GSTEX:DRAFT:{voucher.pk}",
        })
    for voucher in sales_readiness["missing_pos"][:50]:
        exception_items.append({
            "kind": "sales_missing_pos",
            "severity": "danger",
            "label": "Sales POS Missing",
            "title": f"{voucher.voucher_type} {voucher.number or voucher.pk}",
            "reference": voucher.number,
            "date": voucher.date,
            "amount": voucher.total_tax,
            "object": voucher,
            "task_reference": f"GSTEX:SALEPOS:{voucher.pk}",
        })
    for item in sales_readiness["invalid_party_gstin"][:50]:
        voucher = item["voucher"]
        exception_items.append({
            "kind": "sales_invalid_party_gstin",
            "severity": "danger",
            "label": "Invalid Customer GSTIN",
            "title": f"{item['party'].name if item['party'] else 'Customer'} - {voucher.number or voucher.pk}",
            "reference": item["party_gstin"],
            "date": voucher.date,
            "amount": voucher.total_tax,
            "object": voucher,
            "task_reference": f"GSTEX:SALEGSTIN:{voucher.pk}",
        })
    for item in sales_readiness["tax_pos_mismatch"][:50]:
        voucher = item["voucher"]
        exception_items.append({
            "kind": "sales_tax_pos_mismatch",
            "severity": "warning",
            "label": "Tax / POS Mismatch",
            "title": f"{voucher.voucher_type} {voucher.number or voucher.pk}",
            "reference": voucher.place_of_supply,
            "date": voucher.date,
            "amount": voucher.total_tax,
            "object": voucher,
            "task_reference": f"GSTEX:SALETAXPOS:{voucher.pk}",
        })
    for item in sales_readiness["missing_hsn"][:50]:
        voucher = item["voucher"]
        item_names = ", ".join(stock_item.name for stock_item in item["items"][:3])
        exception_items.append({
            "kind": "sales_missing_hsn",
            "severity": "danger",
            "label": "HSN/SAC Missing",
            "title": f"{voucher.voucher_type} {voucher.number or voucher.pk} - {item_names}",
            "reference": voucher.number,
            "date": voucher.date,
            "amount": voucher.total_tax,
            "object": voucher,
            "task_reference": f"GSTEX:SALEHSN:{voucher.pk}",
        })
    if snapshot["e_invoice_enabled"]:
        deadline_cutoff = timezone.localdate() - timedelta(days=snapshot["e_invoice_deadline_days"])
        warning_cutoff = timezone.localdate() - timedelta(days=snapshot["e_invoice_warning_days"])
        for voucher in e_invoice_missing_vouchers[:50]:
            severity = "danger" if voucher.date <= deadline_cutoff else "warning"
            if voucher.date > warning_cutoff:
                severity = "warning"
            exception_items.append({
                "kind": "e_invoice_missing_irn",
                "severity": severity,
                "label": "E-Invoice IRN Missing",
                "title": f"{voucher.voucher_type} {voucher.number or voucher.pk}",
                "reference": voucher.number,
                "date": voucher.date,
                "amount": voucher.total_tax,
                "object": voucher,
                "task_reference": f"GSTEX:EINV:{voucher.pk}",
            })

    execution_candidates = {
        voucher.pk: voucher
        for voucher in [*e_invoice_missing_vouchers, *eway_watch["missing"]]
        if voucher.voucher_type == "Sales"
    }
    execution_context_by_voucher = {
        voucher_id: build_gst_voucher_execution_context(voucher)
        for voucher_id, voucher in execution_candidates.items()
    }
    e_invoice_payload_ready_ids = {
        voucher_id
        for voucher_id, context in execution_context_by_voucher.items()
        if context["e_invoice"]["ready"]
    }
    e_way_bill_payload_ready_ids = {
        voucher_id
        for voucher_id, context in execution_context_by_voucher.items()
        if context["e_way_bill"]["ready"]
    }
    for filing in filings:
        if filing.is_open and filing.due_date and filing.due_date < timezone.localdate():
            exception_items.append({
                "kind": "overdue_gst_filing",
                "severity": "danger",
                "label": "Overdue Filing",
                "title": filing.title,
                "reference": filing.get_filing_type_display(),
                "date": filing.due_date,
                "amount": None,
                "object": filing,
                "task_reference": f"GSTEX:FILING:{filing.pk}",
            })
    for notice in notices[:50]:
        exception_items.append({
            "kind": "open_gst_notice",
            "severity": "danger" if notice.response_due_date and notice.response_due_date < timezone.localdate() else "warning",
            "label": "GST Notice",
            "title": notice.title,
            "reference": notice.reference_number,
            "date": notice.response_due_date,
            "amount": None,
            "object": notice,
            "task_reference": f"GSTEX:NOTICE:{notice.pk}",
        })

    existing_task_refs = set(
        PracticeTask.objects.filter(
            company=company,
            reference__in=[item["task_reference"] for item in exception_items],
        ).values_list("reference", flat=True)
    )
    document_requests = ClientDocumentRequest.objects.filter(
        company=company,
        source_reference__in=[item["task_reference"] for item in exception_items],
    ).select_related("portal_user", "uploaded_submission", "related_task").order_by("status", "due_date", "-created_at")
    document_request_map = {request.source_reference: request for request in document_requests}
    for item in exception_items:
        item["document_request"] = document_request_map.get(item["task_reference"])

    return {
        **snapshot,
        "period_value": period_start.strftime("%Y-%m"),
        "gst_report": gst_report,
        "sales_vouchers": sales_vouchers[:100],
        "e_invoice_missing_vouchers": e_invoice_missing_vouchers[:100],
        "e_invoice_missing_voucher_ids": {voucher.pk for voucher in e_invoice_missing_vouchers},
        "itc_180_watch_rows": itc_180_watch["rows"][:100],
        "rcm_watch_vouchers": rcm_watch["vouchers"][:100],
        "eway_bill_missing_vouchers": eway_watch["missing"][:100],
        "eway_bill_expired_vouchers": eway_watch["expired"][:100],
        "e_invoice_payload_ready_ids": e_invoice_payload_ready_ids,
        "e_way_bill_payload_ready_ids": e_way_bill_payload_ready_ids,
        "matched_2b_entries": matched_2b[:100],
        "missing_in_books_entries": missing_in_books[:100],
        "pending_or_rejected_entries": pending_or_rejected[:100],
        "missing_in_portal_vouchers": missing_in_portal[:100],
        "draft_gst_vouchers_list": draft_gst_vouchers[:100],
        "filing_list": filings,
        "open_task_list": open_tasks[:100],
        "notice_list": notices[:100],
        "exception_items": exception_items,
        "existing_task_refs": existing_task_refs,
        "document_request_map": document_request_map,
        "document_requests": document_requests,
    }


@login_required
def dashboard(request):
    """Main dashboard with key metrics and cash flow chart for the current company."""
    company = request.current_company

    # 1. Basic Stats
    total_vouchers = Voucher.objects.filter(company=company).count()
    total_ledgers  = Ledger.objects.filter(company=company).count()

    from ocr.models import OCRSubmission
    ocr_pending = OCRSubmission.objects.filter(
        company=company, status=OCRSubmission.STATUS_PENDING
    ).count()

    # 2. Optimized Ledger Summaries (Single Query)
    from django.db.models import Count, Sum, Q
    
    # We calculate net balances for each group in one pass over VoucherItems
    # Balance = Opening + CreditSum - DebitSum
    group_stats = Ledger.objects.filter(company=company).values('account_group__nature').annotate(
        opening=Sum('opening_balance'),
        total_dr=Sum(
            'voucher_items__amount',
            filter=Q(
                voucher_items__entry_type='DR',
                voucher_items__voucher__status='APPROVED',
            ),
        ),
        total_cr=Sum(
            'voucher_items__amount',
            filter=Q(
                voucher_items__entry_type='CR',
                voucher_items__voucher__status='APPROVED',
            ),
        ),
    )
    
    # NATURE_CHOICES mapping from AccountGroup
    natures = ["Asset", "Liability", "Income", "Expense", "Equity", "Tax"]
    summary_map = {nature: 0.0 for nature in natures}

    for stat in group_stats:
        nature = stat['account_group__nature']
        if not nature: continue
        opening = stat['opening'] or Decimal("0.00")
        dr = stat['total_dr'] or Decimal("0.00")
        cr = stat['total_cr'] or Decimal("0.00")
        
        # Accounting logic: Net Credit balance
        net = opening + cr - dr
        summary_map[nature] = float(net)

    # 3. KPI Calculations
    # Cash/Bank: Asset group + Name filter
    cash_ledgers = Ledger.objects.filter(
        company=company, 
        account_group__nature="Asset", 
        name__iregex=r"cash|bank"
    )
    cash_balance = sum(Decimal("0.00") - l.current_balance() for l in cash_ledgers)
    
    # Receivables (Asset group net Dr)
    receivables = abs(min(summary_map.get("Asset", 0.0), 0.0))
    # Payables (Liability group net Cr)
    payables = max(summary_map.get("Liability", 0.0), 0.0)
    # Profit (Income - Expense)
    # Income is Cr+, Expense is Dr+ (stored as net Cr-)
    net_profit = summary_map.get("Income", 0.0) + summary_map.get("Expense", 0.0)

    # 4. Cash flow chart
    cf = get_monthly_cash_flow(company, months=12)
    has_cf_data = any(cf["inflow"]) or any(cf["outflow"])

    # 5. Compliance Health logic
    from vouchers import compliance
    issues = compliance.get_compliance_issues(company)
    danger_count = len([i for i in issues if i['level'] == 'danger'])
    warning_count = len([i for i in issues if i['level'] == 'warning'])
    health_score = max(0, 100 - (danger_count * 10) - (warning_count * 3))
    svg_offset = 440 * (1 - (health_score / 100))

    # 6. Additional Module Metrics
    from orders.models import Order
    from inventory.models import StockItem
    from payroll.models import Employee
    from fixedassets.models import FixedAsset

    # Orders Summary
    total_orders = Order.objects.filter(company=company).count()
    open_orders = Order.objects.filter(company=company, status__in=['Draft', 'Confirmed', 'Partially Fulfilled']).count()
    
    # Inventory Summary
    stock_items = StockItem.objects.filter(company=company, is_active=True)
    low_stock_count = sum(1 for item in stock_items if item.is_low_stock())
    total_stock_value = sum(item.closing_stock_value() for item in stock_items)

    # Payroll Summary
    active_employees = Employee.objects.filter(company=company, is_active=True).count()
    
    # Fixed Assets
    total_assets = FixedAsset.objects.filter(company=company).count()

    today = _date.today()
    due_soon = today + timedelta(days=7)

    # 7. CA practice command metrics
    open_task_qs = PracticeTask.objects.filter(company=company).exclude(
        status__in=[PracticeTask.STATUS_DONE, PracticeTask.STATUS_CANCELLED]
    )
    task_counts = {
        "open": open_task_qs.count(),
        "overdue": open_task_qs.filter(due_date__lt=today).count(),
        "due_soon": open_task_qs.filter(due_date__gte=today, due_date__lte=due_soon).count(),
        "blocked": open_task_qs.filter(status=PracticeTask.STATUS_BLOCKED).count(),
        "critical": open_task_qs.filter(priority=PracticeTask.PRIORITY_CRITICAL).count(),
    }
    task_type_labels = dict(PracticeTask.TASK_TYPE_CHOICES)
    task_type_total = task_counts["open"] or 1
    task_type_rows = [
        {
            "label": task_type_labels.get(row["task_type"], row["task_type"]),
            "count": row["count"],
            "percent": round((row["count"] / task_type_total) * 100),
        }
        for row in open_task_qs.values("task_type").annotate(count=Count("id")).order_by("-count", "task_type")[:6]
    ]
    priority_tasks = open_task_qs.select_related("assigned_to").order_by("due_date", "-updated_at")[:6]

    open_filings_qs = ComplianceFiling.objects.filter(company=company).exclude(
        status__in=[ComplianceFiling.STATUS_FILED, ComplianceFiling.STATUS_CANCELLED]
    )
    gst_filing_types = [
        ComplianceFiling.TYPE_GSTR1,
        ComplianceFiling.TYPE_GSTR3B,
        ComplianceFiling.TYPE_GSTR9,
        ComplianceFiling.TYPE_GSTR9C,
        ComplianceFiling.TYPE_GST_IMS,
    ]
    tds_filing_types = [
        ComplianceFiling.TYPE_TDS_PAYMENT,
        ComplianceFiling.TYPE_TDS_24Q,
        ComplianceFiling.TYPE_TDS_26Q,
        ComplianceFiling.TYPE_TDS_27Q,
        ComplianceFiling.TYPE_FORM16,
    ]
    filing_counts = {
        "open": open_filings_qs.count(),
        "overdue": open_filings_qs.filter(due_date__lt=today).count(),
        "due_soon": open_filings_qs.filter(due_date__gte=today, due_date__lte=due_soon).count(),
        "ready_for_review": open_filings_qs.filter(status=ComplianceFiling.STATUS_READY_FOR_REVIEW).count(),
        "client_pending": open_filings_qs.filter(status=ComplianceFiling.STATUS_CLIENT_PENDING).count(),
        "blocked": open_filings_qs.filter(status=ComplianceFiling.STATUS_BLOCKED).count(),
        "gst": open_filings_qs.filter(filing_type__in=gst_filing_types).count(),
        "tds": open_filings_qs.filter(filing_type__in=tds_filing_types).count(),
    }
    upcoming_filings = open_filings_qs.select_related("assigned_to", "reviewer").order_by("due_date", "-updated_at")[:6]

    open_notices_qs = ComplianceNotice.objects.filter(company=company).exclude(status=ComplianceNotice.STATUS_CLOSED)
    notice_counts = {
        "open": open_notices_qs.count(),
        "overdue": open_notices_qs.filter(response_due_date__lt=today).count(),
        "due_soon": open_notices_qs.filter(response_due_date__gte=today, response_due_date__lte=due_soon).count(),
        "escalated": open_notices_qs.filter(status=ComplianceNotice.STATUS_ESCALATED).count(),
        "data_pending": open_notices_qs.filter(status=ComplianceNotice.STATUS_DATA_PENDING).count(),
    }

    gst_review_qs = GSTPeriodReview.objects.filter(company=company).exclude(status=GSTPeriodReview.STATUS_SIGNED_OFF)
    filing_review_qs = FilingReview.objects.filter(company=company).exclude(status=FilingReview.STATUS_APPROVED)
    open_post_filing_qs = GSTPostFilingTracker.objects.filter(company=company).filter(
        Q(gstr1_status__in=[
            GSTPostFilingTracker.STATUS_NOT_CHECKED,
            GSTPostFilingTracker.STATUS_PENDING,
            GSTPostFilingTracker.STATUS_UNDER_NOTICE,
        ])
        | Q(gstr3b_status__in=[
            GSTPostFilingTracker.STATUS_NOT_CHECKED,
            GSTPostFilingTracker.STATUS_PENDING,
            GSTPostFilingTracker.STATUS_UNDER_NOTICE,
        ])
        | Q(ims_status=GSTPostFilingTracker.IMS_EXCEPTIONS)
        | Q(payment_status__in=[
            GSTPostFilingTracker.PAYMENT_PENDING,
            GSTPostFilingTracker.PAYMENT_SHORT_PAID,
        ])
    )
    gst_command = {
        "period_reviews_open": gst_review_qs.count(),
        "filing_reviews_pending": filing_review_qs.count(),
        "packs_ready": GSTFilingPack.objects.filter(company=company, status=GSTFilingPack.STATUS_READY).count(),
        "post_filing_open": open_post_filing_qs.count(),
        "itc_at_risk": open_post_filing_qs.aggregate(total=Sum("itc_at_risk"))["total"] or Decimal("0.00"),
    }

    try:
        from portal.models import ClientDocumentRequest
    except ImportError:
        ClientDocumentRequest = None

    client_requests = {"open": 0, "overdue": 0, "uploaded": 0, "due_soon": 0}
    recent_client_requests = []
    if ClientDocumentRequest:
        client_request_qs = ClientDocumentRequest.objects.filter(company=company).exclude(
            status=ClientDocumentRequest.STATUS_CANCELLED
        )
        open_client_request_qs = client_request_qs.filter(status=ClientDocumentRequest.STATUS_OPEN)
        client_requests = {
            "open": open_client_request_qs.count(),
            "overdue": open_client_request_qs.filter(due_date__lt=today).count(),
            "due_soon": open_client_request_qs.filter(due_date__gte=today, due_date__lte=due_soon).count(),
            "uploaded": client_request_qs.filter(status=ClientDocumentRequest.STATUS_UPLOADED).count(),
        }
        recent_client_requests = client_request_qs.select_related("portal_user").order_by("status", "due_date", "-updated_at")[:5]

    try:
        statutory_profile = company.statutory_profile
    except CompanyStatutoryProfile.DoesNotExist:
        statutory_profile = None

    readiness_items = [
        {
            "label": "GST profile",
            "ready": bool(company.gstin) or bool(statutory_profile and not statutory_profile.gst_registered),
            "detail": company.gstin or "GSTIN pending",
        },
        {
            "label": "TDS/TRACES profile",
            "ready": bool(company.tan) or bool(statutory_profile and not statutory_profile.tds_applicable),
            "detail": company.tan or "TAN pending",
        },
        {
            "label": "Statutory calendar",
            "ready": bool(statutory_profile),
            "detail": "GST/TDS rules configured" if statutory_profile else "Rules not configured",
        },
        {
            "label": "WhatsApp intake",
            "ready": bool(company.whatsapp_intake_number),
            "detail": company.whatsapp_intake_number or "Client intake number pending",
        },
        {
            "label": "Invoice email",
            "ready": bool(company.invoice_email_from_address),
            "detail": company.invoice_email_from_address or "Sender email pending",
        },
        {
            "label": "Client portal",
            "ready": bool(company.portal_token),
            "detail": "Upload link active" if company.portal_token else "Portal token pending",
        },
        {
            "label": "Bank/payment profile",
            "ready": any([company.bank_name, company.account_number, company.ifsc_code, company.upi_id]),
            "detail": company.bank_name or company.upi_id or "Payment details pending",
        },
    ]
    readiness_ready = sum(1 for item in readiness_items if item["ready"])
    readiness_score = round((readiness_ready / len(readiness_items)) * 100)
    setup_readiness = {
        "score": readiness_score,
        "ring_offset": round(314 * (1 - (readiness_score / 100)), 2),
        "ready": readiness_ready,
        "total": len(readiness_items),
        "gaps": len(readiness_items) - readiness_ready,
        "items": readiness_items,
    }

    try:
        engagement = company.engagement
    except ClientEngagement.DoesNotExist:
        engagement = None

    dashboard_alerts = []

    def add_alert(level, title, detail, url, icon, count=None):
        dashboard_alerts.append({
            "level": level,
            "title": title,
            "detail": detail,
            "url": url,
            "icon": icon,
            "count": count,
        })

    if setup_readiness["gaps"]:
        add_alert(
            "warning",
            "Company setup has open gaps",
            f"{setup_readiness['gaps']} setup item(s) need attention before client self-service feels complete.",
            reverse("core:setup_wizard"),
            "bi-sliders",
            setup_readiness["gaps"],
        )
    if danger_count:
        add_alert(
            "danger",
            "Compliance exceptions need review",
            f"{danger_count} high-risk voucher or master-data issue(s) are affecting the health score.",
            reverse("core:compliance_health"),
            "bi-shield-exclamation",
            danger_count,
        )
    if task_counts["overdue"] or task_counts["critical"] or task_counts["blocked"]:
        add_alert(
            "danger" if task_counts["overdue"] or task_counts["blocked"] else "warning",
            "Practice work queue is hot",
            f"{task_counts['overdue']} overdue, {task_counts['blocked']} blocked, {task_counts['critical']} critical task(s).",
            reverse("core:practice_tasks"),
            "bi-kanban",
            task_counts["open"],
        )
    if filing_counts["overdue"] or filing_counts["blocked"]:
        add_alert(
            "danger",
            "Statutory filings are blocked or overdue",
            f"{filing_counts['overdue']} overdue and {filing_counts['blocked']} blocked filing(s).",
            reverse("core:compliance_filings"),
            "bi-calendar2-x",
            filing_counts["overdue"] + filing_counts["blocked"],
        )
    elif filing_counts["due_soon"] or filing_counts["ready_for_review"]:
        add_alert(
            "warning",
            "Filing review window is active",
            f"{filing_counts['due_soon']} due in 7 days and {filing_counts['ready_for_review']} ready for review.",
            reverse("core:filing_review_center"),
            "bi-clipboard-check",
            filing_counts["due_soon"] + filing_counts["ready_for_review"],
        )
    if notice_counts["overdue"] or notice_counts["escalated"]:
        add_alert(
            "danger",
            "Notice response risk",
            f"{notice_counts['overdue']} overdue and {notice_counts['escalated']} escalated notice(s).",
            reverse("core:compliance_notices"),
            "bi-envelope-exclamation",
            notice_counts["overdue"] + notice_counts["escalated"],
        )
    if client_requests["overdue"] or client_requests["uploaded"]:
        add_alert(
            "warning",
            "Client documents need action",
            f"{client_requests['overdue']} overdue request(s), {client_requests['uploaded']} uploaded item(s) waiting closure.",
            reverse("portal:client_requests"),
            "bi-inbox",
            client_requests["open"] + client_requests["uploaded"],
        )
    if not dashboard_alerts:
        add_alert(
            "success",
            "No urgent operating blockers",
            "Setup, filings, notices, client requests, and practice work are clear for this company.",
            reverse("core:ca_command_center"),
            "bi-check2-circle",
        )

    context = {
        "company": company,
        "total_vouchers":  total_vouchers,
        "total_ledgers":   total_ledgers,
        "ocr_pending":     ocr_pending,
        "recent_vouchers": Voucher.objects.filter(company=company).order_by("-date", "-id")[:8],
        "ledger_summary":  summary_map,
        "kpis": {
            "cash": cash_balance,
            "receivables": receivables,
            "payables": payables,
            "profit": net_profit,
        },
        "module_stats": {
            "orders": {"total": total_orders, "open": open_orders},
            "inventory": {"low_stock": low_stock_count, "value": total_stock_value},
            "payroll": {"employees": active_employees},
            "assets": {"total": total_assets},
        },
        "practice": {
            "tasks": task_counts,
            "task_type_rows": task_type_rows,
            "priority_tasks": priority_tasks,
        },
        "statutory": {
            "filings": filing_counts,
            "notices": notice_counts,
            "upcoming_filings": upcoming_filings,
            "gst": gst_command,
        },
        "client_requests": client_requests,
        "recent_client_requests": recent_client_requests,
        "setup_readiness": setup_readiness,
        "engagement": engagement,
        "dashboard_alerts": dashboard_alerts[:6],
        "compliance": {
            "issues": issues[:5], # Just show top 5 on dashboard
            "danger_count": danger_count,
            "warning_count": warning_count,
            "health_score": health_score,
            "svg_offset": svg_offset,
        },
        "today":           _date.today(),
        "due_soon":        due_soon,
        "cf_labels":       json.dumps(cf["labels"]),
        "cf_inflow":       json.dumps(cf["inflow"]),
        "cf_outflow":      json.dumps(cf["outflow"]),
        "has_cf_data":     has_cf_data,
    }
    return render(request, "dashboard.html", context)


@login_required
def setup_wizard(request):
    """Guided first-run checklist for getting a company production-ready."""
    company = request.current_company
    can_write = getattr(request, "current_company_role", None) != "Viewer"

    if request.method == "POST":
        action = request.POST.get("action")
        if not can_write:
            messages.error(request, "You do not have permission to update setup for this company.")
            return redirect("core:setup_wizard")

        if action == "create_statutory_profile":
            _profile, created = CompanyStatutoryProfile.objects.get_or_create(company=company)
            if created:
                messages.success(request, "Default statutory profile created. Review GST and TDS rules before filing.")
            else:
                messages.info(request, "Statutory profile already exists.")
            return redirect("core:setup_wizard")

        if action == "create_portal_token":
            if not company.portal_token:
                company.portal_token = secrets.token_urlsafe(32)
                company.save(update_fields=["portal_token"])
                messages.success(request, "Client portal upload token created.")
            else:
                messages.info(request, "Client portal upload token already exists.")
            return redirect("core:setup_wizard")

        messages.error(request, "Unknown setup action.")
        return redirect("core:setup_wizard")

    try:
        statutory_profile = company.statutory_profile
    except CompanyStatutoryProfile.DoesNotExist:
        statutory_profile = None

    setup_steps = [
        {
            "key": "identity",
            "title": "Company identity",
            "icon": "bi-building",
            "ready": bool(company.gstin) and bool(company.tan),
            "detail": "GSTIN and TAN are needed for GST, TDS, TRACES, and statutory exports.",
            "status": f"GSTIN: {company.gstin or 'Pending'} | TAN: {company.tan or 'Pending'}",
            "action_label": "Open Company Settings",
            "action_url": reverse("core:company_settings"),
        },
        {
            "key": "statutory",
            "title": "Statutory calendar rules",
            "icon": "bi-calendar2-week",
            "ready": bool(statutory_profile),
            "detail": "Configure GST frequency, TDS applicability, due-day rules, MSME watch, and filing assumptions.",
            "status": "GST/TDS rules configured" if statutory_profile else "Default statutory profile not created",
            "action_label": "Create Default Profile",
            "post_action": "create_statutory_profile",
        },
        {
            "key": "whatsapp",
            "title": "WhatsApp document intake",
            "icon": "bi-whatsapp",
            "ready": bool(company.whatsapp_intake_number),
            "detail": "Clients need one clear WhatsApp number for document submission and request follow-ups.",
            "status": company.whatsapp_intake_number or "Client WhatsApp intake number pending",
            "action_label": "Open App Settings",
            "action_url": reverse("core:app_settings"),
        },
        {
            "key": "email",
            "title": "Invoice and reminder email",
            "icon": "bi-envelope-at",
            "ready": bool(company.invoice_email_from_address),
            "detail": "Set invoice sender, reply-to, subject, and body before one-click invoice email goes live.",
            "status": company.invoice_email_from_address or "Invoice sender email pending",
            "action_label": "Open App Settings",
            "action_url": reverse("core:app_settings"),
        },
        {
            "key": "portal",
            "title": "Client portal upload link",
            "icon": "bi-person-lines-fill",
            "ready": bool(company.portal_token),
            "detail": "Create the secure upload token used by client request rooms and external upload links.",
            "status": "Portal token active" if company.portal_token else "Portal upload token pending",
            "action_label": "Create Portal Token",
            "post_action": "create_portal_token",
        },
        {
            "key": "bank",
            "title": "Bank and payment profile",
            "icon": "bi-bank",
            "ready": any([company.bank_name, company.account_number, company.ifsc_code, company.upi_id]),
            "detail": "Bank, IFSC, account, and UPI details power invoices, QR payment, and payment communication.",
            "status": company.bank_name or company.upi_id or "Payment details pending",
            "action_label": "Open Company Settings",
            "action_url": reverse("core:company_settings"),
        },
        {
            "key": "go_live",
            "title": "Production trust checks",
            "icon": "bi-shield-lock",
            "ready": True,
            "detail": "Review security, observability, backup discipline, go-live evidence, and audit trail before rollout.",
            "status": "Production trust center available",
            "action_label": "Open Trust Center",
            "action_url": reverse("core:production_trust_center"),
        },
    ]

    ready_count = sum(1 for step in setup_steps if step["ready"])
    score = round((ready_count / len(setup_steps)) * 100)

    return render(request, "core/setup_wizard.html", {
        "company": company,
        "setup_steps": setup_steps,
        "ready_count": ready_count,
        "total_steps": len(setup_steps),
        "score": score,
        "can_write": can_write,
        "title": "Setup Wizard",
    })


@login_required
def demo_workspace(request):
    """Sales demo mode with one-click rich sample data for CA partner demos."""
    if request.method == "POST":
        result = seed_demo_workspace(user=request.user)
        request.session["current_company_id"] = result.primary_company.pk
        messages.success(
            request,
            "Demo workspace is ready. The demo company is now selected.",
        )
        return redirect("core:dashboard")

    snapshot = demo_workspace_snapshot()
    has_workspace = bool(snapshot.primary_company)
    has_access = bool(
        snapshot.primary_company
        and UserCompanyAccess.objects.filter(
            user=request.user,
            company=snapshot.primary_company,
        ).exists()
    )
    demo_stats = [
        {"label": "Demo companies", "value": snapshot.counts["companies"], "icon": "bi-buildings"},
        {"label": "Ledgers", "value": snapshot.counts["ledgers"], "icon": "bi-journal-text"},
        {"label": "Vouchers", "value": snapshot.counts["vouchers"], "icon": "bi-receipt"},
        {"label": "Filings", "value": snapshot.counts["filings"], "icon": "bi-calendar-check"},
        {"label": "Client requests", "value": snapshot.counts["document_requests"], "icon": "bi-inbox"},
        {"label": "Market proof", "value": snapshot.counts["market_proof"], "icon": "bi-award"},
    ]
    demo_stories = [
        {
            "title": "CA command story",
            "icon": "bi-kanban",
            "detail": "Open tasks, overdue client requests, due filings, notices, and partner review signals across multiple clients.",
        },
        {
            "title": "Accounting story",
            "icon": "bi-calculator",
            "detail": "Approved invoices, purchases, receipts, payments, journals, GST tax lines, TDS, outstanding balances, and cash movement.",
        },
        {
            "title": "Client portal story",
            "icon": "bi-person-lines-fill",
            "detail": "WhatsApp intake, invoice email settings, secure portal token, open requests, overdue requests, and uploaded client documents.",
        },
        {
            "title": "Sales proof story",
            "icon": "bi-graph-up-arrow",
            "detail": "Pilot feedback, adoption evidence, case studies, market proof records, and Tally-exit talking points.",
        },
    ]
    demo_route = [
        {
            "step": "01",
            "title": "Open the operating snapshot",
            "time": "2 min",
            "screen": "Dashboard",
            "url": reverse("core:dashboard"),
            "talk_track": "Show that the CA sees setup gaps, filing pressure, notices, collections, and client requests without opening ten modules.",
        },
        {
            "step": "02",
            "title": "Move into daily CA work",
            "time": "3 min",
            "screen": "CA Command Center",
            "url": reverse("core:ca_command_center"),
            "talk_track": "Position this as the replacement for scattered Excel trackers, WhatsApp follow-ups, and partner status calls.",
        },
        {
            "step": "03",
            "title": "Show GST execution depth",
            "time": "4 min",
            "screen": "GST Filing Pack",
            "url": reverse("core:gst_filing_pack"),
            "talk_track": "Walk through filing readiness, review, pack generation, post-filing closure, and evidence discipline.",
        },
        {
            "step": "04",
            "title": "Show client document chase",
            "time": "3 min",
            "screen": "Client Requests",
            "url": reverse("portal:client_requests"),
            "talk_track": "Show how clients can submit documents through portal, WhatsApp, and email-led workflows without staff-only admin access.",
        },
        {
            "step": "05",
            "title": "Close with proof",
            "time": "3 min",
            "screen": "Market Proof",
            "url": reverse("core:market_proof_pack"),
            "talk_track": "Use pilot evidence, case studies, readiness scores, and external proof to make the sales discussion defensible.",
        },
    ]
    objection_cards = [
        {
            "title": "We already use Tally",
            "icon": "bi-box-arrow-right",
            "answer": "Keep Tally exit controlled: import, validate, reconcile, run in parallel, and then shift CA operations into one command layer.",
            "url": reverse("migration:exit_control"),
            "cta": "Open Tally Exit",
        },
        {
            "title": "Clients will not upload documents",
            "icon": "bi-inbox",
            "answer": "Use requests, reminders, WhatsApp intake, email identity, and portal health to make client chasing visible and owned.",
            "url": reverse("core:client_portal_health"),
            "cta": "Open Portal Health",
        },
        {
            "title": "How do we trust filings?",
            "icon": "bi-shield-check",
            "answer": "Use filing review, GST packs, post-filing ARN tracking, evidence vault, and audit logs before partner signoff.",
            "url": reverse("core:filing_review_center"),
            "cta": "Open Review Center",
        },
    ]
    close_checklist = [
        "Demo company selected before the call",
        "Dashboard alerts visible",
        "Client requests and reminders populated",
        "GST filing pack has ready data",
        "Market proof has evidence records",
    ]

    return render(
        request,
        "core/demo_workspace.html",
        {
            "title": "Demo Mode",
            "has_workspace": has_workspace,
            "has_access": has_access,
            "primary_company": snapshot.primary_company,
            "demo_companies": snapshot.companies,
            "demo_stats": demo_stats,
            "demo_stories": demo_stories,
            "demo_route": demo_route,
            "objection_cards": objection_cards,
            "close_checklist": close_checklist,
        },
    )


@login_required
def audit_log(request):
    """
    Audit Trail — paginated, filterable list of all AuditLog entries for the
    current company.  Accessible to all roles (read-only).
    """
    company = request.current_company

    qs = AuditLog.objects.filter(company=company).select_related("user")

    # Filters
    action_filter     = request.GET.get("action", "").strip()
    model_filter      = request.GET.get("model", "").strip()
    user_filter       = request.GET.get("user", "").strip()

    if action_filter:
        qs = qs.filter(action=action_filter)
    if model_filter:
        qs = qs.filter(model_name__icontains=model_filter)
    if user_filter:
        qs = qs.filter(user__email__icontains=user_filter)

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="audit-trail-{company.short_code or company.pk}-{timezone.localdate():%Y%m%d}.csv"'
        )
        writer = csv.writer(response)
        writer.writerow([
            "Timestamp",
            "Action",
            "Model",
            "Record ID",
            "Object",
            "User",
            "Old Data",
            "New Data",
        ])
        for log in qs.order_by("timestamp", "pk"):
            writer.writerow([
                timezone.localtime(log.timestamp).isoformat(),
                log.get_action_display(),
                log.model_name,
                log.record_id,
                log.object_repr,
                log.user.email if log.user else "System",
                json.dumps(log.old_data or {}, default=str, sort_keys=True),
                json.dumps(log.new_data or {}, default=str, sort_keys=True),
            ])
        return response

    # Distinct model names for the filter dropdown
    model_names = (
        AuditLog.objects.filter(company=company)
        .values_list("model_name", flat=True)
        .distinct()
        .order_by("model_name")
    )

    paginator = Paginator(qs, 50)
    page_obj  = paginator.get_page(request.GET.get("page"))
    export_query = request.GET.copy()
    export_query["export"] = "csv"
    export_query.pop("page", None)

    return render(request, "core/audit_log.html", {
        "page_obj":      page_obj,
        "action_filter": action_filter,
        "model_filter":  model_filter,
        "user_filter":   user_filter,
        "model_names":   model_names,
        "action_choices": AuditLog.ACTION_CHOICES,
        "total_count":   qs.count(),
        "export_query":  export_query.urlencode(),
    })


def _export_query(request):
    query = request.GET.copy()
    query["export"] = "csv"
    query.pop("page", None)
    return query.urlencode()


@login_required
def practice_task_list(request):
    companies = _companies_for_user(request.user)
    tasks = PracticeTask.objects.filter(company__in=companies).select_related(
        "company", "assigned_to"
    )

    status_filter = request.GET.get("status", "open")
    type_filter = request.GET.get("type", "")
    assigned_filter = request.GET.get("assigned", "")
    q = request.GET.get("q", "").strip()

    if status_filter == "open":
        tasks = tasks.exclude(status__in=[PracticeTask.STATUS_DONE, PracticeTask.STATUS_CANCELLED])
    elif status_filter:
        tasks = tasks.filter(status=status_filter)
    if type_filter:
        tasks = tasks.filter(task_type=type_filter)
    if assigned_filter == "me":
        tasks = tasks.filter(assigned_to=request.user)
    elif assigned_filter == "unassigned":
        tasks = tasks.filter(assigned_to__isnull=True)
    if q:
        tasks = tasks.filter(
            Q(title__icontains=q)
            | Q(company__name__icontains=q)
            | Q(reference__icontains=q)
            | Q(description__icontains=q)
        )

    today = timezone.localdate()
    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="practice_work_queue.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Client",
            "Task",
            "Type",
            "Priority",
            "Status",
            "Owner",
            "Due Date",
            "Days Overdue",
            "Period Start",
            "Period End",
            "Reference",
            "Description",
        ])
        for task in tasks:
            days_overdue = ""
            if task.due_date and task.is_open and task.due_date < today:
                days_overdue = (today - task.due_date).days
            writer.writerow([
                task.company.name,
                task.title,
                task.get_task_type_display(),
                task.get_priority_display(),
                task.get_status_display(),
                task.assigned_to.email if task.assigned_to else "",
                task.due_date.isoformat() if task.due_date else "",
                days_overdue,
                task.period_start.isoformat() if task.period_start else "",
                task.period_end.isoformat() if task.period_end else "",
                task.reference,
                task.description,
            ])
        return response

    summary = {
        "open": PracticeTask.objects.filter(company__in=companies)
        .exclude(status__in=[PracticeTask.STATUS_DONE, PracticeTask.STATUS_CANCELLED])
        .count(),
        "overdue": PracticeTask.objects.filter(company__in=companies, due_date__lt=today)
        .exclude(status__in=[PracticeTask.STATUS_DONE, PracticeTask.STATUS_CANCELLED])
        .count(),
        "blocked": PracticeTask.objects.filter(company__in=companies, status=PracticeTask.STATUS_BLOCKED).count(),
        "critical": PracticeTask.objects.filter(company__in=companies, priority=PracticeTask.PRIORITY_CRITICAL)
        .exclude(status__in=[PracticeTask.STATUS_DONE, PracticeTask.STATUS_CANCELLED])
        .count(),
    }

    return render(request, "core/practice_task_list.html", {
        "tasks": tasks[:500],
        "summary": summary,
        "status_filter": status_filter,
        "type_filter": type_filter,
        "assigned_filter": assigned_filter,
        "q": q,
        "task_types": PracticeTask.TASK_TYPE_CHOICES,
        "status_choices": PracticeTask.STATUS_CHOICES,
        "today": today,
        "export_query": _export_query(request),
    })


@login_required
def practice_task_create(request):
    companies = _companies_for_user(request.user)
    initial = {}
    company_id = request.GET.get("company")
    if company_id and companies.filter(pk=company_id).exists():
        initial["company"] = company_id

    form = PracticeTaskForm(
        request.POST or None,
        companies=companies,
        users=_task_users_for_companies(companies),
        initial=initial,
    )
    if request.method == "POST" and form.is_valid():
        task = form.save(commit=False)
        if not _can_manage_company(request.user, task.company):
            messages.error(request, "You do not have permission to create tasks for this company.")
            return redirect("core:practice_tasks")
        task.created_by = request.user
        task.save()
        messages.success(request, "Practice task created.")
        return redirect("core:practice_tasks")

    return render(request, "core/practice_task_form.html", {"form": form, "title": "New Practice Task"})


@login_required
def practice_task_update(request, pk):
    companies = _companies_for_user(request.user)
    task = get_object_or_404(PracticeTask, pk=pk, company__in=companies)
    if not _can_manage_company(request.user, task.company):
        messages.error(request, "You do not have permission to update this task.")
        return redirect("core:practice_tasks")

    form = PracticeTaskForm(
        request.POST or None,
        instance=task,
        companies=companies,
        users=_task_users_for_companies(companies),
    )
    if request.method == "POST" and form.is_valid():
        updated = form.save(commit=False)
        if updated.status == PracticeTask.STATUS_DONE and not updated.completed_at:
            updated.completed_at = timezone.now()
            updated.completed_by = request.user
        elif updated.status != PracticeTask.STATUS_DONE:
            updated.completed_at = None
            updated.completed_by = None
        updated.save()
        messages.success(request, "Practice task updated.")
        return redirect("core:practice_tasks")

    return render(request, "core/practice_task_form.html", {"form": form, "task": task, "title": "Edit Practice Task"})


@login_required
@require_POST
def practice_task_set_status(request, pk):
    companies = _companies_for_user(request.user)
    task = get_object_or_404(PracticeTask, pk=pk, company__in=companies)
    if not _can_manage_company(request.user, task.company):
        messages.error(request, "You do not have permission to update this task.")
        return redirect("core:practice_tasks")

    status = request.POST.get("status")
    allowed = {choice[0] for choice in PracticeTask.STATUS_CHOICES}
    if status not in allowed:
        messages.error(request, "Invalid task status.")
        return redirect("core:practice_tasks")

    task.status = status
    update_fields = ["status", "updated_at"]
    if status == PracticeTask.STATUS_DONE:
        task.completed_at = timezone.now()
        task.completed_by = request.user
        update_fields += ["completed_at", "completed_by"]
    elif task.completed_at or task.completed_by_id:
        task.completed_at = None
        task.completed_by = None
        update_fields += ["completed_at", "completed_by"]
    task.save(update_fields=update_fields)
    messages.success(request, "Task status updated.")
    return redirect(request.META.get("HTTP_REFERER") or "core:practice_tasks")


@login_required
def accounting_close_workbench(request):
    companies = _companies_for_user(request.user)
    company = _resolve_selected_company(request, companies)
    if company is None:
        messages.error(request, "Create or assign a company before running the close workbench.")
        return redirect("core:select_company")

    period_start, period_end = _parse_month_period(request.POST.get("period") or request.GET.get("period"))
    period_value = period_start.strftime("%Y-%m")

    report = build_close_workbench(company, period_start, period_end)
    if request.method == "GET" and request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="accounting_close_{period_value}.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Client",
            "Period",
            "Close Score",
            "Close Status",
            "Check Code",
            "Check",
            "Severity",
            "Count",
            "Amount",
            "Description",
            "Action",
        ])
        for check in report["checks"]:
            writer.writerow([
                company.name,
                period_value,
                report["score"],
                report["close_status"],
                check.code,
                check.title,
                check.severity,
                check.count,
                check.amount,
                check.description,
                check.action_label,
            ])
        return response

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create_tasks":
            if not _can_manage_company(request.user, company):
                messages.error(request, "You do not have permission to create close tasks for this company.")
            elif report["issues"]:
                result = create_close_tasks(report, request.user)
                messages.success(
                    request,
                    f"Close tasks ready: {result['created']} created, {result['existing']} already existed.",
                )
            else:
                messages.info(request, "No close issues need tasks for this period.")
            return redirect(f"{reverse('core:accounting_close')}?period={period_value}&company={company.pk}")

    return render(request, "core/accounting_close_workbench.html", {
        "report": report,
        "companies": companies,
        "selected_company": company,
        "period_value": period_value,
        "period_start": period_start,
        "period_end": period_end,
        "export_query": urlencode({"period": period_value, "company": company.pk, "export": "csv"}),
        "title": "Accounting Close Workbench",
    })


@login_required
def filing_readiness(request):
    companies = _companies_for_user(request.user)
    company = _resolve_selected_company(request, companies)
    if company is None:
        messages.error(request, "Create or assign a company before running filing readiness.")
        return redirect("core:select_company")

    period_start, period_end = _parse_month_period(request.POST.get("period") or request.GET.get("period"))
    period_value = period_start.strftime("%Y-%m")
    report = build_filing_readiness(company, period_start, period_end)

    if request.method == "GET" and request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="filing_readiness_{period_value}.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Client",
            "Period",
            "Readiness Score",
            "Readiness Status",
            "Check Code",
            "Check",
            "Severity",
            "Count",
            "Amount",
            "Description",
            "Action",
        ])
        for check in report["checks"]:
            writer.writerow([
                company.name,
                period_value,
                report["score"],
                report["status"],
                check.code,
                check.title,
                check.severity,
                check.count,
                check.amount,
                check.description,
                check.action_label,
            ])
        return response

    if request.method == "POST":
        if not _can_manage_company(request.user, company):
            messages.error(request, "You do not have permission to update filing readiness for this company.")
            return redirect(f"{reverse('core:filing_readiness')}?period={period_value}&company={company.pk}")

        action = request.POST.get("action")
        if action == "create_tasks":
            if report["issues"]:
                result = create_filing_readiness_tasks(report, request.user)
                messages.success(
                    request,
                    f"Readiness tasks ready: {result['created']} created, {result['existing']} already existed.",
                )
            else:
                messages.info(request, "No filing readiness issues need tasks for this period.")
        elif action == "review_status":
            try:
                review = save_filing_readiness_review(
                    report,
                    request.user,
                    request.POST.get("status"),
                    request.POST.get("notes", ""),
                )
            except ValueError as exc:
                messages.error(request, str(exc))
            else:
                messages.success(request, f"Filing readiness marked {review.get_status_display()} for {company.name}.")
        return redirect(f"{reverse('core:filing_readiness')}?period={period_value}&company={company.pk}")

    return render(request, "core/filing_readiness.html", {
        "report": report,
        "companies": companies,
        "selected_company": company,
        "period_value": period_value,
        "period_start": period_start,
        "period_end": period_end,
        "review_status_choices": [
            (GSTPeriodReview.STATUS_IN_REVIEW, "In Review"),
            (GSTPeriodReview.STATUS_SIGNED_OFF, "Signed Off"),
            (GSTPeriodReview.STATUS_REOPENED, "Reopened"),
        ],
        "can_manage": _can_manage_company(request.user, company),
        "export_query": urlencode({"period": period_value, "company": company.pk, "export": "csv"}),
        "title": "Filing Readiness",
    })


@login_required
def filing_review_center(request):
    companies = _companies_for_user(request.user)
    if not companies.exists():
        messages.error(request, "Create or assign a company before running filing reviews.")
        return redirect("core:select_company")

    period_start, period_end = _parse_month_period(request.POST.get("period") or request.GET.get("period"))
    period_value = period_start.strftime("%Y-%m")
    review_type = normalise_review_type(request.POST.get("review_type") or request.GET.get("review_type"))
    selected_company = _resolve_selected_company(request, companies)

    if request.method == "POST":
        if selected_company is None:
            messages.error(request, "Select a company before updating a filing review.")
            return redirect("core:filing_review_center")
        if not _can_manage_company(request.user, selected_company):
            messages.error(request, "You do not have permission to update filing review for this company.")
            return redirect(_filing_review_center_url(selected_company, period_value, review_type))

        summary = build_filing_review(selected_company, period_start, period_end, review_type)
        action = request.POST.get("action", "")
        notes = request.POST.get("notes", "")
        try:
            if action == "start_review":
                review = start_review(summary, request.user, notes)
                messages.success(request, f"Review started for {selected_company.name}: {review.get_status_display()}.")
            elif action == "mark_reviewed":
                review = mark_reviewed(summary, request.user, notes)
                messages.success(request, f"Review marked {review.get_status_display()} for {selected_company.name}.")
            elif action == "send_back":
                review = send_back_review(summary, request.user, notes)
                messages.success(request, f"Review sent back for {selected_company.name}.")
            elif action == "approve":
                review, ready_count = approve_review(summary, request.user, notes)
                messages.success(
                    request,
                    f"{selected_company.name} approved for filing. {ready_count} GST filing workflows are ready for review.",
                )
            elif action == "reopen":
                review = reopen_review(summary, request.user, notes)
                messages.success(request, f"Review reopened for {selected_company.name}.")
            elif action == "waive":
                review = waive_blocker(
                    summary,
                    request.POST.get("code", ""),
                    request.user,
                    request.POST.get("waiver_note", ""),
                )
                messages.success(request, f"Blocker waiver saved for {selected_company.name}.")
            elif action == "unwaive":
                review = unwaive_blocker(summary, request.POST.get("code", ""), request.user)
                messages.success(request, f"Blocker waiver removed for {selected_company.name}.")
            elif action == "create_tasks":
                result = create_review_blocker_tasks(summary, request.user)
                messages.success(
                    request,
                    f"Review blocker tasks ready: {result['created']} created, {result['existing']} already existed.",
                )
            else:
                messages.error(request, "Invalid filing review action.")
        except ValueError as exc:
            messages.error(request, str(exc))
        return redirect(_filing_review_center_url(selected_company, period_value, review_type))

    rows = build_filing_review_rows(companies, period_start, period_end, review_type)
    if selected_company is None and rows:
        selected_company = rows[0]["company"]
    selected_summary = next(
        (row for row in rows if selected_company and row["company"].pk == selected_company.pk),
        None,
    )
    if selected_summary is None and selected_company:
        selected_summary = build_filing_review(selected_company, period_start, period_end, review_type)

    totals = {
        "companies": len(rows),
        "blocked": sum(1 for row in rows if row["unwaived_critical_count"]),
        "ready": sum(1 for row in rows if row["ready_to_file"]),
        "approved": sum(1 for row in rows if row["review"] and row["review"].is_approved),
        "critical": sum(row["unwaived_critical_count"] for row in rows),
        "warnings": sum(row["unwaived_warning_count"] for row in rows),
        "waived": sum(row["waived_count"] for row in rows),
    }

    return render(request, "core/filing_review_center.html", {
        "rows": rows,
        "selected": selected_summary,
        "companies": companies,
        "selected_company": selected_company,
        "period_value": period_value,
        "period_start": period_start,
        "period_end": period_end,
        "review_type": review_type,
        "review_type_choices": review_type_choices(),
        "totals": totals,
        "can_manage": bool(selected_company and _can_manage_company(request.user, selected_company)),
        "status_choices": FilingReview.STATUS_CHOICES,
        "title": "Filing Review Center",
    })


def _filing_review_center_url(company, period_value, review_type):
    return f"{reverse('core:filing_review_center')}?{urlencode({'period': period_value, 'company': company.pk, 'review_type': review_type})}"


@login_required
def statutory_export_center(request):
    from tds.models import TDSReturnWorkpaper
    from tds.workbench import fy_options as tds_fy_options

    params = request.POST if request.method == "POST" else request.GET
    companies = _companies_for_user(request.user)
    selected_company_id = (params.get("company") or "").strip()
    if selected_company_id and companies.filter(pk=selected_company_id).exists():
        scoped_companies = companies.filter(pk=selected_company_id)
    else:
        scoped_companies = companies
        selected_company_id = ""

    period_start, period_end = parse_gst_export_period(params.get("period"))
    period_value = period_start.strftime("%Y-%m")
    tds_filters = parse_tds_export_filters(params)
    center = build_statutory_export_center(scoped_companies, period_start, period_end, tds_filters)

    query_args = {
        "period": period_value,
        "fy": tds_filters["fy_start"],
        "quarter": tds_filters["quarter"],
        "form_type": tds_filters["form_type"],
    }
    if selected_company_id:
        query_args["company"] = selected_company_id
    base_query = urlencode(query_args)

    if request.method == "GET" and request.GET.get("export") == "csv":
        return statutory_export_csv_response(center, period_value, tds_filters)

    if request.method == "POST":
        action = request.POST.get("action", "")
        if action == "create_tasks":
            manageable_ids = set(
                _manageable_companies_for_user(request.user)
                .filter(pk__in=[row["company"].pk for row in center["rows"]])
                .values_list("pk", flat=True)
            )
            result = create_statutory_export_tasks(center["rows"], request.user, manageable_ids)
            messages.success(
                request,
                f"Statutory export tasks ready: {result['created']} created, {result['existing']} already existed.",
            )
        else:
            messages.error(request, "Invalid statutory export action.")
        return redirect(f"{reverse('core:statutory_export_center')}?{base_query}")

    return render(request, "core/statutory_export_center.html", {
        "rows": center["rows"],
        "totals": center["totals"],
        "companies": companies,
        "selected_company_id": selected_company_id,
        "period_value": period_value,
        "period_start": period_start,
        "period_end": period_end,
        "tds_filters": tds_filters,
        "fy_options": tds_fy_options(tds_filters["fy_start"]),
        "quarter_options": TDSReturnWorkpaper.QUARTER_CHOICES,
        "form_type_options": TDSReturnWorkpaper.FORM_TYPE_CHOICES,
        "base_query": base_query,
        "export_query": f"{base_query}&export=csv",
        "title": "Statutory Export Center",
    })


@login_required
def gst_filing_pack(request):
    companies = _companies_for_user(request.user)
    company = _resolve_selected_company(request, companies)
    if company is None:
        messages.error(request, "Create or assign a company before preparing a GST filing pack.")
        return redirect("core:select_company")

    period_start, period_end = _parse_month_period(request.POST.get("period") or request.GET.get("period"))
    period_value = period_start.strftime("%Y-%m")
    pack = build_gst_filing_pack(company, period_start, period_end)
    can_manage = _can_manage_company(request.user, company)

    if request.method == "POST":
        if not can_manage:
            messages.error(request, "You do not have permission to update GST filing pack for this company.")
            return redirect(_gst_filing_pack_url(company, period_value))

        action = request.POST.get("action", "")
        notes = request.POST.get("notes", "")
        try:
            if action == "generate_pack":
                record = save_gst_filing_pack(pack, request.user, notes)
                messages.success(request, f"GST filing pack generated for {company.name}: {record.get_status_display()}.")
            elif action == "mark_filed":
                if not pack["pack_record"]:
                    raise ValueError("Generate the filing pack before marking it filed.")
                record = mark_gst_filing_pack_filed(
                    pack["pack_record"],
                    request.user,
                    request.POST.get("arn_ack_number", ""),
                    notes,
                )
                messages.success(request, f"GST filing pack archived as filed with {record.arn_ack_number}.")
            elif action == "reopen_pack":
                if not pack["pack_record"]:
                    raise ValueError("No filing pack exists to reopen.")
                pack["pack_record"].status = GSTFilingPack.STATUS_REOPENED
                pack["pack_record"].filed_by = None
                pack["pack_record"].filed_at = None
                pack["pack_record"].arn_ack_number = ""
                pack["pack_record"].save()
                messages.success(request, "GST filing pack reopened.")
            else:
                messages.error(request, "Invalid GST filing pack action.")
        except ValueError as exc:
            messages.error(request, str(exc))
        return redirect(_gst_filing_pack_url(company, period_value))

    return render(request, "core/gst_filing_pack.html", {
        "pack": pack,
        "companies": companies,
        "selected_company": company,
        "period_value": period_value,
        "period_start": period_start,
        "period_end": period_end,
        "can_manage": can_manage,
        "title": "GST Filing Pack",
    })


@login_required
def gst_filing_pack_download(request, kind):
    companies = _companies_for_user(request.user)
    company = _resolve_selected_company(request, companies)
    if company is None:
        messages.error(request, "Select a company before downloading a GST filing pack.")
        return redirect("core:select_company")

    period_start, period_end = _parse_month_period(request.GET.get("period"))
    period_value = period_start.strftime("%Y-%m")
    pack = build_gst_filing_pack(company, period_start, period_end)
    if not pack["can_generate"]:
        messages.error(request, "Approve the review and clear critical validations before downloading the final filing pack.")
        return redirect(_gst_filing_pack_url(company, period_value))

    safe_name = "".join(ch if ch.isalnum() else "_" for ch in company.name).strip("_") or "company"
    if kind == "xlsx":
        payload = filing_pack_xlsx_bytes(pack)
        response = HttpResponse(
            payload,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="GST_Filing_Pack_{safe_name}_{period_value}.xlsx"'
        return response
    if kind == "json":
        response = HttpResponse(draft_json_bytes(pack), content_type="application/json")
        response["Content-Disposition"] = f'attachment; filename="GST_Filing_Draft_{safe_name}_{period_value}.json"'
        return response
    if kind == "gstr1":
        response = HttpResponse(portal_gstr1_json_bytes(pack), content_type="application/json")
        response["Content-Disposition"] = f'attachment; filename="GSTR1_Portal_{safe_name}_{period_value}.json"'
        return response

    raise Http404("Unknown GST filing pack download.")


def _gst_filing_pack_url(company, period_value):
    return f"{reverse('core:gst_filing_pack')}?{urlencode({'period': period_value, 'company': company.pk})}"


@login_required
def gst_post_filing_dashboard(request):
    period_start, period_end = _parse_month_period(request.GET.get("period"))
    period_value = period_start.strftime("%Y-%m")
    companies = _companies_for_user(request.user)
    selected_company_id = request.GET.get("company", "").strip()
    if selected_company_id and companies.filter(pk=selected_company_id).exists():
        companies = companies.filter(pk=selected_company_id)
    dashboard = build_gst_post_filing_dashboard(companies, period_start, period_end)

    return render(request, "core/gst_post_filing_dashboard.html", {
        "dashboard": dashboard,
        "rows": dashboard["rows"],
        "totals": dashboard["totals"],
        "period_start": period_start,
        "period_end": period_end,
        "period_value": period_value,
        "company_filter_options": _companies_for_user(request.user),
        "selected_company_id": selected_company_id,
        "title": "GST Post-Filing Dashboard",
    })


@login_required
def gst_post_filing(request):
    companies = _companies_for_user(request.user)
    company = _resolve_selected_company(request, companies)
    if company is None:
        messages.error(request, "Create or assign a company before tracking GST post-filing work.")
        return redirect("core:select_company")

    period_start, period_end = _parse_month_period(request.POST.get("period") or request.GET.get("period"))
    period_value = period_start.strftime("%Y-%m")
    can_manage = _can_manage_company(request.user, company)
    center = build_gst_post_filing_center(company, period_start, period_end)

    if request.method == "POST":
        if not can_manage:
            messages.error(request, "You do not have permission to update GST post-filing work for this company.")
            return redirect(_gst_post_filing_url(company, period_value))

        action = request.POST.get("action", "")
        try:
            if action == "save_tracker":
                tracker = save_gst_post_filing_tracker(company, period_start, period_end, request.user, request.POST)
                messages.success(request, f"GST post-filing tracker updated for {tracker.company.name}.")
            elif action == "create_notice":
                notice = create_gst_notice_from_post_filing(company, period_start, period_end, request.user, request.POST)
                messages.success(request, f"GST notice room opened: {notice.title}.")
            elif action == "upload_evidence":
                document = upload_gst_evidence(
                    company,
                    period_start,
                    period_end,
                    request.user,
                    request.POST,
                    request.FILES.get("evidence_file"),
                )
                messages.success(request, f"GST evidence uploaded: {document.title}.")
            elif action == "update_notice":
                notice = get_object_or_404(
                    ComplianceNotice,
                    pk=request.POST.get("notice_id"),
                    company=company,
                    notice_type=ComplianceNotice.TYPE_GST,
                )
                update_gst_notice_from_post_filing(notice, request.user, request.POST)
                messages.success(request, f"GST notice updated: {notice.title}.")
            else:
                messages.error(request, "Invalid GST post-filing action.")
        except ValueError as exc:
            messages.error(request, str(exc))
        return redirect(_gst_post_filing_url(company, period_value))

    return render(request, "core/gst_post_filing.html", {
        "center": center,
        "companies": companies,
        "selected_company": company,
        "period_start": period_start,
        "period_end": period_end,
        "period_value": period_value,
        "can_manage": can_manage,
        "return_status_choices": GSTPostFilingTracker.RETURN_STATUS_CHOICES,
        "ims_status_choices": GSTPostFilingTracker.IMS_STATUS_CHOICES,
        "payment_status_choices": GSTPostFilingTracker.PAYMENT_STATUS_CHOICES,
        "evidence_type_choices": GSTEvidenceDocument.EVIDENCE_TYPE_CHOICES,
        "return_type_choices": GSTEvidenceDocument.RETURN_TYPE_CHOICES,
        "notice_status_choices": ComplianceNotice.STATUS_CHOICES,
        "notice_priority_choices": PracticeTask.PRIORITY_CHOICES,
        "gst_filing_type_choices": [
            (ComplianceFiling.TYPE_GSTR1, "GSTR-1"),
            (ComplianceFiling.TYPE_GSTR3B, "GSTR-3B"),
            (ComplianceFiling.TYPE_GST_IMS, "GST IMS Review"),
        ],
        "today": timezone.localdate(),
        "title": "GST Post-Filing Center",
    })


def _gst_post_filing_url(company, period_value):
    return f"{reverse('core:gst_post_filing')}?{urlencode({'period': period_value, 'company': company.pk})}"


GST_CLIENT_CHASE_KINDS = {
    "2b_missing_in_books",
    "purchase_missing_in_2b",
    "draft_gst_voucher",
    "sales_missing_pos",
    "sales_invalid_party_gstin",
    "sales_tax_pos_mismatch",
    "sales_missing_hsn",
    "itc_180_reversal_due",
    "rcm_missing_tax",
    "eway_bill_missing",
    "eway_bill_expired",
    "open_gst_notice",
}


def _gst_exception_metadata(kind, obj, period_start, note=""):
    from gstr2b.models import PortalGSTR2BEntry
    from portal.models import ClientDocumentRequest

    due_date = timezone.localdate() + timedelta(days=2)
    priority = PracticeTask.PRIORITY_HIGH
    description = note.strip()

    if kind == "2b_missing_in_books" and isinstance(obj, PortalGSTR2BEntry):
        return {
            "reference": f"GSTEX:2B:{obj.pk}",
            "task_title": f"Resolve 2B missing in books: {obj.invoice_number}",
            "request_title": f"Upload purchase invoice: {obj.invoice_number}",
            "document_type": ClientDocumentRequest.TYPE_GST_INVOICE,
            "due_date": due_date,
            "priority": priority,
            "description": description or (
                f"Supplier {obj.supplier_name or obj.gstin}, GSTIN {obj.gstin}, "
                f"invoice {obj.invoice_number}, tax Rs.{obj.tax_amount}. Verify the bill and book it or reject it in IMS."
            ),
        }
    if kind == "purchase_missing_in_2b" and isinstance(obj, Voucher):
        return {
            "reference": f"GSTEX:VCH:{obj.pk}",
            "task_title": f"Resolve purchase missing in 2B: {obj.number or obj.pk}",
            "request_title": f"Confirm vendor filing for purchase {obj.number or obj.pk}",
            "document_type": ClientDocumentRequest.TYPE_GST_INVOICE,
            "due_date": due_date,
            "priority": priority,
            "description": description or (
                f"Approved purchase dated {obj.date:%d %b %Y} has unclaimed ITC for {period_start:%b %Y}. "
                "Ask for supplier filing confirmation or corrected invoice details."
            ),
        }
    if kind == "itc_180_reversal_due" and isinstance(obj, Voucher):
        return {
            "reference": f"GSTEX:ITC180:{obj.pk}",
            "task_title": f"Resolve ITC 180-day reversal: {obj.number or obj.pk}",
            "request_title": f"Upload payment proof for purchase {obj.number or obj.pk}",
            "document_type": ClientDocumentRequest.TYPE_OTHER,
            "due_date": due_date,
            "priority": PracticeTask.PRIORITY_CRITICAL,
            "description": description or (
                f"Purchase dated {obj.date:%d %b %Y} has claimed ITC of Rs.{obj.total_tax} "
                "with unpaid balance beyond the 180-day payment watch. Reverse proportionate ITC in GSTR-3B "
                "or upload payment proof and keep the evidence with the GST review."
            ),
        }
    if kind == "rcm_missing_tax" and isinstance(obj, Voucher):
        return {
            "reference": f"GSTEX:RCM:{obj.pk}",
            "task_title": f"Add RCM tax details for purchase {obj.number or obj.pk}",
            "request_title": f"Confirm RCM treatment for purchase {obj.number or obj.pk}",
            "document_type": ClientDocumentRequest.TYPE_GST_INVOICE,
            "due_date": due_date,
            "priority": PracticeTask.PRIORITY_CRITICAL,
            "description": description or (
                f"Purchase dated {obj.date:%d %b %Y} is marked reverse charge but has no GST tax amount. "
                "Add CGST/SGST/IGST liability details or remove RCM if it does not apply."
            ),
        }
    if kind == "eway_bill_missing" and isinstance(obj, Voucher):
        return {
            "reference": f"GSTEX:EWAY:{obj.pk}",
            "task_title": f"Add e-way bill details for {obj.voucher_type} {obj.number or obj.pk}",
            "request_title": f"Upload e-way bill for {obj.voucher_type} {obj.number or obj.pk}",
            "document_type": ClientDocumentRequest.TYPE_GST_INVOICE,
            "due_date": due_date,
            "priority": PracticeTask.PRIORITY_HIGH,
            "description": description or (
                f"{obj.voucher_type} dated {obj.date:%d %b %Y} has movement details and value above Rs.50,000 "
                "but no e-way bill number. Add the e-way bill or document why it is not required."
            ),
        }
    if kind == "eway_bill_expired" and isinstance(obj, Voucher):
        return {
            "reference": f"GSTEX:EWAYEXP:{obj.pk}",
            "task_title": f"Review expired e-way bill for {obj.voucher_type} {obj.number or obj.pk}",
            "request_title": f"Upload updated e-way bill for {obj.voucher_type} {obj.number or obj.pk}",
            "document_type": ClientDocumentRequest.TYPE_GST_INVOICE,
            "due_date": due_date,
            "priority": PracticeTask.PRIORITY_HIGH,
            "description": description or (
                f"E-way bill {obj.e_way_bill_no} for {obj.voucher_type} {obj.number or obj.pk} has expired validity. "
                "Review extension, cancellation, or movement evidence."
            ),
        }
    if kind == "draft_gst_voucher" and isinstance(obj, Voucher):
        return {
            "reference": f"GSTEX:DRAFT:{obj.pk}",
            "task_title": f"Approve draft GST voucher: {obj.voucher_type} {obj.number or obj.pk}",
            "request_title": f"Upload support for draft {obj.voucher_type} {obj.number or obj.pk}",
            "document_type": ClientDocumentRequest.TYPE_GST_INVOICE,
            "due_date": due_date,
            "priority": priority,
            "description": description or "GST-sensitive voucher must be approved before filing review. Ask for missing support or approval details.",
        }
    if kind == "e_invoice_missing_irn" and isinstance(obj, Voucher):
        return {
            "reference": f"GSTEX:EINV:{obj.pk}",
            "task_title": f"Generate IRN for {obj.voucher_type} {obj.number or obj.pk}",
            "request_title": f"Confirm IRN for {obj.voucher_type} {obj.number or obj.pk}",
            "document_type": ClientDocumentRequest.TYPE_GST_INVOICE,
            "due_date": due_date,
            "priority": PracticeTask.PRIORITY_CRITICAL,
            "description": description or (
                f"{obj.voucher_type} dated {obj.date:%d %b %Y} has no e-invoice IRN in books. "
                "Generate IRN through the configured GST provider or document why e-invoicing does not apply."
            ),
        }
    if kind == "sales_missing_pos" and isinstance(obj, Voucher):
        return {
            "reference": f"GSTEX:SALEPOS:{obj.pk}",
            "task_title": f"Add place of supply for {obj.voucher_type} {obj.number or obj.pk}",
            "request_title": f"Confirm place of supply for invoice {obj.number or obj.pk}",
            "document_type": ClientDocumentRequest.TYPE_GST_INVOICE,
            "due_date": due_date,
            "priority": PracticeTask.PRIORITY_HIGH,
            "description": description or (
                f"{obj.voucher_type} dated {obj.date:%d %b %Y} has GST tax but no place of supply. "
                "Confirm destination state and update the voucher before filing."
            ),
        }
    if kind == "sales_invalid_party_gstin" and isinstance(obj, Voucher):
        return {
            "reference": f"GSTEX:SALEGSTIN:{obj.pk}",
            "task_title": f"Correct customer GSTIN for {obj.voucher_type} {obj.number or obj.pk}",
            "request_title": f"Confirm customer GSTIN for invoice {obj.number or obj.pk}",
            "document_type": ClientDocumentRequest.TYPE_GST_INVOICE,
            "due_date": due_date,
            "priority": PracticeTask.PRIORITY_HIGH,
            "description": description or "Customer GSTIN on the sales ledger is invalid. Correct the party master before filing GSTR-1.",
        }
    if kind == "sales_tax_pos_mismatch" and isinstance(obj, Voucher):
        return {
            "reference": f"GSTEX:SALETAXPOS:{obj.pk}",
            "task_title": f"Review GST split vs POS for {obj.voucher_type} {obj.number or obj.pk}",
            "request_title": f"Confirm tax treatment for invoice {obj.number or obj.pk}",
            "document_type": ClientDocumentRequest.TYPE_GST_INVOICE,
            "due_date": due_date,
            "priority": priority,
            "description": description or "CGST/SGST/IGST split does not align with place of supply. Review tax treatment before filing.",
        }
    if kind == "sales_missing_hsn" and isinstance(obj, Voucher):
        return {
            "reference": f"GSTEX:SALEHSN:{obj.pk}",
            "task_title": f"Add HSN/SAC for sales invoice {obj.number or obj.pk}",
            "request_title": f"Confirm HSN/SAC for invoice {obj.number or obj.pk}",
            "document_type": ClientDocumentRequest.TYPE_GST_INVOICE,
            "due_date": due_date,
            "priority": PracticeTask.PRIORITY_HIGH,
            "description": description or (
                "One or more stock items on this sales invoice do not have HSN/SAC codes. "
                "Update the item master before preparing GSTR-1 HSN summary."
            ),
        }
    if kind == "overdue_gst_filing" and isinstance(obj, ComplianceFiling):
        return {
            "reference": f"GSTEX:FILING:{obj.pk}",
            "task_title": f"Complete overdue GST filing: {obj.title}",
            "request_title": f"GST filing support: {obj.title}",
            "document_type": ClientDocumentRequest.TYPE_OTHER,
            "due_date": obj.due_date or due_date,
            "priority": PracticeTask.PRIORITY_CRITICAL,
            "description": description or "Filing is overdue in the GST review packet.",
        }
    if kind == "open_gst_notice" and isinstance(obj, ComplianceNotice):
        notice_overdue = obj.response_due_date and obj.response_due_date < timezone.localdate()
        return {
            "reference": f"GSTEX:NOTICE:{obj.pk}",
            "task_title": f"Respond to GST notice: {obj.title}",
            "request_title": f"Upload evidence for GST notice: {obj.title}",
            "document_type": ClientDocumentRequest.TYPE_GST_NOTICE,
            "due_date": obj.response_due_date or due_date,
            "priority": PracticeTask.PRIORITY_CRITICAL if notice_overdue else priority,
            "description": description or obj.description or "Upload notice evidence, reply draft support, or acknowledgement details.",
        }
    return None


def _gst_exception_contact_defaults(company, kind, obj):
    from gstr2b.models import PortalGSTR2BEntry
    from portal.models import PortalUser

    ledger = None
    if kind == "2b_missing_in_books" and isinstance(obj, PortalGSTR2BEntry):
        ledger = Ledger.objects.filter(company=company, gstin=obj.gstin).order_by("name").first()
    elif kind in {
        "purchase_missing_in_2b",
        "draft_gst_voucher",
        "itc_180_reversal_due",
        "rcm_missing_tax",
        "eway_bill_missing",
        "eway_bill_expired",
        "e_invoice_missing_irn",
        "sales_missing_pos",
        "sales_invalid_party_gstin",
        "sales_tax_pos_mismatch",
        "sales_missing_hsn",
    } and isinstance(obj, Voucher):
        ledger_qs = (
            obj.items
            .select_related("ledger", "ledger__account_group")
            .filter(ledger__email__isnull=False)
            .exclude(ledger__email="")
        )
        if obj.voucher_type == "Purchase":
            ledger = ledger_qs.filter(ledger__account_group__nature="Liability").first() or ledger_qs.first()
        elif obj.voucher_type == "Sales":
            ledger = ledger_qs.filter(ledger__account_group__nature="Asset").first() or ledger_qs.first()
        else:
            ledger = ledger_qs.first()
        ledger = ledger.ledger if ledger else None

    portal_user = None
    recipient_email = ""
    if ledger:
        portal_user = PortalUser.objects.filter(linked_ledger=ledger, is_active=True).order_by("name").first()
        recipient_email = (portal_user.email if portal_user else "") or (ledger.email or "")

    return {
        "portal_user": portal_user,
        "recipient_email": recipient_email,
        "recipient_whatsapp_number": "",
    }


def _create_gst_exception_followup(company, user, period_start, period_end, kind, obj, action_type="resolve", note=""):
    from portal.models import ClientDocumentRequest

    metadata = _gst_exception_metadata(kind, obj, period_start, note=note)
    if not metadata:
        return {
            "created_task": False,
            "existing_task": False,
            "created_request": False,
            "existing_request": False,
            "skipped": True,
        }

    title = metadata["task_title"]
    description = metadata["description"]
    if action_type == "client_chase":
        title = f"Client chase: {title}"
        description = (description + "\n\nAsk the client for supporting invoice/document and upload evidence.").strip()

    task, task_created = PracticeTask.objects.get_or_create(
        company=company,
        reference=metadata["reference"],
        defaults={
            "title": title,
            "task_type": PracticeTask.TYPE_GST,
            "priority": metadata["priority"],
            "status": PracticeTask.STATUS_OPEN,
            "due_date": metadata["due_date"],
            "period_start": period_start,
            "period_end": period_end,
            "created_by": user,
            "description": description,
        },
    )
    if task_created:
        AuditLog.objects.create(
            company=company,
            user=user,
            action=AuditLog.ACTION_CREATE,
            model_name="PracticeTask",
            record_id=task.pk,
            object_repr=task.title[:200],
            old_data={},
            new_data={
                "reference": task.reference,
                "title": task.title,
                "task_type": task.task_type,
                "priority": task.priority,
                "due_date": task.due_date.isoformat() if task.due_date else "",
                "source": "gst_exception",
                "kind": kind,
                "action_type": action_type,
            },
        )

    request_created = False
    request_existing = False
    if action_type == "client_chase":
        contact = _gst_exception_contact_defaults(company, kind, obj)
        doc_request = ClientDocumentRequest.objects.filter(
            company=company,
            source_reference=metadata["reference"],
        ).order_by("-created_at").first()
        if doc_request:
            request_existing = True
            update_fields = []
            if not doc_request.related_task_id:
                doc_request.related_task = task
                update_fields.append("related_task")
            if contact["portal_user"] and not doc_request.portal_user_id:
                doc_request.portal_user = contact["portal_user"]
                update_fields.append("portal_user")
            if contact["recipient_email"] and not doc_request.recipient_email:
                doc_request.recipient_email = contact["recipient_email"]
                update_fields.append("recipient_email")
            if contact["recipient_whatsapp_number"] and not doc_request.recipient_whatsapp_number:
                doc_request.recipient_whatsapp_number = contact["recipient_whatsapp_number"]
                update_fields.append("recipient_whatsapp_number")
            if update_fields:
                doc_request.save(update_fields=[*update_fields, "updated_at"])
        else:
            ClientDocumentRequest.objects.create(
                company=company,
                portal_user=contact["portal_user"],
                recipient_email=contact["recipient_email"],
                recipient_whatsapp_number=contact["recipient_whatsapp_number"],
                source_reference=metadata["reference"],
                title=metadata["request_title"],
                document_type=metadata["document_type"],
                status=ClientDocumentRequest.STATUS_OPEN,
                due_date=metadata["due_date"],
                notes=description,
                related_task=task,
                requested_by=user,
            )
            request_created = True

    return {
        "created_task": task_created,
        "existing_task": not task_created,
        "created_request": request_created,
        "existing_request": request_existing,
        "skipped": False,
    }


@login_required
def gst_workbench(request):
    period_start, period_end = _parse_month_period(request.GET.get("period"))
    period_value = period_start.strftime("%Y-%m")
    companies = _companies_for_user(request.user)
    selected_company_id = request.GET.get("company", "").strip()
    if selected_company_id and companies.filter(pk=selected_company_id).exists():
        companies = companies.filter(pk=selected_company_id)

    rows = [
        _gst_workbench_snapshot(company, period_start, period_end)
        for company in companies
    ]
    rows.sort(key=lambda row: (-row["risk_score"], row["company"].name))

    totals = {
        "companies": len(rows),
        "risk_clients": sum(1 for row in rows if row["risk_score"] >= 50),
        "missing_in_books": sum(row["missing_in_books"] for row in rows),
        "missing_in_portal": sum(row["missing_in_portal"] for row in rows),
        "pending_2b": sum(row["pending_2b"] for row in rows),
        "overdue_filings": sum(row["overdue_filings"] for row in rows),
        "open_notices": sum(row["open_notices"] for row in rows),
        "client_chase_open": sum(row["client_chase_open"] for row in rows),
        "client_chase_uploaded": sum(row["client_chase_uploaded"] for row in rows),
        "client_chase_overdue": sum(row["client_chase_overdue"] for row in rows),
        "e_invoice_missing_irn": sum(row["e_invoice_missing_irn"] for row in rows),
        "e_invoice_expired": sum(row["e_invoice_expired"] for row in rows),
        "e_invoice_due_soon": sum(row["e_invoice_due_soon"] for row in rows),
        "sales_missing_pos": sum(row["sales_missing_pos"] for row in rows),
        "sales_invalid_party_gstin": sum(row["sales_invalid_party_gstin"] for row in rows),
        "sales_tax_pos_mismatch": sum(row["sales_tax_pos_mismatch"] for row in rows),
        "itc_180_overdue": sum(row["itc_180_overdue"] for row in rows),
        "itc_180_due_soon": sum(row["itc_180_due_soon"] for row in rows),
        "itc_180_reversal_due": sum(row["itc_180_reversal_due"] for row in rows),
        "itc_180_reversal_itc": sum((row["itc_180_reversal_itc"] for row in rows), Decimal("0.00")),
        "rcm_purchase_count": sum(row["rcm_purchase_count"] for row in rows),
        "rcm_missing_tax": sum(row["rcm_missing_tax"] for row in rows),
        "rcm_tax_amount": sum((row["rcm_tax_amount"] for row in rows), Decimal("0.00")),
        "eway_missing": sum(row["eway_missing"] for row in rows),
        "eway_expired": sum(row["eway_expired"] for row in rows),
        "eway_due_soon": sum(row["eway_due_soon"] for row in rows),
        "signed_off": sum(1 for row in rows if row["review"] and row["review"].is_signed_off),
        "output_tax": sum((row["output_tax"] for row in rows), Decimal("0.00")),
        "itc": sum((row["itc"] for row in rows), Decimal("0.00")),
        "net_tax_payable": sum((row["net_tax_payable"] for row in rows), Decimal("0.00")),
    }
    company_filter_options = _companies_for_user(request.user)

    return render(request, "core/gst_workbench.html", {
        "rows": rows,
        "totals": totals,
        "period_start": period_start,
        "period_end": period_end,
        "period_value": period_value,
        "company_filter_options": company_filter_options,
        "selected_company_id": selected_company_id,
        "title": "GST Workbench",
    })


def _gst_workbench_detail_csv(context):
    company = context["company"]
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="gst-review-{company.short_code or company.pk}-{context["period_value"]}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow([
        "Section",
        "Severity",
        "Title",
        "Description",
        "Count",
        "Reference",
        "Date",
        "Amount",
        "Task Reference",
        "Task Created",
        "Client Request Status",
    ])
    for blocker in context["signoff_blockers"]:
        writer.writerow([
            "Sign-off Blocker",
            "blocker",
            blocker["title"],
            blocker["description"],
            blocker["count"],
            "",
            "",
            "",
            blocker["code"],
            "",
            "",
        ])
    for item in context["exception_items"]:
        doc_request = item.get("document_request")
        writer.writerow([
            "Exception",
            item["severity"],
            item["label"],
            item["title"],
            "",
            item["reference"] or "",
            item["date"].isoformat() if item["date"] else "",
            item["amount"] or "",
            item["task_reference"],
            "yes" if item["task_reference"] in context["existing_task_refs"] else "no",
            doc_request.get_status_display() if doc_request else "",
        ])
    return response


def _execution_pack_voucher_name(voucher):
    raw = voucher.number or str(voucher.pk)
    cleaned = "".join(ch for ch in raw if ch.isalnum() or ch in ("-", "_"))
    return cleaned or str(voucher.pk)


def _gst_workbench_execution_zip(context):
    from integrations.gst import (
        GSTIntegrationError,
        build_e_invoice_payload,
        build_e_way_bill_payload,
        dump_gst_payload_json,
    )

    manifest = []
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for voucher in context["e_invoice_missing_vouchers"]:
            filename = f"e_invoice/{_execution_pack_voucher_name(voucher)}_irp.json"
            try:
                payload = build_e_invoice_payload(voucher)
            except GSTIntegrationError as exc:
                manifest.append(["e_invoice", voucher.number or voucher.pk, "blocked", str(exc), ""])
                continue
            archive.writestr(filename, dump_gst_payload_json(payload))
            manifest.append(["e_invoice", voucher.number or voucher.pk, "ready", "", filename])

        for voucher in context["eway_bill_missing_vouchers"]:
            filename = f"e_way_bill/{_execution_pack_voucher_name(voucher)}_ewb.json"
            try:
                payload = build_e_way_bill_payload(voucher)
            except GSTIntegrationError as exc:
                manifest.append(["e_way_bill", voucher.number or voucher.pk, "blocked", str(exc), ""])
                continue
            archive.writestr(filename, dump_gst_payload_json(payload))
            manifest.append(["e_way_bill", voucher.number or voucher.pk, "ready", "", filename])

        manifest_buffer = io.StringIO()
        writer = csv.writer(manifest_buffer)
        writer.writerow(["service", "voucher", "status", "message", "file"])
        writer.writerows(manifest)
        archive.writestr("manifest.csv", manifest_buffer.getvalue())

    return buffer.getvalue()


@login_required
def gst_workbench_detail(request, company_id, period):
    period_start, period_end = _parse_month_period(period)
    companies = _companies_for_user(request.user)
    company = get_object_or_404(companies, pk=company_id)
    context = _gst_period_detail_context(company, period_start, period_end)
    if request.GET.get("export") == "csv":
        return _gst_workbench_detail_csv(context)
    context.update({
        "title": f"GST Review - {company.name}",
    })
    return render(request, "core/gst_workbench_detail.html", context)


@login_required
def gst_workbench_execution_pack(request, company_id, period):
    period_start, period_end = _parse_month_period(period)
    companies = _companies_for_user(request.user)
    company = get_object_or_404(companies, pk=company_id)
    context = _gst_period_detail_context(company, period_start, period_end)
    payload = _gst_workbench_execution_zip(context)
    response = HttpResponse(payload, content_type="application/zip")
    filename = f"gst-execution-{company.short_code or company.pk}-{context['period_value']}.zip"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_POST
def gst_workbench_create_filings(request):
    period_start, period_end = _parse_month_period(request.POST.get("period"))
    companies = _companies_for_user(request.user)
    company = get_object_or_404(companies, pk=request.POST.get("company_id"))
    if not _can_manage_company(request.user, company):
        messages.error(request, "You do not have permission to create GST filings for this company.")
        return redirect("core:gst_workbench")

    created = _ensure_gst_period_filings(company, period_start, period_end, user=request.user)
    if created:
        messages.success(request, f"Created {len(created)} GST filing workflows for {company.name}.")
    else:
        messages.info(request, "GST filing workflows already exist for this period.")
    return redirect(f"{reverse('core:gst_workbench')}?period={period_start:%Y-%m}&company={company.pk}")


@login_required
@require_POST
def gst_workbench_chase_pack(request):
    period_start, period_end = _parse_month_period(request.POST.get("period"))
    companies = _companies_for_user(request.user)
    company = get_object_or_404(companies, pk=request.POST.get("company_id"))
    if not _can_manage_company(request.user, company):
        messages.error(request, "You do not have permission to create client chase requests for this company.")
        return redirect("core:gst_workbench")

    _ensure_gst_period_filings(company, period_start, period_end, user=request.user)
    context = _gst_period_detail_context(company, period_start, period_end)
    chase_items = [
        item for item in context["exception_items"]
        if item["kind"] in GST_CLIENT_CHASE_KINDS
    ]

    result = {
        "created_tasks": 0,
        "existing_tasks": 0,
        "created_requests": 0,
        "existing_requests": 0,
        "skipped": 0,
    }
    for item in chase_items:
        outcome = _create_gst_exception_followup(
            company=company,
            user=request.user,
            period_start=period_start,
            period_end=period_end,
            kind=item["kind"],
            obj=item["object"],
            action_type="client_chase",
        )
        result["created_tasks"] += 1 if outcome["created_task"] else 0
        result["existing_tasks"] += 1 if outcome["existing_task"] else 0
        result["created_requests"] += 1 if outcome["created_request"] else 0
        result["existing_requests"] += 1 if outcome["existing_request"] else 0
        result["skipped"] += 1 if outcome["skipped"] else 0

    if not chase_items:
        messages.info(request, "No GST client chase exceptions were found for this period.")
    else:
        messages.success(
            request,
            (
                "GST client chase pack ready: "
                f"{result['created_requests']} upload requests created, "
                f"{result['existing_requests']} already existed; "
                f"{result['created_tasks']} tasks created."
            ),
        )

    return _gst_detail_redirect(company, period_start)


@login_required
@require_POST
def gst_workbench_signoff(request):
    period_start, period_end = _parse_month_period(request.POST.get("period"))
    companies = _companies_for_user(request.user)
    company = get_object_or_404(companies, pk=request.POST.get("company_id"))
    if not _can_manage_company(request.user, company):
        messages.error(request, "You do not have permission to review GST for this company.")
        return redirect("core:gst_workbench")

    status = request.POST.get("status")
    allowed = {
        GSTPeriodReview.STATUS_IN_REVIEW,
        GSTPeriodReview.STATUS_SIGNED_OFF,
        GSTPeriodReview.STATUS_REOPENED,
    }
    if status not in allowed:
        messages.error(request, "Invalid GST review status.")
        return redirect(f"{reverse('core:gst_workbench')}?period={period_start:%Y-%m}&company={company.pk}")

    row = _gst_workbench_snapshot(company, period_start, period_end)
    if status == GSTPeriodReview.STATUS_SIGNED_OFF and row["signoff_blockers"]:
        blocker_titles = ", ".join(
            f"{blocker['title']} ({blocker['count']})"
            for blocker in row["signoff_blockers"][:4]
        )
        remaining = len(row["signoff_blockers"]) - 4
        if remaining > 0:
            blocker_titles = f"{blocker_titles}, +{remaining} more"
        messages.error(request, f"GST sign-off blocked. Clear these first: {blocker_titles}.")
        return _gst_detail_redirect(company, period_start)

    review, created = GSTPeriodReview.objects.get_or_create(
        company=company,
        period_start=period_start,
        period_end=period_end,
        defaults={"prepared_by": request.user},
    )
    old_data = {} if created else {
        "status": review.status,
        "risk_score": review.risk_score,
        "notes": review.notes,
        "reviewed_by_id": review.reviewed_by_id,
    }
    review.status = status
    review.risk_score = row["risk_score"]
    review.summary_snapshot = row["snapshot"]
    review.notes = request.POST.get("notes", "").strip()
    if not review.prepared_by_id:
        review.prepared_by = request.user
    if status == GSTPeriodReview.STATUS_SIGNED_OFF:
        review.reviewed_by = request.user
        review.reviewed_at = timezone.now()
    elif status == GSTPeriodReview.STATUS_REOPENED:
        review.reviewed_by = None
        review.reviewed_at = None
    review.save()
    AuditLog.objects.create(
        company=company,
        user=request.user,
        action=AuditLog.ACTION_CREATE if created else AuditLog.ACTION_UPDATE,
        model_name="GSTPeriodReview",
        record_id=review.pk,
        object_repr=f"{company.name} GST review {period_start:%Y-%m}",
        old_data=old_data,
        new_data={
            "status": review.status,
            "risk_score": review.risk_score,
            "notes": review.notes,
            "reviewed_by_id": review.reviewed_by_id,
            "period_start": period_start.isoformat(),
            "period_end": period_end.isoformat(),
            "can_sign_off": row["can_sign_off"],
            "signoff_blockers": row["snapshot"].get("signoff_blockers", []),
        },
    )

    messages.success(request, f"GST review marked {review.get_status_display()} for {company.name}.")
    return redirect(f"{reverse('core:gst_workbench')}?period={period_start:%Y-%m}&company={company.pk}")


def _gst_detail_redirect(company, period_start):
    return redirect("core:gst_workbench_detail", company_id=company.pk, period=period_start.strftime("%Y-%m"))


def _resolve_selected_company(request, companies):
    company_id = (request.POST.get("company_id") or request.GET.get("company") or "").strip()
    if company_id and companies.filter(pk=company_id).exists():
        return companies.get(pk=company_id)
    current_company = getattr(request, "current_company", None)
    if current_company and companies.filter(pk=current_company.pk).exists():
        return current_company
    return companies.first()


def _parse_as_of_date(raw_value=None):
    if raw_value:
        try:
            return _date.fromisoformat(raw_value)
        except (TypeError, ValueError):
            pass
    return timezone.localdate()


@login_required
def collections_command_center(request):
    params = request.POST if request.method == "POST" else request.GET
    companies = _companies_for_user(request.user)
    selected_company_id = (params.get("company") or "").strip()
    if selected_company_id and companies.filter(pk=selected_company_id).exists():
        scoped_companies = companies.filter(pk=selected_company_id)
    else:
        scoped_companies = companies
        selected_company_id = ""

    as_of_date = _parse_as_of_date(params.get("as_of"))
    risk_filter = (params.get("risk") or "active").strip()
    valid_risks = {value for value, _ in collection_risk_filter_choices()}
    if risk_filter not in valid_risks:
        risk_filter = "active"

    center = build_collections_center(scoped_companies, as_of_date=as_of_date, risk_filter=risk_filter)
    query_args = {"as_of": as_of_date.isoformat(), "risk": risk_filter}
    if selected_company_id:
        query_args["company"] = selected_company_id
    base_query = urlencode(query_args)

    if request.method == "GET" and request.GET.get("export") == "csv":
        return collections_csv_response(center)

    if request.method == "POST":
        manageable_ids = set(
            _manageable_companies_for_user(request.user)
            .filter(pk__in=[row["company"].pk for row in center["rows"]])
            .values_list("pk", flat=True)
        )
        selected_ids = request.POST.getlist("invoice_ids")
        action = request.POST.get("action", "")
        if action == "create_tasks":
            result = create_collection_center_tasks(center["rows"], request.user, manageable_ids, selected_ids)
            messages.success(
                request,
                f"Collection tasks ready: {result['created']} created, {result['existing']} already existed.",
            )
        elif action == "send_email":
            result = send_collection_emails(request, center["rows"], request.user, manageable_ids, selected_ids)
            if result.get("empty"):
                messages.error(request, "Select at least one invoice with an email address.")
            else:
                if result["sent"]:
                    messages.success(request, f"Sent {result['sent']} payment reminder email(s).")
                if result["skipped"]:
                    messages.warning(request, f"Skipped {result['skipped']} invoice(s) without access or email.")
                if result["failed"]:
                    messages.error(request, f"{result['failed']} payment reminder email(s) failed.")
        else:
            messages.error(request, "Invalid collections action.")
        return redirect(f"{reverse('core:collections_command_center')}?{base_query}")

    return render(request, "core/collections_command_center.html", {
        "rows": center["rows"],
        "party_rows": center["party_rows"],
        "totals": center["totals"],
        "companies": companies,
        "selected_company_id": selected_company_id,
        "as_of_date": as_of_date,
        "risk_filter": risk_filter,
        "risk_filter_choices": collection_risk_filter_choices(),
        "base_query": base_query,
        "export_query": f"{base_query}&export=csv",
        "title": "Collections Center",
    })


@login_required
def bank_reco_autopilot(request):
    params = request.POST if request.method == "POST" else request.GET
    companies = _companies_for_user(request.user)
    selected_company_id = (params.get("company") or "").strip()
    if selected_company_id and companies.filter(pk=selected_company_id).exists():
        scoped_companies = companies.filter(pk=selected_company_id)
    else:
        scoped_companies = companies
        selected_company_id = ""

    as_of_date = _parse_as_of_date(params.get("as_of"))
    focus = (params.get("focus") or "attention").strip()
    valid_focuses = {value for value, _ in bank_reco_focus_choices()}
    if focus not in valid_focuses:
        focus = "attention"

    center = build_bank_reco_autopilot(scoped_companies, as_of_date=as_of_date, focus=focus)
    query_args = {"as_of": as_of_date.isoformat(), "focus": focus}
    if selected_company_id:
        query_args["company"] = selected_company_id
    base_query = urlencode(query_args)

    if request.method == "GET":
        export_kind = request.GET.get("export")
        if export_kind == "csv":
            return bank_reco_autopilot_csv_response(center)
        if export_kind == "zip":
            return bank_reco_working_paper_zip_response(center)

    if request.method == "POST":
        manageable_ids = set(
            _manageable_companies_for_user(request.user)
            .filter(pk__in=[row["company"].pk for row in center["statement_rows"]])
            .values_list("pk", flat=True)
        )
        action = request.POST.get("action", "")
        selected_keys = request.POST.getlist("work_ids")
        selected_statement_ids = {
            int(value) for value in request.POST.getlist("statement_ids") if str(value).isdigit()
        }

        if action == "create_tasks":
            result = create_bank_reco_tasks(center["statement_rows"], request.user, manageable_ids, selected_keys)
            messages.success(
                request,
                f"Bank reconciliation tasks ready: {result['created']} created, {result['existing']} already existed.",
            )
            if result["skipped"]:
                messages.warning(request, f"Skipped {result['skipped']} item(s) without write access.")
        elif action == "auto_match":
            auto_match_ids = []
            for row in center["statement_rows"]:
                statement = row.get("statement")
                if not statement or row["company"].pk not in manageable_ids:
                    continue
                if selected_statement_ids and statement.pk not in selected_statement_ids:
                    continue
                if not selected_statement_ids and not row["pending_count"]:
                    continue
                auto_match_ids.append(statement.pk)
            matched = 0
            refreshed = 0
            for statement in BankStatement.objects.filter(pk__in=auto_match_ids, company__in=scoped_companies):
                matched += _auto_match(statement)
                refreshed += 1
            if refreshed:
                messages.success(request, f"Auto-match ran on {refreshed} statement(s); {matched} row(s) matched.")
            else:
                messages.info(request, "No eligible bank statements were selected for auto-match.")
        elif action == "post_auto_ready":
            result = post_bank_reco_auto_ready_vouchers(
                center["statement_rows"],
                request.user,
                manageable_ids,
                selected_keys,
            )
            if result["created"]:
                messages.success(request, f"Posted and reconciled {result['created']} auto-ready bank voucher(s).")
            else:
                messages.info(request, "No auto-ready bank rows were posted.")
            if result["skipped"]:
                messages.warning(request, f"Skipped {result['skipped']} statement(s) without write access or eligible rows.")
            if result["failed"]:
                messages.warning(request, f"{result['failed']} bank row(s) could not be posted.")
                for error in result["errors"]:
                    messages.warning(request, error)
            if result["task_closed"]:
                messages.success(request, f"Completed {result['task_closed']} bank reconciliation task(s).")
        elif action == "ask_client":
            result = create_bank_reco_client_requests(
                center["statement_rows"],
                request.user,
                manageable_ids,
                selected_keys,
                as_of_date=as_of_date,
            )
            if result["created"]:
                messages.success(
                    request,
                    f"Client bank requests ready: {result['created']} created, {result['existing']} already existed.",
                )
            elif result["existing"]:
                messages.info(request, f"Client bank requests already existed for {result['existing']} item(s).")
            else:
                messages.info(request, "No eligible bank rows needed client clarification.")
            if result["task_created"] or result["task_existing"]:
                messages.success(
                    request,
                    f"Linked request tasks: {result['task_created']} created, {result['task_existing']} already existed.",
                )
            if result["skipped"]:
                messages.warning(request, f"Skipped {result['skipped']} item(s) without write access or eligible bank rows.")
        elif action == "close_evidence":
            result = close_bank_reco_uploaded_evidence(
                center["statement_rows"],
                request.user,
                manageable_ids,
                selected_keys,
            )
            if result["closed"]:
                messages.success(
                    request,
                    f"Closed {result['closed']} uploaded bank evidence request(s).",
                )
            else:
                messages.info(request, "No uploaded bank evidence was selected for closure.")
            if result["task_closed"]:
                messages.success(request, f"Completed {result['task_closed']} linked task(s).")
            if result["skipped"]:
                messages.warning(request, f"Skipped {result['skipped']} item(s) without write access or uploaded evidence.")
        else:
            messages.error(request, "Invalid bank reconciliation action.")
        return redirect(f"{reverse('core:bank_reco_autopilot')}?{base_query}")

    return render(request, "core/bank_reco_autopilot.html", {
        "company_rows": center["company_rows"],
        "statement_rows": center["statement_rows"],
        "totals": center["totals"],
        "companies": companies,
        "selected_company_id": selected_company_id,
        "as_of_date": as_of_date,
        "focus": focus,
        "focus_choices": bank_reco_focus_choices(),
        "base_query": base_query,
        "export_query": f"{base_query}&export=csv",
        "review_pack_query": f"{base_query}&export=zip",
        "title": "Bank Reco Autopilot",
    })


@login_required
@require_POST
def gst_workbench_2b_action(request, pk):
    from gstr2b.models import PortalGSTR2BEntry

    companies = _companies_for_user(request.user)
    entry = get_object_or_404(PortalGSTR2BEntry, pk=pk, company__in=companies)
    if not _can_manage_company(request.user, entry.company):
        messages.error(request, "You do not have permission to update this GSTR-2B action.")
        return redirect("core:gst_workbench")

    action = request.POST.get("action_status", "new")
    allowed = {choice[0] for choice in PortalGSTR2BEntry.ACTION_STATUS_CHOICES}
    if action not in allowed:
        messages.error(request, "Invalid GSTR-2B action.")
        return _gst_detail_redirect(entry.company, entry.invoice_date.replace(day=1))

    entry.action_status = action
    entry.action_note = request.POST.get("action_note", "").strip()[:300]
    entry.save(update_fields=["action_status", "action_note", "updated_at"])
    messages.success(request, f"Marked {entry.invoice_number} as {entry.get_action_status_display()}.")
    return _gst_detail_redirect(entry.company, entry.invoice_date.replace(day=1))


@login_required
@require_POST
def gst_workbench_exception_task(request):
    from gstr2b.models import PortalGSTR2BEntry

    period_start, period_end = _parse_month_period(request.POST.get("period"))
    companies = _companies_for_user(request.user)
    company = get_object_or_404(companies, pk=request.POST.get("company_id"))
    if not _can_manage_company(request.user, company):
        messages.error(request, "You do not have permission to create GST exception tasks for this company.")
        return redirect("core:gst_workbench")

    kind = request.POST.get("kind", "")
    object_id = request.POST.get("object_id")
    action_type = request.POST.get("action_type", "resolve")
    description = request.POST.get("note", "").strip()
    target_object = None

    if kind == "2b_missing_in_books":
        target_object = get_object_or_404(PortalGSTR2BEntry, pk=object_id, company=company)
    elif kind == "purchase_missing_in_2b":
        target_object = get_object_or_404(Voucher, pk=object_id, company=company, voucher_type="Purchase")
    elif kind == "itc_180_reversal_due":
        target_object = get_object_or_404(Voucher, pk=object_id, company=company, voucher_type="Purchase")
    elif kind == "rcm_missing_tax":
        target_object = get_object_or_404(Voucher, pk=object_id, company=company, voucher_type="Purchase", reverse_charge=True)
    elif kind == "eway_bill_missing":
        target_object = get_object_or_404(Voucher, pk=object_id, company=company, voucher_type__in=["Sales", "Purchase", "Sales Return", "Purchase Return"])
    elif kind == "eway_bill_expired":
        target_object = get_object_or_404(Voucher, pk=object_id, company=company, voucher_type__in=["Sales", "Purchase", "Sales Return", "Purchase Return"])
    elif kind == "draft_gst_voucher":
        target_object = get_object_or_404(Voucher, pk=object_id, company=company, voucher_type__in=["Sales", "Purchase"])
    elif kind == "e_invoice_missing_irn":
        target_object = get_object_or_404(Voucher, pk=object_id, company=company, voucher_type__in=["Sales", "Sales Return"])
    elif kind in {"sales_missing_pos", "sales_invalid_party_gstin", "sales_tax_pos_mismatch"}:
        target_object = get_object_or_404(Voucher, pk=object_id, company=company, voucher_type__in=["Sales", "Sales Return"])
    elif kind == "overdue_gst_filing":
        target_object = get_object_or_404(ComplianceFiling, pk=object_id, company=company)
    elif kind == "open_gst_notice":
        target_object = get_object_or_404(ComplianceNotice, pk=object_id, company=company)

    if not target_object:
        messages.error(request, "Unknown GST exception type.")
        return _gst_detail_redirect(company, period_start)

    outcome = _create_gst_exception_followup(
        company=company,
        user=request.user,
        period_start=period_start,
        period_end=period_end,
        kind=kind,
        obj=target_object,
        action_type=action_type,
        note=description,
    )
    if outcome["created_task"]:
        messages.success(request, "GST exception task created.")
    else:
        messages.info(request, "GST exception task already exists.")

    if action_type == "client_chase" and outcome["created_request"]:
        messages.success(request, "Client upload request created.")
    elif action_type == "client_chase" and outcome["existing_request"]:
        messages.info(request, "Client upload request already exists.")
    return _gst_detail_redirect(company, period_start)


@login_required
@require_POST
def gst_workbench_document_request_status(request, pk):
    from portal.models import ClientDocumentRequest

    companies = _companies_for_user(request.user)
    doc_request = get_object_or_404(
        ClientDocumentRequest.objects.select_related("company", "related_task"),
        pk=pk,
        company__in=companies,
    )
    if not _can_manage_company(request.user, doc_request.company):
        messages.error(request, "You do not have permission to update this client request.")
        return redirect("core:gst_workbench")

    period_start, _ = _parse_month_period(request.POST.get("period"))
    action = request.POST.get("action", "")
    now = timezone.now()

    if action == "close_reviewed":
        if doc_request.status != ClientDocumentRequest.STATUS_UPLOADED:
            messages.error(request, "Only uploaded client requests can be closed as reviewed from GST Workbench.")
            return _gst_detail_redirect(doc_request.company, period_start)

        doc_request.status = ClientDocumentRequest.STATUS_CLOSED
        doc_request.closed_at = now
        doc_request.save(update_fields=["status", "closed_at", "updated_at"])

        if doc_request.related_task_id:
            task = doc_request.related_task
            task.status = PracticeTask.STATUS_DONE
            task.completed_by = request.user
            task.completed_at = now
            close_note = f"Client evidence reviewed and closed by {getattr(request.user, 'email', request.user)}."
            task.description = (task.description + "\n\n" + close_note).strip()
            task.save(update_fields=["status", "completed_by", "completed_at", "description", "updated_at"])

        messages.success(request, "Client evidence marked reviewed and request closed.")
    elif action == "reopen":
        doc_request.status = ClientDocumentRequest.STATUS_OPEN
        doc_request.closed_at = None
        doc_request.save(update_fields=["status", "closed_at", "updated_at"])
        if doc_request.related_task_id and doc_request.related_task.status == PracticeTask.STATUS_DONE:
            task = doc_request.related_task
            task.status = PracticeTask.STATUS_IN_PROGRESS
            task.completed_by = None
            task.completed_at = None
            task.save(update_fields=["status", "completed_by", "completed_at", "updated_at"])
        messages.success(request, "Client request reopened.")
    else:
        messages.error(request, "Invalid client request action.")

    return _gst_detail_redirect(doc_request.company, period_start)


@login_required
def compliance_calendar(request):
    companies = _companies_for_user(request.user)
    manageable_companies = _manageable_companies_for_user(request.user)
    users = _task_users_for_companies(manageable_companies)
    today = timezone.localdate()
    initial = {
        "companies": list(manageable_companies.values_list("pk", flat=True)),
        "from_date": today.replace(day=1),
        "months": 3,
    }

    form = ComplianceCalendarGenerationForm(
        request.POST or None,
        companies=manageable_companies,
        users=users,
        initial=initial,
    )
    result = None
    if request.method == "POST" and form.is_valid():
        action = request.POST.get("action", "preview")
        dry_run = action != "generate"
        cleaned = form.cleaned_data
        try:
            result = generate_compliance_calendar(
                companies=cleaned["companies"],
                months=cleaned["months"],
                from_date=cleaned["from_date"],
                assigned_to=cleaned["assigned_to"],
                reviewer=cleaned["reviewer"],
                created_by=request.user,
                dry_run=dry_run,
                include_ims=cleaned["include_ims"],
                include_gstr1=cleaned["include_gstr1"],
                include_gstr3b=cleaned["include_gstr3b"],
                include_tds_payment=cleaned["include_tds_payment"],
                include_tds_returns=cleaned["include_tds_returns"],
                ims_review_day=cleaned["ims_review_day"],
                gstr1_day=cleaned["gstr1_day"],
                gstr3b_day=cleaned["gstr3b_day"],
                tds_payment_day=cleaned["tds_payment_day"],
                gstr9_due=cleaned["gstr9_due"],
                gstr9c_due=cleaned["gstr9c_due"],
                itr_due=cleaned["itr_due"],
                tax_audit_due=cleaned["tax_audit_due"],
                mca_aoc4_due=cleaned["mca_aoc4_due"],
                mca_mgt7_due=cleaned["mca_mgt7_due"],
            )
        except ValueError as exc:
            form.add_error(None, str(exc))
        else:
            if dry_run:
                messages.info(
                    request,
                    f"Preview ready: {result['created']} would be created; {result['existing']} already exist.",
                )
            else:
                messages.success(
                    request,
                    f"Generated {result['created']} filings/tasks; {result['existing']} already existed.",
                )

    open_filings = ComplianceFiling.objects.filter(company__in=companies).exclude(
        status__in=[ComplianceFiling.STATUS_FILED, ComplianceFiling.STATUS_CANCELLED]
    )
    next_30 = today + timedelta(days=30)
    summary = {
        "open": open_filings.count(),
        "overdue": open_filings.filter(due_date__lt=today).count(),
        "next_30": open_filings.filter(due_date__gte=today, due_date__lte=next_30).count(),
        "unassigned": open_filings.filter(assigned_to__isnull=True).count(),
        "manageable_clients": manageable_companies.count(),
    }
    upcoming_filings = open_filings.filter(due_date__gte=today).select_related(
        "company", "assigned_to", "reviewer"
    ).order_by("due_date", "company__name", "filing_type")[:25]
    overdue_filings = open_filings.filter(due_date__lt=today).select_related(
        "company", "assigned_to", "reviewer"
    ).order_by("due_date", "company__name", "filing_type")[:15]

    if request.method == "GET" and request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="compliance_calendar_open_filings.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Client",
            "Filing",
            "Type",
            "Priority",
            "Status",
            "Due Date",
            "Days Overdue",
            "Owner",
            "Reviewer",
            "Period Start",
            "Period End",
            "ARN / Ack",
        ])
        for filing in open_filings.select_related("company", "assigned_to", "reviewer").order_by("due_date", "company__name", "filing_type"):
            days_overdue = ""
            if filing.due_date and filing.due_date < today:
                days_overdue = (today - filing.due_date).days
            writer.writerow([
                filing.company.name,
                filing.title,
                filing.get_filing_type_display(),
                filing.get_priority_display(),
                filing.get_status_display(),
                filing.due_date.isoformat() if filing.due_date else "",
                days_overdue,
                filing.assigned_to.email if filing.assigned_to else "",
                filing.reviewer.email if filing.reviewer else "",
                filing.period_start.isoformat() if filing.period_start else "",
                filing.period_end.isoformat() if filing.period_end else "",
                filing.arn_ack_number,
            ])
        return response

    return render(request, "core/compliance_calendar.html", {
        "form": form,
        "result": result,
        "summary": summary,
        "upcoming_filings": upcoming_filings,
        "overdue_filings": overdue_filings,
        "today": today,
        "export_query": urlencode({"export": "csv"}),
        "title": "Compliance Calendar",
    })


@login_required
def compliance_filing_list(request):
    companies = _companies_for_user(request.user)
    filings = ComplianceFiling.objects.filter(company__in=companies).select_related(
        "company", "assigned_to", "reviewer", "related_task"
    )

    status_filter = request.GET.get("status", "open")
    type_filter = request.GET.get("type", "")
    assigned_filter = request.GET.get("assigned", "")
    q = request.GET.get("q", "").strip()

    if status_filter == "open":
        filings = filings.exclude(status__in=[ComplianceFiling.STATUS_FILED, ComplianceFiling.STATUS_CANCELLED])
    elif status_filter:
        filings = filings.filter(status=status_filter)
    if type_filter:
        filings = filings.filter(filing_type=type_filter)
    if assigned_filter == "me":
        filings = filings.filter(Q(assigned_to=request.user) | Q(reviewer=request.user))
    elif assigned_filter == "unassigned":
        filings = filings.filter(assigned_to__isnull=True, reviewer__isnull=True)
    if q:
        filings = filings.filter(
            Q(title__icontains=q)
            | Q(company__name__icontains=q)
            | Q(arn_ack_number__icontains=q)
            | Q(portal_status__icontains=q)
            | Q(notes__icontains=q)
        )

    today = timezone.localdate()
    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="compliance_filings.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Client",
            "Filing",
            "Type",
            "Priority",
            "Status",
            "Owner",
            "Reviewer",
            "Due Date",
            "Days Overdue",
            "Period Start",
            "Period End",
            "ARN / Ack Number",
            "Portal Status",
            "Source",
            "Source Reference",
            "Notes",
        ])
        for filing in filings:
            days_overdue = ""
            if filing.due_date and filing.is_open and filing.due_date < today:
                days_overdue = (today - filing.due_date).days
            writer.writerow([
                filing.company.name,
                filing.title,
                filing.get_filing_type_display(),
                filing.get_priority_display(),
                filing.get_status_display(),
                filing.assigned_to.email if filing.assigned_to else "",
                filing.reviewer.email if filing.reviewer else "",
                filing.due_date.isoformat() if filing.due_date else "",
                days_overdue,
                filing.period_start.isoformat() if filing.period_start else "",
                filing.period_end.isoformat() if filing.period_end else "",
                filing.arn_ack_number,
                filing.portal_status,
                filing.get_source_display(),
                filing.source_reference,
                filing.notes,
            ])
        return response

    base = ComplianceFiling.objects.filter(company__in=companies)
    open_base = base.exclude(status__in=[ComplianceFiling.STATUS_FILED, ComplianceFiling.STATUS_CANCELLED])
    summary = {
        "open": open_base.count(),
        "overdue": open_base.filter(due_date__lt=today).count(),
        "client_pending": base.filter(status=ComplianceFiling.STATUS_CLIENT_PENDING).count(),
        "review": base.filter(status=ComplianceFiling.STATUS_READY_FOR_REVIEW).count(),
        "filed": base.filter(status=ComplianceFiling.STATUS_FILED).count(),
    }

    return render(request, "core/compliance_filing_list.html", {
        "filings": filings[:500],
        "summary": summary,
        "status_filter": status_filter,
        "type_filter": type_filter,
        "assigned_filter": assigned_filter,
        "q": q,
        "filing_types": ComplianceFiling.FILING_TYPE_CHOICES,
        "status_choices": ComplianceFiling.STATUS_CHOICES,
        "today": today,
        "export_query": _export_query(request),
    })


@login_required
def compliance_filing_create(request):
    companies = _companies_for_user(request.user)
    initial = {}
    company_id = request.GET.get("company")
    filing_type = request.GET.get("type")
    if company_id and companies.filter(pk=company_id).exists():
        initial["company"] = company_id
    if filing_type in {choice[0] for choice in ComplianceFiling.FILING_TYPE_CHOICES}:
        initial["filing_type"] = filing_type

    form = ComplianceFilingForm(
        request.POST or None,
        companies=companies,
        users=_task_users_for_companies(companies),
        initial=initial,
    )
    if request.method == "POST" and form.is_valid():
        filing = form.save(commit=False)
        if not _can_manage_company(request.user, filing.company):
            messages.error(request, "You do not have permission to create filings for this company.")
            return redirect("core:compliance_filings")
        filing.created_by = request.user
        if filing.status == ComplianceFiling.STATUS_FILED:
            filing.filed_at = timezone.now()
            filing.filed_by = request.user
        filing.save()
        sync_task_for_filing(filing, user=request.user)
        messages.success(request, "Compliance filing created.")
        return redirect("core:compliance_filings")

    return render(request, "core/compliance_filing_form.html", {"form": form, "title": "New Compliance Filing"})


@login_required
def compliance_filing_update(request, pk):
    companies = _companies_for_user(request.user)
    filing = get_object_or_404(ComplianceFiling, pk=pk, company__in=companies)
    if not _can_manage_company(request.user, filing.company):
        messages.error(request, "You do not have permission to update this filing.")
        return redirect("core:compliance_filings")

    form = ComplianceFilingForm(
        request.POST or None,
        instance=filing,
        companies=companies,
        users=_task_users_for_companies(companies),
    )
    if request.method == "POST" and form.is_valid():
        updated = form.save(commit=False)
        if updated.status == ComplianceFiling.STATUS_FILED and not updated.filed_at:
            updated.filed_at = timezone.now()
            updated.filed_by = request.user
        elif updated.status != ComplianceFiling.STATUS_FILED:
            updated.filed_at = None
            updated.filed_by = None
        updated.save()
        sync_task_for_filing(updated, user=request.user)
        messages.success(request, "Compliance filing updated.")
        return redirect("core:compliance_filings")

    return render(request, "core/compliance_filing_form.html", {"form": form, "filing": filing, "title": "Edit Compliance Filing"})


@login_required
@require_POST
def compliance_filing_set_status(request, pk):
    companies = _companies_for_user(request.user)
    filing = get_object_or_404(ComplianceFiling, pk=pk, company__in=companies)
    if not _can_manage_company(request.user, filing.company):
        messages.error(request, "You do not have permission to update this filing.")
        return redirect("core:compliance_filings")

    status = request.POST.get("status")
    allowed = {choice[0] for choice in ComplianceFiling.STATUS_CHOICES}
    if status not in allowed:
        messages.error(request, "Invalid filing status.")
        return redirect("core:compliance_filings")

    set_filing_status(filing, status, user=request.user)
    messages.success(request, "Filing status updated.")
    return redirect(request.META.get("HTTP_REFERER") or "core:compliance_filings")


@login_required
def compliance_notice_list(request):
    companies = _companies_for_user(request.user)
    notices = ComplianceNotice.objects.filter(company__in=companies).select_related(
        "company", "assigned_to", "related_filing", "related_task"
    )

    status_filter = request.GET.get("status", "open")
    type_filter = request.GET.get("type", "")
    assigned_filter = request.GET.get("assigned", "")
    q = request.GET.get("q", "").strip()

    if status_filter == "open":
        notices = notices.exclude(status=ComplianceNotice.STATUS_CLOSED)
    elif status_filter:
        notices = notices.filter(status=status_filter)
    if type_filter:
        notices = notices.filter(notice_type=type_filter)
    if assigned_filter == "me":
        notices = notices.filter(assigned_to=request.user)
    elif assigned_filter == "unassigned":
        notices = notices.filter(assigned_to__isnull=True)
    if q:
        notices = notices.filter(
            Q(title__icontains=q)
            | Q(company__name__icontains=q)
            | Q(reference_number__icontains=q)
            | Q(portal_status__icontains=q)
            | Q(description__icontains=q)
            | Q(response_summary__icontains=q)
        )

    today = timezone.localdate()
    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="compliance_notices.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Client",
            "Notice",
            "Type",
            "Priority",
            "Status",
            "Owner",
            "Issue Date",
            "Response Due Date",
            "Days Overdue",
            "Reference Number",
            "Portal Status",
            "Related Filing",
            "Description",
            "Response Summary",
        ])
        for notice in notices:
            days_overdue = ""
            if notice.response_due_date and notice.is_open and notice.response_due_date < today:
                days_overdue = (today - notice.response_due_date).days
            writer.writerow([
                notice.company.name,
                notice.title,
                notice.get_notice_type_display(),
                notice.get_priority_display(),
                notice.get_status_display(),
                notice.assigned_to.email if notice.assigned_to else "",
                notice.issue_date.isoformat() if notice.issue_date else "",
                notice.response_due_date.isoformat() if notice.response_due_date else "",
                days_overdue,
                notice.reference_number,
                notice.portal_status,
                notice.related_filing.title if notice.related_filing else "",
                notice.description,
                notice.response_summary,
            ])
        return response

    base = ComplianceNotice.objects.filter(company__in=companies)
    open_base = base.exclude(status=ComplianceNotice.STATUS_CLOSED)
    summary = {
        "open": open_base.count(),
        "overdue": open_base.filter(response_due_date__lt=today).count(),
        "data_pending": base.filter(status=ComplianceNotice.STATUS_DATA_PENDING).count(),
        "response_ready": base.filter(status=ComplianceNotice.STATUS_RESPONSE_READY).count(),
        "escalated": base.filter(status=ComplianceNotice.STATUS_ESCALATED).count(),
    }

    return render(request, "core/compliance_notice_list.html", {
        "notices": notices[:500],
        "summary": summary,
        "status_filter": status_filter,
        "type_filter": type_filter,
        "assigned_filter": assigned_filter,
        "q": q,
        "notice_types": ComplianceNotice.NOTICE_TYPE_CHOICES,
        "status_choices": ComplianceNotice.STATUS_CHOICES,
        "today": today,
        "export_query": _export_query(request),
    })


@login_required
def compliance_notice_create(request):
    companies = _companies_for_user(request.user)
    initial = {}
    company_id = request.GET.get("company")
    notice_type = request.GET.get("type")
    if company_id and companies.filter(pk=company_id).exists():
        initial["company"] = company_id
    if notice_type in {choice[0] for choice in ComplianceNotice.NOTICE_TYPE_CHOICES}:
        initial["notice_type"] = notice_type

    form = ComplianceNoticeForm(
        request.POST or None,
        companies=companies,
        users=_task_users_for_companies(companies),
        initial=initial,
    )
    if request.method == "POST" and form.is_valid():
        notice = form.save(commit=False)
        if not _can_manage_company(request.user, notice.company):
            messages.error(request, "You do not have permission to create notices for this company.")
            return redirect("core:compliance_notices")
        notice.created_by = request.user
        if notice.status == ComplianceNotice.STATUS_CLOSED:
            notice.closed_at = timezone.now()
            notice.closed_by = request.user
        notice.save()
        sync_task_for_notice(notice, user=request.user)
        messages.success(request, "Compliance notice created.")
        return redirect("core:compliance_notices")

    return render(request, "core/compliance_notice_form.html", {"form": form, "title": "New Compliance Notice"})


@login_required
def compliance_notice_update(request, pk):
    companies = _companies_for_user(request.user)
    notice = get_object_or_404(ComplianceNotice, pk=pk, company__in=companies)
    if not _can_manage_company(request.user, notice.company):
        messages.error(request, "You do not have permission to update this notice.")
        return redirect("core:compliance_notices")

    form = ComplianceNoticeForm(
        request.POST or None,
        instance=notice,
        companies=companies,
        users=_task_users_for_companies(companies),
    )
    if request.method == "POST" and form.is_valid():
        updated = form.save(commit=False)
        if updated.status == ComplianceNotice.STATUS_CLOSED and not updated.closed_at:
            updated.closed_at = timezone.now()
            updated.closed_by = request.user
        elif updated.status != ComplianceNotice.STATUS_CLOSED:
            updated.closed_at = None
            updated.closed_by = None
        updated.save()
        sync_task_for_notice(updated, user=request.user)
        messages.success(request, "Compliance notice updated.")
        return redirect("core:compliance_notices")

    return render(request, "core/compliance_notice_form.html", {"form": form, "notice": notice, "title": "Edit Compliance Notice"})


@login_required
@require_POST
def compliance_notice_set_status(request, pk):
    companies = _companies_for_user(request.user)
    notice = get_object_or_404(ComplianceNotice, pk=pk, company__in=companies)
    if not _can_manage_company(request.user, notice.company):
        messages.error(request, "You do not have permission to update this notice.")
        return redirect("core:compliance_notices")

    status = request.POST.get("status")
    allowed = {choice[0] for choice in ComplianceNotice.STATUS_CHOICES}
    if status not in allowed:
        messages.error(request, "Invalid notice status.")
        return redirect("core:compliance_notices")

    set_notice_status(notice, status, user=request.user)
    messages.success(request, "Notice status updated.")
    return redirect(request.META.get("HTTP_REFERER") or "core:compliance_notices")


@login_required
def app_settings(request):
    """
    In-app settings that normal company users can manage without Django admin access.
    """
    company = request.current_company
    try:
        statutory_profile = company.statutory_profile
    except CompanyStatutoryProfile.DoesNotExist:
        statutory_profile = CompanyStatutoryProfile(company=company)

    if request.method == "POST":
        action = request.POST.get("action") or "save"
        if action == "add_rule_override":
            override_form = StatutoryRuleOverrideForm(request.POST)
            if override_form.is_valid():
                override = override_form.save(commit=False)
                override.company = company
                override.created_by = request.user
                override.save()
                messages.success(request, "Statutory rule override added.")
                return redirect("core:app_settings")
            form = AppSettingsForm(instance=company)
            profile_form = CompanyStatutoryProfileForm(instance=statutory_profile)
        elif action == "deactivate_rule_override":
            override = get_object_or_404(
                StatutoryRuleOverride,
                pk=request.POST.get("override_id"),
                company=company,
                is_active=True,
            )
            override.is_active = False
            override.save(update_fields=["is_active", "updated_at"])
            messages.success(request, "Statutory rule override deactivated.")
            return redirect("core:app_settings")
        else:
            form = AppSettingsForm(request.POST, instance=company)
            profile_form = CompanyStatutoryProfileForm(request.POST, instance=statutory_profile)
            override_form = StatutoryRuleOverrideForm()

        if action not in {"add_rule_override", "deactivate_rule_override"} and form.is_valid() and profile_form.is_valid():
            company = form.save()
            statutory_profile = profile_form.save(commit=False)
            statutory_profile.company = company
            statutory_profile.updated_by = request.user
            statutory_profile.save()
            if action == "send_test_email":
                recipient = getattr(request.user, "email", "")
                if not recipient:
                    messages.error(request, "Your user account has no email address for the test email.")
                    return redirect("core:app_settings")
                try:
                    _send_app_settings_test_email(company, recipient)
                except Exception as exc:
                    messages.error(request, f"Test email failed: {exc}")
                    return redirect("core:app_settings")
                messages.success(request, f"App settings saved and test email sent to {recipient}.")
            elif action == "run_compliance_autopilot":
                months = normalize_autopilot_months(request.POST.get("autopilot_months"))
                result = run_compliance_autopilot(
                    companies=[company],
                    months=months,
                    from_date=timezone.localdate(),
                    created_by=request.user,
                )
                messages.success(
                    request,
                    (
                        "App settings saved. Compliance Autopilot created "
                        f"{result['created']} filing/task workflow(s); {result['existing']} already existed."
                    ),
                )
                if result["profile_warnings"]:
                    messages.warning(
                        request,
                        f"{len(result['profile_warnings'])} statutory profile warning(s) need review.",
                    )
            else:
                messages.success(request, "App settings saved successfully.")
            return redirect("core:app_settings")
    else:
        form = AppSettingsForm(instance=company)
        profile_form = CompanyStatutoryProfileForm(instance=statutory_profile)
        override_form = StatutoryRuleOverrideForm()

    return render(request, "core/app_settings.html", {
        "form": form,
        "profile_form": profile_form,
        "override_form": override_form,
        "rule_overrides": StatutoryRuleOverride.objects.filter(company=company).order_by("-is_active", "rule_type", "-period_start", "-updated_at")[:20],
        "company": company,
        "statutory_profile": statutory_profile,
        "whatsapp_webhook_url": request.build_absolute_uri(
            reverse("integrations:whatsapp_document_webhook")
        ),
        "whatsapp_test_url": _app_settings_whatsapp_test_url(company),
        "autopilot_months": 3,
    })


class _EmailTemplateContext(dict):
    def __missing__(self, key):
        return "{" + key + "}"


def _render_app_email_template(template, context):
    return (template or "").format_map(_EmailTemplateContext(context))


def _app_settings_sender(company):
    if company.invoice_email_from_address:
        return formataddr((company.invoice_email_from_name or company.name, company.invoice_email_from_address))
    return settings.DEFAULT_FROM_EMAIL


def _app_settings_reply_to(company):
    if company.invoice_email_reply_to:
        return [company.invoice_email_reply_to]
    if company.invoice_email_from_address:
        return [company.invoice_email_from_address]
    return None


def _send_app_settings_test_email(company, recipient):
    context = {
        "company_name": company.name,
        "client_name": "Sample Client",
        "voucher_number": "AV-TEST-001",
        "voucher_date": timezone.localdate().strftime("%d %b %Y"),
        "amount": "Rs.1,180.00",
        "outstanding": "Rs.1,180.00",
        "due_date": timezone.localdate().strftime("%d %b %Y"),
        "aging_line": "This is a test message from App Settings.",
    }
    subject = _render_app_email_template(
        company.invoice_email_subject or "Invoice {voucher_number} from {company_name}",
        context,
    )
    body = _render_app_email_template(
        company.invoice_email_body or "Dear {client_name},\n\nPlease find invoice {voucher_number} attached.",
        context,
    )
    email = EmailMessage(
        subject=subject,
        body=body,
        from_email=_app_settings_sender(company),
        to=[recipient],
        reply_to=_app_settings_reply_to(company),
    )
    email.send(fail_silently=False)


def _app_settings_whatsapp_test_url(company):
    if not company.whatsapp_intake_number:
        return ""
    message = quote(f"Test document intake message for {company.name}.")
    return f"https://wa.me/{company.whatsapp_intake_number.lstrip('+')}?text={message}"


@login_required
def production_trust_center(request):
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "Only staff users can access Production Trust.")
        return redirect("core:dashboard")

    include_deploy = request.GET.get("deploy") == "1" or request.POST.get("deploy") == "1"
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "run_backup":
            try:
                result = run_operational_backup(
                    include_sessions=request.POST.get("include_sessions") == "1",
                    encrypt=request.POST.get("encrypt_backup") == "1",
                    prune=request.POST.get("prune_backups") == "1",
                )
            except Exception as exc:
                messages.error(request, f"Backup drill failed: {exc}")
            else:
                manifest = result.get("manifest")
                if manifest:
                    messages.success(request, f"Backup drill completed: {manifest['name']}.")
                else:
                    messages.success(request, "Backup drill completed.")
        elif action == "run_scheduled_backup":
            try:
                result = run_scheduled_backup(
                    include_sessions=request.POST.get("include_sessions") == "1",
                    encrypt=request.POST.get("encrypt_backup") == "1",
                    prune=request.POST.get("prune_backups") == "1",
                    copy_offsite=request.POST.get("copy_offsite") == "1",
                    mode="manual-scheduled",
                )
            except Exception as exc:
                messages.error(request, f"Scheduled backup failed: {exc}")
            else:
                evidence = result.get("scheduled_evidence", {})
                payload = evidence.get("payload", {})
                messages.success(
                    request,
                    f"Scheduled backup evidence recorded: {evidence.get('name', '-')}; offsite {payload.get('offsite_status', '-')}.",
                )
        elif action == "create_backup_policy_tasks":
            try:
                trust_context = build_production_trust_context(include_deploy=include_deploy)
                result = create_backup_policy_tasks(
                    company=getattr(request, "current_company", None),
                    user=request.user,
                    watchdog=trust_context["backup_policy"],
                )
            except Exception as exc:
                messages.error(request, f"Backup policy tasks were not created: {exc}")
            else:
                messages.success(
                    request,
                    (
                        "Backup policy tasks synced: "
                        f"{result['created']} created, {result['updated']} updated, {result['closed']} closed."
                    ),
                )
        elif action == "create_scheduled_backup_tasks":
            try:
                trust_context = build_production_trust_context(include_deploy=include_deploy)
                result = create_scheduled_backup_tasks(
                    company=getattr(request, "current_company", None),
                    user=request.user,
                    watchdog=trust_context["scheduled_backup"],
                )
            except Exception as exc:
                messages.error(request, f"Scheduled backup tasks were not created: {exc}")
            else:
                messages.success(
                    request,
                    (
                        "Scheduled backup tasks synced: "
                        f"{result['created']} created, {result['updated']} updated, {result['closed']} closed."
                    ),
                )
        elif action == "record_restore_drill":
            try:
                result = record_restore_drill(
                    manifest_name=request.POST.get("manifest_name", ""),
                    outcome=request.POST.get("outcome", "failed"),
                    checks=request.POST,
                    target_environment=request.POST.get("target_environment", ""),
                    notes=request.POST.get("notes", ""),
                    unresolved_findings=request.POST.get("unresolved_findings") or 0,
                    finding_notes=request.POST.get("finding_notes", ""),
                    user=request.user,
                    company=getattr(request, "current_company", None),
                )
            except Exception as exc:
                messages.error(request, f"Restore drill evidence was not recorded: {exc}")
            else:
                if result.get("task"):
                    messages.warning(request, "Restore drill evidence recorded and a follow-up task was opened.")
                elif result.get("closed_task_count"):
                    messages.success(request, f"Restore drill evidence recorded. Closed {result['closed_task_count']} restore follow-up task(s).")
                else:
                    messages.success(request, "Restore drill evidence recorded.")
        elif action == "verify_restore_rehearsal":
            try:
                result = verify_backup_restore_rehearsal(
                    manifest_name=request.POST.get("manifest_name", ""),
                    target_environment=request.POST.get("target_environment", "") or "archive rehearsal",
                    user=request.user,
                    company=getattr(request, "current_company", None),
                )
            except Exception as exc:
                messages.error(request, f"Restore rehearsal failed: {exc}")
            else:
                verification = result.get("verification", {})
                if result.get("passed"):
                    messages.success(
                        request,
                        (
                            "Restore rehearsal passed: "
                            f"{verification.get('object_count', 0)} object(s), "
                            f"{verification.get('model_count', 0)} model(s), evidence {result.get('name', '-')}."
                        ),
                    )
                else:
                    messages.warning(
                        request,
                        f"Restore rehearsal recorded with findings: {', '.join(result.get('findings', [])) or result.get('name', '-')}.",
                    )
        else:
            messages.error(request, "Unknown Production Trust action.")
        redirect_url = reverse("core:production_trust_center")
        if include_deploy:
            redirect_url = f"{redirect_url}?deploy=1"
        return redirect(redirect_url)

    return render(request, "core/production_trust.html", {
        **build_production_trust_context(include_deploy=include_deploy),
        "title": "Production Trust",
    })


@login_required
def operations_monitor(request):
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "Only staff users can access Operations Monitor.")
        return redirect("core:dashboard")

    include_deploy = request.GET.get("deploy") == "1" or request.POST.get("deploy") == "1"
    companies = Company.objects.all().order_by("name")
    selected_company_id = (request.GET.get("company") or request.POST.get("company") or "").strip()
    if selected_company_id and companies.filter(pk=selected_company_id).exists():
        scoped_companies = companies.filter(pk=selected_company_id)
        current_company = scoped_companies.first()
    else:
        scoped_companies = companies
        current_company = getattr(request, "current_company", None) or companies.first()
        selected_company_id = ""

    monitor = build_operations_monitor(
        scoped_companies,
        current_company=current_company,
        include_deploy=include_deploy,
    )

    if request.method == "POST":
        action = request.POST.get("action", "")
        if action == "sync_operations_tasks":
            result = create_operations_monitor_tasks(request.user, monitor)
            messages.success(
                request,
                (
                    "Operations Monitor tasks synced: "
                    f"{result['created']} created, {result['updated']} updated, {result['closed']} closed."
                ),
            )
        else:
            messages.error(request, "Unknown Operations Monitor action.")
        redirect_url = reverse("core:operations_monitor")
        query_args = {}
        if include_deploy:
            query_args["deploy"] = "1"
        if selected_company_id:
            query_args["company"] = selected_company_id
        if query_args:
            redirect_url = f"{redirect_url}?{urlencode(query_args)}"
        return redirect(redirect_url)

    return render(request, "core/operations_monitor.html", {
        "monitor": monitor,
        "companies": companies,
        "selected_company_id": selected_company_id,
        "include_deploy": include_deploy,
        "title": "Operations Monitor",
    })


@login_required
def system_observability(request):
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "Only staff users can access System Observability.")
        return redirect("core:dashboard")

    include_deploy = request.GET.get("deploy") == "1" or request.POST.get("deploy") == "1"
    report = build_system_observability(
        company=getattr(request, "current_company", None),
        include_deploy=include_deploy,
    )
    if request.method == "POST":
        action = request.POST.get("action", "")
        if action == "sync_observability_tasks":
            try:
                result = create_system_observability_tasks(
                    getattr(request, "current_company", None),
                    request.user,
                    report,
                )
            except Exception as exc:
                messages.error(request, f"System Observability tasks were not created: {exc}")
            else:
                messages.success(
                    request,
                    (
                        "System Observability tasks synced: "
                        f"{result['created']} created, {result['updated']} updated, {result['closed']} closed."
                    ),
                )
        elif action == "run_observability_backup_drill":
            try:
                result = run_operational_backup(
                    include_sessions=request.POST.get("include_sessions") == "1",
                    encrypt=request.POST.get("encrypt_backup") == "1",
                    prune=request.POST.get("prune_backups") == "1",
                )
            except Exception as exc:
                messages.error(request, f"Backup drill failed: {exc}")
            else:
                manifest = result.get("manifest")
                if manifest:
                    messages.success(request, f"Backup evidence created: {manifest['name']}.")
                else:
                    messages.success(request, "Backup drill completed.")
        elif action == "run_observability_scheduled_backup":
            try:
                result = run_scheduled_backup(
                    include_sessions=request.POST.get("include_sessions") == "1",
                    encrypt=request.POST.get("encrypt_backup") == "1",
                    prune=request.POST.get("prune_backups") == "1",
                    copy_offsite=request.POST.get("copy_offsite") == "1",
                    mode="observability-remediation",
                )
            except Exception as exc:
                messages.error(request, f"Scheduled backup failed: {exc}")
            else:
                evidence = result.get("scheduled_evidence", {})
                payload = evidence.get("payload", {})
                messages.success(
                    request,
                    f"Scheduled backup evidence recorded: {evidence.get('name', '-')}; offsite {payload.get('offsite_status', '-')}.",
                )
        else:
            messages.error(request, "Unknown System Observability action.")
        redirect_url = reverse("core:system_observability")
        if include_deploy:
            redirect_url = f"{redirect_url}?deploy=1"
        return redirect(redirect_url)

    return render(request, "core/system_observability.html", {
        "report": report,
        "backup_encryption_available": bool(getattr(settings, "BACKUP_ENCRYPTION_KEY", "") or getattr(settings, "BACKUP_ENCRYPTION_PASSPHRASE", "")),
        "backup_offsite_configured": bool(getattr(settings, "BACKUP_OFFSITE_DIR", "")),
        "include_deploy": include_deploy,
        "title": "System Observability",
    })


@login_required
def system_observability_api(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({"ok": False, "error": "staff_required"}, status=403)

    include_deploy = request.GET.get("deploy") == "1"
    report = build_system_observability(
        company=getattr(request, "current_company", None),
        include_deploy=include_deploy,
    )
    return JsonResponse(observability_public_payload(report))


@login_required
def go_live_certificate(request):
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "Only staff users can access the Go-Live Certificate.")
        return redirect("core:dashboard")

    include_deploy = request.GET.get("deploy", request.POST.get("deploy", "1")) != "0"
    certificate = build_go_live_certificate(
        company=getattr(request, "current_company", None),
        include_deploy=include_deploy,
    )
    if request.method == "POST":
        action = request.POST.get("action", "")
        if action == "sync_go_live_tasks":
            try:
                result = create_go_live_remediation_tasks(
                    getattr(request, "current_company", None),
                    request.user,
                    certificate,
                )
            except Exception as exc:
                messages.error(request, f"Go-Live tasks were not created: {exc}")
            else:
                messages.success(
                    request,
                    (
                        "Go-Live remediation tasks synced: "
                        f"{result['created']} created, {result['updated']} updated, {result['closed']} closed."
                    ),
                )
        else:
            messages.error(request, "Unknown Go-Live Certificate action.")
        redirect_url = reverse("core:go_live_certificate")
        if not include_deploy:
            redirect_url = f"{redirect_url}?deploy=0"
        return redirect(redirect_url)

    return render(request, "core/go_live_certificate.html", {
        "certificate": certificate,
        "include_deploy": include_deploy,
        "title": "Go-Live Certificate",
    })


@login_required
def go_live_certificate_api(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({"ok": False, "error": "staff_required"}, status=403)

    include_deploy = request.GET.get("deploy", "1") != "0"
    certificate = build_go_live_certificate(
        company=getattr(request, "current_company", None),
        include_deploy=include_deploy,
    )
    return JsonResponse(go_live_certificate_payload(certificate))


@login_required
def go_live_evidence_pack_download(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({"ok": False, "error": "staff_required"}, status=403)

    company = getattr(request, "current_company", None)
    if not company:
        return JsonResponse({"ok": False, "error": "company_required"}, status=400)

    include_deploy = request.GET.get("deploy", "1") != "0"
    pack = build_go_live_evidence_pack(
        company=company,
        user=request.user,
        include_deploy=include_deploy,
    )
    response = HttpResponse(go_live_evidence_pack_bytes(pack), content_type="application/json; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{go_live_evidence_pack_filename(pack)}"'
    return response


@login_required
def security_control(request):
    company = getattr(request, "current_company", None)
    if not company:
        messages.error(request, "Select a company before opening Security Control.")
        return redirect("core:select_company")
    if not _can_manage_security_control(request.user, company, getattr(request, "current_company_role", "")):
        messages.error(request, "Only company admins and staff users can access Security Control.")
        return redirect("core:dashboard")

    assessment = build_security_control(company)
    if request.method == "POST":
        action = request.POST.get("action", "")
        if action == "sync_security_tasks":
            result = create_security_control_tasks(company, request.user, assessment)
            messages.success(
                request,
                (
                    "Security Control tasks synced: "
                    f"{result['created']} created, {result['updated']} updated, {result['closed']} closed."
                ),
            )
        else:
            messages.error(request, "Unknown Security Control action.")
        return redirect("core:security_control")

    return render(request, "core/security_control.html", {
        "assessment": assessment,
        "company": company,
        "title": "Security Control",
    })


def _can_manage_security_control(user, company, current_role=""):
    if user.is_staff or user.is_superuser:
        return True
    if current_role == "Admin":
        return True
    return UserCompanyAccess.objects.filter(user=user, company=company, role="Admin").exists()


@login_required
@admin_required
def company_settings(request):
    """
    Admin-only view: edit the current company's profile and banking / UPI details.
    """
    company = request.current_company

    if request.method == "POST":
        form = CompanySettingsForm(request.POST, instance=company)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                "Company settings saved successfully."
            )
            return redirect("core:company_settings")
    else:
        form = CompanySettingsForm(instance=company)

    return render(request, "core/company_settings.html", {
        "form":    form,
        "company": company,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Bank Reconciliation views
# ─────────────────────────────────────────────────────────────────────────────

import csv
import io
from decimal import Decimal, InvalidOperation
from datetime import datetime

from django.db import transaction
from django.http import JsonResponse

from .models import BankStatement, BankStatementRow
from .forms import BankStatementForm


# ── helpers ──────────────────────────────────────────────────────────────────

_DATE_FORMATS = [
    "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d",
    "%d-%b-%Y", "%d-%b-%y", "%b %d, %Y",
    "%m/%d/%Y",
]


def _parse_date(raw):
    raw = raw.strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(raw):
    """Convert '1,23,456.78' or '-1234.00' to Decimal, or Decimal('0') on fail."""
    if not raw or not raw.strip():
        return Decimal("0.00")
    cleaned = raw.strip().replace(",", "").replace(" ", "")
    try:
        val = Decimal(cleaned)
        return abs(val)   # always positive; caller decides debit/credit
    except InvalidOperation:
        return Decimal("0.00")


def _find_col(headers, *candidates):
    """Case-insensitive column header search, returns index or None."""
    headers_lower = [h.strip().lower() for h in headers]
    for cand in candidates:
        if cand.lower() in headers_lower:
            return headers_lower.index(cand.lower())
    return None


def _parse_csv(file_data):
    """
    Parse CSV bytes into a list of dicts:
      {row_number, date, description, debit, credit, balance}

    Supports common bank CSV layouts:
      • Separate Debit / Credit columns
      • Single Amount column (negative = debit)
      • With or without Balance column
    """
    text = file_data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows_out = []
    headers  = None

    for i, row in enumerate(reader):
        if not any(c.strip() for c in row):
            continue  # skip blank lines

        if headers is None:
            # The first non-blank row with at least 3 cells is the header
            if len(row) >= 3:
                headers = row
            continue

        idx_date  = _find_col(headers, "date", "txn date", "transaction date", "value date")
        idx_desc  = _find_col(headers, "description", "narration", "particulars", "remarks", "details")
        idx_debit = _find_col(headers, "debit", "withdrawal", "withdrawals", "dr", "debit amount")
        idx_credit= _find_col(headers, "credit", "deposit", "deposits", "cr", "credit amount")
        idx_amt   = _find_col(headers, "amount", "transaction amount")
        idx_bal   = _find_col(headers, "balance", "closing balance", "running balance")

        def cell(idx):
            if idx is None or idx >= len(row):
                return ""
            return row[idx].strip()

        raw_date = cell(idx_date)
        parsed   = _parse_date(raw_date)
        if not parsed:
            continue  # skip rows with unparseable dates

        description = cell(idx_desc) or "—"

        if idx_debit is not None and idx_credit is not None:
            debit  = _parse_amount(cell(idx_debit))
            credit = _parse_amount(cell(idx_credit))
        elif idx_amt is not None:
            amt = cell(idx_amt).strip()
            if amt.startswith("-"):
                debit  = _parse_amount(amt)
                credit = Decimal("0.00")
            else:
                debit  = Decimal("0.00")
                credit = _parse_amount(amt)
        else:
            debit  = Decimal("0.00")
            credit = Decimal("0.00")

        balance = _parse_amount(cell(idx_bal)) if idx_bal is not None else None

        rows_out.append({
            "row_number":  i,
            "date":        parsed,
            "description": description,
            "debit":       debit,
            "credit":      credit,
            "balance":     balance,
        })

    return rows_out


def _bank_row_abs_amount(row):
    return row.debit if row.debit > 0 else row.credit


def _bank_row_direction(row):
    if row.debit > 0:
        return "debit"
    if row.credit > 0:
        return "credit"
    return "zero"


def _normalize_bank_description(value):
    import re

    normalized = re.sub(r"[^a-z0-9]+", " ", (value or "").lower()).strip()
    return " ".join(normalized.split())


_BANK_NARRATION_NOISE_WORDS = {
    "account",
    "acct",
    "bank",
    "by",
    "chq",
    "credit",
    "cr",
    "debit",
    "dr",
    "from",
    "imps",
    "inb",
    "info",
    "neft",
    "payment",
    "paytm",
    "pos",
    "ref",
    "rtgs",
    "to",
    "transaction",
    "trf",
    "upi",
    "utr",
}


def _bank_description_signature(value):
    words = []
    for word in _normalize_bank_description(value).split():
        if word in _BANK_NARRATION_NOISE_WORDS:
            continue
        if len(word) < 3:
            continue
        if any(char.isdigit() for char in word):
            continue
        words.append(word)
    return " ".join(words[:12])


_BANK_REFERENCE_STOPWORDS = {
    "ACCOUNT",
    "ACCT",
    "BANK",
    "CLIENT",
    "CREDIT",
    "CUSTOMER",
    "DEBIT",
    "FROM",
    "IMPS",
    "NEFT",
    "PAYMENT",
    "RECEIPT",
    "REF",
    "RTGS",
    "TRANSFER",
    "TRANSACTION",
    "UPI",
    "VENDOR",
}


def _bank_reference_tokens(*values):
    import re

    tokens = set()
    for value in values:
        text = str(value or "").upper()
        compact = re.sub(r"[^A-Z0-9]", "", text)
        if len(compact) >= 8 and not compact.isdigit():
            tokens.add(compact[:40])
        for token in re.findall(r"[A-Z0-9]{4,}", text):
            if token in _BANK_REFERENCE_STOPWORDS:
                continue
            if token.isdigit() and len(token) < 6:
                continue
            if token.isalpha() and len(token) < 5:
                continue
            tokens.add(token[:40])
    return tokens


def _duplicate_group_key(row):
    amount = _bank_row_abs_amount(row)
    if amount <= 0:
        return ""
    account_id = row.statement.account_ledger_id or "none"
    description = _normalize_bank_description(row.description)[:60]
    return f"{account_id}:{row.date:%Y%m%d}:{_bank_row_direction(row)}:{amount}:{description}"


def _refresh_duplicate_flags(statement):
    rows = list(
        BankStatementRow.objects.filter(statement__company=statement.company)
        .select_related("statement")
        .only(
            "id",
            "statement_id",
            "statement__company_id",
            "statement__account_ledger_id",
            "date",
            "description",
            "debit",
            "credit",
            "potential_duplicate",
            "duplicate_group_key",
        )
    )
    groups = {}
    for row in rows:
        key = _duplicate_group_key(row)
        if key:
            groups.setdefault(key, []).append(row.pk)

    for row in rows:
        key = _duplicate_group_key(row)
        is_duplicate = bool(key and len(groups.get(key, [])) > 1)
        if row.potential_duplicate != is_duplicate or row.duplicate_group_key != key:
            row.potential_duplicate = is_duplicate
            row.duplicate_group_key = key
            row.save(update_fields=["potential_duplicate", "duplicate_group_key"])


def _infer_party_ledger_from_bank_history_row(history_row, ledger_by_id):
    if history_row.suggested_ledger_id in ledger_by_id:
        return ledger_by_id[history_row.suggested_ledger_id]
    if not history_row.matched_voucher_id:
        return None

    expected_entry_type = "DR" if history_row.debit > 0 else "CR"
    amount = _bank_row_abs_amount(history_row)
    if amount <= 0:
        return None

    for item in history_row.matched_voucher.items.all():
        if item.entry_type != expected_entry_type:
            continue
        if item.amount != amount:
            continue
        if item.ledger_id == history_row.statement.account_ledger_id:
            continue
        return ledger_by_id.get(item.ledger_id)
    return None


def _suggest_bank_ledger_from_history(row, ledgers):
    signature = _bank_description_signature(row.description)
    if len(signature) < 5:
        return None, 0, ""

    ledger_by_id = {ledger.pk: ledger for ledger in ledgers if ledger.is_active}
    history = (
        BankStatementRow.objects.filter(
            statement__company=row.statement.company,
            is_reconciled=True,
        )
        .exclude(pk=row.pk)
        .select_related("statement", "suggested_ledger", "matched_voucher")
        .prefetch_related("matched_voucher__items__ledger")
        .order_by("-date", "-id")[:300]
    )
    if row.debit > 0:
        history = [item for item in history if item.debit > 0]
    elif row.credit > 0:
        history = [item for item in history if item.credit > 0]
    else:
        return None, 0, ""

    candidates = {}
    for history_row in history:
        ledger = _infer_party_ledger_from_bank_history_row(history_row, ledger_by_id)
        if not ledger or ledger.pk == row.statement.account_ledger_id:
            continue
        history_signature = _bank_description_signature(history_row.description)
        if not history_signature:
            continue
        score = int(difflib.SequenceMatcher(None, signature, history_signature).ratio() * 100)
        if score < 76:
            continue
        bucket = candidates.setdefault(
            ledger.pk,
            {"ledger": ledger, "support": 0, "best_score": 0, "sample": history_signature},
        )
        bucket["support"] += 1
        if score > bucket["best_score"]:
            bucket["best_score"] = score
            bucket["sample"] = history_signature

    if not candidates:
        return None, 0, ""

    ranked = sorted(candidates.values(), key=lambda item: (item["best_score"], item["support"]), reverse=True)
    top = ranked[0]
    second = ranked[1] if len(ranked) > 1 else None
    if second and top["best_score"] - second["best_score"] < 8:
        return None, 0, ""
    if top["support"] < 2 and top["best_score"] < 92:
        return None, 0, ""

    confidence = min(92, max(74, top["best_score"] + min(8, top["support"] * 3)))
    reason = f"Learned bank rule: {top['support']} similar row(s)"
    return top["ledger"], confidence, reason


def _suggest_bank_ledger(row, ledgers):
    learned_ledger, learned_confidence, learned_reason = _suggest_bank_ledger_from_history(row, ledgers)
    if learned_ledger:
        return learned_ledger, learned_confidence, learned_reason

    desc = row.description.upper()
    keywords = {
        "CASH": "Cash",
        "ATM": "Cash",
        "FUEL": "Fuel",
        "PETROL": "Fuel",
        "SALARY": "Salary",
        "RENT": "Rent",
        "ELECTRICITY": "Electricity",
        "INT": "Bank Interest",
        "TAX": "GST Payable",
    }
    for keyword, ledger_name in keywords.items():
        if keyword in desc:
            ledger = next((item for item in ledgers if ledger_name.upper() in item.name.upper()), None)
            if ledger:
                return ledger, 88, f"Keyword match: {keyword}"

    best_ledger = None
    best_score = 0.0
    for ledger in ledgers:
        score = difflib.SequenceMatcher(None, desc, ledger.name.upper()).ratio()
        if score > best_score:
            best_score = score
            best_ledger = ledger

    if best_ledger and best_score >= 0.6:
        return best_ledger, min(84, int(best_score * 100)), "Ledger name similarity"
    return None, 0, ""


def _voucher_match_text(voucher_item):
    voucher = voucher_item.voucher
    return " ".join([
        voucher.number or "",
        voucher.source_reference or "",
        voucher.narration or "",
        voucher_item.narration or "",
    ])


def _rank_bank_voucher_candidates(row, voucher_items):
    row_tokens = _bank_reference_tokens(row.description)
    row_description = _normalize_bank_description(row.description)
    ranked = []

    for item in voucher_items:
        voucher = item.voucher
        voucher_text = _voucher_match_text(item)
        voucher_tokens = _bank_reference_tokens(voucher_text)
        shared_tokens = sorted(row_tokens & voucher_tokens)
        date_delta = abs((voucher.date - row.date).days)
        description_score = 0
        if row_description:
            voucher_description = _normalize_bank_description(voucher_text)
            if voucher_description:
                description_score = int(difflib.SequenceMatcher(None, row_description, voucher_description).ratio() * 100)

        score = 76 if date_delta <= 3 else 62
        reason = "Amount, direction, bank ledger, and date window"
        if shared_tokens:
            score = 98 if date_delta <= 7 else 94
            reason = f"Reference token match: {', '.join(shared_tokens[:2])}"
        elif description_score >= 70 and date_delta <= 3:
            score = 88
            reason = "Narration similarity with voucher"
        elif description_score >= 55 and date_delta <= 3:
            score = 82
            reason = "Partial narration similarity with voucher"

        ranked.append({
            "voucher_item": item,
            "score": score,
            "reason": reason,
            "shared_tokens": shared_tokens,
            "date_delta": date_delta,
        })

    ranked.sort(key=lambda item: (item["score"], -item["date_delta"]), reverse=True)
    return ranked


def _bank_voucher_match_decision(row, voucher_items):
    ranked = _rank_bank_voucher_candidates(row, voucher_items)
    if not ranked:
        return None

    top = ranked[0]
    second_score = ranked[1]["score"] if len(ranked) > 1 else 0
    if len(ranked) == 1 and top["date_delta"] <= 3:
        return {
            "action": "auto",
            "voucher_item": top["voucher_item"],
            "confidence": 100,
            "reason": "Exact amount, direction, bank ledger, and date window",
        }
    if top["shared_tokens"] and top["score"] >= 94 and top["score"] - second_score >= 8:
        return {
            "action": "auto",
            "voucher_item": top["voucher_item"],
            "confidence": top["score"],
            "reason": top["reason"],
        }
    return {
        "action": "review",
        "voucher_item": top["voucher_item"],
        "confidence": min(top["score"], 84) if len(ranked) > 1 else top["score"],
        "reason": f"{len(ranked)} voucher candidates; best: {top['reason']}",
    }


def _auto_match(statement):
    """
    Auto-match BankStatementRow entries to Vouchers and suggest Ledgers.
    Strategy:
      1. Exact Amount + Date Match (High Confidence) -> Auto-reconcile
      2. Fuzzy Match on Description -> Suggest Ledger
    """
    from vouchers.models import VoucherItem
    from ledger.models import Ledger
    if not statement.account_ledger_id:
        return 0

    _refresh_duplicate_flags(statement)
    matched = 0
    all_ledgers = list(Ledger.objects.filter(company=statement.company, is_active=True))
    
    for row in statement.rows.filter(is_reconciled=False).select_related("statement"):
        amount = _bank_row_abs_amount(row)
        if amount <= 0:
            continue

        # 1. Exact amount + Date match (±3 days)
        from datetime import timedelta
        row_tokens = _bank_reference_tokens(row.description)
        has_numbered_reference = any(any(char.isdigit() for char in token) for token in row_tokens)
        date_window_days = 15 if has_numbered_reference else 3
        date_min = row.date - timedelta(days=date_window_days)
        date_max = row.date + timedelta(days=date_window_days)

        # Look for existing VoucherItems on the same bank ledger
        voucher_matches = list(VoucherItem.objects.filter(
            ledger=statement.account_ledger,
            voucher__company=statement.company,
            voucher__status="APPROVED",
            voucher__date__gte=date_min,
            voucher__date__lte=date_max,
            entry_type="CR" if row.debit > 0 else "DR",
            amount=amount,
        ).select_related("voucher").order_by("voucher__date", "voucher_id"))

        if row.potential_duplicate:
            row.match_confidence = max(row.match_confidence, 50)
            row.match_reason = "Potential duplicate bank row; manual review required"
            row.save(update_fields=["match_confidence", "match_reason"])
            continue

        decision = _bank_voucher_match_decision(row, voucher_matches)
        if decision and decision["action"] == "auto":
            v_match = decision["voucher_item"]
            row.is_reconciled    = True
            row.matched_voucher  = v_match.voucher
            row.match_confidence = decision["confidence"]
            row.match_reason = decision["reason"]
            row.save(update_fields=["is_reconciled", "matched_voucher", "match_confidence", "match_reason"])
            matched += 1
            continue
        if decision:
            row.match_confidence = decision["confidence"]
            row.match_reason = decision["reason"]
            row.save(update_fields=["match_confidence", "match_reason"])
            continue

        # 2. If no exact voucher match, suggest a ledger based on description
        best_ledger, confidence, reason = _suggest_bank_ledger(row, all_ledgers)
        if best_ledger:
            row.suggested_ledger = best_ledger
            row.match_confidence = confidence
            row.match_reason = reason
            row.save(update_fields=["suggested_ledger", "match_confidence", "match_reason"])

    return matched


def _voucher_matches_bank_row(row, voucher):
    """Return True when a voucher has the expected bank line for a statement row."""
    if not row.statement.account_ledger_id:
        return False

    from datetime import timedelta
    from vouchers.models import VoucherItem

    amount = _bank_row_abs_amount(row)
    if amount <= 0:
        return False

    expected_entry_type = "CR" if row.debit > 0 else "DR"
    return VoucherItem.objects.filter(
        voucher=voucher,
        ledger=row.statement.account_ledger,
        entry_type=expected_entry_type,
        amount=amount,
        voucher__date__gte=row.date - timedelta(days=3),
        voucher__date__lte=row.date + timedelta(days=3),
    ).exists()


# ── Views ────────────────────────────────────────────────────────────────────

@login_required
def bank_statement_list(request):
    company    = request.current_company
    statements = BankStatement.objects.filter(
        company=company
    ).select_related("account_ledger").order_by("-statement_date")

    return render(request, "core/bank_statement_list.html", {
        "statements": statements,
    })


def _parse_excel(file_content):
    """
    Parses XLSX bank statement.
    Expects first row as header.
    """
    import io
    from openpyxl import load_workbook
    from datetime import datetime
    
    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active
    rows_out = []
    
    # Simple mapping: assume column order Date, Desc, Dr, Cr, Bal
    # Real systems use flexible mapping, but here we provide a standard foundational one.
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not any(row): continue
        
        date_val = row[0]
        if isinstance(date_val, str):
            try: date_val = datetime.strptime(date_val, "%Y-%m-%d").date()
            except: date_val = _date.today()
        
        rows_out.append({
            "row_number":  i,
            "date":        date_val or _date.today(),
            "description": str(row[1] or ""),
            "debit":       Decimal(str(row[2] or "0.00")),
            "credit":      Decimal(str(row[3] or "0.00")),
            "balance":     Decimal(str(row[4] or "0.00")) if len(row) > 4 else None,
        })
    return rows_out


@login_required
@write_required
def bank_statement_upload(request):
    company = request.current_company
    if request.method == "POST":
        form = BankStatementForm(request.POST, request.FILES, company=company)
        if form.is_valid():
            with transaction.atomic():
                stmt = form.save(commit=False)
                stmt.company = company
                stmt.save()

                file = request.FILES["statement_file"]
                if file.name.endswith(".csv"):
                    rows_data = _parse_csv(file.read())
                elif file.name.lower().endswith((".pdf", ".png", ".jpg", ".jpeg")):
                    from ocr import ocr_utils
                    rows_data = ocr_utils.extract_bank_statement_rows(file.read())
                else:
                    rows_data = _parse_excel(file.read())

                for rd in rows_data:
                    BankStatementRow.objects.create(
                        statement   = stmt,
                        date        = rd["date"],
                        description = rd["description"],
                        debit       = rd["debit"],
                        credit      = rd["credit"],
                        balance     = rd.get("balance"),
                        row_number  = rd["row_number"],
                    )

                auto_matched = _auto_match(stmt)

            messages.success(
                request,
                f"Uploaded {stmt.total_rows} rows. {auto_matched} auto-matched."
            )
            return redirect("core:bank_statement_detail", pk=stmt.pk)
    else:
        form = BankStatementForm(company=company)

    return render(request, "core/bank_statement_upload.html", {
        "form":  form,
        "title": "Upload Bank Statement",
    })


@login_required
def bank_reconciliation_report(request):
    """
    Step 6: Reconciliation Report.
    Summary of Matched vs Unmatched for the company.
    """
    company = request.current_company
    rows = BankStatementRow.objects.filter(statement__company=company)
    
    matched = rows.filter(is_reconciled=True)
    unmatched = rows.filter(is_reconciled=False)
    
    total_matched = sum(r.amount for r in matched)
    total_unmatched = sum(r.amount for r in unmatched)

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="bank_reconciliation_report.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Statement Date",
            "Bank Account",
            "Row Date",
            "Description",
            "Debit",
            "Credit",
            "Balance",
            "Status",
            "Matched Voucher",
            "Suggested Ledger",
            "Confidence",
            "Match Reason",
            "Potential Duplicate",
        ])
        for row in rows.select_related(
            "statement",
            "statement__account_ledger",
            "matched_voucher",
            "suggested_ledger",
        ).order_by("is_reconciled", "date", "statement__statement_date", "row_number"):
            writer.writerow([
                row.statement.statement_date.isoformat(),
                row.statement.account_ledger.name if row.statement.account_ledger else "",
                row.date.isoformat(),
                row.description,
                f"{row.debit:.2f}",
                f"{row.credit:.2f}",
                f"{row.balance:.2f}" if row.balance is not None else "",
                "Reconciled" if row.is_reconciled else "Pending",
                str(row.matched_voucher) if row.matched_voucher else "",
                row.suggested_ledger.name if row.suggested_ledger else "",
                row.match_confidence,
                row.match_reason,
                "Yes" if row.potential_duplicate else "No",
            ])
        return response
    
    return render(request, "core/bank_reconciliation_report.html", {
        "matched_count": matched.count(),
        "unmatched_count": unmatched.count(),
        "total_matched": total_matched,
        "total_unmatched": total_unmatched,
        "net_difference": total_unmatched,
    })


@login_required
def bank_statement_detail(request, pk):
    company   = request.current_company
    statement = get_object_or_404(BankStatement, pk=pk, company=company)
    _refresh_duplicate_flags(statement)
    rows      = list(statement.rows.select_related("matched_voucher", "suggested_ledger").order_by("row_number"))

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="bank_reconciliation_{statement.pk}.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Row Number",
            "Date",
            "Description",
            "Debit",
            "Credit",
            "Balance",
            "Status",
            "Matched Voucher",
            "Suggested Ledger",
            "Confidence",
            "Match Reason",
            "Potential Duplicate",
        ])
        for row in rows:
            writer.writerow([
                row.row_number,
                row.date.isoformat(),
                row.description,
                row.debit,
                row.credit,
                row.balance if row.balance is not None else "",
                "Reconciled" if row.is_reconciled else "Pending",
                str(row.matched_voucher) if row.matched_voucher else "",
                row.suggested_ledger.name if row.suggested_ledger else "",
                row.match_confidence,
                row.match_reason,
                "Yes" if row.potential_duplicate else "No",
            ])
        return response

    total_debit  = sum(r.debit  for r in rows)
    total_credit = sum(r.credit for r in rows)
    reconciled   = sum(1 for r in rows if r.is_reconciled)
    unreconciled = len(rows) - reconciled
    suggested = sum(1 for r in rows if not r.is_reconciled and r.suggested_ledger_id)
    duplicates = sum(1 for r in rows if r.potential_duplicate)
    high_confidence = sum(1 for r in rows if not r.is_reconciled and r.match_confidence >= 70)
    bulk_eligible = 0
    for row in rows:
        row.bulk_create_eligible = bool(
            not row.is_reconciled
            and row.suggested_ledger_id
            and row.match_confidence >= 70
            and not row.potential_duplicate
        )
        if row.bulk_create_eligible:
            bulk_eligible += 1

    # Load vouchers for manual-match dropdown (date range of statement)
    from vouchers.models import Voucher as V
    vouchers_qs = V.objects.filter(
        company=company,
        status="APPROVED",
    ).order_by("-date")[:500]

    return render(request, "core/bank_statement_detail.html", {
        "statement":    statement,
        "rows":         rows,
        "total_debit":  total_debit,
        "total_credit": total_credit,
        "reconciled":   reconciled,
        "unreconciled": unreconciled,
        "suggested":    suggested,
        "duplicates":   duplicates,
        "high_confidence": high_confidence,
        "bulk_eligible": bulk_eligible,
        "vouchers_qs":  vouchers_qs,
    })


@login_required
@write_required
@require_POST
def bank_statement_reconcile(request, row_pk):
    """AJAX POST: manually reconcile / unreconcile a row."""
    company = request.current_company
    row     = get_object_or_404(
        BankStatementRow, pk=row_pk, statement__company=company
    )
    data       = json.loads(request.body)
    action     = data.get("action")      # "match" | "unmatch"
    voucher_pk = data.get("voucher_id")

    if action == "match" and voucher_pk:
        from vouchers.models import Voucher as V
        voucher = get_object_or_404(
            V, pk=voucher_pk, company=company, status="APPROVED"
        )
        if not _voucher_matches_bank_row(row, voucher):
            return JsonResponse({
                "error": (
                    "Selected voucher does not match this bank row's account, "
                    "amount, direction, and date window."
                )
            }, status=400)
        row.is_reconciled   = True
        row.matched_voucher = voucher
        row.match_confidence = 100
        row.match_reason = "Manual voucher match"
        row.save(update_fields=["is_reconciled", "matched_voucher", "match_confidence", "match_reason"])
        return JsonResponse({"ok": True, "voucher": str(voucher)})

    elif action == "unmatch":
        row.is_reconciled   = False
        row.matched_voucher = None
        row.match_confidence = 0
        row.match_reason = ""
        row.save(update_fields=["is_reconciled", "matched_voucher", "match_confidence", "match_reason"])
        return JsonResponse({"ok": True})

    return JsonResponse({"error": "Bad request"}, status=400)


@login_required
@write_required
@require_POST
def bank_statement_auto_match(request, pk):
    """Re-run auto-match on a statement (POST)."""
    company   = request.current_company
    statement = get_object_or_404(BankStatement, pk=pk, company=company)
    matched = _auto_match(statement)
    messages.success(request, f"Auto-match complete - {matched} row(s) matched.")
    return redirect("core:bank_statement_detail", pk=pk)


@login_required
@admin_required
def bank_statement_delete(request, pk):
    company   = request.current_company
    statement = get_object_or_404(BankStatement, pk=pk, company=company)
    if request.method == "POST":
        statement.delete()
        messages.success(request, "Bank statement deleted.")
        return redirect("core:bank_statement_list")
    return render(request, "core/bank_statement_confirm_delete.html", {
        "statement": statement,
    })


def _create_voucher_for_bank_row(row, party_ledger, user):
    from vouchers.models import Voucher, VoucherItem

    if row.is_reconciled:
        raise ValueError("Bank row is already reconciled.")
    if not row.statement.account_ledger_id:
        raise ValueError("Statement has no bank account ledger linked.")

    amount = _bank_row_abs_amount(row)
    if amount <= 0:
        raise ValueError("Bank row has no debit or credit amount.")

    is_withdrawal = row.debit > 0
    voucher_type = "Payment" if is_withdrawal else "Receipt"

    voucher = Voucher.objects.create(
        company=row.statement.company,
        date=row.date,
        voucher_type=voucher_type,
        narration=f"Reconciled from bank statement: {row.description}",
    )
    VoucherItem.objects.create(
        voucher=voucher,
        ledger=party_ledger,
        entry_type="DR" if is_withdrawal else "CR",
        amount=amount,
        narration=row.description,
    )
    VoucherItem.objects.create(
        voucher=voucher,
        ledger=row.statement.account_ledger,
        entry_type="CR" if is_withdrawal else "DR",
        amount=amount,
        narration=row.description,
    )
    voucher.validate_balance()
    voucher.approve(user)

    row.is_reconciled = True
    row.matched_voucher = voucher
    row.match_confidence = 100
    row.match_reason = "Voucher created from bank row suggestion"
    row.save(update_fields=["is_reconciled", "matched_voucher", "match_confidence", "match_reason"])
    return voucher


@login_required
@write_required
@require_POST
def bank_statement_bulk_action(request, pk):
    from ledger.models import Ledger

    company = request.current_company
    statement = get_object_or_404(BankStatement, pk=pk, company=company)
    action = request.POST.get("action")
    row_ids = request.POST.getlist("row_ids")

    if action != "create_selected_vouchers":
        messages.error(request, "Unknown bank reconciliation bulk action.")
        return redirect("core:bank_statement_detail", pk=statement.pk)

    _refresh_duplicate_flags(statement)
    created = 0
    skipped = 0
    rows = BankStatementRow.objects.filter(
        pk__in=row_ids,
        statement=statement,
        is_reconciled=False,
    ).select_related("suggested_ledger", "statement")

    with transaction.atomic():
        for row in rows:
            if row.potential_duplicate or not row.suggested_ledger_id or row.match_confidence < 70:
                skipped += 1
                continue
            party_ledger = Ledger.objects.get(pk=row.suggested_ledger_id, company=company)
            _create_voucher_for_bank_row(row, party_ledger, request.user)
            created += 1

    if created:
        messages.success(request, f"Created and reconciled {created} voucher(s) from selected bank rows.")
    if skipped:
        messages.warning(request, f"Skipped {skipped} row(s) because they were duplicate, low-confidence, or missing a ledger suggestion.")
    if not created and not skipped:
        messages.info(request, "No eligible bank rows were selected.")
    return redirect("core:bank_statement_detail", pk=statement.pk)


@login_required
def protected_media(request, path):
    """
    Serve company-owned media files after path normalization and DB ownership checks.
    """
    if not request.current_company:
        return HttpResponseForbidden("Please select a company first.")

    try:
        media_path = _normalize_media_path(path)
        full_path = Path(safe_join(settings.MEDIA_ROOT, media_path))
    except (SuspiciousFileOperation, ValueError):
        return HttpResponseForbidden("Access denied.")

    if not _media_path_belongs_to_company(media_path, request.current_company):
        return HttpResponseForbidden("You do not have permission to access this file.")

    if not full_path.is_file():
        raise Http404("File not found.")

    return FileResponse(open(full_path, "rb"))


def _normalize_media_path(path):
    normalized = str(path or "").replace("\\", "/").lstrip("/")
    pure_path = PurePosixPath(normalized)
    if (
        not normalized
        or pure_path.is_absolute()
        or any(part in {"", ".", ".."} for part in pure_path.parts)
    ):
        raise ValueError("Invalid media path.")
    return pure_path.as_posix()


def _media_path_belongs_to_company(media_path, company):
    from migration.models import ImportSession
    from ocr.models import OCRSubmission

    return (
        OCRSubmission.objects.filter(company=company, file=media_path).exists()
        or GSTEvidenceDocument.objects.filter(company=company, file=media_path).exists()
        or Voucher.objects.filter(company=company, document=media_path).exists()
        or ImportSession.objects.filter(company=company, file=media_path).exists()
    )


@login_required
@write_required
@require_POST
def create_voucher_from_suggestion(request, row_pk):
    """
    Creates a new Voucher based on a bank statement row and a suggested ledger.
    Then reconciles that row to the new voucher.
    """
    from ledger.models import Ledger

    company = request.current_company
    row = get_object_or_404(BankStatementRow, pk=row_pk, statement__company=company)

    try:
        data = json.loads(request.body)
        ledger_id = data.get("ledger_id")
        party_ledger = Ledger.objects.get(pk=ledger_id, company=company)
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)

    if not row.statement.account_ledger:
        return JsonResponse({"ok": False, "error": "Statement has no bank account ledger linked."}, status=400)
    if row.potential_duplicate:
        return JsonResponse({"ok": False, "error": "This bank row looks duplicated. Review it manually before creating a voucher."}, status=400)

    with transaction.atomic():
        try:
            vch = _create_voucher_for_bank_row(row, party_ledger, request.user)
        except ValueError as exc:
            return JsonResponse({"ok": False, "error": str(exc)}, status=400)

    return JsonResponse({
        "ok": True,
        "voucher": f"{vch.number} — {vch.voucher_type}"
    })


@login_required
def compliance_dashboard(request):
    """
    Shows a 'Health Check' dashboard for CAs to identify GST/TDS/Audit issues.
    """
    from vouchers import compliance
    company = request.current_company

    issues = compliance.get_compliance_issues(company)

    # Group issues by type
    by_type = {}
    for issue in issues:
        by_type.setdefault(issue['type'], []).append(issue)

    danger_count = len([i for i in issues if i['level'] == 'danger'])
    warning_count = len([i for i in issues if i['level'] == 'warning'])

    # Calculate a "Health Score" (start at 100, subtract for issues)
    # Danger = -10 pts, Warning = -3 pts
    score = 100 - (danger_count * 10) - (warning_count * 3)
    score = max(0, score)

    # SVG offset for the gauge (circumference approx 440)
    # 0% health = 440 offset, 100% health = 0 offset
    svg_offset = 440 * (1 - (score / 100))

    return render(request, "core/compliance_dashboard.html", {
        "issues": issues,
        "by_type": by_type,
        "danger_count": danger_count,
        "warning_count": warning_count,
        "health_score": score,
        "svg_offset": svg_offset,
        "title": "Compliance Health Check"
    })


def _client_fee_proxy(company):
    try:
        engagement = company.engagement
    except ClientEngagement.DoesNotExist:
        engagement = None
    if engagement and engagement.monthly_retainer > 0:
        return {
            "amount": engagement.monthly_retainer,
            "source": "Engagement retainer",
            "plan": engagement.get_service_package_display(),
            "usage_percent": None,
            "engagement": engagement,
        }

    subscription = getattr(company, "subscription", None)
    if not subscription:
        return {
            "amount": Decimal("0.00"),
            "source": "No fee data",
            "plan": "",
            "usage_percent": None,
            "engagement": engagement,
        }

    amount = Decimal(str(subscription.last_payment_amount or "0.00"))
    source = "Last payment"
    if amount <= 0:
        try:
            latest_payment = subscription.payments.filter(status="completed").first()
        except Exception:
            latest_payment = None
        if latest_payment:
            amount = Decimal(str(latest_payment.amount or "0.00"))
            source = "Latest completed payment"

    if amount <= 0:
        plan_defaults = {
            "basic": Decimal("2500.00"),
            "pro": Decimal("7500.00"),
            "enterprise": Decimal("20000.00"),
        }
        amount = plan_defaults.get(subscription.plan, Decimal("0.00"))
        source = "Plan proxy" if amount else "No fee data"

    return {
        "amount": amount,
        "source": source,
        "plan": subscription.get_plan_display(),
        "usage_percent": subscription.usage_percentage(),
        "engagement": engagement,
    }


def _workload_band(units):
    if units >= Decimal("70"):
        return "Overloaded"
    if units >= Decimal("35"):
        return "Heavy"
    if units >= Decimal("15"):
        return "Normal"
    return "Light"


def _workload_badge_class(band):
    return {
        "Overloaded": "bg-danger",
        "Heavy": "bg-warning text-dark",
        "Normal": "bg-info text-dark",
        "Light": "bg-success",
    }.get(band, "bg-secondary")


def _realization_band(fee_proxy, workload_units):
    if fee_proxy <= 0:
        return "No Fee Data", Decimal("0.00"), "bg-secondary"
    if workload_units <= 0:
        return "Scope Headroom", fee_proxy, "bg-success"

    per_unit = fee_proxy / workload_units
    if per_unit < Decimal("300"):
        return "Underpriced", per_unit, "bg-danger"
    if per_unit < Decimal("700"):
        return "Watch", per_unit, "bg-warning text-dark"
    return "Healthy", per_unit, "bg-success"


def _pricing_action(row):
    if row["realization_band"] == "No Fee Data":
        return {
            "label": "Capture Fee",
            "tone": "critical",
            "detail": "No recurring fee/payment signal exists. Add client fee data before pricing decisions.",
            "url": row["client_url"],
        }
    if row["realization_band"] == "Underpriced":
        return {
            "label": "Renegotiate",
            "tone": "critical",
            "detail": "Workload is high for the visible fee proxy. Review scope, exclusions, and monthly retainer.",
            "url": row["tasks_url"],
        }
    if row["workload_band"] == "Overloaded":
        return {
            "label": "Add Capacity",
            "tone": "warning",
            "detail": "Client workload is above capacity. Assign staff or reduce backlog before due dates slip.",
            "url": row["tasks_url"],
        }
    if row["overdue_work"] > 0:
        return {
            "label": "Clear Overdues",
            "tone": "warning",
            "detail": "Overdue work exists. Clear this before accepting more client scope.",
            "url": row["tasks_url"],
        }
    if row["workload_units"] < Decimal("10") and row["fee_proxy"] > 0:
        return {
            "label": "Grow Scope",
            "tone": "info",
            "detail": "Low current workload. This client may be suitable for advisory or automation upsell.",
            "url": row["client_url"],
        }
    return {
        "label": "On Track",
        "tone": "success",
        "detail": "Commercial load and operational work are within the current watch range.",
        "url": row["client_url"],
    }


def _can_manage_company_engagement(user, company):
    if user.is_superuser:
        return True
    return UserCompanyAccess.objects.filter(
        user=user,
        company=company,
        role__in=["Admin", "Accountant"],
    ).exists()


_ENGAGEMENT_ALERT_RANK = {
    "critical": 0,
    "high": 1,
    "warning": 2,
    "info": 3,
}


def _engagement_alert_badge_class(severity):
    return {
        "critical": "bg-danger",
        "high": "bg-warning text-dark",
        "warning": "bg-info text-dark",
        "info": "bg-secondary",
    }.get(severity, "bg-secondary")


def _engagement_alerts(company, engagement, commercial, today):
    edit_url = reverse("core:client_engagement_update", args=[company.pk])
    tasks_url = commercial.get("tasks_url") or f"{reverse('core:practice_tasks')}?{urlencode({'company': company.pk, 'status': 'open'})}"
    alerts = []

    def add(key, severity, title, detail, action_label, action_url, due_days=7):
        alerts.append({
            "key": key,
            "severity": severity,
            "rank": _ENGAGEMENT_ALERT_RANK.get(severity, 9),
            "badge_class": _engagement_alert_badge_class(severity),
            "title": title,
            "detail": detail,
            "action_label": action_label,
            "action_url": action_url,
            "task_reference": f"ENGAGE:{company.pk}:{key}",
            "due_date": today + timedelta(days=due_days),
        })

    if not engagement:
        add(
            "missing_profile",
            "critical",
            "Engagement missing",
            "No retainer, owner, renewal, or scope profile is configured for this client.",
            "Configure",
            edit_url,
            due_days=2,
        )
        return alerts

    monthly_retainer = engagement.monthly_retainer or Decimal("0.00")
    scope_summary = (engagement.scope_summary or "").strip()
    out_of_scope = (engagement.out_of_scope or "").strip()
    high_workload = commercial.get("workload_band") in {"Heavy", "Overloaded"}
    underpriced = commercial.get("realization_band") in {"No Fee Data", "Underpriced"}
    overdue_work = commercial.get("overdue_work", 0)
    open_work = commercial.get("open_work", 0)

    if monthly_retainer <= 0:
        add(
            "missing_retainer",
            "high",
            "Retainer missing",
            "Monthly retainer is zero, so workload cannot be priced or escalated cleanly.",
            "Add Fee",
            edit_url,
            due_days=5,
        )

    if not scope_summary:
        add(
            "missing_scope",
            "high",
            "Scope not captured",
            "Scope summary is empty. This increases scope creep and billing dispute risk.",
            "Define Scope",
            edit_url,
            due_days=5,
        )
    elif high_workload and len(scope_summary) < 40:
        add(
            "thin_scope",
            "warning",
            "Scope too thin for workload",
            "Workload is high, but the scope note is too short to protect the engagement.",
            "Tighten Scope",
            edit_url,
            due_days=7,
        )

    if not out_of_scope and engagement.service_package != ClientEngagement.PACKAGE_CUSTOM:
        add(
            "missing_exclusions",
            "warning",
            "Exclusions missing",
            "Out-of-scope work is not documented, making add-on billing harder to defend.",
            "Add Exclusions",
            edit_url,
            due_days=10,
        )

    if engagement.renewal_date:
        renewal_days = (engagement.renewal_date - today).days
        if renewal_days < 0:
            add(
                "renewal_overdue",
                "critical",
                "Renewal overdue",
                f"Renewal was due {abs(renewal_days)} day(s) ago.",
                "Renew Now",
                edit_url,
                due_days=2,
            )
        elif renewal_days <= 45:
            add(
                "renewal_due",
                "high",
                "Renewal due",
                f"Renewal is due in {renewal_days} day(s). Review price, scope, and owners before renewal.",
                "Review Renewal",
                edit_url,
                due_days=7,
            )
    elif engagement.status in {ClientEngagement.STATUS_ACTIVE, ClientEngagement.STATUS_ONBOARDING}:
        add(
            "renewal_date_missing",
            "warning",
            "Renewal date missing",
            "Active engagement has no renewal date, so the firm has no renewal trigger.",
            "Set Renewal",
            edit_url,
            due_days=10,
        )

    if engagement.risk_rating in {ClientEngagement.RISK_CRITICAL, ClientEngagement.RISK_HIGH}:
        if not engagement.partner_owner_id:
            add(
                "partner_owner_missing",
                "high",
                "Partner owner missing",
                "High-risk engagement has no partner owner assigned.",
                "Assign Owner",
                edit_url,
                due_days=3,
            )
        if not engagement.manager_owner_id:
            add(
                "manager_owner_missing",
                "warning",
                "Manager owner missing",
                "High-risk engagement has no manager owner assigned for day-to-day control.",
                "Assign Manager",
                edit_url,
                due_days=5,
            )

    if engagement.status in {ClientEngagement.STATUS_PAUSED, ClientEngagement.STATUS_EXITING} and overdue_work:
        add(
            "inactive_with_overdues",
            "critical",
            "Paused/exiting client has overdues",
            f"{overdue_work} overdue item(s) remain open while the engagement is paused or exiting.",
            "Clear Overdues",
            tasks_url,
            due_days=2,
        )

    if high_workload and underpriced:
        add(
            "underpriced_workload",
            "critical" if commercial.get("workload_band") == "Overloaded" else "high",
            "Underpriced workload",
            f"{commercial.get('workload_band')} workload with {commercial.get('realization_band')} realization needs pricing action.",
            "Renegotiate",
            edit_url,
            due_days=3,
        )

    if (
        engagement.service_package in {ClientEngagement.PACKAGE_BASIC, ClientEngagement.PACKAGE_GST_TDS}
        and high_workload
    ):
        add(
            "scope_creep_risk",
            "high" if commercial.get("workload_band") == "Overloaded" else "warning",
            "Scope creep risk",
            f"{engagement.get_service_package_display()} package is carrying {commercial.get('workload_band')} workload.",
            "Review Scope",
            edit_url,
            due_days=5,
        )

    if open_work >= 10 and not engagement.last_reviewed_at:
        add(
            "commercial_review_missing",
            "warning",
            "Commercial review missing",
            f"{open_work} open item(s) exist, but no last reviewed date is captured.",
            "Record Review",
            edit_url,
            due_days=10,
        )

    alerts.sort(key=lambda alert: (alert["rank"], alert["title"]))
    return alerts


def _safe_post_redirect(request, fallback_url):
    next_url = request.POST.get("next") or request.GET.get("next") or fallback_url
    if url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return next_url
    return fallback_url


@login_required
def client_engagements(request):
    companies = _companies_for_user(request.user).select_related("engagement").order_by("name")
    q = request.GET.get("q", "").strip()
    status_filter = request.GET.get("status", "all").strip() or "all"
    risk_filter = request.GET.get("risk", "all").strip() or "all"
    today = timezone.localdate()
    renewal_cutoff = today + timedelta(days=45)

    rows = []
    totals = {
        "clients": 0,
        "configured": 0,
        "monthly_retainer": Decimal("0.00"),
        "renewals_due": 0,
        "high_risk": 0,
        "missing_scope": 0,
        "alerts": 0,
        "critical_alerts": 0,
        "actionable_alerts": 0,
        "scope_creep_alerts": 0,
    }
    priority_alerts = []

    for company in companies:
        if q and q.lower() not in company.name.lower() and q.lower() not in (company.gstin or "").lower():
            continue
        try:
            engagement = company.engagement
        except ClientEngagement.DoesNotExist:
            engagement = None

        effective_status = engagement.status if engagement else "missing"
        effective_risk = engagement.risk_rating if engagement else "missing"
        if status_filter != "all" and effective_status != status_filter:
            continue
        if risk_filter != "all" and effective_risk != risk_filter:
            continue

        renewal_due = bool(engagement and engagement.renewal_date and engagement.renewal_date <= renewal_cutoff)
        missing_scope = bool(not engagement or not engagement.scope_summary.strip())
        monthly_retainer = engagement.monthly_retainer if engagement else Decimal("0.00")
        commercial = _client_workload_snapshot(company, today)
        alerts = _engagement_alerts(company, engagement, commercial, today)
        rows.append({
            "company": company,
            "engagement": engagement,
            "monthly_retainer": monthly_retainer,
            "commercial": commercial,
            "alerts": alerts,
            "top_alerts": alerts[:3],
            "extra_alert_count": max(0, len(alerts) - 3),
            "critical_alert_count": len([alert for alert in alerts if alert["severity"] == "critical"]),
            "actionable_alert_count": len([alert for alert in alerts if alert["severity"] in {"critical", "high"}]),
            "renewal_due": renewal_due,
            "missing_scope": missing_scope,
            "can_manage": _can_manage_company_engagement(request.user, company),
            "client_360_url": reverse("core:client_360", args=[company.pk]),
            "edit_url": reverse("core:client_engagement_update", args=[company.pk]),
        })
        for alert in alerts:
            priority_alerts.append({
                **alert,
                "company": company,
                "client_360_url": reverse("core:client_360", args=[company.pk]),
            })

    alert_refs = [
        alert["task_reference"]
        for row in rows
        for alert in row["alerts"]
        if alert["severity"] in {"critical", "high"}
    ]
    existing_alert_refs = set()
    if alert_refs:
        existing_alert_refs = set(
            PracticeTask.objects.filter(
                company__in=[row["company"] for row in rows],
                reference__in=alert_refs,
            )
            .exclude(status__in=[PracticeTask.STATUS_DONE, PracticeTask.STATUS_CANCELLED])
            .values_list("reference", flat=True)
        )
    for row in rows:
        row["open_alert_task_count"] = 0
        for alert in row["alerts"]:
            alert["task_exists"] = alert["task_reference"] in existing_alert_refs
            if alert["task_exists"]:
                row["open_alert_task_count"] += 1

    rows.sort(key=lambda row: (
        0 if not row["engagement"] else 1,
        -row["critical_alert_count"],
        -row["actionable_alert_count"],
        0 if row["renewal_due"] else 1,
        0 if row["engagement"] and row["engagement"].risk_rating in {ClientEngagement.RISK_CRITICAL, ClientEngagement.RISK_HIGH} else 1,
        row["company"].name,
    ))
    priority_alerts.sort(key=lambda alert: (alert["rank"], alert["company"].name, alert["title"]))

    for row in rows:
        engagement = row["engagement"]
        totals["clients"] += 1
        totals["monthly_retainer"] += row["monthly_retainer"]
        totals["alerts"] += len(row["alerts"])
        totals["critical_alerts"] += row["critical_alert_count"]
        totals["actionable_alerts"] += row["actionable_alert_count"]
        if any(alert["key"] == "scope_creep_risk" for alert in row["alerts"]):
            totals["scope_creep_alerts"] += 1
        if engagement:
            totals["configured"] += 1
            if engagement.risk_rating in {ClientEngagement.RISK_CRITICAL, ClientEngagement.RISK_HIGH}:
                totals["high_risk"] += 1
        if row["renewal_due"]:
            totals["renewals_due"] += 1
        if row["missing_scope"]:
            totals["missing_scope"] += 1

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="client-engagements-{today:%Y%m%d}.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Client",
            "GSTIN",
            "Status",
            "Package",
            "Monthly Retainer",
            "Billing Cycle",
            "Renewal Date",
            "Risk",
            "Partner Owner",
            "Manager Owner",
            "Alerts",
            "Scope Summary",
            "Out Of Scope",
        ])
        for row in rows:
            engagement = row["engagement"]
            writer.writerow([
                row["company"].name,
                row["company"].gstin or "",
                engagement.get_status_display() if engagement else "Missing",
                engagement.get_service_package_display() if engagement else "",
                engagement.monthly_retainer if engagement else "",
                engagement.get_billing_cycle_display() if engagement else "",
                engagement.renewal_date.isoformat() if engagement and engagement.renewal_date else "",
                engagement.get_risk_rating_display() if engagement else "",
                engagement.partner_owner.email if engagement and engagement.partner_owner else "",
                engagement.manager_owner.email if engagement and engagement.manager_owner else "",
                "; ".join(alert["title"] for alert in row["alerts"]),
                engagement.scope_summary if engagement else "",
                engagement.out_of_scope if engagement else "",
            ])
        return response

    return render(request, "core/client_engagements.html", {
        "rows": rows,
        "totals": totals,
        "q": q,
        "status_filter": status_filter,
        "risk_filter": risk_filter,
        "status_choices": [("all", "All Statuses"), ("missing", "Missing"), *ClientEngagement.STATUS_CHOICES],
        "risk_choices": [("all", "All Risks"), ("missing", "Missing"), *ClientEngagement.RISK_CHOICES],
        "export_query": _export_query(request),
        "priority_alerts": priority_alerts[:8],
        "title": "Client Engagements",
    })


@login_required
@require_POST
def client_engagement_alert_tasks(request):
    today = timezone.localdate()
    fallback_url = reverse("core:client_engagements")
    redirect_url = _safe_post_redirect(request, fallback_url)
    companies = _manageable_companies_for_user(request.user).select_related("engagement").order_by("name")
    company_id = request.POST.get("company_id", "").strip()
    if company_id:
        if not company_id.isdigit():
            messages.error(request, "Invalid client selected for engagement alert tasks.")
            return redirect(redirect_url)
        companies = companies.filter(pk=company_id)

    created = 0
    existing = 0
    actionable = 0
    for company in companies:
        try:
            engagement = company.engagement
        except ClientEngagement.DoesNotExist:
            engagement = None
        commercial = _client_workload_snapshot(company, today)
        alerts = _engagement_alerts(company, engagement, commercial, today)
        assigned_to = None
        if engagement:
            assigned_to = engagement.manager_owner or engagement.partner_owner

        for alert in alerts:
            if alert["severity"] not in {"critical", "high"}:
                continue
            actionable += 1
            if PracticeTask.objects.filter(company=company, reference=alert["task_reference"]).exists():
                existing += 1
                continue

            priority = (
                PracticeTask.PRIORITY_CRITICAL
                if alert["severity"] == "critical"
                else PracticeTask.PRIORITY_HIGH
            )
            title = f"{alert['title']} - {company.name}"[:160]
            task = PracticeTask.objects.create(
                company=company,
                title=title,
                task_type=PracticeTask.TYPE_OTHER,
                priority=priority,
                status=PracticeTask.STATUS_OPEN,
                due_date=alert["due_date"],
                assigned_to=assigned_to,
                created_by=request.user,
                reference=alert["task_reference"],
                description=(
                    f"{alert['detail']}\n\n"
                    f"Recommended action: {alert['action_label']}.\n"
                    f"Generated from Client Engagement Alerts."
                ),
            )
            created += 1
            AuditLog.objects.create(
                company=company,
                user=request.user,
                action=AuditLog.ACTION_CREATE,
                model_name="PracticeTask",
                record_id=task.pk,
                object_repr=task.title,
                old_data={},
                new_data={
                    "reference": task.reference,
                    "source": "client_engagement_alert",
                    "severity": alert["severity"],
                },
            )

    if created:
        messages.success(request, f"Created {created} engagement alert task(s).")
    elif actionable:
        messages.info(request, "Engagement alert tasks already exist for the current high-priority alerts.")
    else:
        messages.info(request, "No high-priority engagement alerts need tasks right now.")
    if existing and created:
        messages.info(request, f"Skipped {existing} alert task(s) that already exist.")
    return redirect(redirect_url)


@login_required
def client_engagement_update(request, company_id):
    companies = _companies_for_user(request.user)
    company = get_object_or_404(companies, pk=company_id)
    if not _can_manage_company_engagement(request.user, company):
        messages.error(request, "You do not have permission to edit this client engagement.")
        return redirect("core:client_engagements")

    try:
        engagement = company.engagement
        created = False
    except ClientEngagement.DoesNotExist:
        engagement = ClientEngagement(company=company)
        created = True
    users = get_user_model().objects.filter(company_access__company=company, is_active=True).distinct().order_by("email")
    if request.method == "POST":
        before = {
            "status": engagement.status,
            "service_package": engagement.service_package,
            "monthly_retainer": str(engagement.monthly_retainer),
            "billing_cycle": engagement.billing_cycle,
            "renewal_date": engagement.renewal_date.isoformat() if engagement.renewal_date else "",
            "risk_rating": engagement.risk_rating,
        }
        form = ClientEngagementForm(request.POST, instance=engagement, users=users)
        if form.is_valid():
            updated = form.save(commit=False)
            updated.company = company
            updated.save()
            AuditLog.objects.create(
                company=company,
                user=request.user,
                action=AuditLog.ACTION_CREATE if created else AuditLog.ACTION_UPDATE,
                model_name="ClientEngagement",
                record_id=updated.pk,
                object_repr=f"{company.name} engagement",
                old_data={} if created else before,
                new_data={
                    "status": updated.status,
                    "service_package": updated.service_package,
                    "monthly_retainer": str(updated.monthly_retainer),
                    "billing_cycle": updated.billing_cycle,
                    "renewal_date": updated.renewal_date.isoformat() if updated.renewal_date else "",
                    "risk_rating": updated.risk_rating,
                },
            )
            messages.success(request, "Client engagement saved.")
            next_url = request.POST.get("next") or reverse("core:client_engagements")
            return redirect(next_url)
    else:
        form = ClientEngagementForm(instance=engagement, users=users)

    return render(request, "core/client_engagement_form.html", {
        "company": company,
        "engagement": engagement,
        "form": form,
        "created": created,
        "next": request.GET.get("next", ""),
        "title": f"{company.name} Engagement",
    })


def _month_bounds_from_request(value):
    today = timezone.localdate()
    if value:
        try:
            year, month = [int(part) for part in value.split("-", 1)]
            start = _date(year, month, 1)
        except (TypeError, ValueError):
            start = today.replace(day=1)
    else:
        start = today.replace(day=1)
    end = start.replace(day=calendar.monthrange(start.year, start.month)[1])
    return start, end


def _client_workload_snapshot(company, today):
    from .models import BankStatementRow
    from portal.models import ClientDocumentRequest

    window_start = today - timedelta(days=30)
    tasks_open_qs = PracticeTask.objects.filter(company=company).exclude(
        status__in=[PracticeTask.STATUS_DONE, PracticeTask.STATUS_CANCELLED]
    )
    filings_open_qs = ComplianceFiling.objects.filter(company=company).exclude(
        status__in=[ComplianceFiling.STATUS_FILED, ComplianceFiling.STATUS_CANCELLED]
    )
    notices_open_qs = ComplianceNotice.objects.filter(company=company).exclude(
        status=ComplianceNotice.STATUS_CLOSED
    )
    document_requests_qs = ClientDocumentRequest.objects.filter(company=company).exclude(
        status__in=[
            ClientDocumentRequest.STATUS_CLOSED,
            ClientDocumentRequest.STATUS_CANCELLED,
        ]
    )

    open_tasks = tasks_open_qs.count()
    overdue_tasks = tasks_open_qs.filter(due_date__lt=today).count()
    critical_tasks = tasks_open_qs.filter(priority=PracticeTask.PRIORITY_CRITICAL).count()
    blocked_tasks = tasks_open_qs.filter(status=PracticeTask.STATUS_BLOCKED).count()
    open_filings = filings_open_qs.count()
    overdue_filings = filings_open_qs.filter(due_date__lt=today).count()
    open_notices = notices_open_qs.count()
    overdue_notices = notices_open_qs.filter(response_due_date__lt=today).count()
    open_requests = document_requests_qs.filter(status=ClientDocumentRequest.STATUS_OPEN).count()
    uploads_for_review = document_requests_qs.filter(status=ClientDocumentRequest.STATUS_UPLOADED).count()
    overdue_requests = document_requests_qs.filter(
        status=ClientDocumentRequest.STATUS_OPEN,
        due_date__lt=today,
    ).count()
    unreconciled_bank_rows = BankStatementRow.objects.filter(
        statement__company=company,
        is_reconciled=False,
    ).count()
    voucher_volume = Voucher.objects.filter(
        company=company,
        date__gte=window_start,
        date__lte=today,
    ).count()

    fee = _client_fee_proxy(company)
    workload_units = (
        Decimal(open_tasks * 2)
        + Decimal(overdue_tasks * 4)
        + Decimal(critical_tasks * 5)
        + Decimal(blocked_tasks * 3)
        + Decimal(open_filings * 3)
        + Decimal(overdue_filings * 5)
        + Decimal(open_notices * 4)
        + Decimal(overdue_notices * 6)
        + Decimal(open_requests) * Decimal("1.5")
        + Decimal(uploads_for_review * 2)
        + Decimal(overdue_requests * 3)
        + min(Decimal(unreconciled_bank_rows) * Decimal("0.25"), Decimal("25"))
        + min(Decimal(voucher_volume) * Decimal("0.20"), Decimal("30"))
    ).quantize(Decimal("0.01"))
    workload_band = _workload_band(workload_units)
    realization_band, realization_per_unit, realization_badge_class = _realization_band(
        fee["amount"],
        workload_units,
    )
    tasks_url = f"{reverse('core:practice_tasks')}?{urlencode({'company': company.pk, 'status': 'open'})}"
    client_url = f"{reverse('core:switch_company', args=[company.pk])}?{urlencode({'next': reverse('core:dashboard')})}"
    row = {
        "company": company,
        "fee_proxy": fee["amount"],
        "fee_source": fee["source"],
        "plan": fee["plan"],
        "usage_percent": fee["usage_percent"],
        "engagement": fee["engagement"],
        "workload_units": workload_units,
        "workload_band": workload_band,
        "workload_badge_class": _workload_badge_class(workload_band),
        "realization_band": realization_band,
        "realization_per_unit": realization_per_unit.quantize(Decimal("0.01")),
        "realization_badge_class": realization_badge_class,
        "open_tasks": open_tasks,
        "overdue_tasks": overdue_tasks,
        "critical_tasks": critical_tasks,
        "blocked_tasks": blocked_tasks,
        "open_filings": open_filings,
        "overdue_filings": overdue_filings,
        "open_notices": open_notices,
        "overdue_notices": overdue_notices,
        "open_requests": open_requests,
        "uploads_for_review": uploads_for_review,
        "overdue_requests": overdue_requests,
        "unreconciled_bank_rows": unreconciled_bank_rows,
        "voucher_volume": voucher_volume,
        "overdue_work": overdue_tasks + overdue_filings + overdue_notices + overdue_requests,
        "open_work": open_tasks + open_filings + open_notices + open_requests + uploads_for_review,
        "tasks_url": tasks_url,
        "client_url": client_url,
        "client_360_url": reverse("core:client_360", args=[company.pk]),
    }
    row["pricing_action"] = _pricing_action(row)
    return row


_PARTNER_REVIEW_RANK = {
    "critical": 0,
    "high": 1,
    "warning": 2,
    "info": 3,
}


def _partner_review_badge_class(severity):
    return {
        "critical": "bg-danger",
        "high": "bg-warning text-dark",
        "warning": "bg-info text-dark",
        "info": "bg-secondary",
    }.get(severity, "bg-secondary")


def _partner_review_focus_match(item, focus):
    if focus == "all":
        return True
    if focus == "blocked":
        return item["severity"] == "critical" or item["decision_type"] == "block"
    if focus == "approvals":
        return item["decision_type"] in {"approve", "signoff"}
    if focus == "commercial":
        return item["gate"] in {"Engagement", "Billing"}
    if focus == "communications":
        return item["gate"] in {"Client Comms", "Notice"}
    return True


def _add_partner_review_item(
    items,
    *,
    company,
    gate,
    severity,
    title,
    reason,
    decision,
    action_label,
    action_url,
    count=1,
    due_date=None,
    owner=None,
    evidence=None,
    decision_type="review",
):
    items.append({
        "company": company,
        "gate": gate,
        "severity": severity,
        "rank": _PARTNER_REVIEW_RANK.get(severity, 9),
        "badge_class": _partner_review_badge_class(severity),
        "title": title,
        "reason": reason,
        "decision": decision,
        "decision_type": decision_type,
        "action_label": action_label,
        "action_url": action_url,
        "count": count,
        "due_date": due_date,
        "owner": owner,
        "evidence": evidence or [],
    })


@login_required
def partner_review_cockpit(request):
    from portal.models import ClientDocumentRequest

    companies = _companies_for_user(request.user).select_related("engagement").order_by("name")
    q = request.GET.get("q", "").strip()
    focus = request.GET.get("focus", "all").strip() or "all"
    focus_options = [
        ("all", "All Sign-offs"),
        ("blocked", "Blocked / Critical"),
        ("approvals", "Ready for Approval"),
        ("commercial", "Commercial"),
        ("communications", "Client Comms"),
    ]
    if focus not in {value for value, _label in focus_options}:
        focus = "all"

    today = timezone.localdate()
    period_start, period_end = _month_bounds_from_request(request.GET.get("period", "").strip())
    period_value = period_start.strftime("%Y-%m")
    next_7 = today + timedelta(days=7)
    all_items = []
    rows = []
    totals = {
        "clients": 0,
        "clients_with_items": 0,
        "items": 0,
        "critical": 0,
        "high": 0,
        "ready": 0,
        "blocked": 0,
        "commercial": 0,
        "communications": 0,
        "filings": 0,
        "notices": 0,
        "billing": 0,
    }

    for company in companies:
        if q and q.lower() not in company.name.lower() and q.lower() not in (company.gstin or "").lower():
            continue

        totals["clients"] += 1
        try:
            engagement = company.engagement
        except ClientEngagement.DoesNotExist:
            engagement = None
        owner = None
        if engagement:
            owner = engagement.partner_owner or engagement.manager_owner

        commercial = _client_workload_snapshot(company, today)
        review_url = f"{reverse('core:filing_review_center')}?{urlencode({'period': period_value, 'company': company.pk, 'review_type': FilingReview.TYPE_GST_MONTHLY})}"
        gst_url = reverse("core:gst_workbench_detail", args=[company.pk, period_value])
        company_items = []

        filing_review = build_filing_review(
            company,
            period_start,
            period_end,
            FilingReview.TYPE_GST_MONTHLY,
        )
        review = filing_review["review"]
        review_status = review.status if review else ""
        if filing_review["unwaived_critical_count"]:
            first_issue = next(
                (item for item in filing_review["issue_checks"] if item["severity"] == "critical" and not item["is_waived"]),
                None,
            )
            _add_partner_review_item(
                company_items,
                company=company,
                gate="Filing",
                severity="critical",
                title="GST filing blocked",
                reason=(
                    f"{filing_review['unwaived_critical_count']} unwaived blocker(s) must be cleared before filing approval."
                ),
                decision="Block filing until the correction or waiver is documented.",
                action_label="Open Review Center",
                action_url=review_url,
                count=filing_review["unwaived_critical_count"],
                due_date=today,
                owner=owner,
                evidence=[first_issue["title"]] if first_issue else [],
                decision_type="block",
            )
        elif filing_review["ready_to_file"] and (not review or not review.is_approved):
            _add_partner_review_item(
                company_items,
                company=company,
                gate="Filing",
                severity="high",
                title="GST filing ready for partner sign-off",
                reason=f"Readiness is {filing_review['readiness_score']}% with no unwaived critical blockers.",
                decision="Approve for filing, send back, or add a waiver note before returns are filed.",
                action_label="Approve Review",
                action_url=review_url,
                count=1,
                due_date=next_7,
                owner=owner,
                evidence=[review.get_status_display() if review else "Review not started"],
                decision_type="approve",
            )
        elif review_status == FilingReview.STATUS_SENT_BACK:
            _add_partner_review_item(
                company_items,
                company=company,
                gate="Filing",
                severity="warning",
                title="Filing review sent back",
                reason="Corrections are pending after partner review.",
                decision="Track correction closure before the filing pack is regenerated.",
                action_label="Open Review Center",
                action_url=review_url,
                count=1,
                owner=owner,
                evidence=[review.notes] if review and review.notes else [],
                decision_type="review",
            )

        gst_snapshot = _gst_workbench_snapshot(company, period_start, period_end)
        if gst_snapshot["signoff_blockers"]:
            first_blocker = gst_snapshot["signoff_blockers"][0]
            _add_partner_review_item(
                company_items,
                company=company,
                gate="GST",
                severity="critical" if gst_snapshot["risk_score"] >= 50 else "warning",
                title="GST sign-off blocked",
                reason=f"{len(gst_snapshot['signoff_blockers'])} GST blocker(s), starting with {first_blocker['title']}.",
                decision="Hold GST sign-off until the blocker is cleared or formally accepted.",
                action_label="Open GST Workbench",
                action_url=gst_url,
                count=len(gst_snapshot["signoff_blockers"]),
                due_date=today if gst_snapshot["risk_score"] >= 50 else None,
                owner=owner,
                evidence=[first_blocker["description"]],
                decision_type="block",
            )

        ready_filings = ComplianceFiling.objects.filter(
            company=company,
            status=ComplianceFiling.STATUS_READY_FOR_REVIEW,
        ).exclude(status__in=[ComplianceFiling.STATUS_FILED, ComplianceFiling.STATUS_CANCELLED])
        ready_filing_count = ready_filings.count()
        if ready_filing_count:
            next_due = ready_filings.order_by("due_date").values_list("due_date", flat=True).first()
            _add_partner_review_item(
                company_items,
                company=company,
                gate="Filing",
                severity="high",
                title="Statutory filing ready for sign-off",
                reason=f"{ready_filing_count} filing workflow(s) are marked ready for review.",
                decision="Approve, file, or send back before the statutory due date.",
                action_label="Open Filings",
                action_url=f"{reverse('core:compliance_filings')}?{urlencode({'company': company.pk, 'status': ComplianceFiling.STATUS_READY_FOR_REVIEW})}",
                count=ready_filing_count,
                due_date=next_due,
                owner=owner,
                decision_type="signoff",
            )

        overdue_filings = ComplianceFiling.objects.filter(
            company=company,
            due_date__lt=today,
        ).exclude(status__in=[ComplianceFiling.STATUS_FILED, ComplianceFiling.STATUS_CANCELLED])
        if overdue_filings.exists():
            _add_partner_review_item(
                company_items,
                company=company,
                gate="Filing",
                severity="critical",
                title="Filing overdue",
                reason=f"{overdue_filings.count()} filing workflow(s) have crossed due date.",
                decision="Partner intervention is required before further client-facing commitments.",
                action_label="Open Filings",
                action_url=f"{reverse('core:compliance_filings')}?{urlencode({'company': company.pk, 'status': 'open'})}",
                count=overdue_filings.count(),
                due_date=today,
                owner=owner,
                decision_type="block",
            )

        response_ready_notices = ComplianceNotice.objects.filter(
            company=company,
            status=ComplianceNotice.STATUS_RESPONSE_READY,
        )
        if response_ready_notices.exists():
            next_due = response_ready_notices.order_by("response_due_date").values_list("response_due_date", flat=True).first()
            _add_partner_review_item(
                company_items,
                company=company,
                gate="Notice",
                severity="high",
                title="Notice response ready for partner sign-off",
                reason=f"{response_ready_notices.count()} notice response(s) are ready to be approved or revised.",
                decision="Approve the response, revise it, or escalate before it is sent.",
                action_label="Open Notices",
                action_url=f"{reverse('core:compliance_notices')}?{urlencode({'company': company.pk, 'status': ComplianceNotice.STATUS_RESPONSE_READY})}",
                count=response_ready_notices.count(),
                due_date=next_due,
                owner=owner,
                decision_type="approve",
            )

        urgent_notices = ComplianceNotice.objects.filter(company=company).exclude(
            status=ComplianceNotice.STATUS_CLOSED
        ).filter(Q(response_due_date__lt=today) | Q(status=ComplianceNotice.STATUS_ESCALATED))
        if urgent_notices.exists():
            _add_partner_review_item(
                company_items,
                company=company,
                gate="Notice",
                severity="critical",
                title="Notice intervention required",
                reason=f"{urgent_notices.count()} open notice(s) are overdue or escalated.",
                decision="Intervene before the client response position is finalised.",
                action_label="Open Notices",
                action_url=f"{reverse('core:compliance_notices')}?{urlencode({'company': company.pk, 'status': 'open'})}",
                count=urgent_notices.count(),
                due_date=today,
                owner=owner,
                decision_type="block",
            )

        pending_sales = Voucher.objects.filter(
            company=company,
            voucher_type="Sales",
            status__in=["DRAFT", "PENDING"],
        )
        if pending_sales.exists():
            pending_amount = pending_sales.aggregate(total=Sum("outstanding_amount"))["total"] or Decimal("0.00")
            _add_partner_review_item(
                company_items,
                company=company,
                gate="Billing",
                severity="high" if engagement and engagement.risk_rating in {ClientEngagement.RISK_HIGH, ClientEngagement.RISK_CRITICAL} else "warning",
                title="Sales invoices awaiting approval",
                reason=f"{pending_sales.count()} sales invoice(s) are not approved yet.",
                decision="Approve or send back before invoice email/WhatsApp dispatch.",
                action_label="Open Vouchers",
                action_url=f"{reverse('core:switch_company', args=[company.pk])}?{urlencode({'next': reverse('vouchers:list')})}",
                count=pending_sales.count(),
                owner=owner,
                evidence=[f"Pending amount signal Rs. {pending_amount:.2f}"],
                decision_type="approve",
            )

        overdue_receivables = Voucher.objects.filter(
            company=company,
            voucher_type="Sales",
            status="APPROVED",
            outstanding_amount__gt=0,
            due_date__lt=today,
        )
        if overdue_receivables.exists() and engagement and engagement.risk_rating in {ClientEngagement.RISK_HIGH, ClientEngagement.RISK_CRITICAL}:
            overdue_amount = overdue_receivables.aggregate(total=Sum("outstanding_amount"))["total"] or Decimal("0.00")
            _add_partner_review_item(
                company_items,
                company=company,
                gate="Billing",
                severity="warning",
                title="High-risk client has overdue receivables",
                reason=f"{overdue_receivables.count()} invoice(s) overdue for Rs. {overdue_amount:.2f}.",
                decision="Approve collection tone or pause new scope until payment position is clear.",
                action_label="Open Outstanding",
                action_url=f"{reverse('core:switch_company', args=[company.pk])}?{urlencode({'next': reverse('vouchers:outstanding')})}",
                count=overdue_receivables.count(),
                due_date=today,
                owner=owner,
                decision_type="review",
            )

        for alert in _engagement_alerts(company, engagement, commercial, today):
            if alert["severity"] not in {"critical", "high"}:
                continue
            _add_partner_review_item(
                company_items,
                company=company,
                gate="Engagement",
                severity=alert["severity"],
                title=alert["title"],
                reason=alert["detail"],
                decision="Review commercial risk before accepting more work or renewing scope.",
                action_label=alert["action_label"],
                action_url=alert["action_url"],
                count=1,
                due_date=alert["due_date"],
                owner=owner,
                evidence=[alert["task_reference"]],
                decision_type="review",
            )

        uploaded_requests = ClientDocumentRequest.objects.filter(
            company=company,
            status=ClientDocumentRequest.STATUS_UPLOADED,
        )
        if uploaded_requests.exists():
            _add_partner_review_item(
                company_items,
                company=company,
                gate="Client Comms",
                severity="warning",
                title="Client uploads need review before reply",
                reason=f"{uploaded_requests.count()} uploaded document request(s) are waiting for CA review.",
                decision="Review evidence before replying to the client or closing the request.",
                action_label="Open Client Requests",
                action_url=f"{reverse('portal:client_requests')}?{urlencode({'company': company.pk, 'status': ClientDocumentRequest.STATUS_UPLOADED})}",
                count=uploaded_requests.count(),
                owner=owner,
                decision_type="review",
            )

        overdue_requests = ClientDocumentRequest.objects.filter(
            company=company,
            status=ClientDocumentRequest.STATUS_OPEN,
            due_date__lt=today,
        )
        if overdue_requests.exists():
            _add_partner_review_item(
                company_items,
                company=company,
                gate="Client Comms",
                severity="warning",
                title="Client chase overdue",
                reason=f"{overdue_requests.count()} client request(s) are overdue.",
                decision="Approve escalation tone and reminder channel before chasing again.",
                action_label="Open Reminders",
                action_url=f"{reverse('portal:client_request_reminders')}?{urlencode({'company': company.pk, 'kind': 'overdue'})}",
                count=overdue_requests.count(),
                due_date=today,
                owner=owner,
                decision_type="review",
            )

        company_items.sort(key=lambda item: (
            item["rank"],
            item["due_date"] or _date.max,
            -item["count"],
            item["title"],
        ))
        filtered_items = [item for item in company_items if _partner_review_focus_match(item, focus)]
        if filtered_items:
            rows.append({
                "company": company,
                "engagement": engagement,
                "owner": owner,
                "items": filtered_items[:6],
                "top_item": filtered_items[0],
                "item_count": len(filtered_items),
                "critical_count": sum(1 for item in filtered_items if item["severity"] == "critical"),
                "high_count": sum(1 for item in filtered_items if item["severity"] == "high"),
                "client_360_url": reverse("core:client_360", args=[company.pk]),
                "review_url": review_url,
                "gst_url": gst_url,
            })
            all_items.extend(filtered_items)

    all_items.sort(key=lambda item: (
        item["rank"],
        item["due_date"] or _date.max,
        -item["count"],
        item["company"].name,
        item["title"],
    ))
    rows.sort(key=lambda row: (
        row["top_item"]["rank"],
        row["top_item"]["due_date"] or _date.max,
        -row["item_count"],
        row["company"].name,
    ))

    totals["items"] = len(all_items)
    totals["clients_with_items"] = len(rows)
    totals["critical"] = sum(1 for item in all_items if item["severity"] == "critical")
    totals["high"] = sum(1 for item in all_items if item["severity"] == "high")
    totals["ready"] = sum(1 for item in all_items if item["decision_type"] in {"approve", "signoff"})
    totals["blocked"] = sum(1 for item in all_items if item["decision_type"] == "block" or item["severity"] == "critical")
    totals["commercial"] = sum(1 for item in all_items if item["gate"] in {"Engagement", "Billing"})
    totals["communications"] = sum(1 for item in all_items if item["gate"] in {"Client Comms", "Notice"})
    totals["filings"] = sum(1 for item in all_items if item["gate"] in {"Filing", "GST"})
    totals["notices"] = sum(1 for item in all_items if item["gate"] == "Notice")
    totals["billing"] = sum(1 for item in all_items if item["gate"] == "Billing")

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="partner-review-cockpit-{today:%Y%m%d}.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Client",
            "GSTIN",
            "Gate",
            "Severity",
            "Sign-off Item",
            "Reason",
            "Exact Action",
            "Action Link Label",
            "Count",
            "Due Date",
            "Owner",
            "Evidence",
        ])
        for item in all_items:
            writer.writerow([
                item["company"].name,
                item["company"].gstin or "",
                item["gate"],
                item["severity"],
                item["title"],
                item["reason"],
                item["decision"],
                item["action_label"],
                item["count"],
                item["due_date"].isoformat() if item["due_date"] else "",
                item["owner"].email if item["owner"] else "",
                "; ".join(str(value) for value in item["evidence"]),
            ])
        return response

    return render(request, "core/partner_review_cockpit.html", {
        "rows": rows,
        "items": all_items[:12],
        "totals": totals,
        "focus": focus,
        "focus_options": focus_options,
        "q": q,
        "period_value": period_value,
        "period_start": period_start,
        "period_end": period_end,
        "today": today,
        "export_query": _export_query(request),
        "title": "Partner Review Cockpit",
    })


@login_required
def client_360(request, company_id):
    from vouchers import compliance
    from gstr2b.models import PortalGSTR2BEntry
    from portal.models import ClientDocumentRequest
    from .models import BankStatementRow

    companies = _companies_for_user(request.user)
    company = get_object_or_404(companies, pk=company_id)
    today = timezone.localdate()
    period_start, period_end = _month_bounds_from_request(request.GET.get("period", "").strip())
    period_value = period_start.strftime("%Y-%m")

    commercial = _client_workload_snapshot(company, today)
    issues = compliance.get_compliance_issues(company)
    danger_count = len([issue for issue in issues if issue.get("level") == "danger"])
    warning_count = len([issue for issue in issues if issue.get("level") == "warning"])
    compliance_score = max(0, 100 - (danger_count * 10) - (warning_count * 3))
    gst_snapshot = _gst_workbench_snapshot(company, period_start, period_end)

    open_tasks_qs = PracticeTask.objects.filter(company=company).exclude(
        status__in=[PracticeTask.STATUS_DONE, PracticeTask.STATUS_CANCELLED]
    ).select_related("assigned_to").order_by("due_date", "-priority", "title")
    open_filings_qs = ComplianceFiling.objects.filter(company=company).exclude(
        status__in=[ComplianceFiling.STATUS_FILED, ComplianceFiling.STATUS_CANCELLED]
    ).select_related("assigned_to", "reviewer").order_by("due_date", "-priority", "title")
    open_notices_qs = ComplianceNotice.objects.filter(company=company).exclude(
        status=ComplianceNotice.STATUS_CLOSED
    ).select_related("assigned_to").order_by("response_due_date", "-priority", "title")
    client_requests_qs = ClientDocumentRequest.objects.filter(company=company).exclude(
        status__in=[ClientDocumentRequest.STATUS_CLOSED, ClientDocumentRequest.STATUS_CANCELLED]
    ).select_related("portal_user").order_by("due_date", "status", "title")

    sales_outstanding_qs = Voucher.objects.filter(
        company=company,
        voucher_type="Sales",
        status="APPROVED",
        outstanding_amount__gt=0,
    )
    overdue_sales_qs = sales_outstanding_qs.filter(due_date__lt=today)
    purchase_outstanding_qs = Voucher.objects.filter(
        company=company,
        voucher_type="Purchase",
        status="APPROVED",
        outstanding_amount__gt=0,
    )
    accounting = {
        "voucher_volume_30d": commercial["voucher_volume"],
        "draft_vouchers": Voucher.objects.filter(company=company, status__in=["DRAFT", "PENDING"]).count(),
        "sales_outstanding_count": sales_outstanding_qs.count(),
        "sales_outstanding_amount": sales_outstanding_qs.aggregate(total=Sum("outstanding_amount"))["total"] or Decimal("0.00"),
        "overdue_sales_count": overdue_sales_qs.count(),
        "overdue_sales_amount": overdue_sales_qs.aggregate(total=Sum("outstanding_amount"))["total"] or Decimal("0.00"),
        "purchase_outstanding_count": purchase_outstanding_qs.count(),
        "purchase_outstanding_amount": purchase_outstanding_qs.aggregate(total=Sum("outstanding_amount"))["total"] or Decimal("0.00"),
        "unreconciled_bank_rows": BankStatementRow.objects.filter(statement__company=company, is_reconciled=False).count(),
        "ledgers": Ledger.objects.filter(company=company).count(),
    }
    gstr2b_summary = {
        "missing_in_books": PortalGSTR2BEntry.objects.filter(company=company, match_status="missing_in_books").count(),
        "pending_action": PortalGSTR2BEntry.objects.filter(company=company, action_status="pending").count(),
        "rejected": PortalGSTR2BEntry.objects.filter(company=company, action_status="rejected").count(),
    }

    next_actions = []
    engagement_alerts = _engagement_alerts(company, commercial["engagement"], commercial, today)
    actionable_engagement_alerts = len([
        alert for alert in engagement_alerts if alert["severity"] in {"critical", "high"}
    ])

    def add_action(severity, title, detail, label, url, count=1):
        next_actions.append({
            "severity": severity,
            "rank": {"critical": 0, "warning": 1, "info": 2, "success": 3}.get(severity, 4),
            "title": title,
            "detail": detail,
            "label": label,
            "url": url,
            "count": count,
        })

    for alert in engagement_alerts[:4]:
        add_action(
            "critical" if alert["severity"] == "critical" else "warning",
            alert["title"],
            alert["detail"],
            alert["action_label"],
            alert["action_url"],
        )
    if commercial["pricing_action"]["label"] not in {"On Track", "Grow Scope"}:
        add_action(
            "critical" if commercial["pricing_action"]["tone"] == "critical" else "warning",
            commercial["pricing_action"]["label"],
            commercial["pricing_action"]["detail"],
            commercial["pricing_action"]["label"],
            commercial["pricing_action"]["url"],
        )
    if commercial["overdue_tasks"]:
        add_action("critical", "Practice tasks overdue", f"{commercial['overdue_tasks']} task(s) are overdue.", "Open Work Queue", commercial["tasks_url"], commercial["overdue_tasks"])
    if commercial["overdue_filings"]:
        add_action("critical", "Filings overdue", f"{commercial['overdue_filings']} filing(s) are overdue.", "Open Filings", f"{reverse('core:compliance_filings')}?{urlencode({'company': company.pk})}", commercial["overdue_filings"])
    if commercial["overdue_notices"]:
        add_action("critical", "Notice responses overdue", f"{commercial['overdue_notices']} notice response(s) are overdue.", "Open Notices", f"{reverse('core:compliance_notices')}?{urlencode({'company': company.pk})}", commercial["overdue_notices"])
    if commercial["overdue_requests"]:
        add_action("critical", "Client document chase overdue", f"{commercial['overdue_requests']} document request(s) are overdue.", "Open Reminders", f"{reverse('portal:client_request_reminders')}?{urlencode({'company': company.pk})}", commercial["overdue_requests"])
    if gst_snapshot["risk_score"]:
        add_action("warning", "GST review needs attention", f"GST risk score is {gst_snapshot['risk_score']} for {period_value}.", "Open GST Workbench", reverse("core:gst_workbench_detail", args=[company.pk, period_value]), gst_snapshot["risk_score"])
    if accounting["unreconciled_bank_rows"]:
        add_action("warning", "Bank reconciliation pending", f"{accounting['unreconciled_bank_rows']} bank row(s) are unreconciled.", "Open Bank", f"{reverse('core:switch_company', args=[company.pk])}?{urlencode({'next': reverse('core:bank_statement_list')})}", accounting["unreconciled_bank_rows"])
    if accounting["overdue_sales_count"]:
        add_action("warning", "Receivables overdue", f"{accounting['overdue_sales_count']} sales invoice(s) are overdue.", "Open Outstanding", f"{reverse('core:switch_company', args=[company.pk])}?{urlencode({'next': reverse('vouchers:outstanding')})}", accounting["overdue_sales_count"])
    if not next_actions:
        add_action("success", "Client is on track", "No urgent exception is visible in current operational data.", "Open Client", commercial["client_url"])
    next_actions.sort(key=lambda item: (item["rank"], -item["count"], item["title"]))

    recent_vouchers = Voucher.objects.filter(company=company).order_by("-date", "-created_at")[:8]
    recent_audit_logs = AuditLog.objects.filter(company=company).select_related("user").order_by("-timestamp")[:8]

    return render(request, "core/client_360.html", {
        "company": company,
        "period_value": period_value,
        "period_start": period_start,
        "period_end": period_end,
        "commercial": commercial,
        "compliance": {
            "score": compliance_score,
            "danger_count": danger_count,
            "warning_count": warning_count,
            "issues": issues[:8],
        },
        "gst_snapshot": gst_snapshot,
        "accounting": accounting,
        "gstr2b_summary": gstr2b_summary,
        "engagement_alerts": engagement_alerts,
        "actionable_engagement_alerts": actionable_engagement_alerts,
        "next_actions": next_actions[:8],
        "open_tasks": open_tasks_qs[:8],
        "open_filings": open_filings_qs[:8],
        "open_notices": open_notices_qs[:8],
        "client_requests": client_requests_qs[:8],
        "recent_vouchers": recent_vouchers,
        "recent_audit_logs": recent_audit_logs,
        "today": today,
        "title": f"{company.name} Client 360",
    })


@login_required
def ca_client_profitability(request):
    from .models import BankStatementRow
    from portal.models import ClientDocumentRequest

    companies = _companies_for_user(request.user)
    today = timezone.localdate()
    window_start = today - timedelta(days=30)
    q = request.GET.get("q", "").strip()
    band_filter = request.GET.get("band", "all").strip() or "all"

    rows = []
    totals = {
        "clients": 0,
        "fee_proxy": Decimal("0.00"),
        "workload_units": Decimal("0.00"),
        "overloaded": 0,
        "underpriced": 0,
        "no_fee": 0,
        "overdue_work": 0,
        "open_work": 0,
        "avg_realization": Decimal("0.00"),
    }

    for company in companies:
        if q and q.lower() not in company.name.lower() and q.lower() not in (company.gstin or "").lower():
            continue

        tasks_open_qs = PracticeTask.objects.filter(company=company).exclude(
            status__in=[PracticeTask.STATUS_DONE, PracticeTask.STATUS_CANCELLED]
        )
        open_tasks = tasks_open_qs.count()
        overdue_tasks = tasks_open_qs.filter(due_date__lt=today).count()
        critical_tasks = tasks_open_qs.filter(priority=PracticeTask.PRIORITY_CRITICAL).count()
        blocked_tasks = tasks_open_qs.filter(status=PracticeTask.STATUS_BLOCKED).count()

        filings_open_qs = ComplianceFiling.objects.filter(company=company).exclude(
            status__in=[ComplianceFiling.STATUS_FILED, ComplianceFiling.STATUS_CANCELLED]
        )
        open_filings = filings_open_qs.count()
        overdue_filings = filings_open_qs.filter(due_date__lt=today).count()

        notices_open_qs = ComplianceNotice.objects.filter(company=company).exclude(
            status=ComplianceNotice.STATUS_CLOSED
        )
        open_notices = notices_open_qs.count()
        overdue_notices = notices_open_qs.filter(response_due_date__lt=today).count()

        document_requests_qs = ClientDocumentRequest.objects.filter(company=company).exclude(
            status__in=[
                ClientDocumentRequest.STATUS_CLOSED,
                ClientDocumentRequest.STATUS_CANCELLED,
            ]
        )
        open_requests = document_requests_qs.filter(status=ClientDocumentRequest.STATUS_OPEN).count()
        uploads_for_review = document_requests_qs.filter(status=ClientDocumentRequest.STATUS_UPLOADED).count()
        overdue_requests = document_requests_qs.filter(
            status=ClientDocumentRequest.STATUS_OPEN,
            due_date__lt=today,
        ).count()

        unreconciled_bank_rows = BankStatementRow.objects.filter(
            statement__company=company,
            is_reconciled=False,
        ).count()
        voucher_volume = Voucher.objects.filter(
            company=company,
            date__gte=window_start,
            date__lte=today,
        ).count()

        fee = _client_fee_proxy(company)
        workload_units = (
            Decimal(open_tasks * 2)
            + Decimal(overdue_tasks * 4)
            + Decimal(critical_tasks * 5)
            + Decimal(blocked_tasks * 3)
            + Decimal(open_filings * 3)
            + Decimal(overdue_filings * 5)
            + Decimal(open_notices * 4)
            + Decimal(overdue_notices * 6)
            + Decimal(open_requests) * Decimal("1.5")
            + Decimal(uploads_for_review * 2)
            + Decimal(overdue_requests * 3)
            + min(Decimal(unreconciled_bank_rows) * Decimal("0.25"), Decimal("25"))
            + min(Decimal(voucher_volume) * Decimal("0.20"), Decimal("30"))
        ).quantize(Decimal("0.01"))
        workload_band = _workload_band(workload_units)
        realization_band, realization_per_unit, realization_badge_class = _realization_band(
            fee["amount"],
            workload_units,
        )
        overdue_work = overdue_tasks + overdue_filings + overdue_notices + overdue_requests
        tasks_url = f"{reverse('core:practice_tasks')}?{urlencode({'company': company.pk, 'status': 'open'})}"
        client_url = f"{reverse('core:switch_company', args=[company.pk])}?{urlencode({'next': reverse('core:dashboard')})}"

        row = {
            "company": company,
            "fee_proxy": fee["amount"],
            "fee_source": fee["source"],
            "plan": fee["plan"],
            "usage_percent": fee["usage_percent"],
            "engagement": fee["engagement"],
            "workload_units": workload_units,
            "workload_band": workload_band,
            "workload_badge_class": _workload_badge_class(workload_band),
            "realization_band": realization_band,
            "realization_per_unit": realization_per_unit.quantize(Decimal("0.01")),
            "realization_badge_class": realization_badge_class,
            "open_tasks": open_tasks,
            "overdue_tasks": overdue_tasks,
            "critical_tasks": critical_tasks,
            "blocked_tasks": blocked_tasks,
            "open_filings": open_filings,
            "overdue_filings": overdue_filings,
            "open_notices": open_notices,
            "overdue_notices": overdue_notices,
            "open_requests": open_requests,
            "uploads_for_review": uploads_for_review,
            "overdue_requests": overdue_requests,
            "unreconciled_bank_rows": unreconciled_bank_rows,
            "voucher_volume": voucher_volume,
            "overdue_work": overdue_work,
            "open_work": open_tasks + open_filings + open_notices + open_requests + uploads_for_review,
            "tasks_url": tasks_url,
            "client_url": client_url,
            "client_360_url": reverse("core:client_360", args=[company.pk]),
        }
        row["pricing_action"] = _pricing_action(row)

        if (
            band_filter != "all"
            and row["realization_band"].lower().replace(" ", "_") != band_filter
            and row["workload_band"].lower() != band_filter
        ):
            continue

        rows.append(row)

    rows.sort(key=lambda item: (
        0 if item["realization_band"] in {"No Fee Data", "Underpriced"} else 1,
        0 if item["workload_band"] == "Overloaded" else 1,
        -item["workload_units"],
        item["company"].name,
    ))

    for row in rows:
        totals["clients"] += 1
        totals["fee_proxy"] += row["fee_proxy"]
        totals["workload_units"] += row["workload_units"]
        totals["open_work"] += row["open_work"]
        totals["overdue_work"] += row["overdue_work"]
        if row["workload_band"] == "Overloaded":
            totals["overloaded"] += 1
        if row["realization_band"] == "Underpriced":
            totals["underpriced"] += 1
        if row["realization_band"] == "No Fee Data":
            totals["no_fee"] += 1

    if totals["workload_units"] > 0:
        totals["avg_realization"] = (totals["fee_proxy"] / totals["workload_units"]).quantize(Decimal("0.01"))

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="ca-client-profitability-{today:%Y%m%d}.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Client",
            "GSTIN",
            "Fee Proxy",
            "Fee Source",
            "Workload Units",
            "Workload Band",
            "Realization Per Unit",
            "Realization Band",
            "Open Work",
            "Overdue Work",
            "Open Tasks",
            "Critical Tasks",
            "Blocked Tasks",
            "Open Filings",
            "Open Notices",
            "Open Requests",
            "Uploads For Review",
            "Unreconciled Bank Rows",
            "Voucher Volume 30D",
            "Recommended Action",
            "Action Detail",
        ])
        for row in rows:
            writer.writerow([
                row["company"].name,
                row["company"].gstin or "",
                row["fee_proxy"],
                row["fee_source"],
                row["workload_units"],
                row["workload_band"],
                row["realization_per_unit"],
                row["realization_band"],
                row["open_work"],
                row["overdue_work"],
                row["open_tasks"],
                row["critical_tasks"],
                row["blocked_tasks"],
                row["open_filings"],
                row["open_notices"],
                row["open_requests"],
                row["uploads_for_review"],
                row["unreconciled_bank_rows"],
                row["voucher_volume"],
                row["pricing_action"]["label"],
                row["pricing_action"]["detail"],
            ])
        return response

    band_options = [
        ("all", "All Clients"),
        ("underpriced", "Underpriced"),
        ("no_fee_data", "No Fee Data"),
        ("overloaded", "Overloaded"),
        ("heavy", "Heavy Workload"),
        ("healthy", "Healthy"),
    ]

    return render(request, "core/ca_client_profitability.html", {
        "rows": rows,
        "totals": totals,
        "q": q,
        "band_filter": band_filter,
        "band_options": band_options,
        "export_query": _export_query(request),
        "title": "Client Profitability",
    })


@login_required
def ca_approval_inbox(request):
    context = build_ca_approval_inbox(request.user, request.GET)
    if request.GET.get("export") == "csv":
        return approval_inbox_csv_response(context["items"], context["today"])
    return render(request, "core/ca_approval_inbox.html", context)


@login_required
def statutory_exposure(request):
    context = build_statutory_exposure(request.user, request.GET if request.method == "GET" else request.POST)
    if request.method == "POST":
        result = create_statutory_exposure_tasks(context["items"], request.user)
        if result["created"]:
            messages.success(request, f"Created {result['created']} statutory recovery task(s).")
        elif result["existing"]:
            messages.info(request, "Statutory recovery tasks already exist for the current filters.")
        elif result.get("skipped"):
            messages.warning(request, "No tasks were created because your role is read-only for the matching clients.")
        else:
            messages.info(request, "No actionable statutory recovery tasks matched the current filters.")
        if result["created"] and result["existing"]:
            messages.info(request, f"{result['existing']} existing task(s) were left unchanged.")
        if result.get("skipped") and (result["created"] or result["existing"]):
            messages.warning(request, f"{result['skipped']} item(s) were skipped because your role is read-only for those clients.")
        redirect_url = reverse("core:statutory_exposure")
        if context["task_query"]:
            redirect_url = f"{redirect_url}?{context['task_query']}"
        return redirect(redirect_url)
    if request.GET.get("export") == "csv":
        return statutory_exposure_csv_response(context["items"], context["period_value"])
    return render(request, "core/statutory_exposure.html", context)


@login_required
def client_operating_readiness(request):
    context = build_operating_readiness(request.user, request.GET if request.method == "GET" else request.POST)
    if request.method == "POST":
        result = create_operating_readiness_tasks(context["rows"], request.user)
        if result["created"]:
            messages.success(request, f"Created {result['created']} operating readiness task(s).")
        elif result["existing"]:
            messages.info(request, "Operating readiness tasks already exist for the current filters.")
        else:
            messages.info(request, "No operating readiness gaps need new tasks for the current filters.")
        if result["created"] and result["existing"]:
            messages.info(request, f"{result['existing']} existing task(s) were left unchanged.")
        if result["skipped"]:
            messages.warning(request, f"{result['skipped']} gap(s) were skipped because your role is read-only for those clients.")
        redirect_url = reverse("core:client_operating_readiness")
        if context["export_query"]:
            redirect_url = f"{redirect_url}?{context['export_query']}"
        return redirect(redirect_url)
    if request.GET.get("export") == "csv":
        return operating_readiness_csv_response(context["rows"])
    return render(request, "core/client_operating_readiness.html", {
        **context,
        "title": "Operating Readiness",
    })


@login_required
def client_pilot_launch(request):
    context = build_pilot_launch_control(request.user, request.GET if request.method == "GET" else request.POST)
    if request.method == "POST":
        result = create_pilot_launch_tasks(context["rows"], request.user)
        if result["created"]:
            messages.success(request, f"Created {result['created']} pilot launch task(s).")
        elif result["existing"]:
            messages.info(request, "Pilot launch tasks already exist for the current filters.")
        else:
            messages.info(request, "No launch blockers need new tasks for the current filters.")
        if result["created"] and result["existing"]:
            messages.info(request, f"{result['existing']} existing task(s) were left unchanged.")
        if result["skipped"]:
            messages.warning(request, f"{result['skipped']} gate(s) were skipped because your role is read-only for those clients.")
        redirect_url = reverse("core:client_pilot_launch")
        if context["export_query"]:
            redirect_url = f"{redirect_url}?{context['export_query']}"
        return redirect(redirect_url)
    if request.GET.get("export") == "csv":
        return pilot_launch_csv_response(context["rows"])
    return render(request, "core/client_pilot_launch.html", {
        **context,
        "title": "Pilot Launch Control",
    })


@login_required
def client_success_cockpit(request):
    context = build_client_success_cockpit(request.user, request.GET if request.method == "GET" else request.POST)
    if request.method == "POST":
        result = create_client_success_tasks(context["rows"], request.user)
        if result["created"]:
            messages.success(request, f"Created {result['created']} client success task(s).")
        elif result["existing"]:
            messages.info(request, "Client success tasks already exist for the current filters.")
        else:
            messages.info(request, "No client success risks need new tasks for the current filters.")
        if result["created"] and result["existing"]:
            messages.info(request, f"{result['existing']} existing task(s) were left unchanged.")
        if result["skipped"]:
            messages.warning(request, f"{result['skipped']} gate(s) were skipped because your role is read-only for those clients.")
        redirect_url = reverse("core:client_success_cockpit")
        if context["export_query"]:
            redirect_url = f"{redirect_url}?{context['export_query']}"
        return redirect(redirect_url)
    if request.GET.get("export") == "csv":
        return client_success_csv_response(context["rows"])
    return render(request, "core/client_success_cockpit.html", {
        **context,
        "title": "Client Success Cockpit",
    })


@login_required
def client_portal_health(request):
    context = build_client_portal_health(request.user, request.GET if request.method == "GET" else request.POST)
    if request.method == "POST":
        result = create_client_portal_health_tasks(context["rows"], request.user)
        if result["created"]:
            messages.success(request, f"Created {result['created']} client portal task(s).")
        elif result["existing"]:
            messages.info(request, "Client portal tasks already exist for the current filters.")
        else:
            messages.info(request, "No client portal risks need new tasks for the current filters.")
        if result["created"] and result["existing"]:
            messages.info(request, f"{result['existing']} existing task(s) were left unchanged.")
        if result["skipped"]:
            messages.warning(request, f"{result['skipped']} gate(s) were skipped because your role is read-only for those clients.")
        redirect_url = reverse("core:client_portal_health")
        if context["export_query"]:
            redirect_url = f"{redirect_url}?{context['export_query']}"
        return redirect(redirect_url)
    if request.GET.get("export") == "csv":
        return client_portal_health_csv_response(context["rows"])
    return render(request, "core/client_portal_health.html", {
        **context,
        "title": "Client Portal Health",
    })


@login_required
def pilot_adoption_evidence(request):
    context = build_pilot_adoption_evidence(request.user, request.GET if request.method == "GET" else request.POST)
    if request.method == "POST":
        result = create_pilot_adoption_tasks(context["rows"], request.user)
        if result["created"]:
            messages.success(request, f"Created {result['created']} pilot adoption task(s).")
        elif result["existing"]:
            messages.info(request, "Pilot adoption tasks already exist for the current filters.")
        else:
            messages.info(request, "No pilot adoption evidence gaps need new tasks for the current filters.")
        if result["created"] and result["existing"]:
            messages.info(request, f"{result['existing']} existing task(s) were left unchanged.")
        if result["skipped"]:
            messages.warning(request, f"{result['skipped']} gate(s) were skipped because your role is read-only for those clients.")
        redirect_url = reverse("core:pilot_adoption_evidence")
        if context["export_query"]:
            redirect_url = f"{redirect_url}?{context['export_query']}"
        return redirect(redirect_url)
    if request.GET.get("export") == "csv":
        return pilot_adoption_csv_response(context["rows"])
    return render(request, "core/pilot_adoption_evidence.html", {
        **context,
        "title": "Pilot Adoption Evidence",
    })


@login_required
def pilot_feedback_register(request):
    context = build_pilot_feedback_register(request.user, request.GET if request.method == "GET" else request.POST)
    manageable_companies = _manageable_companies_for_user(request.user)
    feedback_users = _task_users_for_companies(_companies_for_user(request.user))
    redirect_url = reverse("core:pilot_feedback_register")
    filter_query = pilot_feedback_filter_query(request.POST if request.method == "POST" else request.GET)
    if filter_query:
        redirect_url = f"{redirect_url}?{filter_query}"

    if request.method == "POST":
        action = request.POST.get("action") or "create_feedback"
        if action == "create_feedback":
            form = PilotFeedbackForm(request.POST, companies=manageable_companies, users=feedback_users)
            if form.is_valid():
                feedback = form.save(commit=False)
                feedback.recorded_by = request.user
                if feedback.status in {PilotFeedback.STATUS_RESOLVED, PilotFeedback.STATUS_DISMISSED}:
                    feedback.resolved_at = timezone.now()
                feedback.save()
                if form.cleaned_data.get("create_follow_up_task") and feedback.is_open:
                    _, created_task = create_pilot_feedback_follow_up(feedback, request.user)
                    task_message = " Follow-up task created." if created_task else " Existing follow-up task linked."
                else:
                    task_message = ""
                messages.success(request, f"Pilot feedback captured for {feedback.company.name}.{task_message}")
                return redirect(redirect_url)
        elif action == "resolve_feedback":
            feedback = get_object_or_404(PilotFeedback, pk=request.POST.get("feedback_id"), company__in=manageable_companies)
            resolve_pilot_feedback(feedback, request.user)
            messages.success(request, "Pilot feedback marked resolved.")
            return redirect(redirect_url)
        elif action == "reopen_feedback":
            feedback = get_object_or_404(PilotFeedback, pk=request.POST.get("feedback_id"), company__in=manageable_companies)
            reopen_pilot_feedback(feedback)
            messages.success(request, "Pilot feedback reopened.")
            return redirect(redirect_url)
        else:
            form = PilotFeedbackForm(companies=manageable_companies, users=feedback_users)
            messages.warning(request, "Pilot feedback action was not recognised.")
    else:
        form = PilotFeedbackForm(companies=manageable_companies, users=feedback_users)

    if request.GET.get("export") == "csv":
        return pilot_feedback_csv_response(context["rows"])
    return render(request, "core/pilot_feedback_register.html", {
        **context,
        "feedback_form": form,
        "title": "Pilot Feedback Register",
    })


@login_required
def market_proof_pack(request):
    current_company = getattr(request, "current_company", None)
    context = build_market_proof_pack(
        request.user,
        request.GET if request.method == "GET" else request.POST,
        current_company=current_company,
    )
    if request.method == "POST":
        result = create_market_proof_tasks(context, request.user)
        if result["created"]:
            messages.success(request, f"Created {result['created']} market proof task(s).")
        elif result["existing"]:
            messages.info(request, "Market proof tasks already exist for the current filters.")
        else:
            messages.info(request, "No market proof gaps need new tasks for the current filters.")
        if result["created"] and result["existing"]:
            messages.info(request, f"{result['existing']} existing task(s) were left unchanged.")
        if result["skipped"]:
            messages.warning(request, f"{result['skipped']} proof gap(s) were skipped because your role is read-only or no current company is selected.")
        redirect_url = reverse("core:market_proof_pack")
        if context["export_query"]:
            redirect_url = f"{redirect_url}?{context['export_query']}"
        return redirect(redirect_url)
    if request.GET.get("export") == "csv":
        return market_proof_csv_response(context)
    return render(request, "core/market_proof_pack.html", {
        **context,
        "title": "Market Proof Pack",
    })


@login_required
def market_proof_evidence_pack_download(request):
    pack = build_market_proof_evidence_pack(
        request.user,
        request.GET,
        current_company=getattr(request, "current_company", None),
    )
    return market_proof_evidence_pack_response(pack)


@login_required
def market_external_evidence(request):
    context = build_market_external_evidence_register(request.user, request.GET if request.method == "GET" else request.POST)
    manageable_companies = _manageable_companies_for_user(request.user)
    evidence_users = _task_users_for_companies(_companies_for_user(request.user))
    redirect_url = reverse("core:market_external_evidence")
    filter_query = market_external_evidence_filter_query(request.POST if request.method == "POST" else request.GET)
    if filter_query:
        redirect_url = f"{redirect_url}?{filter_query}"

    if request.method == "POST":
        action = request.POST.get("action") or "create_evidence"
        if action == "create_evidence":
            form = MarketProofExternalEvidenceForm(request.POST, companies=manageable_companies, users=evidence_users)
            if form.is_valid():
                evidence = form.save(commit=False)
                evidence.created_by = request.user
                if evidence.status == MarketProofExternalEvidence.STATUS_VERIFIED:
                    evidence.verified_by = request.user
                    evidence.verified_at = timezone.now()
                evidence.save()
                task_message = ""
                if form.cleaned_data.get("create_follow_up_task") and not evidence.is_verified:
                    _, created_task = create_external_evidence_follow_up(evidence, request.user)
                    task_message = " Follow-up task created." if created_task else " Existing follow-up task linked."
                messages.success(request, f"External evidence captured for {evidence.company.name}.{task_message}")
                return redirect(redirect_url)
        elif action == "verify_evidence":
            evidence = get_object_or_404(MarketProofExternalEvidence, pk=request.POST.get("evidence_id"), company__in=manageable_companies)
            if not (evidence.evidence_reference.strip() or evidence.artifact_sha256.strip() or evidence.evidence_url.strip()):
                create_external_evidence_follow_up(evidence, request.user)
                messages.warning(request, "Evidence needs a reference, URL, or SHA-256 hash before verification.")
            else:
                verify_external_evidence(evidence, request.user)
                messages.success(request, "External evidence verified.")
            return redirect(redirect_url)
        elif action == "reject_evidence":
            evidence = get_object_or_404(MarketProofExternalEvidence, pk=request.POST.get("evidence_id"), company__in=manageable_companies)
            reject_external_evidence(evidence, request.user)
            messages.warning(request, "External evidence rejected and a follow-up task was queued.")
            return redirect(redirect_url)
        elif action == "reopen_evidence":
            evidence = get_object_or_404(MarketProofExternalEvidence, pk=request.POST.get("evidence_id"), company__in=manageable_companies)
            reopen_external_evidence(evidence)
            messages.info(request, "External evidence reopened for review.")
            return redirect(redirect_url)
        else:
            form = MarketProofExternalEvidenceForm(companies=manageable_companies, users=evidence_users)
            messages.warning(request, "External evidence action was not recognised.")
    else:
        form = MarketProofExternalEvidenceForm(companies=manageable_companies, users=evidence_users)

    if request.GET.get("export") == "csv":
        return market_external_evidence_csv_response(context["rows"])
    return render(request, "core/market_external_evidence.html", {
        **context,
        "evidence_form": form,
        "title": "External Evidence Register",
    })


@login_required
def market_case_studies(request):
    context = build_market_case_study_register(request.user, request.GET if request.method == "GET" else request.POST)
    manageable_companies = _manageable_companies_for_user(request.user)
    case_users = _task_users_for_companies(_companies_for_user(request.user))
    redirect_url = reverse("core:market_case_studies")
    filter_query = market_case_study_filter_query(request.POST if request.method == "POST" else request.GET)
    if filter_query:
        redirect_url = f"{redirect_url}?{filter_query}"

    if request.method == "POST":
        action = request.POST.get("action") or "create_case_study"
        if action == "create_case_study":
            form = MarketProofCaseStudyForm(request.POST, companies=manageable_companies, users=case_users)
            if form.is_valid():
                case_study = form.save(commit=False)
                case_study.created_by = request.user
                if case_study.status in {MarketProofCaseStudy.STATUS_APPROVED, MarketProofCaseStudy.STATUS_PUBLISHED}:
                    case_study.approved_by = request.user
                    case_study.approved_at = timezone.now()
                if case_study.status == MarketProofCaseStudy.STATUS_PUBLISHED:
                    case_study.published_at = timezone.now()
                case_study.save()
                if form.cleaned_data.get("create_follow_up_task"):
                    _, created_task = create_case_study_follow_up(case_study, request.user)
                    task_message = " Proof follow-up task created." if created_task else ""
                else:
                    task_message = ""
                messages.success(request, f"Market case study captured for {case_study.company.name}.{task_message}")
                return redirect(redirect_url)
        elif action == "approve_case_study":
            case_study = get_object_or_404(MarketProofCaseStudy, pk=request.POST.get("case_study_id"), company__in=manageable_companies)
            approve_case_study(case_study, request.user)
            messages.success(request, "Market case study approved.")
            return redirect(redirect_url)
        elif action == "publish_case_study":
            case_study = get_object_or_404(MarketProofCaseStudy, pk=request.POST.get("case_study_id"), company__in=manageable_companies)
            if case_study.is_publishable:
                publish_case_study(case_study, request.user)
                messages.success(request, "Market case study marked published.")
            else:
                create_case_study_follow_up(case_study, request.user)
                messages.warning(request, "Case study still needs consent, quote, evidence, approval, and metric proof before publishing.")
            return redirect(redirect_url)
        else:
            form = MarketProofCaseStudyForm(companies=manageable_companies, users=case_users)
            messages.warning(request, "Market case study action was not recognised.")
    else:
        form = MarketProofCaseStudyForm(companies=manageable_companies, users=case_users)

    if request.GET.get("export") == "csv":
        return market_case_study_csv_response(context["rows"])
    return render(request, "core/market_case_studies.html", {
        **context,
        "case_study_form": form,
        "title": "Market Case Studies",
    })


@login_required
@require_POST
def ca_command_center_autopilot(request):
    months = normalize_autopilot_months(request.POST.get("months"))
    companies = list(_manageable_companies_for_user(request.user))
    if not companies:
        messages.warning(request, "No writable client workspaces are available for Compliance Autopilot.")
        return redirect("core:ca_command_center")

    result = run_compliance_autopilot(
        companies=companies,
        months=months,
        from_date=timezone.localdate(),
        created_by=request.user,
    )
    messages.success(
        request,
        (
            f"Compliance Autopilot prepared {result['created']} filing/task workflow(s) "
            f"for {result['companies']} client(s); {result['existing']} already existed."
        ),
    )
    if result["profile_warnings"]:
        messages.warning(
            request,
            f"{len(result['profile_warnings'])} statutory profile warning(s) need review in App Settings.",
        )
    return redirect("core:ca_command_center")


@login_required
def ca_command_center(request):
    """
    Cross-client command center for CA teams.

    This is intentionally metric-first: it shows where staff attention is
    needed before filing or review, without forcing the CA to enter each
    client workspace one by one.
    """
    from vouchers import compliance
    from ocr.models import OCRSubmission
    from gstr2b.models import PortalGSTR2BEntry
    from portal.models import ClientDocumentRequest
    from reports.utils import get_msme_payable_watch
    from tds.models import TDSReturnWorkpaper
    from tds.workbench import build_tds_deposit_watch, build_tds_return_workbench, default_return_period
    from .models import BankStatementRow, ComplianceFiling, ComplianceNotice, PracticeTask

    if request.user.is_superuser:
        companies = Company.objects.all().order_by("name")
    else:
        companies = (
            Company.objects.filter(user_access__user=request.user)
            .distinct()
            .order_by("name")
        )

    rows = []
    totals = {
        "companies": 0,
        "critical": 0,
        "warnings": 0,
        "ocr_pending": 0,
        "gstr2b_missing": 0,
        "bank_unreconciled": 0,
        "pending_vouchers": 0,
        "open_tasks": 0,
        "overdue_tasks": 0,
        "open_filings": 0,
        "overdue_filings": 0,
        "filings_due_soon": 0,
        "open_notices": 0,
        "overdue_notices": 0,
        "notices_due_soon": 0,
        "client_requests": 0,
        "client_uploads": 0,
        "gst_signoff_blockers": 0,
        "e_invoice_missing_irn": 0,
        "e_invoice_expired": 0,
        "e_invoice_due_soon": 0,
        "sales_missing_pos": 0,
        "sales_invalid_party_gstin": 0,
        "sales_tax_pos_mismatch": 0,
        "sales_missing_hsn": 0,
        "itc_180_reversal_due": 0,
        "itc_180_overdue": 0,
        "itc_180_due_soon": 0,
        "itc_180_reversal_itc": Decimal("0.00"),
        "rcm_purchase_count": 0,
        "rcm_missing_tax": 0,
        "rcm_tax_amount": Decimal("0.00"),
        "eway_missing": 0,
        "eway_expired": 0,
        "eway_due_soon": 0,
        "tds_deposit_overdue": 0,
        "tds_deposit_due_today": 0,
        "tds_deposit_due_soon": 0,
        "tds_deposit_pending_amount": Decimal("0.00"),
        "tds_return_critical": 0,
        "tds_return_warnings": 0,
        "msme_overdue": 0,
        "msme_due_soon": 0,
        "msme_overdue_amount": Decimal("0.00"),
        "msme_interest_exposure": Decimal("0.00"),
        "receivables_overdue": 0,
        "receivables_overdue_amount": Decimal("0.00"),
        "health_score_total": 0,
        "health_critical": 0,
        "health_watch": 0,
        "health_stable": 0,
        "health_clean": 0,
        "avg_health": 0,
    }
    today = timezone.localdate()
    period_start = today.replace(day=1)
    next_7 = today + timedelta(days=7)
    period_end = period_start.replace(day=calendar.monthrange(period_start.year, period_start.month)[1])
    period_value = period_start.strftime("%Y-%m")
    tds_return_fy_start, tds_return_quarter = default_return_period(today)
    tds_return_form_type = TDSReturnWorkpaper.FORM_26Q
    next_actions = []

    def add_next_action(*, company, severity, title, description, action_label, action_url, due_date=None, count=1):
        severity_rank = {"critical": 0, "warning": 1, "info": 2}.get(severity, 3)
        next_actions.append({
            "company": company,
            "severity": severity,
            "severity_rank": severity_rank,
            "title": title,
            "description": description,
            "action_label": action_label,
            "action_url": action_url,
            "due_date": due_date,
            "count": count,
        })

    def health_band(score):
        if score < 50:
            return "Critical"
        if score < 75:
            return "Watch"
        if score < 90:
            return "Stable"
        return "Clean"

    def health_badge_class(score):
        if score < 50:
            return "bg-danger"
        if score < 75:
            return "bg-warning text-dark"
        if score < 90:
            return "bg-info text-dark"
        return "bg-success"

    for company in companies:
        issues = compliance.get_compliance_issues(company)
        danger_count = len([issue for issue in issues if issue.get("level") == "danger"])
        warning_count = len([issue for issue in issues if issue.get("level") == "warning"])
        health_score = max(0, 100 - (danger_count * 10) - (warning_count * 3))

        ocr_pending = OCRSubmission.objects.filter(
            company=company,
            status__in=[
                OCRSubmission.STATUS_PENDING,
                OCRSubmission.STATUS_PROCESSING,
                OCRSubmission.STATUS_ERROR,
            ],
        ).count()
        gstr2b_missing = PortalGSTR2BEntry.objects.filter(
            company=company,
            match_status="missing_in_books",
        ).count()
        unclaimed_itc = Voucher.objects.filter(
            company=company,
            voucher_type="Purchase",
            status="APPROVED",
            is_itc_claimed=False,
        ).count()
        bank_unreconciled = BankStatementRow.objects.filter(
            statement__company=company,
            is_reconciled=False,
        ).count()
        pending_vouchers = Voucher.objects.filter(
            company=company,
            status__in=["DRAFT", "PENDING"],
        ).count()
        overdue_receivables_qs = Voucher.objects.filter(
            company=company,
            voucher_type="Sales",
            status="APPROVED",
            outstanding_amount__gt=0,
            due_date__lt=today,
        )
        overdue_receivables = overdue_receivables_qs.count()
        overdue_receivables_amount = overdue_receivables_qs.aggregate(
            total=Sum("outstanding_amount")
        )["total"] or Decimal("0.00")
        open_tasks = PracticeTask.objects.filter(company=company).exclude(
            status__in=[PracticeTask.STATUS_DONE, PracticeTask.STATUS_CANCELLED]
        ).count()
        overdue_tasks = PracticeTask.objects.filter(company=company, due_date__lt=today).exclude(
            status__in=[PracticeTask.STATUS_DONE, PracticeTask.STATUS_CANCELLED]
        ).count()
        open_filings = ComplianceFiling.objects.filter(company=company).exclude(
            status__in=[ComplianceFiling.STATUS_FILED, ComplianceFiling.STATUS_CANCELLED]
        ).count()
        overdue_filings = ComplianceFiling.objects.filter(company=company, due_date__lt=today).exclude(
            status__in=[ComplianceFiling.STATUS_FILED, ComplianceFiling.STATUS_CANCELLED]
        ).count()
        filings_due_soon = ComplianceFiling.objects.filter(
            company=company,
            due_date__gte=today,
            due_date__lte=next_7,
        ).exclude(status__in=[ComplianceFiling.STATUS_FILED, ComplianceFiling.STATUS_CANCELLED]).count()
        open_notices = ComplianceNotice.objects.filter(company=company).exclude(
            status=ComplianceNotice.STATUS_CLOSED
        ).count()
        overdue_notices = ComplianceNotice.objects.filter(company=company, response_due_date__lt=today).exclude(
            status=ComplianceNotice.STATUS_CLOSED
        ).count()
        notices_due_soon = ComplianceNotice.objects.filter(
            company=company,
            response_due_date__gte=today,
            response_due_date__lte=next_7,
        ).exclude(status=ComplianceNotice.STATUS_CLOSED).count()
        client_requests_qs = ClientDocumentRequest.objects.filter(company=company).exclude(
            status__in=[
                ClientDocumentRequest.STATUS_CLOSED,
                ClientDocumentRequest.STATUS_CANCELLED,
            ]
        )
        client_requests = client_requests_qs.filter(status=ClientDocumentRequest.STATUS_OPEN).count()
        client_uploads = client_requests_qs.filter(status=ClientDocumentRequest.STATUS_UPLOADED).count()
        overdue_client_requests = client_requests_qs.filter(
            status=ClientDocumentRequest.STATUS_OPEN,
            due_date__lt=today,
        ).count()
        gst_snapshot = _gst_workbench_snapshot(company, period_start, period_end)
        gst_signoff_blockers = len(gst_snapshot["signoff_blockers"])
        e_invoice_missing_irn = gst_snapshot["e_invoice_missing_irn"]
        e_invoice_expired = gst_snapshot["e_invoice_expired"]
        e_invoice_due_soon = gst_snapshot["e_invoice_due_soon"]
        sales_missing_pos = gst_snapshot["sales_missing_pos"]
        sales_invalid_party_gstin = gst_snapshot["sales_invalid_party_gstin"]
        sales_tax_pos_mismatch = gst_snapshot["sales_tax_pos_mismatch"]
        sales_missing_hsn = gst_snapshot["sales_missing_hsn"]
        itc_180_reversal_due = gst_snapshot["itc_180_reversal_due"]
        itc_180_overdue = gst_snapshot["itc_180_overdue"]
        itc_180_due_soon = gst_snapshot["itc_180_due_soon"]
        itc_180_reversal_itc = gst_snapshot["itc_180_reversal_itc"]
        rcm_purchase_count = gst_snapshot["rcm_purchase_count"]
        rcm_missing_tax = gst_snapshot["rcm_missing_tax"]
        rcm_tax_amount = gst_snapshot["rcm_tax_amount"]
        eway_missing = gst_snapshot["eway_missing"]
        eway_expired = gst_snapshot["eway_expired"]
        eway_due_soon = gst_snapshot["eway_due_soon"]
        tds_watch = build_tds_deposit_watch(company)
        tds_summary = tds_watch["summary"]
        tds_payable = tds_summary["pending_amount"]
        tds_pending_url = f"{reverse('tds:entry_list')}?deposited=0"
        tds_deposit_url = (
            f"{reverse('core:switch_company', args=[company.pk])}?"
            f"{urlencode({'next': tds_pending_url})}"
        )
        tds_return_watch = build_tds_return_workbench(
            company,
            tds_return_fy_start,
            tds_return_quarter,
            tds_return_form_type,
        )
        tds_return_summary = tds_return_watch["summary"]
        tds_return_has_work = bool(tds_return_summary["entry_count"] or tds_return_watch["workpaper"])
        tds_return_critical = tds_return_summary["critical_issue_count"] if tds_return_has_work else 0
        tds_return_warnings = tds_return_summary["warning_issue_count"] if tds_return_has_work else 0
        tds_return_path = (
            f"{reverse('tds:return_workbench')}?"
            f"{urlencode({'fy': tds_return_fy_start, 'quarter': tds_return_quarter, 'form_type': tds_return_form_type})}"
        )
        tds_return_url = (
            f"{reverse('core:switch_company', args=[company.pk])}?"
            f"{urlencode({'next': tds_return_path})}"
        )
        msme_watch = get_msme_payable_watch(company, as_of_date=today)
        msme_summary = msme_watch["summary"]
        msme_report_url = (
            f"{reverse('core:switch_company', args=[company.pk])}?"
            f"{urlencode({'next': reverse('reports:msme_overdue')})}"
        )

        totals["companies"] += 1
        totals["critical"] += danger_count
        totals["warnings"] += warning_count
        totals["ocr_pending"] += ocr_pending
        totals["gstr2b_missing"] += gstr2b_missing
        totals["bank_unreconciled"] += bank_unreconciled
        totals["pending_vouchers"] += pending_vouchers
        totals["open_tasks"] += open_tasks
        totals["overdue_tasks"] += overdue_tasks
        totals["open_filings"] += open_filings
        totals["overdue_filings"] += overdue_filings
        totals["filings_due_soon"] += filings_due_soon
        totals["open_notices"] += open_notices
        totals["overdue_notices"] += overdue_notices
        totals["notices_due_soon"] += notices_due_soon
        totals["client_requests"] += client_requests
        totals["client_uploads"] += client_uploads
        totals["gst_signoff_blockers"] += gst_signoff_blockers
        totals["e_invoice_missing_irn"] += e_invoice_missing_irn
        totals["e_invoice_expired"] += e_invoice_expired
        totals["e_invoice_due_soon"] += e_invoice_due_soon
        totals["sales_missing_pos"] += sales_missing_pos
        totals["sales_invalid_party_gstin"] += sales_invalid_party_gstin
        totals["sales_tax_pos_mismatch"] += sales_tax_pos_mismatch
        totals["sales_missing_hsn"] += sales_missing_hsn
        totals["itc_180_reversal_due"] += itc_180_reversal_due
        totals["itc_180_overdue"] += itc_180_overdue
        totals["itc_180_due_soon"] += itc_180_due_soon
        totals["itc_180_reversal_itc"] += itc_180_reversal_itc
        totals["rcm_purchase_count"] += rcm_purchase_count
        totals["rcm_missing_tax"] += rcm_missing_tax
        totals["rcm_tax_amount"] += rcm_tax_amount
        totals["eway_missing"] += eway_missing
        totals["eway_expired"] += eway_expired
        totals["eway_due_soon"] += eway_due_soon
        totals["tds_deposit_overdue"] += tds_summary["overdue_count"]
        totals["tds_deposit_due_today"] += tds_summary["due_today_count"]
        totals["tds_deposit_due_soon"] += tds_summary["due_soon_count"]
        totals["tds_deposit_pending_amount"] += tds_payable
        totals["tds_return_critical"] += tds_return_critical
        totals["tds_return_warnings"] += tds_return_warnings
        totals["msme_overdue"] += msme_summary["overdue_count"]
        totals["msme_due_soon"] += msme_summary["due_soon_count"]
        totals["msme_overdue_amount"] += msme_summary["overdue_amount"]
        totals["msme_interest_exposure"] += msme_summary["interest_liability"]
        totals["receivables_overdue"] += overdue_receivables
        totals["receivables_overdue_amount"] += overdue_receivables_amount

        if overdue_client_requests:
            add_next_action(
                company=company,
                severity="critical",
                title="Client document requests overdue",
                description=f"{overdue_client_requests} client upload request(s) are past due.",
                action_label="Open Reminder Queue",
                action_url=f"{reverse('portal:client_request_reminders')}?{urlencode({'company': company.pk, 'kind': 'overdue'})}",
                due_date=today,
                count=overdue_client_requests,
            )
        if client_uploads:
            add_next_action(
                company=company,
                severity="warning",
                title="Client evidence uploaded for review",
                description=f"{client_uploads} uploaded document(s) need CA review and closure.",
                action_label="Open GST Review",
                action_url=reverse("core:gst_workbench_detail", args=[company.pk, period_value]),
                count=client_uploads,
            )
        if ocr_pending:
            ocr_url = (
                f"{reverse('core:switch_company', args=[company.pk])}?"
                f"{urlencode({'next': reverse('ocr:list')})}"
            )
            add_next_action(
                company=company,
                severity="warning",
                title="OCR documents awaiting review",
                description=f"{ocr_pending} OCR document(s) are pending, processing, or errored.",
                action_label="Open OCR Queue",
                action_url=ocr_url,
                count=ocr_pending,
            )
        if bank_unreconciled:
            bank_url = (
                f"{reverse('core:switch_company', args=[company.pk])}?"
                f"{urlencode({'next': reverse('core:bank_statement_list')})}"
            )
            add_next_action(
                company=company,
                severity="warning",
                title="Bank reconciliation pending",
                description=f"{bank_unreconciled} imported bank row(s) are not reconciled.",
                action_label="Open Bank Reconciliation",
                action_url=bank_url,
                count=bank_unreconciled,
            )
        if overdue_receivables:
            outstanding_url = f"{reverse('vouchers:outstanding')}?type=Sales&status=outstanding"
            receivables_url = (
                f"{reverse('core:switch_company', args=[company.pk])}?"
                f"{urlencode({'next': outstanding_url})}"
            )
            add_next_action(
                company=company,
                severity="warning",
                title="Receivables overdue",
                description=f"{overdue_receivables} sales invoice(s) are past due.",
                action_label="Open Outstanding",
                action_url=receivables_url,
                due_date=today,
                count=overdue_receivables,
            )
        if overdue_notices:
            add_next_action(
                company=company,
                severity="critical",
                title="Notice response overdue",
                description=f"{overdue_notices} compliance notice(s) have crossed their response date.",
                action_label="Open Notices",
                action_url=f"{reverse('core:compliance_notices')}?{urlencode({'company': company.pk})}",
                due_date=today,
                count=overdue_notices,
            )
        elif notices_due_soon:
            add_next_action(
                company=company,
                severity="warning",
                title="Notice response due soon",
                description=f"{notices_due_soon} notice response(s) are due in the next 7 days.",
                action_label="Open Notices",
                action_url=f"{reverse('core:compliance_notices')}?{urlencode({'company': company.pk})}",
                count=notices_due_soon,
            )
        if overdue_filings:
            add_next_action(
                company=company,
                severity="critical",
                title="Filing overdue",
                description=f"{overdue_filings} filing workflow(s) are overdue.",
                action_label="Open Filings",
                action_url=f"{reverse('core:compliance_filings')}?{urlencode({'company': company.pk})}",
                due_date=today,
                count=overdue_filings,
            )
        elif filings_due_soon:
            add_next_action(
                company=company,
                severity="warning",
                title="Filing due soon",
                description=f"{filings_due_soon} filing workflow(s) are due in the next 7 days.",
                action_label="Open Filings",
                action_url=f"{reverse('core:compliance_filings')}?{urlencode({'company': company.pk})}",
                count=filings_due_soon,
            )
        if overdue_tasks:
            add_next_action(
                company=company,
                severity="critical",
                title="Practice tasks overdue",
                description=f"{overdue_tasks} practice task(s) are overdue.",
                action_label="Open Work Queue",
                action_url=f"{reverse('core:practice_tasks')}?{urlencode({'company': company.pk, 'status': 'open'})}",
                due_date=today,
                count=overdue_tasks,
            )
        if e_invoice_expired:
            add_next_action(
                company=company,
                severity="critical",
                title="E-invoice IRP deadline crossed",
                description=f"{e_invoice_expired} sales document(s) are past the configured IRP reporting window.",
                action_label="Open GST Review",
                action_url=reverse("core:gst_workbench_detail", args=[company.pk, period_value]),
                due_date=today,
                count=e_invoice_expired,
            )
        elif e_invoice_due_soon:
            add_next_action(
                company=company,
                severity="warning",
                title="E-invoice IRN due soon",
                description=f"{e_invoice_due_soon} sales document(s) need IRN before the reporting window closes.",
                action_label="Open GST Review",
                action_url=reverse("core:gst_workbench_detail", args=[company.pk, period_value]),
                count=e_invoice_due_soon,
            )
        sales_readiness_issues = sales_missing_pos + sales_invalid_party_gstin + sales_tax_pos_mismatch + sales_missing_hsn
        if sales_readiness_issues:
            add_next_action(
                company=company,
                severity="critical" if sales_missing_pos or sales_invalid_party_gstin or sales_missing_hsn else "warning",
                title="GSTR-1 sales readiness blocked",
                description=(
                    f"{sales_missing_pos} missing POS, {sales_invalid_party_gstin} invalid GSTIN, "
                    f"{sales_missing_hsn} missing HSN/SAC, {sales_tax_pos_mismatch} tax/POS mismatch."
                ),
                action_label="Open GST Review",
                action_url=reverse("core:gst_workbench_detail", args=[company.pk, period_value]),
                count=sales_readiness_issues,
            )
        if itc_180_overdue:
            add_next_action(
                company=company,
                severity="critical",
                title="ITC 180-day reversal due",
                description=f"{itc_180_overdue} purchase bill(s) have claimed ITC with unpaid balance past 180 days.",
                action_label="Open GST Review",
                action_url=reverse("core:gst_workbench_detail", args=[company.pk, period_value]),
                due_date=today,
                count=itc_180_overdue,
            )
        elif itc_180_due_soon:
            add_next_action(
                company=company,
                severity="warning",
                title="ITC 180-day watch due soon",
                description=f"{itc_180_due_soon} purchase bill(s) need payment proof or ITC reversal before period close.",
                action_label="Open GST Review",
                action_url=reverse("core:gst_workbench_detail", args=[company.pk, period_value]),
                count=itc_180_due_soon,
            )
        if rcm_missing_tax:
            add_next_action(
                company=company,
                severity="critical",
                title="RCM purchase tax missing",
                description=f"{rcm_missing_tax} reverse-charge purchase(s) have no GST tax amount for GSTR-3B.",
                action_label="Open GST Review",
                action_url=reverse("core:gst_workbench_detail", args=[company.pk, period_value]),
                due_date=today,
                count=rcm_missing_tax,
            )
        if eway_missing or eway_expired:
            count = eway_missing + eway_expired
            add_next_action(
                company=company,
                severity="critical",
                title="E-way bill movement issue",
                description=f"{eway_missing} missing and {eway_expired} expired e-way bill item(s) need review.",
                action_label="Open GST Review",
                action_url=reverse("core:gst_workbench_detail", args=[company.pk, period_value]),
                due_date=today,
                count=count,
            )
        elif eway_due_soon:
            add_next_action(
                company=company,
                severity="warning",
                title="E-way bill validity due soon",
                description=f"{eway_due_soon} e-way bill(s) expire within 24 hours.",
                action_label="Open GST Review",
                action_url=reverse("core:gst_workbench_detail", args=[company.pk, period_value]),
                count=eway_due_soon,
            )
        if tds_summary["overdue_count"] or tds_summary["due_today_count"]:
            count = tds_summary["overdue_count"] + tds_summary["due_today_count"]
            add_next_action(
                company=company,
                severity="critical",
                title="TDS deposit due",
                description=f"{count} TDS deposit item(s) need challan closure now.",
                action_label="Open TDS Register",
                action_url=tds_deposit_url,
                due_date=today,
                count=count,
            )
        elif tds_summary["due_soon_count"]:
            add_next_action(
                company=company,
                severity="warning",
                title="TDS deposit due soon",
                description=f"{tds_summary['due_soon_count']} TDS deposit item(s) are due in the next few days.",
                action_label="Open TDS Register",
                action_url=tds_deposit_url,
                count=tds_summary["due_soon_count"],
            )
        if tds_return_critical:
            add_next_action(
                company=company,
                severity="critical",
                title="TDS return blockers open",
                description=f"{tds_return_critical} critical return-readiness issue(s) need correction.",
                action_label="Open TDS Workpaper",
                action_url=tds_return_url,
                due_date=tds_return_summary["due_date"],
                count=tds_return_critical,
            )
        elif tds_return_warnings:
            add_next_action(
                company=company,
                severity="warning",
                title="TDS return review pending",
                description=f"{tds_return_warnings} TDS return-readiness warning(s) remain open.",
                action_label="Open TDS Workpaper",
                action_url=tds_return_url,
                due_date=tds_return_summary["due_date"],
                count=tds_return_warnings,
            )
        if msme_summary["overdue_count"]:
            add_next_action(
                company=company,
                severity="critical",
                title="MSME payment overdue",
                description=f"{msme_summary['overdue_count']} MSME payable(s) have crossed the watch due date.",
                action_label="Open MSME Report",
                action_url=msme_report_url,
                due_date=today,
                count=msme_summary["overdue_count"],
            )
        elif msme_summary["due_soon_count"]:
            add_next_action(
                company=company,
                severity="warning",
                title="MSME payment due soon",
                description=f"{msme_summary['due_soon_count']} MSME payable(s) are close to the 45-day watch window.",
                action_label="Open MSME Report",
                action_url=msme_report_url,
                count=msme_summary["due_soon_count"],
            )
        if gst_signoff_blockers:
            first_blocker = gst_snapshot["signoff_blockers"][0]
            add_next_action(
                company=company,
                severity="warning",
                title="GST sign-off blocked",
                description=f"{gst_signoff_blockers} blocker(s), starting with {first_blocker['title']}.",
                action_label="Open GST Review",
                action_url=reverse("core:gst_workbench_detail", args=[company.pk, period_value]),
                count=gst_signoff_blockers,
            )

        risk_drivers = []

        def add_driver(severity, label, detail, action_label, action_url, count=1, due_date=None):
            risk_drivers.append({
                "severity": severity,
                "severity_rank": {"critical": 0, "warning": 1, "info": 2}.get(severity, 3),
                "label": label,
                "detail": detail,
                "action_label": action_label,
                "action_url": action_url,
                "count": count,
                "due_date": due_date,
            })

        gst_review_url = reverse("core:gst_workbench_detail", args=[company.pk, period_value])
        tasks_url = f"{reverse('core:practice_tasks')}?{urlencode({'company': company.pk, 'status': 'open'})}"
        filings_url = f"{reverse('core:compliance_filings')}?{urlencode({'company': company.pk})}"
        notices_url = f"{reverse('core:compliance_notices')}?{urlencode({'company': company.pk})}"
        client_reminders_url = f"{reverse('portal:client_request_reminders')}?{urlencode({'company': company.pk})}"

        if overdue_client_requests:
            add_driver("critical", "Client document chase overdue", f"{overdue_client_requests} request(s) past due.", "Open reminders", client_reminders_url, overdue_client_requests, today)
        if overdue_notices:
            add_driver("critical", "Notice response overdue", f"{overdue_notices} notice response(s) crossed due date.", "Open notices", notices_url, overdue_notices, today)
        if overdue_filings:
            add_driver("critical", "Filing overdue", f"{overdue_filings} filing workflow(s) overdue.", "Open filings", filings_url, overdue_filings, today)
        if overdue_tasks:
            add_driver("critical", "Practice tasks overdue", f"{overdue_tasks} task(s) overdue.", "Open work queue", tasks_url, overdue_tasks, today)
        if e_invoice_expired:
            add_driver("critical", "IRP deadline crossed", f"{e_invoice_expired} e-invoice document(s) past reporting window.", "Open GST review", gst_review_url, e_invoice_expired, today)
        if sales_missing_pos or sales_invalid_party_gstin or sales_missing_hsn:
            add_driver(
                "critical",
                "GSTR-1 master data blocked",
                f"{sales_missing_pos} POS, {sales_invalid_party_gstin} GSTIN, {sales_missing_hsn} HSN/SAC issue(s).",
                "Open GST review",
                gst_review_url,
                sales_missing_pos + sales_invalid_party_gstin + sales_missing_hsn,
            )
        if itc_180_overdue:
            add_driver("critical", "ITC 180-day reversal due", f"{itc_180_overdue} purchase bill(s) past 180-day payment watch.", "Open GST review", gst_review_url, itc_180_overdue, today)
        if rcm_missing_tax:
            add_driver("critical", "RCM tax missing", f"{rcm_missing_tax} reverse-charge purchase(s) missing GST tax.", "Open GST review", gst_review_url, rcm_missing_tax, today)
        if eway_missing or eway_expired:
            add_driver("critical", "E-way bill exception", f"{eway_missing} missing, {eway_expired} expired.", "Open GST review", gst_review_url, eway_missing + eway_expired, today)
        if tds_summary["overdue_count"] or tds_summary["due_today_count"]:
            count = tds_summary["overdue_count"] + tds_summary["due_today_count"]
            add_driver("critical", "TDS deposit due now", f"{count} TDS deposit item(s) need challan closure.", "Open TDS register", tds_deposit_url, count, today)
        if tds_return_critical:
            add_driver("critical", "TDS return blockers", f"{tds_return_critical} return-readiness blocker(s).", "Open TDS workpaper", tds_return_url, tds_return_critical, tds_return_summary["due_date"])
        if msme_summary["overdue_count"]:
            add_driver("critical", "MSME payable overdue", f"{msme_summary['overdue_count']} MSME payable(s) overdue.", "Open MSME report", msme_report_url, msme_summary["overdue_count"], today)
        if client_uploads:
            add_driver("warning", "Client upload pending review", f"{client_uploads} uploaded document(s) need CA review.", "Open requests", f"{reverse('portal:client_requests')}?{urlencode({'company': company.pk, 'status': ClientDocumentRequest.STATUS_UPLOADED})}", client_uploads)
        if bank_unreconciled:
            add_driver("warning", "Bank reconciliation pending", f"{bank_unreconciled} bank row(s) unreconciled.", "Open bank reconciliation", bank_url, bank_unreconciled)
        if overdue_receivables:
            add_driver("warning", "Receivables overdue", f"{overdue_receivables} sales invoice(s) overdue.", "Open outstanding", receivables_url, overdue_receivables, today)
        if gst_signoff_blockers:
            add_driver("warning", "GST sign-off blocked", f"{gst_signoff_blockers} GST period blocker(s).", "Open GST review", gst_review_url, gst_signoff_blockers)
        if e_invoice_due_soon:
            add_driver("warning", "IRN due soon", f"{e_invoice_due_soon} e-invoice document(s) approaching deadline.", "Open GST review", gst_review_url, e_invoice_due_soon)
        if sales_tax_pos_mismatch:
            add_driver("warning", "GSTR-1 tax/POS mismatch", f"{sales_tax_pos_mismatch} tax split issue(s).", "Open GST review", gst_review_url, sales_tax_pos_mismatch)
        if itc_180_due_soon:
            add_driver("warning", "ITC 180-day due soon", f"{itc_180_due_soon} purchase bill(s) approaching reversal watch.", "Open GST review", gst_review_url, itc_180_due_soon)
        if eway_due_soon:
            add_driver("warning", "E-way bill expiring soon", f"{eway_due_soon} e-way bill(s) expire soon.", "Open GST review", gst_review_url, eway_due_soon)
        if tds_summary["due_soon_count"]:
            add_driver("warning", "TDS deposit due soon", f"{tds_summary['due_soon_count']} TDS deposit item(s) due soon.", "Open TDS register", tds_deposit_url, tds_summary["due_soon_count"])
        if tds_return_warnings:
            add_driver("warning", "TDS return warnings", f"{tds_return_warnings} return-readiness warning(s).", "Open TDS workpaper", tds_return_url, tds_return_warnings, tds_return_summary["due_date"])
        if msme_summary["due_soon_count"]:
            add_driver("warning", "MSME payable due soon", f"{msme_summary['due_soon_count']} MSME payable(s) approaching watch window.", "Open MSME report", msme_report_url, msme_summary["due_soon_count"])
        if ocr_pending:
            add_driver("info", "OCR queue pending", f"{ocr_pending} document(s) pending or errored.", "Open OCR queue", ocr_url, ocr_pending)
        if gstr2b_missing:
            add_driver("info", "2B missing in books", f"{gstr2b_missing} portal invoice(s) not booked.", "Open 2B console", f"{reverse('gstr2b:results')}?{urlencode({'period': period_value, 'status': 'missing_in_books'})}", gstr2b_missing)
        if pending_vouchers:
            voucher_url = (
                f"{reverse('core:switch_company', args=[company.pk])}?"
                f"{urlencode({'next': reverse('vouchers:list')})}"
            )
            add_driver("info", "Draft/pending vouchers", f"{pending_vouchers} voucher(s) not approved.", "Open vouchers", voucher_url, pending_vouchers)

        risk_drivers.sort(key=lambda item: (item["severity_rank"], item["due_date"] or today, -item["count"], item["label"]))
        primary_driver = risk_drivers[0] if risk_drivers else {
            "label": "Monitor",
            "detail": "No urgent exception is open.",
            "action_label": "Open client",
            "action_url": f"{reverse('core:switch_company', args=[company.pk])}?{urlencode({'next': reverse('core:dashboard')})}",
        }

        base_health_score = health_score
        operational_penalty = min(
            55,
            min(10, ocr_pending * 2)
            + min(12, bank_unreconciled)
            + min(12, pending_vouchers * 2)
            + min(15, overdue_receivables * 5)
            + min(20, gst_signoff_blockers * 4)
            + min(15, e_invoice_expired * 8)
            + min(15, sales_readiness_issues * 5)
            + min(15, itc_180_reversal_due * 6)
            + min(15, rcm_missing_tax * 8)
            + min(15, (eway_missing + eway_expired) * 6)
            + min(15, tds_summary["attention_count"] * 5)
            + min(15, tds_return_critical * 8)
            + min(10, tds_return_warnings * 3)
            + min(15, msme_summary["overdue_count"] * 8),
        )
        health_score = max(0, base_health_score - operational_penalty)
        band = health_band(health_score)
        totals["health_score_total"] += health_score
        if band == "Critical":
            totals["health_critical"] += 1
        elif band == "Watch":
            totals["health_watch"] += 1
        elif band == "Stable":
            totals["health_stable"] += 1
        else:
            totals["health_clean"] += 1

        rows.append({
            "company": company,
            "health_score": health_score,
            "health_band": band,
            "health_badge_class": health_badge_class(health_score),
            "base_health_score": base_health_score,
            "operational_penalty": operational_penalty,
            "risk_drivers": risk_drivers[:5],
            "risk_driver_summary": "; ".join(f"{item['label']} ({item['count']})" for item in risk_drivers[:5]),
            "primary_action_label": primary_driver["action_label"],
            "primary_action_url": primary_driver["action_url"],
            "primary_action_title": primary_driver["label"],
            "primary_action_detail": primary_driver["detail"],
            "danger_count": danger_count,
            "warning_count": warning_count,
            "ocr_pending": ocr_pending,
            "gstr2b_missing": gstr2b_missing,
            "unclaimed_itc": unclaimed_itc,
            "bank_unreconciled": bank_unreconciled,
            "pending_vouchers": pending_vouchers,
            "receivables_overdue": overdue_receivables,
            "receivables_overdue_amount": overdue_receivables_amount,
            "open_tasks": open_tasks,
            "overdue_tasks": overdue_tasks,
            "open_filings": open_filings,
            "overdue_filings": overdue_filings,
            "filings_due_soon": filings_due_soon,
            "open_notices": open_notices,
            "overdue_notices": overdue_notices,
            "notices_due_soon": notices_due_soon,
            "client_requests": client_requests,
            "client_uploads": client_uploads,
            "gst_signoff_blockers": gst_signoff_blockers,
            "e_invoice_missing_irn": e_invoice_missing_irn,
            "e_invoice_expired": e_invoice_expired,
            "e_invoice_due_soon": e_invoice_due_soon,
            "sales_missing_pos": sales_missing_pos,
            "sales_invalid_party_gstin": sales_invalid_party_gstin,
            "sales_tax_pos_mismatch": sales_tax_pos_mismatch,
            "sales_missing_hsn": sales_missing_hsn,
            "itc_180_reversal_due": itc_180_reversal_due,
            "itc_180_overdue": itc_180_overdue,
            "itc_180_due_soon": itc_180_due_soon,
            "itc_180_reversal_itc": itc_180_reversal_itc,
            "rcm_purchase_count": rcm_purchase_count,
            "rcm_missing_tax": rcm_missing_tax,
            "rcm_tax_amount": rcm_tax_amount,
            "eway_missing": eway_missing,
            "eway_expired": eway_expired,
            "eway_due_soon": eway_due_soon,
            "tds_deposit_overdue": tds_summary["overdue_count"],
            "tds_deposit_due_today": tds_summary["due_today_count"],
            "tds_deposit_due_soon": tds_summary["due_soon_count"],
            "tds_return_critical": tds_return_critical,
            "tds_return_warnings": tds_return_warnings,
            "tds_return_has_work": tds_return_has_work,
            "tds_return_readiness_score": tds_return_summary["readiness_score"] if tds_return_has_work else None,
            "msme_overdue": msme_summary["overdue_count"],
            "msme_due_soon": msme_summary["due_soon_count"],
            "msme_overdue_amount": msme_summary["overdue_amount"],
            "msme_interest_exposure": msme_summary["interest_liability"],
            "tds_payable": tds_payable,
            "needs_attention": (
                danger_count
                + warning_count
                + ocr_pending
                + gstr2b_missing
                + bank_unreconciled
                + pending_vouchers
                + overdue_receivables
                + unclaimed_itc
                + open_tasks
                + overdue_tasks
                + open_filings
                + overdue_filings
                + filings_due_soon
                + open_notices
                + overdue_notices
                + notices_due_soon
                + client_requests
                + client_uploads
                + gst_signoff_blockers
                + e_invoice_missing_irn
                + sales_readiness_issues
                + itc_180_reversal_due
                + rcm_missing_tax
                + eway_missing
                + eway_expired
                + eway_due_soon
                + tds_summary["attention_count"]
                + tds_return_critical
                + tds_return_warnings
                + msme_summary["overdue_count"]
                + msme_summary["due_soon_count"]
            ),
        })

    rows.sort(key=lambda row: (row["health_score"], -row["needs_attention"], row["company"].name))
    next_actions.sort(key=lambda item: (item["severity_rank"], item["due_date"] or today, -item["count"], item["company"].name))
    if totals["companies"]:
        totals["avg_health"] = round(totals["health_score_total"] / totals["companies"])

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="ca-command-center-{today:%Y%m%d}.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Section",
            "Company",
            "Severity",
            "Title",
            "Description",
            "Count",
            "Due Date",
            "Action",
        ])
        for item in next_actions:
            writer.writerow([
                "Next Best Action",
                item["company"].name,
                item["severity"],
                item["title"],
                item["description"],
                item["count"],
                item["due_date"].isoformat() if item["due_date"] else "",
                item["action_label"],
            ])
        writer.writerow([])
        writer.writerow([
            "Section",
            "Company",
            "Health Score",
            "Health Band",
            "Primary Action",
            "Risk Drivers",
            "Critical",
            "Warnings",
            "OCR Pending",
            "2B Missing",
            "Bank Unreconciled",
            "Overdue Receivables",
            "Filings Due Soon",
            "Notices Due Soon",
            "GST Sign-off Blockers",
            "Sales Missing HSN/SAC",
            "ITC 180-Day Due",
            "ITC 180-Day Reversal ITC",
            "RCM Missing Tax",
            "RCM Tax Amount",
            "E-Way Missing",
            "E-Way Expired",
            "TDS Payable",
            "TDS Return Critical",
            "TDS Return Warnings",
            "MSME Overdue",
        ])
        for row in rows:
            writer.writerow([
                "Client Work Queue",
                row["company"].name,
                row["health_score"],
                row["health_band"],
                row["primary_action_title"],
                row["risk_driver_summary"],
                row["danger_count"],
                row["warning_count"],
                row["ocr_pending"],
                row["gstr2b_missing"],
                row["bank_unreconciled"],
                row["receivables_overdue"],
                row["filings_due_soon"],
                row["notices_due_soon"],
                row["gst_signoff_blockers"],
                row["sales_missing_hsn"],
                row["itc_180_reversal_due"],
                row["itc_180_reversal_itc"],
                row["rcm_missing_tax"],
                row["rcm_tax_amount"],
                row["eway_missing"],
                row["eway_expired"],
                row["tds_payable"],
                row["tds_return_critical"],
                row["tds_return_warnings"],
                row["msme_overdue"],
            ])
        return response

    monthly_due_dates = [
        {"label": "GSTR-1", "date": today.replace(day=11), "owner": "GST template"},
        {"label": "GSTR-3B", "date": today.replace(day=20), "owner": "GST template"},
        {"label": "TDS Payment", "date": today.replace(day=7), "owner": "TDS template"},
    ]

    return render(request, "core/ca_command_center.html", {
        "rows": rows,
        "totals": totals,
        "next_actions": next_actions[:12],
        "monthly_due_dates": monthly_due_dates,
        "period_value": period_value,
        "title": "CA Command Center",
    })
