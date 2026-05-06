import json
import re
from collections import defaultdict
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO

from django.db import transaction
from django.urls import reverse
from django.utils import timezone

from reports.utils import get_gst_report
from vouchers.models import Voucher

from .gst_portal_schema import (
    build_gstr1_portal_payload,
    summarize_schema_issues,
    validate_gstr1_portal_payload,
)
from .models import FilingReview, GSTFilingPack, PracticeTask


GSTIN_RE = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]$")
GSTIN_FIND_RE = re.compile(r"\b([0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z])\b")
ZERO = Decimal("0.00")


def b2cl_limit_for_period(period_end):
    return Decimal("100000.00") if period_end and period_end >= date(2024, 8, 1) else Decimal("250000.00")


def build_gst_filing_pack(company, period_start, period_end):
    report = get_gst_report(company, period_start, period_end)
    period_value = period_start.strftime("%Y-%m")
    approved_review = _approved_review(company, period_start, period_end)
    pack_record = (
        GSTFilingPack.objects.filter(
            company=company,
            period_start=period_start,
            period_end=period_end,
        )
        .select_related("review", "generated_by", "filed_by")
        .first()
    )

    sales_vouchers = list(_sales_vouchers(company, period_start, period_end))
    purchase_vouchers = list(_purchase_vouchers(company, period_start, period_end))
    sales_rows = [_sales_row(voucher, company, period_end) for voucher in sales_vouchers]
    hsn_rows = _hsn_summary(sales_rows)
    b2cs_rows = _b2cs_summary(sales_rows)
    document_rows = _document_summary(sales_vouchers)
    gstr1 = {
        "b2b": [row for row in sales_rows if row["section"] == "B2B"],
        "b2cl": [row for row in sales_rows if row["section"] == "B2CL"],
        "b2cs": b2cs_rows,
        "hsn": hsn_rows,
        "documents": document_rows,
        "nil_exempt": [],
    }
    validations = _validations(
        company=company,
        period_start=period_start,
        period_end=period_end,
        sales_rows=sales_rows,
        sales_vouchers=sales_vouchers,
        purchase_vouchers=purchase_vouchers,
        approved_review=approved_review,
    )

    portal_payload = build_gstr1_portal_payload({
        "company": company,
        "period_end": period_end,
        "gstr1": gstr1,
    })
    schema_issues = validate_gstr1_portal_payload(
        portal_payload,
        period_start=period_start,
        period_end=period_end,
    )
    schema_validation = summarize_schema_issues(schema_issues)
    validations.append(_schema_validation(company, period_value, schema_validation))

    critical_count = sum(1 for item in validations if item["severity"] == "critical")
    warning_count = sum(1 for item in validations if item["severity"] == "warning")

    can_generate = bool(approved_review and critical_count == 0)
    if pack_record and pack_record.is_filed:
        status = "Filed"
    elif can_generate:
        status = "Ready for filing pack"
    elif approved_review:
        status = "Blocked by validations"
    else:
        status = "Review approval pending"

    pack = {
        "company": company,
        "period_start": period_start,
        "period_end": period_end,
        "period_value": period_value,
        "approved_review": approved_review,
        "pack_record": pack_record,
        "status": status,
        "can_generate": can_generate,
        "critical_count": critical_count,
        "warning_count": warning_count,
        "validations": validations,
        "schema_validation": schema_validation,
        "report": report,
        "gstr1": gstr1,
        "gstr3b": _gstr3b_summary(report),
        "portal_payload": portal_payload,
        "summary_snapshot": _summary_snapshot(
            company=company,
            period_start=period_start,
            period_end=period_end,
            report=report,
            sales_rows=sales_rows,
            hsn_rows=hsn_rows,
            b2cs_rows=b2cs_rows,
        ),
    }
    pack["draft_payload"] = _draft_payload(pack)
    return pack


@transaction.atomic
def save_gst_filing_pack(pack, user, notes=""):
    if not pack["approved_review"]:
        raise ValueError("Approve the filing review before generating the final GST filing pack.")
    if pack["critical_count"]:
        raise ValueError("Resolve critical filing pack validations before generating the final pack.")

    record, _ = GSTFilingPack.objects.get_or_create(
        company=pack["company"],
        period_start=pack["period_start"],
        period_end=pack["period_end"],
        defaults={"generated_by": user},
    )
    record.review = pack["approved_review"]
    record.status = GSTFilingPack.STATUS_READY
    record.summary_snapshot = pack["summary_snapshot"]
    record.validation_snapshot = {
        "critical_count": pack["critical_count"],
        "warning_count": pack["warning_count"],
        "validations": pack["validations"],
        "schema_validation": pack["schema_validation"],
        "generated_at": timezone.now().isoformat(),
    }
    record.notes = notes.strip()
    if not record.generated_by_id:
        record.generated_by = user
    record.save()
    return record


@transaction.atomic
def mark_gst_filing_pack_filed(pack_record, user, arn_ack_number, notes=""):
    arn_ack_number = arn_ack_number.strip()
    if not arn_ack_number:
        raise ValueError("ARN or acknowledgement number is required.")
    pack_record.status = GSTFilingPack.STATUS_FILED
    pack_record.arn_ack_number = arn_ack_number
    pack_record.filed_by = user
    pack_record.filed_at = timezone.now()
    if notes.strip():
        pack_record.notes = notes.strip()
    pack_record.save()
    return pack_record


def draft_json_bytes(pack):
    return json.dumps(pack["draft_payload"], indent=2, ensure_ascii=False).encode("utf-8")


def portal_gstr1_json_bytes(pack):
    return json.dumps(pack["portal_payload"], indent=2, ensure_ascii=False, default=str).encode("utf-8")


def filing_pack_xlsx_bytes(pack):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True)

    ws["A1"] = "GST Filing Pack"
    ws["A1"].font = title_font
    ws["A2"] = pack["company"].name
    ws["A3"] = f"Period: {pack['period_start']:%d %b %Y} to {pack['period_end']:%d %b %Y}"
    ws["A5"] = "Status"
    ws["B5"] = pack["status"]
    ws["A6"] = "GSTIN"
    ws["B6"] = pack["company"].gstin or ""
    ws["A7"] = "Critical validations"
    ws["B7"] = pack["critical_count"]
    ws["A8"] = "Warnings"
    ws["B8"] = pack["warning_count"]
    ws["A9"] = "Net tax payable"
    ws["B9"] = _float(pack["report"]["net_tax_payable"])
    for row in range(5, 10):
        ws.cell(row=row, column=1).font = label_font
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).border = border

    _write_table(
        wb.create_sheet("Validations"),
        ["Severity", "Area", "Count", "Description", "Action"],
        [
            [item["severity"], item["title"], item["count"], item["description"], item["action_label"]]
            for item in pack["validations"]
        ],
        header_fill,
        header_font,
        border,
    )
    _write_table(
        wb.create_sheet("Portal Schema Issues"),
        ["Severity", "Path", "Message"],
        [
            [item["severity"], item["path"], item["message"]]
            for item in pack["schema_validation"]["issues"]
        ],
        header_fill,
        header_font,
        border,
    )
    _write_table(
        wb.create_sheet("GSTR1_B2B"),
        _invoice_headers(),
        [_invoice_row(row) for row in pack["gstr1"]["b2b"]],
        header_fill,
        header_font,
        border,
    )
    _write_table(
        wb.create_sheet("GSTR1_B2CL"),
        _invoice_headers(),
        [_invoice_row(row) for row in pack["gstr1"]["b2cl"]],
        header_fill,
        header_font,
        border,
    )
    _write_table(
        wb.create_sheet("GSTR1_B2CS"),
        ["Supply Type", "POS", "Rate", "Taxable Value", "CGST", "SGST", "IGST", "Total Tax"],
        [
            [
                row["supply_type"],
                row["pos"],
                _float(row["rate"]),
                _float(row["taxable_value"]),
                _float(row["cgst"]),
                _float(row["sgst"]),
                _float(row["igst"]),
                _float(row["total_tax"]),
            ]
            for row in pack["gstr1"]["b2cs"]
        ],
        header_fill,
        header_font,
        border,
    )
    _write_table(
        wb.create_sheet("GSTR3B"),
        ["Table", "Description", "Taxable Value", "IGST", "CGST", "SGST", "Total"],
        [
            [
                row["table"],
                row["description"],
                _float(row["taxable_value"]),
                _float(row["igst"]),
                _float(row["cgst"]),
                _float(row["sgst"]),
                _float(row["total"]),
            ]
            for row in pack["gstr3b"]
        ],
        header_fill,
        header_font,
        border,
    )
    _write_table(
        wb.create_sheet("HSN"),
        ["Section", "HSN/SAC", "Description", "UQC", "Quantity", "Rate", "Taxable Value", "CGST", "SGST", "IGST", "Total Tax"],
        [
            [
                row["section"],
                row["hsn_code"],
                row["description"],
                row["uqc"],
                _float(row["quantity"]),
                _float(row["rate"]),
                _float(row["taxable_value"]),
                _float(row["cgst"]),
                _float(row["sgst"]),
                _float(row["igst"]),
                _float(row["total_tax"]),
            ]
            for row in pack["gstr1"]["hsn"]
        ],
        header_fill,
        header_font,
        border,
    )
    _write_table(
        wb.create_sheet("Documents"),
        ["Document Type", "From", "To", "Total", "Cancelled", "Net Issued"],
        [
            [
                row["document_type"],
                row["from_no"],
                row["to_no"],
                row["total"],
                row["cancelled"],
                row["net_issued"],
            ]
            for row in pack["gstr1"]["documents"]
        ],
        header_fill,
        header_font,
        border,
    )

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 42)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _write_table(ws, headers, rows, header_fill, header_font, border):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    for row_index, row in enumerate(rows, 2):
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=row_index, column=col, value=value)
            cell.border = border
            if isinstance(value, float):
                cell.number_format = "#,##0.00"


def _invoice_headers():
    return [
        "Invoice No",
        "Date",
        "Party",
        "GSTIN",
        "Section",
        "POS",
        "Rate",
        "Taxable Value",
        "CGST",
        "SGST",
        "IGST",
        "Total Tax",
        "Invoice Value",
        "IRN",
        "E-Way Bill",
    ]


def _invoice_row(row):
    return [
        row["invoice_number"],
        row["date"].isoformat(),
        row["party_name"],
        row["party_gstin"],
        row["section"],
        row["pos"],
        _float(row["rate"]),
        _float(row["taxable_value"]),
        _float(row["cgst"]),
        _float(row["sgst"]),
        _float(row["igst"]),
        _float(row["total_tax"]),
        _float(row["invoice_value"]),
        row["irn"],
        row["eway_bill_no"],
    ]


def _approved_review(company, period_start, period_end):
    return (
        FilingReview.objects.filter(
            company=company,
            review_type=FilingReview.TYPE_GST_MONTHLY,
            period_start=period_start,
            period_end=period_end,
            status=FilingReview.STATUS_APPROVED,
        )
        .select_related("approved_by", "reviewed_by")
        .first()
    )


def _sales_vouchers(company, period_start, period_end):
    return (
        Voucher.objects.filter(
            company=company,
            voucher_type="Sales",
            date__gte=period_start,
            date__lte=period_end,
            status="APPROVED",
        )
        .prefetch_related(
            "items__ledger__account_group",
            "items__stock_item__hsn_sac",
            "items__stock_item__tax_rate",
        )
        .order_by("date", "number", "id")
    )


def _purchase_vouchers(company, period_start, period_end):
    return (
        Voucher.objects.filter(
            company=company,
            voucher_type="Purchase",
            date__gte=period_start,
            date__lte=period_end,
            status="APPROVED",
        )
        .prefetch_related("items__ledger__account_group")
        .order_by("date", "number", "id")
    )


def _sales_row(voucher, company, period_end):
    own_state = (company.gstin or "")[:2]
    party = _party_ledger(voucher, entry_type="DR", preferred_nature="Asset")
    party_gstin = _party_gstin(voucher, party, company)
    pos = voucher.place_of_supply or (party_gstin[:2] if party_gstin else "")
    taxable_value, cgst, sgst, igst, other_gst = _voucher_tax_parts(voucher, sales=True)
    total_tax = cgst + sgst + igst + other_gst
    invoice_value = taxable_value + total_tax
    b2cl_limit = b2cl_limit_for_period(period_end)
    section = "B2B" if party_gstin else "B2CS"
    if not party_gstin and invoice_value > b2cl_limit and pos and own_state and pos != own_state:
        section = "B2CL"
    rate = _tax_rate(taxable_value, total_tax)
    supply_type = "INTER" if pos and own_state and pos != own_state else "INTRA"
    return {
        "voucher": voucher,
        "invoice_number": voucher.number,
        "date": voucher.date,
        "party_name": party.name if party else "Unidentified party",
        "party_gstin": party_gstin,
        "ledger_gstin": (party.gstin or "").strip().upper() if party and party.gstin else "",
        "section": section,
        "b2cl_limit": b2cl_limit,
        "supply_type": supply_type,
        "pos": pos,
        "taxable_value": taxable_value,
        "cgst": cgst,
        "sgst": sgst,
        "igst": igst,
        "other_gst": other_gst,
        "total_tax": total_tax,
        "invoice_value": invoice_value,
        "rate": rate,
        "reverse_charge": "Y" if voucher.reverse_charge else "N",
        "irn": voucher.e_invoice_irn or "",
        "eway_bill_no": voucher.e_way_bill_no or "",
    }


def _voucher_tax_parts(voucher, *, sales):
    taxable_value = cgst = sgst = igst = other_gst = ZERO
    tax_entry = "CR" if sales else "DR"
    taxable_nature = "Income" if sales else "Expense"
    taxable_entry = "CR" if sales else "DR"
    for item in voucher.items.all():
        if _is_gst_ledger(item.ledger):
            amount = item.amount if item.entry_type == tax_entry else -item.amount
            name = item.ledger.name.upper()
            if "CGST" in name:
                cgst += amount
            elif "SGST" in name or "UTGST" in name:
                sgst += amount
            elif "IGST" in name:
                igst += amount
            else:
                other_gst += amount
        elif item.ledger.account_group.nature == taxable_nature:
            taxable_value += item.amount if item.entry_type == taxable_entry else -item.amount
    return taxable_value, cgst, sgst, igst, other_gst


def _party_ledger(voucher, *, entry_type, preferred_nature):
    fallback = None
    for item in voucher.items.all():
        ledger = item.ledger
        if item.entry_type != entry_type or _is_gst_ledger(ledger):
            continue
        if fallback is None:
            fallback = ledger
        if ledger.account_group.nature == preferred_nature:
            return ledger
    return fallback


def _party_gstin(voucher, party, company):
    if party and party.gstin:
        return party.gstin.strip().upper()
    own_gstin = (company.gstin or "").strip().upper()
    if voucher.narration:
        match = GSTIN_FIND_RE.search(voucher.narration.upper())
        if match and match.group(1) != own_gstin:
            return match.group(1)
    return ""


def _is_gst_ledger(ledger):
    name = ledger.name.upper()
    return ledger.account_group.nature == "Tax" or any(
        keyword in name
        for keyword in (
            "CGST",
            "SGST",
            "IGST",
            "UTGST",
            "VAT",
            "TAX PAYABLE",
            "INPUT TAX",
            "OUTPUT TAX",
            "GST PAYABLE",
            "GST INPUT",
            "GST OUTPUT",
        )
    )


def _tax_rate(taxable_value, total_tax):
    if not taxable_value:
        return ZERO
    return ((total_tax / taxable_value) * Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _hsn_summary(sales_rows):
    grouped = defaultdict(lambda: {
        "section": "",
        "hsn_code": "",
        "description": "",
        "uqc": "",
        "quantity": Decimal("0.000"),
        "rate": ZERO,
        "taxable_value": ZERO,
        "cgst": ZERO,
        "sgst": ZERO,
        "igst": ZERO,
        "total_tax": ZERO,
    })
    for row in sales_rows:
        for item in row["voucher"].items.all():
            if not item.stock_item or item.ledger.account_group.nature != "Income":
                continue
            stock = item.stock_item
            hsn = stock.hsn_sac
            code = hsn.code if hsn else ""
            description = hsn.description if hsn else stock.name
            rate = stock.tax_rate.rate if stock.tax_rate else row["rate"]
            section = "B2B" if row["section"] == "B2B" else "B2C"
            key = (section, code, stock.unit, str(rate))
            bucket = grouped[key]
            bucket.update({
                "section": section,
                "hsn_code": code,
                "description": description,
                "uqc": stock.unit,
                "rate": rate,
            })
            amount = item.amount if item.entry_type == "CR" else -item.amount
            share = amount / row["taxable_value"] if row["taxable_value"] else Decimal("0.00")
            cgst_piece = (row["cgst"] * share).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            sgst_piece = (row["sgst"] * share).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            igst_piece = (row["igst"] * share).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            bucket["quantity"] += item.quantity or Decimal("0.000")
            bucket["taxable_value"] += amount
            bucket["cgst"] += cgst_piece
            bucket["sgst"] += sgst_piece
            bucket["igst"] += igst_piece
            bucket["total_tax"] += cgst_piece + sgst_piece + igst_piece
    return sorted(grouped.values(), key=lambda item: (item["section"], item["hsn_code"], item["description"]))


def _b2cs_summary(sales_rows):
    grouped = defaultdict(lambda: {
        "supply_type": "",
        "pos": "",
        "rate": ZERO,
        "taxable_value": ZERO,
        "cgst": ZERO,
        "sgst": ZERO,
        "igst": ZERO,
        "total_tax": ZERO,
    })
    for row in sales_rows:
        if row["section"] != "B2CS":
            continue
        key = (row["supply_type"], row["pos"] or "NA", str(row["rate"]))
        bucket = grouped[key]
        bucket["supply_type"] = row["supply_type"]
        bucket["pos"] = row["pos"] or "NA"
        bucket["rate"] = row["rate"]
        for field in ("taxable_value", "cgst", "sgst", "igst", "total_tax"):
            bucket[field] += row[field]
    return sorted(grouped.values(), key=lambda item: (item["pos"], item["rate"]))


def _document_summary(sales_vouchers):
    numbers = [voucher.number for voucher in sales_vouchers if voucher.number]
    return [{
        "document_type": "Tax Invoice",
        "from_no": numbers[0] if numbers else "",
        "to_no": numbers[-1] if numbers else "",
        "total": len(numbers),
        "cancelled": 0,
        "net_issued": len(numbers),
    }]


def _validations(*, company, period_start, period_end, sales_rows, sales_vouchers, purchase_vouchers, approved_review):
    validations = []
    period_value = period_start.strftime("%Y-%m")

    validations.append(_validation(
        code="review_approval",
        title="Review Center approval",
        severity="ok" if approved_review else "critical",
        count=0 if approved_review else 1,
        description=(
            "Filing Review Center approval is present."
            if approved_review else "Approve this period in the Filing Review Center before generating the final filing pack."
        ),
        action_label="Open Review Center",
        action_url=f"{reverse('core:filing_review_center')}?period={period_value}&company={company.pk}",
    ))

    own_gstin = (company.gstin or "").strip().upper()
    validations.append(_validation(
        code="company_gstin",
        title="Company GSTIN",
        severity="ok" if GSTIN_RE.match(own_gstin) else "critical",
        count=0 if GSTIN_RE.match(own_gstin) else 1,
        description="Company GSTIN is valid." if GSTIN_RE.match(own_gstin) else "Company GSTIN is missing or invalid.",
        action_label="Open Company Settings",
        action_url=reverse("core:company_settings"),
    ))

    unapproved = Voucher.objects.filter(
        company=company,
        voucher_type__in=["Sales", "Purchase", "Sales Return", "Purchase Return"],
        date__gte=period_start,
        date__lte=period_end,
    ).exclude(status="APPROVED")
    validations.append(_validation(
        code="unapproved_vouchers",
        title="Unapproved GST vouchers",
        severity="critical" if unapproved.exists() else "ok",
        count=unapproved.count(),
        description=(
            f"{unapproved.count()} GST-relevant vouchers are still draft, pending, or rejected."
            if unapproved.exists() else "All GST-relevant vouchers in this period are approved."
        ),
        action_label="Open Vouchers",
        action_url=f"{reverse('vouchers:list')}?start_date={period_start.isoformat()}&end_date={period_end.isoformat()}",
    ))

    missing_pos = [row for row in sales_rows if row["taxable_value"] and not row["pos"]]
    validations.append(_validation(
        code="place_of_supply",
        title="Place of supply",
        severity="critical" if missing_pos else "ok",
        count=len(missing_pos),
        description=(
            f"{len(missing_pos)} sales invoices are missing place of supply."
            if missing_pos else "Place of supply is available for taxable sales invoices."
        ),
        action_label="Open Vouchers",
        action_url=f"{reverse('vouchers:list')}?start_date={period_start.isoformat()}&end_date={period_end.isoformat()}",
    ))

    invalid_gstin_rows = [
        row for row in sales_rows
        if row["ledger_gstin"] and not GSTIN_RE.match(row["ledger_gstin"])
    ]
    validations.append(_validation(
        code="party_gstin",
        title="Party GSTIN",
        severity="critical" if invalid_gstin_rows else "ok",
        count=len(invalid_gstin_rows),
        description=(
            f"{len(invalid_gstin_rows)} party GSTIN values are invalid."
            if invalid_gstin_rows else "Party GSTIN values used for B2B invoices are valid."
        ),
        action_label="Open Ledgers",
        action_url=reverse("ledger:list"),
    ))

    tax_pos_mismatch = [
        row for row in sales_rows
        if row["pos"] and own_gstin
        and (
            (row["pos"] == own_gstin[:2] and row["igst"] > ZERO)
            or (row["pos"] != own_gstin[:2] and (row["cgst"] > ZERO or row["sgst"] > ZERO))
        )
    ]
    validations.append(_validation(
        code="tax_pos_mismatch",
        title="Tax split vs POS",
        severity="warning" if tax_pos_mismatch else "ok",
        count=len(tax_pos_mismatch),
        description=(
            f"{len(tax_pos_mismatch)} invoices have tax split that may not match place of supply."
            if tax_pos_mismatch else "CGST/SGST/IGST split is consistent with place of supply."
        ),
        action_label="Open GST Report",
        action_url=f"{reverse('reports:gst_report')}?start_date={period_start.isoformat()}&end_date={period_end.isoformat()}",
    ))

    missing_hsn = []
    for row in sales_rows:
        for item in row["voucher"].items.all():
            if item.stock_item and item.ledger.account_group.nature == "Income" and not item.stock_item.hsn_sac_id:
                missing_hsn.append(item)
    validations.append(_validation(
        code="hsn_summary",
        title="HSN/SAC summary",
        severity="warning" if missing_hsn else "ok",
        count=len(missing_hsn),
        description=(
            f"{len(missing_hsn)} sales stock lines are missing HSN/SAC."
            if missing_hsn else "HSN/SAC summary can be prepared from item masters."
        ),
        action_label="Open Inventory",
        action_url=reverse("inventory:list"),
    ))

    b2b_without_irn = [row for row in sales_rows if row["section"] == "B2B" and not row["irn"]]
    validations.append(_validation(
        code="einvoice_irn_review",
        title="E-invoice IRN review",
        severity="warning" if b2b_without_irn else "ok",
        count=len(b2b_without_irn),
        description=(
            f"{len(b2b_without_irn)} B2B invoices do not have IRN recorded. Review e-invoice applicability before filing."
            if b2b_without_irn else "B2B invoices have IRN recorded where applicable."
        ),
        action_label="Open Integrations",
        action_url=reverse("integrations:dashboard"),
    ))

    eway_without_number = [
        row for row in sales_rows
        if row["invoice_value"] > Decimal("50000.00")
        and not row["eway_bill_no"]
        and row["voucher"].items.filter(stock_item__isnull=False).exists()
    ]
    validations.append(_validation(
        code="eway_bill_review",
        title="E-way bill review",
        severity="warning" if eway_without_number else "ok",
        count=len(eway_without_number),
        description=(
            f"{len(eway_without_number)} high-value goods invoices do not have e-way bill numbers recorded."
            if eway_without_number else "High-value goods invoices have e-way bill numbers recorded where applicable."
        ),
        action_label="Open Integrations",
        action_url=reverse("integrations:dashboard"),
    ))

    return validations


def _validation(*, code, title, severity, count, description, action_label, action_url):
    return {
        "code": code,
        "title": title,
        "severity": severity,
        "count": count,
        "description": description,
        "action_label": action_label,
        "action_url": action_url,
        "task_type": PracticeTask.TYPE_GST,
        "priority": PracticeTask.PRIORITY_CRITICAL if severity == "critical" else PracticeTask.PRIORITY_HIGH,
    }


def _schema_validation(company, period_value, schema_validation):
    critical = schema_validation["critical_count"]
    warning = schema_validation["warning_count"]
    if critical:
        severity = "critical"
        description = f"{critical} portal JSON schema issue(s) must be fixed before export."
    elif warning:
        severity = "warning"
        description = f"{warning} portal JSON schema warning(s) should be reviewed before upload."
    else:
        severity = "ok"
        description = "GSTR-1 portal JSON schema checks are clear."
    return _validation(
        code="gstr1_portal_schema",
        title="GSTR-1 portal JSON schema",
        severity=severity,
        count=critical + warning,
        description=description,
        action_label="Open Filing Pack",
        action_url=f"{reverse('core:gst_filing_pack')}?period={period_value}&company={company.pk}",
    )


def _gstr3b_summary(report):
    return [
        {
            "table": "3.1(a)",
            "description": "Outward taxable supplies",
            "taxable_value": report["tot_taxable_sales"],
            "igst": report["tot_out_igst"],
            "cgst": report["tot_out_cgst"],
            "sgst": report["tot_out_sgst"],
            "total": report["tot_out_tax"],
        },
        {
            "table": "4(A)",
            "description": "Eligible ITC",
            "taxable_value": report["tot_taxable_purchases"],
            "igst": report["itc_igst"],
            "cgst": report["itc_cgst"],
            "sgst": report["itc_sgst"],
            "total": report["tot_itc"],
        },
        {
            "table": "6.1",
            "description": "Net tax payable",
            "taxable_value": ZERO,
            "igst": ZERO,
            "cgst": ZERO,
            "sgst": ZERO,
            "total": report["net_tax_payable"],
        },
    ]


def _summary_snapshot(*, company, period_start, period_end, report, sales_rows, hsn_rows, b2cs_rows):
    return {
        "company_id": company.pk,
        "company": company.name,
        "gstin": company.gstin or "",
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "gstr1": {
            "invoice_count": len(sales_rows),
            "b2b_count": len([row for row in sales_rows if row["section"] == "B2B"]),
            "b2cl_count": len([row for row in sales_rows if row["section"] == "B2CL"]),
            "b2cl_threshold": str(b2cl_limit_for_period(period_end)),
            "b2cs_count": len([row for row in sales_rows if row["section"] == "B2CS"]),
            "b2cs_summary_count": len(b2cs_rows),
            "hsn_count": len(hsn_rows),
            "taxable_sales": str(report["tot_taxable_sales"]),
            "output_tax": str(report["tot_out_tax"]),
        },
        "gstr3b": {
            "taxable_sales": str(report["tot_taxable_sales"]),
            "taxable_purchases": str(report["tot_taxable_purchases"]),
            "output_tax": str(report["tot_out_tax"]),
            "itc": str(report["tot_itc"]),
            "net_tax_payable": str(report["net_tax_payable"]),
        },
        "generated_at": timezone.now().isoformat(),
    }


def _draft_payload(pack):
    company = pack["company"]
    fp = pack["period_end"].strftime("%m%Y")
    return {
        "gstin": (company.gstin or "").upper(),
        "fp": fp,
        "source": "Akshaya Vistara GST Filing Pack",
        "status": pack["status"],
        "gstr1": {
            "b2b": [_json_invoice(row) for row in pack["gstr1"]["b2b"]],
            "b2cl": [_json_invoice(row) for row in pack["gstr1"]["b2cl"]],
            "b2cs": [
                {
                    "sply_ty": row["supply_type"],
                    "pos": row["pos"],
                    "rt": _float(row["rate"]),
                    "txval": _float(row["taxable_value"]),
                    "iamt": _float(row["igst"]),
                    "camt": _float(row["cgst"]),
                    "samt": _float(row["sgst"]),
                    "csamt": 0,
                }
                for row in pack["gstr1"]["b2cs"]
            ],
            "hsn": [
                {
                    "section": row["section"],
                    "hsn_sc": row["hsn_code"],
                    "desc": row["description"],
                    "uqc": row["uqc"],
                    "qty": _float(row["quantity"]),
                    "rt": _float(row["rate"]),
                    "txval": _float(row["taxable_value"]),
                    "iamt": _float(row["igst"]),
                    "camt": _float(row["cgst"]),
                    "samt": _float(row["sgst"]),
                    "tax": _float(row["total_tax"]),
                }
                for row in pack["gstr1"]["hsn"]
            ],
            "doc_issue": pack["gstr1"]["documents"],
        },
        "gstr3b": {
            "summary": [
                {
                    "table": row["table"],
                    "description": row["description"],
                    "txval": _float(row["taxable_value"]),
                    "iamt": _float(row["igst"]),
                    "camt": _float(row["cgst"]),
                    "samt": _float(row["sgst"]),
                    "total": _float(row["total"]),
                }
                for row in pack["gstr3b"]
            ]
        },
        "validations": pack["validations"],
    }


def _json_invoice(row):
    return {
        "inum": row["invoice_number"],
        "idt": row["date"].strftime("%d-%m-%Y"),
        "party": row["party_name"],
        "ctin": row["party_gstin"],
        "section": row["section"],
        "pos": row["pos"],
        "rchrg": row["reverse_charge"],
        "irn": row["irn"],
        "ewb_no": row["eway_bill_no"],
        "val": _float(row["invoice_value"]),
        "itms": [{
            "num": 1,
            "itm_det": {
                "rt": _float(row["rate"]),
                "txval": _float(row["taxable_value"]),
                "iamt": _float(row["igst"]),
                "camt": _float(row["cgst"]),
                "samt": _float(row["sgst"]),
                "csamt": 0,
            },
        }],
    }


def _float(value):
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return round(float(value), 2)
    return value
