"""
reports/views.py
"""

import csv
import hashlib
import io
import json
from datetime import date

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render

from integrations.models import IntegrationConnector, StatutoryExportLog

from . import utils


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _parse_dates(request):
    """Extract start_date / end_date from GET params with current-month defaults."""
    today = date.today()
    # Default to first day of current month
    default_start = today.replace(day=1)
    default_end = today

    try:
        start_date = date.fromisoformat(request.GET.get("start_date", ""))
    except ValueError:
        start_date = default_start

    try:
        end_date = date.fromisoformat(request.GET.get("end_date", ""))
    except ValueError:
        end_date = default_end

    return start_date, end_date


def _parse_as_of(request):
    """Extract as_of_date from GET params, defaulting to today."""
    try:
        return date.fromisoformat(request.GET.get("as_of_date", ""))
    except ValueError:
        return date.today()


def _xl_header(ws, company_name, title, subtitle, styles):
    """Write the standard 3-row company header into a worksheet, returns next row."""
    from openpyxl.styles import Alignment

    hdr_font   = styles["hdr_font"]
    title_font = styles["title_font"]
    sub_font   = styles["sub_font"]
    hdr_fill   = styles["hdr_fill"]

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value, c.font, c.fill = company_name, hdr_font, hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value, c.font = title, title_font
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:F3")
    c = ws["A3"]
    c.value, c.font = subtitle, sub_font
    c.alignment = Alignment(horizontal="center")

    return 5  # data starts at row 5


# ─────────────────────────────────────────────────────────────────────────────
# REPORT VIEWS
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def reports_home(request):
    return render(request, "reports/reports_home.html")


@login_required
def profit_loss(request):
    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_profit_loss(company, start_date, end_date)
    return render(request, "reports/profit_loss.html", {
        "start_date": start_date,
        "end_date":   end_date,
        **data,
    })


@login_required
def balance_sheet(request):
    company = request.current_company
    as_of_date = _parse_as_of(request)
    data = utils.get_balance_sheet(company, as_of_date)
    return render(request, "reports/balance_sheet.html", {
        "as_of_date": as_of_date,
        **data,
    })


@login_required
def receivables_aging(request):
    company = request.current_company
    as_of_date = _parse_as_of(request)
    data = utils.get_receivables_aging(company, as_of_date)
    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="Receivables_Aging_{company.name}_{as_of_date:%Y%m%d}.csv"'
            .replace(" ", "_")
        )
        writer = csv.writer(response)
        writer.writerow([
            "Bucket",
            "Voucher",
            "Customer",
            "Email",
            "Invoice Date",
            "Due Date",
            "Original",
            "Settled",
            "Outstanding",
            "Days Overdue",
            "Priority",
        ])
        for bucket_key, label in [
            ("current", "0-30"),
            ("thirty", "31-60"),
            ("sixty", "61-90"),
            ("ninety", "90+"),
        ]:
            for entry in data["buckets"][bucket_key]:
                writer.writerow([
                    label,
                    entry["voucher"].number,
                    entry["customer_name"],
                    entry["customer_email"],
                    entry["voucher"].date.isoformat(),
                    entry["due_date"].isoformat(),
                    f"{entry['original']:.2f}",
                    f"{entry['settled']:.2f}",
                    f"{entry['outstanding']:.2f}",
                    entry["days_overdue"],
                    entry["priority"],
                ])
        return response
    return render(request, "reports/receivables_aging.html", {
        "as_of_date": as_of_date,
        **data,
    })


@login_required
def trial_balance(request):
    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_trial_balance(company, start_date, end_date)
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, "reports/partials/trial_balance_content.html", {
            "start_date": start_date,
            "end_date":   end_date,
            **data,
        })

    return render(request, "reports/trial_balance.html", {
        "start_date": start_date,
        "end_date":   end_date,
        **data,
    })


@login_required
def gst_report(request):
    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_gst_report(company, start_date, end_date)
    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            f'attachment; filename="GST_Report_{start_date:%Y%m%d}_{end_date:%Y%m%d}.csv"'
        )
        writer = csv.writer(response)
        writer.writerow(["GST Report", company.name, f"{start_date:%Y-%m-%d} to {end_date:%Y-%m-%d}"])
        writer.writerow([])
        writer.writerow(["GSTR-1 Outward Supplies"])
        writer.writerow([
            "Invoice No",
            "Date",
            "Buyer",
            "Buyer GSTIN",
            "POS",
            "Bucket",
            "Supply Type",
            "Rate %",
            "Taxable Value",
            "CGST",
            "SGST",
            "IGST",
            "Total Tax",
            "Invoice Value",
        ])
        for row in data["gstr1_rows"]:
            writer.writerow([
                row["voucher_number"],
                row["date"].isoformat(),
                row["buyer_name"],
                row["buyer_gstin"],
                row["place_of_supply"],
                row["gstr1_bucket"],
                row["portal_supply_type"],
                row["rate"],
                row["taxable_value"],
                row["cgst"],
                row["sgst"],
                row["igst"],
                row["total_tax"],
                row["invoice_value"],
            ])
        writer.writerow([])
        writer.writerow(["HSN/SAC Summary"])
        writer.writerow(["HSN/SAC", "Description", "UQC", "Qty", "Rate %", "Taxable", "CGST", "SGST", "IGST", "Total Value"])
        for row in data["hsn_summary_rows"]:
            writer.writerow([
                row["hsn_code"],
                row["description"],
                row["uqc"],
                row["quantity"],
                row["rate"],
                row["taxable_value"],
                row["cgst"],
                row["sgst"],
                row["igst"],
                row["total_value"],
            ])
        writer.writerow([])
        writer.writerow(["Document Issue Summary"])
        doc = data["doc_issue_summary"]
        writer.writerow(["Document Type", "From", "To", "Total", "Cancelled", "Net Issued"])
        writer.writerow([
            doc["document_type"],
            doc["from_number"],
            doc["to_number"],
            doc["total_number"],
            doc["cancelled"],
            doc["net_issued"],
        ])
        writer.writerow([])
        writer.writerow(["GSTR-3B Summary"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Taxable Sales", data["tot_taxable_sales"]])
        writer.writerow(["Output CGST", data["tot_out_cgst"]])
        writer.writerow(["Output SGST", data["tot_out_sgst"]])
        writer.writerow(["Output IGST", data["tot_out_igst"]])
        writer.writerow(["Total Output Tax", data["tot_out_tax"]])
        writer.writerow(["Taxable Purchases", data["tot_taxable_purchases"]])
        writer.writerow(["ITC CGST", data["itc_cgst"]])
        writer.writerow(["ITC SGST", data["itc_sgst"]])
        writer.writerow(["ITC IGST", data["itc_igst"]])
        writer.writerow(["Total ITC", data["tot_itc"]])
        writer.writerow(["Net Tax Payable", data["net_tax_payable"]])
        return response

    return render(request, "reports/gst_report.html", {
        "start_date": start_date,
        "end_date":   end_date,
        **data,
    })


@login_required
def project_pnl(request, cost_center_id):
    from costcenter.models import CostCenter
    from django.shortcuts import get_object_or_404
    company = request.current_company
    cost_center = get_object_or_404(CostCenter, id=cost_center_id, company=company)
    data = utils.get_project_pnl(company, cost_center)
    return render(request, "reports/project_pnl.html", data)


# ─────────────────────────────────────────────────────────────────────────────
# PDF EXPORTS  (WeasyPrint — server-side, downloadable)
# ─────────────────────────────────────────────────────────────────────────────

def _render_report_pdf(request, template_name: str, context: dict, filename: str) -> HttpResponse:
    """
    Render a Django template to a WeasyPrint PDF and return as HttpResponse.
    context must already contain all template variables including current_company.
    """
    import weasyprint
    from django.template.loader import render_to_string

    html_str = render_to_string(template_name, context, request=request)
    pdf_bytes = weasyprint.HTML(string=html_str, base_url=request.build_absolute_uri("/")).write_pdf()
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def profit_loss_pdf(request):
    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_profit_loss(company, start_date, end_date)
    filename = (
        f"ProfitLoss_{company.name}_{start_date:%Y%m%d}_{end_date:%Y%m%d}.pdf"
        .replace(" ", "_")
    )
    return _render_report_pdf(
        request,
        "reports/profit_loss_pdf.html",
        {"current_company": company, "start_date": start_date, "end_date": end_date, **data},
        filename,
    )


@login_required
def balance_sheet_pdf(request):
    company = request.current_company
    as_of_date = _parse_as_of(request)
    data = utils.get_balance_sheet(company, as_of_date)
    filename = f"BalanceSheet_{company.name}_{as_of_date:%Y%m%d}.pdf".replace(" ", "_")
    return _render_report_pdf(
        request,
        "reports/balance_sheet_pdf.html",
        {"current_company": company, "as_of_date": as_of_date, **data},
        filename,
    )


@login_required
def trial_balance_pdf(request):
    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_trial_balance(company, start_date, end_date)
    filename = (
        f"TrialBalance_{company.name}_{start_date:%Y%m%d}_{end_date:%Y%m%d}.pdf"
        .replace(" ", "_")
    )
    return _render_report_pdf(
        request,
        "reports/trial_balance_pdf.html",
        {"current_company": company, "start_date": start_date, "end_date": end_date, **data},
        filename,
    )


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORTS
# ─────────────────────────────────────────────────────────────────────────────

def _xl_styles():
    """Return a dict of commonly used openpyxl styles."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "hdr_font":   Font(name="Calibri", bold=True, size=14, color="FFFFFF"),
        "title_font": Font(name="Calibri", bold=True, size=12),
        "sub_font":   Font(name="Calibri", italic=True, size=10, color="555555"),
        "col_font":   Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
        "data_font":  Font(name="Calibri", size=10),
        "total_font": Font(name="Calibri", bold=True, size=10),
        "hdr_fill":   PatternFill("solid", fgColor="1F3864"),
        "col_fill":   PatternFill("solid", fgColor="2F5496"),
        "tot_fill":   PatternFill("solid", fgColor="E2EFDA"),
        "alt_fill":   PatternFill("solid", fgColor="F5F5F5"),
        "border":     border,
        "center":     Alignment(horizontal="center", vertical="center"),
        "right":      Alignment(horizontal="right"),
        "left":       Alignment(horizontal="left"),
        "num_fmt":    '#,##0.00',
    }


@login_required
def export_pl_excel(request):
    """Download Profit & Loss as a formatted .xlsx file."""
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font

    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_profit_loss(company, start_date, end_date)
    st = _xl_styles()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    ws.sheet_view.showGridLines = False

    subtitle = f"Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    row = _xl_header(ws, str(company.name), "Profit & Loss Statement", subtitle, st)

    # ── Income section ───────────────────────────────────────────────────────
    ws.cell(row=row, column=1).value = "INCOME"
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11, color="1A7942")
    row += 1

    ws.cell(row=row, column=1).value = "Ledger Account"
    ws.cell(row=row, column=2).value = "Amount (₹)"
    for col in (1, 2):
        c = ws.cell(row=row, column=col)
        c.font = st["col_font"]
        c.fill = PatternFill("solid", fgColor="1A7942")
        c.alignment = st["center"]
        c.border = st["border"]
    row += 1

    for i, item in enumerate(data["income_items"]):
        fill = st["alt_fill"] if i % 2 == 1 else None
        c1 = ws.cell(row=row, column=1, value=item["name"])
        c2 = ws.cell(row=row, column=2, value=float(item["amount"]))
        c1.font = st["data_font"]
        c2.font = st["data_font"]
        c2.number_format = st["num_fmt"]
        c2.alignment = st["right"]
        c1.border = c2.border = st["border"]
        if fill:
            c1.fill = c2.fill = fill
        row += 1

    # Total income row
    c1 = ws.cell(row=row, column=1, value="Total Income")
    c2 = ws.cell(row=row, column=2, value=float(data["total_income"]))
    c1.font = c2.font = st["total_font"]
    c1.fill = c2.fill = PatternFill("solid", fgColor="D5F0DC")
    c2.number_format = st["num_fmt"]
    c2.alignment = st["right"]
    c1.border = c2.border = st["border"]
    row += 2  # blank line

    # ── Expense section ──────────────────────────────────────────────────────
    ws.cell(row=row, column=1).value = "EXPENSES"
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11, color="CC0000")
    row += 1

    ws.cell(row=row, column=1).value = "Ledger Account"
    ws.cell(row=row, column=2).value = "Amount (₹)"
    for col in (1, 2):
        c = ws.cell(row=row, column=col)
        c.font = st["col_font"]
        c.fill = PatternFill("solid", fgColor="CC0000")
        c.alignment = st["center"]
        c.border = st["border"]
    row += 1

    for i, item in enumerate(data["expense_items"]):
        fill = st["alt_fill"] if i % 2 == 1 else None
        c1 = ws.cell(row=row, column=1, value=item["name"])
        c2 = ws.cell(row=row, column=2, value=float(item["amount"]))
        c1.font = st["data_font"]
        c2.font = st["data_font"]
        c2.number_format = st["num_fmt"]
        c2.alignment = st["right"]
        c1.border = c2.border = st["border"]
        if fill:
            c1.fill = c2.fill = fill
        row += 1

    c1 = ws.cell(row=row, column=1, value="Total Expenses")
    c2 = ws.cell(row=row, column=2, value=float(data["total_expense"]))
    c1.font = c2.font = st["total_font"]
    c1.fill = c2.fill = PatternFill("solid", fgColor="FFE0E0")
    c2.number_format = st["num_fmt"]
    c2.alignment = st["right"]
    c1.border = c2.border = st["border"]
    row += 2

    # ── Net Profit ───────────────────────────────────────────────────────────
    net = data["net_profit"]
    c1 = ws.cell(row=row, column=1, value="NET PROFIT" if net >= 0 else "NET LOSS")
    c2 = ws.cell(row=row, column=2, value=float(net))
    profit_color = "1F6B37" if net >= 0 else "CC0000"
    c1.font = c2.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    c1.fill = c2.fill = PatternFill("solid", fgColor=profit_color)
    c2.number_format = st["num_fmt"]
    c2.alignment = st["right"]
    c1.border = c2.border = st["border"]

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"PL_{start_date}_{end_date}.xlsx"
    resp = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


@login_required
def export_bs_excel(request):
    """Download Balance Sheet as a formatted .xlsx file."""
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font

    company = request.current_company
    as_of_date = _parse_as_of(request)
    data = utils.get_balance_sheet(company, as_of_date)
    st = _xl_styles()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.sheet_view.showGridLines = False

    subtitle = f"As of {as_of_date.strftime('%d %b %Y')}"
    row = _xl_header(ws, str(company.name), "Balance Sheet", subtitle, st)

    def _write_section(label, items, total_label, total_val, header_color, total_color):
        nonlocal row
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11,
                                               color=header_color)
        row += 1
        ws.cell(row=row, column=1).value = "Ledger Account"
        ws.cell(row=row, column=2).value = "Balance (₹)"
        for col in (1, 2):
            c = ws.cell(row=row, column=col)
            c.font = st["col_font"]
            c.fill = PatternFill("solid", fgColor=header_color)
            c.alignment = st["center"]
            c.border = st["border"]
        row += 1

        for i, item in enumerate(items):
            fill = st["alt_fill"] if i % 2 == 1 else None
            c1 = ws.cell(row=row, column=1, value=item["name"])
            c2 = ws.cell(row=row, column=2, value=float(item["balance"]))
            c1.font = c2.font = st["data_font"]
            c2.number_format = st["num_fmt"]
            c2.alignment = st["right"]
            c1.border = c2.border = st["border"]
            if fill:
                c1.fill = c2.fill = fill
            row += 1

        c1 = ws.cell(row=row, column=1, value=total_label)
        c2 = ws.cell(row=row, column=2, value=float(total_val))
        c1.font = c2.font = st["total_font"]
        c1.fill = c2.fill = PatternFill("solid", fgColor=total_color)
        c2.number_format = st["num_fmt"]
        c2.alignment = st["right"]
        c1.border = c2.border = st["border"]
        row += 2

    _write_section("ASSETS", data["asset_items"], "Total Assets",
                   data["total_assets"], "1F497D", "DAEEF3")
    _write_section("LIABILITIES & EQUITY", data["liability_items"],
                   "Total Liabilities & Equity", data["total_liabilities"],
                   "7030A0", "EAD1DC")

    # Balance check row
    balanced = data["difference"] == 0
    msg = "✔ Books are BALANCED" if balanced else f"⚠ Out of balance by ₹{data['difference']:.2f}"
    c = ws.cell(row=row, column=1, value=msg)
    c.font = Font(name="Calibri", bold=True, size=10,
                  color="1F6B37" if balanced else "CC0000")
    ws.merge_cells(f"A{row}:B{row}")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"BalanceSheet_{as_of_date}.xlsx"
    resp = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


@login_required
def export_tb_excel(request):
    """Download Trial Balance as a formatted .xlsx file."""
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font

    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_trial_balance(company, start_date, end_date)
    st = _xl_styles()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    ws.sheet_view.showGridLines = False

    subtitle = f"Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    row = _xl_header(ws, str(company.name), "Trial Balance", subtitle, st)

    # Column headers
    headers = ["Ledger Account", "Group",
               "Opening Dr (₹)", "Opening Cr (₹)",
               "Period Dr (₹)", "Period Cr (₹)",
               "Closing Dr (₹)", "Closing Cr (₹)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = st["col_font"]
        c.fill = st["col_fill"]
        c.alignment = st["center"]
        c.border = st["border"]
    row += 1

    for i, r in enumerate(data["rows"]):
        fill = st["alt_fill"] if i % 2 == 1 else None
        vals = [r["name"], r["group"],
                float(r["opening_dr"]), float(r["opening_cr"]),
                float(r["period_dr"]),  float(r["period_cr"]),
                float(r["closing_dr"]), float(r["closing_cr"])]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = st["data_font"]
            c.border = st["border"]
            if col > 2:
                c.number_format = st["num_fmt"]
                c.alignment = st["right"]
            if fill:
                c.fill = fill
        row += 1

    # Totals row
    totals = ["TOTALS", "",
              float(data["tot_open_dr"]), float(data["tot_open_cr"]),
              float(data["tot_per_dr"]),  float(data["tot_per_cr"]),
              float(data["tot_clos_dr"]), float(data["tot_clos_cr"])]
    for col, v in enumerate(totals, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = st["total_font"]
        c.fill = PatternFill("solid", fgColor="D9E1F2")
        c.border = st["border"]
        if col > 2:
            c.number_format = st["num_fmt"]
            c.alignment = st["right"]

    col_widths = [42, 12, 16, 16, 16, 16, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"TrialBalance_{start_date}_{end_date}.xlsx"
    resp = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


# ─────────────────────────────────────────────────────────────────────────────
# GST JSON EXPORTS  (GSTR-1 and GSTR-3B — machine-readable JSON download)
# ─────────────────────────────────────────────────────────────────────────────

def _float(val):
    """Safely convert Decimal / None → rounded float for JSON output."""
    if val is None:
        return 0.0
    return round(float(val), 2)


def _statutory_json_response(request, payload, filename, export_type, start_date, end_date, row_count, amount_total, validation_summary):
    body = json.dumps(payload, indent=2, ensure_ascii=False)
    body_bytes = body.encode("utf-8")
    connector = IntegrationConnector.objects.filter(
        company=request.current_company,
        connector_type=IntegrationConnector.TYPE_GST,
    ).first()
    StatutoryExportLog.objects.create(
        company=request.current_company,
        connector=connector,
        generated_by=request.user if getattr(request.user, "is_authenticated", False) else None,
        export_type=export_type,
        period_start=start_date,
        period_end=end_date,
        file_name=filename,
        file_sha256=hashlib.sha256(body_bytes).hexdigest(),
        row_count=row_count,
        amount_total=amount_total or 0,
        validation_summary=validation_summary,
    )
    response = HttpResponse(body, content_type="application/json")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def gstr1_export(request):
    """
    Download GSTR-1 JSON file (outward supplies) for a chosen period.

    Format follows the GST portal's JSON schema (simplified):
      { gstin, fp, b2b: [...], b2cs: [...], nil: {...}, doc_issue: {...} }

    GET params: start_date, end_date  (default = current FY)
    """
    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_gst_report(company, start_date, end_date)

    # Filing period — last month of the range in MMYYYY
    fp = end_date.strftime("%m%Y")
    own_gstin = (company.gstin or "").upper()

    # ── B2B section ─────────────────────────────────────────────────────────
    b2b_map = {}  # buyer_gstin → list of invoices
    for row in data["b2b_rows"]:
        buyer = row["buyer_gstin"]
        if buyer not in b2b_map:
            b2b_map[buyer] = []
        b2b_map[buyer].append({
            "inum":  row["voucher_number"],
            "idt":   row["date"].strftime("%d-%m-%Y"),
            "val":   _float(row["invoice_value"]),
            "pos":   row.get("place_of_supply") or "99",
            "rchrg": "N",
            "itms": [{
                "num": 1,
                "itm_det": {
                    "txval": _float(row["taxable_value"]),
                    "rt":    _float(row.get("rate", 0)),
                    "camt":  _float(row["cgst"]),
                    "samt":  _float(row["sgst"]),
                    "iamt":  _float(row["igst"]),
                    "csamt": 0,
                },
            }],
        })

    b2b_section = [
        {"ctin": gstin, "inv": invoices}
        for gstin, invoices in b2b_map.items()
    ]

    # B2CS section: intrastate B2C and interstate B2C up to the period threshold.
    b2cl_map = {}
    for row in data["b2cl_rows"]:
        pos = row.get("place_of_supply") or "99"
        if pos not in b2cl_map:
            b2cl_map[pos] = []
        b2cl_map[pos].append({
            "inum": row["voucher_number"],
            "idt":  row["date"].strftime("%d-%m-%Y"),
            "val":  _float(row["invoice_value"]),
            "itms": [{
                "num": 1,
                "itm_det": {
                    "txval": _float(row["taxable_value"]),
                    "rt":    _float(row.get("rate", 0)),
                    "iamt":  _float(row["igst"]),
                    "csamt": 0,
                },
            }],
        })
    b2cl_section = [
        {"pos": pos, "inv": invoices}
        for pos, invoices in b2cl_map.items()
    ]

    b2cs_map = {}
    for row in data["b2cs_rows"]:
        key = (
            row.get("portal_supply_type") or "INTRA",
            row.get("place_of_supply") or "99",
            row.get("rate", 0),
        )
        if key not in b2cs_map:
            b2cs_map[key] = {
                "sply_ty": key[0],
                "pos":     key[1],
                "typ":     "OE",
                "txval":   0.0,
                "rt":      _float(key[2]),
                "camt":    0.0,
                "samt":    0.0,
                "iamt":    0.0,
                "csamt":   0.0,
            }
        b2cs_map[key]["txval"] += _float(row["taxable_value"])
        b2cs_map[key]["camt"] += _float(row["cgst"])
        b2cs_map[key]["samt"] += _float(row["sgst"])
        b2cs_map[key]["iamt"] += _float(row["igst"])
        b2cs_map[key]["csamt"] += 0.0
    b2cs_section = list(b2cs_map.values())

    hsn_data = []
    for num, row in enumerate((r for r in data.get("hsn_summary_rows", []) if r.get("hsn_code")), 1):
        hsn_data.append({
            "num":    num,
            "hsn_sc": row["hsn_code"],
            "desc":   row["description"],
            "uqc":    row["uqc"],
            "qty":    _float(row["quantity"]),
            "val":    _float(row["total_value"]),
            "txval":  _float(row["taxable_value"]),
            "rt":     _float(row["rate"]),
            "iamt":   _float(row["igst"]),
            "camt":   _float(row["cgst"]),
            "samt":   _float(row["sgst"]),
            "csamt":  _float(row["other_gst"]),
        })

    payload = {
        "gstin":  own_gstin,
        "fp":     fp,
        "b2b":    b2b_section,
        "b2cl":   b2cl_section,
        "b2cs":   b2cs_section,
        "hsn":    {"data": hsn_data},
        "doc_issue": {
            "doc_det": [{
                "doc_num": 1,
                "doc_typ": data["doc_issue_summary"]["document_type"],
                "docs": [{
                    "num": 1,
                    "from": data["doc_issue_summary"]["from_number"],
                    "to": data["doc_issue_summary"]["to_number"],
                    "totnum": data["doc_issue_summary"]["total_number"],
                    "cancel": data["doc_issue_summary"]["cancelled"],
                    "net_issue": data["doc_issue_summary"]["net_issued"],
                }],
            }],
        },
        "_meta": {
            "period_from": start_date.strftime("%d-%m-%Y"),
            "period_to":   end_date.strftime("%d-%m-%Y"),
            "generated_by": "Akshaya Vistara",
            "total_invoices": len(data["gstr1_rows"]),
            "total_taxable_value": _float(data["tot_taxable_sales"]),
            "total_output_tax":    _float(data["tot_out_tax"]),
            "missing_hsn_rows":    len(data.get("missing_hsn_rows", [])),
            "b2cl_threshold":      _float(data["b2cl_threshold"]),
        },
    }

    fname = (
        f"GSTR1_{company.name}_{start_date.strftime('%Y%m')}_{end_date.strftime('%Y%m')}.json"
        .replace(" ", "_")
    )
    return _statutory_json_response(
        request,
        payload,
        fname,
        StatutoryExportLog.TYPE_GSTR1_JSON,
        start_date,
        end_date,
        len(data["gstr1_rows"]),
        data["tot_out_tax"],
        {
            "missing_hsn_rows": len(data.get("missing_hsn_rows", [])),
            "b2b_rows": len(data.get("b2b_rows", [])),
            "b2cl_rows": len(data.get("b2cl_rows", [])),
            "b2cs_rows": len(data.get("b2cs_rows", [])),
            "hsn_rows": len(hsn_data),
        },
    )


@login_required
def gstr3b_export(request):
    """
    Download GSTR-3B JSON file (summary return) for a chosen period.

    Format follows the GST portal's JSON schema (simplified):
      { gstin, ret_period, sup_details, itc_elg, inward_sup, intr_ltfee }

    GET params: start_date, end_date  (default = current FY)
    """
    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_gst_report(company, start_date, end_date)

    fp = end_date.strftime("%m%Y")
    own_gstin = (company.gstin or "").upper()

    payload = {
        "gstin":      own_gstin,
        "ret_period": fp,
        "sup_details": {
            "osup_det": {
                "txval": _float(data["tot_taxable_sales"]),
                "iamt":  _float(data["tot_out_igst"]),
                "camt":  _float(data["tot_out_cgst"]),
                "samt":  _float(data["tot_out_sgst"]),
                "csamt": 0,
            },
            "osup_zero": {"txval": 0, "iamt": 0, "csamt": 0},
            "osup_nil_exmp": {"txval": 0},
            "isup_rev": {"txval": 0, "iamt": 0, "camt": 0, "samt": 0, "csamt": 0},
        },
        "itc_elg": {
            "itc_avl": [
                {
                    "ty":   "IMPG",
                    "iamt": _float(data["itc_igst"]),
                    "camt": _float(data["itc_cgst"]),
                    "samt": _float(data["itc_sgst"]),
                    "csamt": 0,
                },
                {
                    "ty":    "ISRC",
                    "iamt":  0, "camt": 0, "samt": 0, "csamt": 0,
                },
            ],
            "itc_rev": [
                {"ty": "RUL", "iamt": 0, "camt": 0, "samt": 0, "csamt": 0},
            ],
            "itc_net": {
                "iamt": _float(data["itc_igst"]),
                "camt": _float(data["itc_cgst"]),
                "samt": _float(data["itc_sgst"]),
                "csamt": 0,
            },
        },
        "inward_sup": {
            "isup_details": [
                {
                    "ty":   "GST",
                    "inter": 0,
                    "intra": _float(data["tot_taxable_purchases"]),
                },
            ],
        },
        "intr_ltfee": {
            "intr_details": {"camt": 0, "samt": 0, "iamt": 0},
        },
        "_meta": {
            "period_from":         start_date.strftime("%d-%m-%Y"),
            "period_to":           end_date.strftime("%d-%m-%Y"),
            "generated_by":        "Akshaya Vistara",
            "total_output_tax":    _float(data["tot_out_tax"]),
            "total_itc_claimed":   _float(data["tot_itc"]),
            "net_tax_payable":     _float(data["net_tax_payable"]),
        },
    }

    fname = (
        f"GSTR3B_{company.name}_{start_date.strftime('%Y%m')}_{end_date.strftime('%Y%m')}.json"
        .replace(" ", "_")
    )
    return _statutory_json_response(
        request,
        payload,
        fname,
        StatutoryExportLog.TYPE_GSTR3B_JSON,
        start_date,
        end_date,
        1,
        data["net_tax_payable"],
        {
            "total_output_tax": _float(data["tot_out_tax"]),
            "total_itc_claimed": _float(data["tot_itc"]),
            "net_tax_payable": _float(data["net_tax_payable"]),
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# CASH FLOW STATEMENT  (Indirect Method)
# ─────────────────────────────────────────────────────────────────────────────

# Ledger group → cash flow category mapping
_CF_INVESTING = {"Fixed Assets", "Investments", "Capital WIP"}
_CF_FINANCING = {
    "Loans (Liability)", "Bank OD Accounts", "Capital Account",
    "Reserves & Surplus", "Secured Loans", "Unsecured Loans",
}
_CASH_GROUPS  = {"Cash-in-Hand", "Bank Accounts"}


def _classify_cf(group: str) -> str:
    if group in _CF_INVESTING: return "investing"
    if group in _CF_FINANCING: return "financing"
    return "operating"


@login_required
def cash_flow(request):
    from vouchers.models import VoucherItem
    from django.db.models import Sum, Q

    company              = request.current_company
    start_date, end_date = _parse_dates(request)

    items = (
        VoucherItem.objects
        .filter(voucher__company=company, voucher__date__range=(start_date, end_date), voucher__status='APPROVED')
        .values("ledger__name", "ledger__account_group__name")
        .annotate(
            total_dr=Sum("amount", filter=Q(entry_type='DR')),
            total_cr=Sum("amount", filter=Q(entry_type='CR'))
        )
        .order_by("ledger__account_group__name", "ledger__name")
    )

    operating_rows = []
    investing_rows = []
    financing_rows = []
    net_cash_change = 0

    for row in items:
        group = row["ledger__account_group__name"] or "Operating"
        name  = row["ledger__name"]
        dr    = row["total_dr"] or 0
        cr    = row["total_cr"] or 0
        net   = float(dr - cr) # Net Debit for cash flow logic (outflow - inflow usually, but here we'll just use it as change)
        
        if group in _CASH_GROUPS:
            net_cash_change -= net # Cash increase is Dr, so -(-net) = +net
            continue
            
        entry = {"name": name, "group": group, "net": net}
        cat = _classify_cf(group)
        if cat == "investing":   investing_rows.append(entry)
        elif cat == "financing": financing_rows.append(entry)
        else:                    operating_rows.append(entry)

    operating_total  = sum(r["net"] for r in operating_rows)
    investing_total  = sum(r["net"] for r in investing_rows)
    financing_total  = sum(r["net"] for r in financing_rows)

    return render(request, "reports/cash_flow.html", {
        "start_date":      start_date,
        "end_date":        end_date,
        "operating_rows":  operating_rows,
        "investing_rows":  investing_rows,
        "financing_rows":  financing_rows,
        "operating_total": operating_total,
        "investing_total": investing_total,
        "financing_total": financing_total,
        "net_cash_change": net_cash_change,
    })


# ─────────────────────────────────────────────────────────────────────────────
# SIMPLIFIED REPORTS (Phase 5 Cleanup)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def dashboard_financials(request):
    """Simplified Financial Dashboard view."""
    company = request.current_company
    data = utils.get_dashboard_summary(company)
    
    # Financial metrics for dashboard display
    revenue = data['total_revenue']
    expense = data['total_expense']
    profit  = data['net_profit']
    net_margin = (profit / revenue * 100) if revenue > 0 else 0
    expense_ratio = (expense / revenue * 100) if revenue > 0 else 0

    return render(request, "reports/dashboard_financials.html", {
        "revenue": revenue,
        "expense": expense,
        "profit":  profit,
        "liquidity": data['total_assets_liquidity'],
        "net_margin": net_margin,
        "expense_ratio": expense_ratio
    })


@login_required
def profit_loss_simple(request):
    """Simplified Profit & Loss view."""
    company = request.current_company
    data = utils.get_profit_loss_simple(company)
    return render(request, "reports/profit_loss_simple.html", data)


@login_required
def balance_sheet_simple(request):
    """Simplified Balance Sheet view."""
    company = request.current_company
    data = utils.get_balance_sheet_simple(company)
    
    # Initialize total_le
    data['total_le'] = data['total_liabilities'] + data['total_equity']

    # Integrate Net Profit/Loss into Equity for balancing
    pl_data = utils.get_profit_loss_simple(company)
    net_profit = pl_data['net_profit_loss']
    if net_profit != 0:
        data['equity'].append({'ledger_name': 'Retained Earnings (P&L)', 'balance': net_profit})
        data['total_equity'] += net_profit
        data['total_le'] = data['total_liabilities'] + data['total_equity']

    difference = abs(data['total_assets'] - data['total_le'])
    return render(request, "reports/balance_sheet_simple.html", {**data, "difference": difference})


@login_required
def trial_balance_simple(request):
    """Simplified Trial Balance view."""
    company = request.current_company
    report_data = utils.get_trial_balance_simple(company)
    
    totals = {
        'opening': sum(item['opening_balance'] for item in report_data),
        'debit':   sum(item['total_debit'] for item in report_data),
        'credit':  sum(item['total_credit'] for item in report_data),
    }
    totals['difference'] = abs(totals['debit'] - totals['credit'])
    
    return render(request, "reports/trial_balance_simple.html", {
        "report_data": report_data,
        "totals": totals,
    })


@login_required
def report_group_detail(request, group_name):
    """Drill-down: Group → Ledgers."""
    company = request.current_company
    data = utils.get_group_detail(company, group_name)
    return render(request, "reports/group_detail.html", {
        "group_name": group_name,
        **data
    })


@login_required
def report_ledger_detail(request, ledger_id):
    """Drill-down: Ledger → Vouchers (Statement)."""
    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_ledger_history(company, ledger_id, start_date, end_date)
    
    ctx = {
        "start_date": start_date,
        "end_date":   end_date,
        **data
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, "reports/partials/ledger_detail_content.html", ctx)

    return render(request, "reports/ledger_detail.html", ctx)



@login_required
def msme_overdue_report(request):
    """
    Tracks outstanding payments to MSME vendors and assesses risk (45-day rule).
    """
    from .utils import get_msme_payable_watch

    company = request.current_company
    as_of_date = _parse_as_of(request)
    watch = get_msme_payable_watch(company, as_of_date=as_of_date)

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="MSME_Risk_{company.name}_{as_of_date:%Y%m%d}.csv"'
            .replace(" ", "_")
        )
        writer = csv.writer(response)
        writer.writerow([
            "Voucher",
            "Vendor",
            "MSME Registration",
            "Invoice Date",
            "Due Date",
            "Status",
            "Days Outstanding",
            "Days Overdue",
            "Outstanding",
            "Interest Exposure",
        ])
        for row in watch["rows"]:
            voucher = row["voucher"]
            vendor = row["vendor"]
            writer.writerow([
                voucher.number or voucher.pk,
                vendor.name if vendor else "",
                vendor.msme_reg_number if vendor else "",
                voucher.date.isoformat(),
                row["due_date"].isoformat(),
                row["status"],
                row["days_outstanding"],
                row["days_overdue"],
                f"{row['outstanding_amount']:.2f}",
                f"{row['interest_liability']:.2f}",
            ])
        return response

    return render(request, "reports/msme_overdue.html", {
        "watch": watch,
        "rows": watch["rows"],
        "as_of_date": as_of_date,
        "current_company": company,
    })


# -----------------------------------------------------------------------------
# DAY BOOK
# -----------------------------------------------------------------------------

@login_required
def day_book_view(request):
    """
    Chronological register of all transactions.
    Supports date range filtering. AJAX for instant updates.
    """
    company = request.current_company
    start_date, end_date = _parse_dates(request)

    day_book_rows = utils.get_day_book(company, start_date, end_date)

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            f'attachment; filename="Day_Book_{company.name}_{start_date:%Y%m%d}_{end_date:%Y%m%d}.csv"'
            .replace(" ", "_")
        )
        writer = csv.writer(response)
        writer.writerow(["Date", "Voucher No", "Type", "Particulars", "Amount", "Running Balance"])
        for row in day_book_rows:
            writer.writerow([
                row["date"].isoformat(),
                row["number"],
                row["type"],
                row["party"],
                f"{row['amount']:.2f}",
                f"{row['balance']:.2f}",
            ])
        return response

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, "reports/partials/day_book_table.html", {
            "day_book_rows": day_book_rows,
        })

    return render(request, "reports/day_book.html", {
        "day_book_rows": day_book_rows,
        "start_date":    start_date,
        "end_date":      end_date,
        "export_query":  f"start_date={start_date:%Y-%m-%d}&end_date={end_date:%Y-%m-%d}&export=csv",
        "title":         "Day Book",
    })


@login_required
def cash_flow_forecast(request):
    """
    30-Day Cash Flow Forecast.
    Aggregates receivables (Sales, Sales Return) and payables (Purchase, Purchase Return) by due_date.
    Starts with current Cash & Bank balances.
    """
    from vouchers.models import Voucher
    from ledger.models import Ledger
    from datetime import date, timedelta
    from django.db.models import Sum
    from collections import defaultdict
    from decimal import Decimal

    company = request.current_company
    today = date.today()
    end_forecast = today + timedelta(days=30)

    # 1. Starting Cash & Bank Balance
    cash_bank_ledgers = Ledger.objects.filter(
        company=company,
        account_group__name__in=['Cash-in-Hand', 'Bank Accounts']
    )
    # Using float to keep it consistent with the table logic (or Decimal if preferred)
    opening_cash = sum(float(Decimal("0.00") - l.current_balance()) for l in cash_bank_ledgers)

    # 2. Fetch outstanding receivables (Sales - Sales Return)
    # Sales increase receivable, Sales Return decreases it.
    receivables_qs = Voucher.objects.filter(
        company=company,
        voucher_type__in=['Sales', 'Sales Return'],
        due_date__range=(today, end_forecast),
        outstanding_amount__gt=0
    ).values('due_date', 'voucher_type').annotate(total=Sum('outstanding_amount'))

    # 3. Fetch outstanding payables (Purchase - Purchase Return)
    # Purchase increase payable, Purchase Return decreases it.
    payables_qs = Voucher.objects.filter(
        company=company,
        voucher_type__in=['Purchase', 'Purchase Return'],
        due_date__range=(today, end_forecast),
        outstanding_amount__gt=0
    ).values('due_date', 'voucher_type').annotate(total=Sum('outstanding_amount'))

    # 4. Group by date
    daily_data = defaultdict(lambda: {'incoming': 0.0, 'outgoing': 0.0})
    
    for r in receivables_qs:
        val = float(r['total'])
        if r['voucher_type'] == 'Sales':
            daily_data[r['due_date']]['incoming'] += val
        else: # Sales Return
            daily_data[r['due_date']]['incoming'] -= val
            
    for p in payables_qs:
        val = float(p['total'])
        if p['voucher_type'] == 'Purchase':
            daily_data[p['due_date']]['outgoing'] += val
        else: # Purchase Return
            daily_data[p['due_date']]['outgoing'] -= val

    # 5. Build 30-day forecast table
    forecast_table = []
    cumulative = opening_cash
    
    # We include 'Today' in the loop
    for i in range(31):
        day = today + timedelta(days=i)
        day_info = daily_data.get(day, {'incoming': 0.0, 'outgoing': 0.0})
        net = day_info['incoming'] - day_info['outgoing']
        cumulative += net
        forecast_table.append({
            'date': day,
            'incoming': day_info['incoming'],
            'outgoing': day_info['outgoing'],
            'net': net,
            'cumulative': cumulative
        })

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="cash_flow_forecast_30_days.csv"'
        writer = csv.writer(response)
        writer.writerow(["Date", "Incoming Receivables", "Outgoing Payables", "Daily Net", "Cumulative Position"])
        for row in forecast_table:
            writer.writerow([
                row["date"].isoformat(),
                f"{row['incoming']:.2f}",
                f"{row['outgoing']:.2f}",
                f"{row['net']:.2f}",
                f"{row['cumulative']:.2f}",
            ])
        return response

    return render(request, "reports/cash_flow_forecast.html", {
        "forecast_table": forecast_table,
        "opening_cash": opening_cash,
        "today": today,
        "end_forecast": end_forecast,
    })
